#!/usr/bin/env python3
"""One-off correction (2026-07-10): the History sheet in Stocks_Buy_Strategy.xlsx had
wrong RELX data — a 811-share tranche's sell leg was miscopied from an unrelated trade,
a second tranche's quantity was inflated by merging two separate buy lots into one row,
and the still-open 811-share position bought 2026-05-13 was missing entirely. Corrected
against the authoritative source export (Downloads/claude_files/files.zip -> History.xlsx):

  Source History.xlsx RELX rows:
    1) Buy 2026-04-11, Sell 2026-07-06, qty 811, buy 24.50, sell 24.67 -> Closed
    2) Buy 2026-05-13, Sell 2026-07-06, qty  49, buy 23.11, sell 24.67 -> Closed
    3) Buy 2026-05-13, no sell,          qty 811, buy 23.11            -> Open (MISSING)

The new Open row is appended at the very end of the trade log (row 56, right before
the blank/Totals section) rather than inserted inside the chronological "May 2026"
block. A first attempt inserted it in-section instead, but openpyxl's insert_rows()
does not shift merged-cell ranges or per-row self-referencing formulas (only raw
cell values), which silently corrupted several unrelated rows elsewhere in the
sheet — restored from backup and redone this way, which only shifts the blank
spacer + summary section below all trade data, where there are no merges and no
per-row self-references to break.

Not part of the periodic pipeline — run once, by hand, against the live Downloads file.
"""
import openpyxl
from datetime import datetime
from copy import copy

MASTER_PATH = r'C:\Users\Paul\Downloads\Stocks_Buy_Strategy.xlsx'
SHEET = 'History'
NEW_ROW = 56          # right after the last trade (row 55), before the blank/Totals section
OLD_LAST_DATA_ROW = 55
NEW_LAST_DATA_ROW = 56


def bump_range(formula):
    for suffix in [')', ',', '"', '*', '=']:
        formula = formula.replace(f'{OLD_LAST_DATA_ROW}{suffix}', f'{NEW_LAST_DATA_ROW}{suffix}')
    return formula


def main():
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    ws = wb[SHEET]

    # --- 1. Fix row 12: two separate buy lots (811 @ 24.50 on 4/11, and 49 @ 23.11
    #    on 5/13) were wrongly merged into one 860-share row dated 5/13. Correct it
    #    down to just the 49-share tranche; the 811-share tranche is fixed at row 16.
    assert ws.cell(row=12, column=1).value == 'RELX PLC, ORD GBP0.1444 (REL)'
    ws.cell(row=12, column=6, value=49)          # Quantity: 860 -> 49
    ws.cell(row=12, column=10, value=1132.39)     # Cost (£): source value for this tranche
    ws.cell(row=12, column=12, value=1208.83)     # Proceeds (£): 49 * 24.67, matches source

    # --- 2. Fix row 16: sell leg was miscopied from an unrelated trade.
    assert ws.cell(row=16, column=1).value == 'RELX PLC, ORD GBP0.1444 (REL)'
    ws.cell(row=16, column=5, value=datetime(2026, 7, 6))   # Sell Date: 4/29 -> 7/6
    ws.cell(row=16, column=8, value=24.67)                   # Sell Price: 26.28 -> 24.67
    ws.cell(row=16, column=12, value=20007.37)               # Proceeds: 811 * 24.67, matches source

    # --- 3. Append the missing OPEN 811-share tranche at row 56, right before the
    #    blank/Totals section — only rows 56+ shift down by 1, and there are no
    #    merges or per-row self-references down there to break.
    ws.insert_rows(NEW_ROW, amount=1)

    for col in range(1, 17):
        src = ws.cell(row=11, column=col)  # copy style from AUTO, another Open equity row
        dst = ws.cell(row=NEW_ROW, column=col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)

    ws.cell(row=NEW_ROW, column=1, value='RELX PLC, ORD GBP0.1444 (REL)')
    ws.cell(row=NEW_ROW, column=2, value='2000001606')
    ws.cell(row=NEW_ROW, column=3, value='SIPP - Pension Savings Account')
    ws.cell(row=NEW_ROW, column=4, value=datetime(2026, 5, 13))
    ws.cell(row=NEW_ROW, column=5, value=None)
    ws.cell(row=NEW_ROW, column=6, value=811)
    ws.cell(row=NEW_ROW, column=7, value=23.11)
    ws.cell(row=NEW_ROW, column=8, value=None)
    ws.cell(row=NEW_ROW, column=9, value=9)          # Fee: matches other ~£19k Open equity rows (AUTO/SAGE, both 9)
    ws.cell(row=NEW_ROW, column=10, value=18742.21)   # Cost (£): matches source exactly (811*23.11)
    ws.cell(row=NEW_ROW, column=11, value=(
        f'=IF(IFERROR(MID($A{NEW_ROW},FIND("(",$A{NEW_ROW})+1,FIND(")",$A{NEW_ROW})-FIND("(",$A{NEW_ROW})-1),"")="",'
        f'"N/A (fund)",IFERROR(ROUND(($F{NEW_ROW}*(VLOOKUP(IFERROR(MID($A{NEW_ROW},FIND("(",$A{NEW_ROW})+1,'
        f'FIND(")",$A{NEW_ROW})-FIND("(",$A{NEW_ROW})-1),""),\'Stocks Buy Strategy\'!$C:$I,7,FALSE())/100))'
        f'-IF(($F{NEW_ROW}*(VLOOKUP(IFERROR(MID($A{NEW_ROW},FIND("(",$A{NEW_ROW})+1,FIND(")",$A{NEW_ROW})-FIND("(",$A{NEW_ROW})-1),""),'
        f'\'Stocks Buy Strategy\'!$C:$I,7,FALSE())/100))>10000,9,7.5),2),"Price pending (open in Sheets)"))'
    ))
    ws.cell(row=NEW_ROW, column=12, value=None)
    ws.cell(row=NEW_ROW, column=13, value=None)
    ws.cell(row=NEW_ROW, column=14, value=None)
    ws.cell(row=NEW_ROW, column=15, value=f'=TODAY()-D{NEW_ROW}')
    ws.cell(row=NEW_ROW, column=16, value='Open')

    # --- 4. The blank spacer + summary section (rows 56-78 originally) shifted down
    #    by 1 to 57-79; their fixed-range formulas ("...2:...55") need to cover one
    #    more row ("...2:...56") at their new positions.
    for row, col in [(58, 13), (59, 9)]:  # Total Realised P&L, Total Dealing/PTM Fees
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and str(OLD_LAST_DATA_ROW) in cell.value:
            cell.value = bump_range(cell.value)

    for row in range(63, 73):  # Capital Deployed by Month
        cell = ws.cell(row=row, column=2)
        if isinstance(cell.value, str) and str(OLD_LAST_DATA_ROW) in cell.value:
            cell.value = bump_range(cell.value)

    for row in range(76, 80):  # Realised P&L by Month
        cell = ws.cell(row=row, column=2)
        if isinstance(cell.value, str) and str(OLD_LAST_DATA_ROW) in cell.value:
            cell.value = bump_range(cell.value)

    wb.save(MASTER_PATH)
    print('Saved.')


if __name__ == '__main__':
    main()
