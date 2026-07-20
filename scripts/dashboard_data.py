"""Investment Dashboard — data layer (Phase 1).

Reads the master workbook (Stocks_Buy_Strategy.xlsx) + history.db and emits the
JSON the dashboard front end renders: overview.json, portfolio.json, historic.json.

KEY CONSTRAINT (why this file computes rather than reads): the workbook is written
by openpyxl (the pipeline), so every FORMULA cell — all the %/P&L columns, and every
stock's googlefinance() current price — has NO cached value and reads as None locally.
Only literals survive: identity, Holdings (£), Alert Low/High/Source, and the History
transaction inputs (Qty, Buy/Sell price, Fees, Cost, Proceeds). So we take those
literals, pull the freshest CURRENT PRICE per ticker from history.db (the captured
TradingView last price), and derive every percentage ourselves. Never trust a formula
cell's value here.
"""
import os
import sys
import json
import sqlite3
import datetime
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:      # importable from any cwd, incl. as a module
    sys.path.insert(0, SCRIPT_DIR)
import ticker_normalize             # noqa: E402  (needs the path above)

REPO = os.path.dirname(SCRIPT_DIR)
WORKBOOK = os.path.join(os.path.expanduser('~'), 'Downloads', 'Stocks_Buy_Strategy.xlsx')
HISTORY_DB = os.path.join(REPO, 'data', 'history.db')
OUT_DIR = os.path.join(SCRIPT_DIR, 'dashboard_app', 'data')

# --- Investments sheet columns (1-indexed; header row 2, data from row 4) ---
I_NAME, I_TICKER, I_HOLDINGS, I_CURPRICE = 3, 4, 5, 10
I_ALERT_LOW, I_ALERT_LOW_SRC, I_ALERT_HIGH = 13, 14, 16
I_TV = 37
I_TYPE = 39   # 'Type' — Short Term / Strategic (renamed from 'Long Term' 2026-07-19)
# --- Income Funds columns (header row 4, data from row 5) ---
F_NAME, F_CURVAL, F_DIVYLD, F_MONTHLY_REV, F_ANNUAL_REV = 1, 3, 6, 9, 10
# --- History columns (header row 1) ---
H_INV, H_ACCT, H_WRAP, H_BUY, H_SELL = 1, 2, 3, 4, 5
H_COST, H_PROCEEDS = 10, 12
# --- Wealth Summary: month headers on row 3 from col D; investable account rows ---
WS_MONTH_ROW = 3
WS_INVEST_ROWS = [5, 6, 7, 8, 9, 12]     # Investment Account(s) + ISAs + Junior ISA
WS_CASH_ROWS = [10, 11, 13]              # Fidelity Cash Accounts
# Total Portfolio Value / the value-over-time series read the Wealth Summary's own
# **row 33 'Fidelity accounts'** line (user decision 2026-07-19) — the sheet's and the
# Finance Google Sheet's headline Fidelity total. It spans the account block (rows
# 5-13) PLUS the two Fidelity SIPPs, which sit up in the pension blocks: Paul's row 15
# and Susan's row 25. The Accounts widget lists exactly those rows so its total ties
# back to row 33 (verified: 1,938,320 + 1,823,862 = 3,762,182).
WS_FIDELITY_TOTAL_ROW = 33
WS_SIPP_ROWS = {15: 'Paul', 25: 'Susan'}     # row -> account holder (label has no name)
WS_ACCOUNT_ROWS = WS_INVEST_ROWS + WS_CASH_ROWS + list(WS_SIPP_ROWS)
WS_MONTHLY_INCREASE_ROW = 46            # 'Monthly Investment Increase' (literal)
# --- Stocks of Interest section-table columns (A..Q) ---
SOI_STOCK, SOI_TICKER, SOI_PATTERN, SOI_ALOW, SOI_AHIGH = 1, 2, 3, 5, 7
SOI_CHARTNOTE, SOI_RATING, SOI_HOLDINGS, SOI_TARGET, SOI_NOTES, SOI_UPDATED, SOI_TV = 11, 12, 13, 14, 15, 16, 17
# Dashboard watchlist = strictly the 'AT LOWER BOUNDARY — within 5% of alert low'
# section band of Stocks of Interest (user decision 2026-07-17). Only this band.
SOI_WATCHLIST_BAND = 'AT LOWER BOUNDARY'
# --- Base Data columns (header row 2) ---
BD_TICKER, BD_NAME, BD_PE, BD_DIV_PENCE, BD_DIVYLD, BD_EXDIV = 1, 2, 6, 8, 9, 11


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _latest_prices():
    """ticker(upper) -> captured last price from the most recent history.db run."""
    prices = {}
    if not os.path.exists(HISTORY_DB):
        return prices
    con = sqlite3.connect(HISTORY_DB)
    row = con.execute('SELECT MAX(run_id) FROM chart_snapshots').fetchone()
    if row and row[0] is not None:
        for tkr, price in con.execute(
                'SELECT ticker, price FROM chart_snapshots WHERE run_id=? AND price IS NOT NULL',
                (row[0],)):
            if tkr:
                prices[str(tkr).strip().upper()] = price
    con.close()
    return prices


def _price_changes():
    """ticker(upper) -> {'change', 'change_pct'} day-over-day: latest captured
    price vs the most recent captured price from an EARLIER calendar day. Same-day
    re-runs don't count as 'the last change' — we want the daily move. Returns {}
    when there's no prior-day run to compare against."""
    changes = {}
    if not os.path.exists(HISTORY_DB):
        return changes
    con = sqlite3.connect(HISTORY_DB)
    try:
        row = con.execute('SELECT MAX(run_id) FROM chart_snapshots').fetchone()
        if not row or row[0] is None:
            return changes
        latest_run = row[0]
        # Calendar day of the latest run (price_checked_at is an ISO string).
        d = con.execute('SELECT MAX(substr(price_checked_at,1,10)) FROM chart_snapshots '
                        'WHERE run_id=?', (latest_run,)).fetchone()
        latest_day = d[0] if d else None
        latest = {}
        for tkr, price in con.execute(
                'SELECT ticker, price FROM chart_snapshots WHERE run_id=? AND price IS NOT NULL',
                (latest_run,)):
            if tkr:
                latest[str(tkr).strip().upper()] = price
        # Most recent run strictly before latest_day (a genuine prior trading day).
        prow = con.execute(
            'SELECT MAX(run_id) FROM chart_snapshots WHERE substr(price_checked_at,1,10) < ?',
            (latest_day,)).fetchone()
        if not prow or prow[0] is None:
            return changes
        prev = {}
        for tkr, price in con.execute(
                'SELECT ticker, price FROM chart_snapshots WHERE run_id=? AND price IS NOT NULL',
                (prow[0],)):
            if tkr:
                prev[str(tkr).strip().upper()] = price
    finally:
        con.close()
    for tkr, now_p in latest.items():
        old_p = prev.get(tkr)
        if isinstance(old_p, (int, float)) and old_p and isinstance(now_p, (int, float)):
            changes[tkr] = {'change': round(now_p - old_p, 2),
                            'change_pct': round((now_p - old_p) / old_p * 100.0, 2)}
    return changes


def _price_for(ticker, xlsx_literal, latest):
    """Prefer the captured price; fall back to a literal xlsx price (commodities)."""
    if ticker:
        p = latest.get(str(ticker).strip().upper())
        if isinstance(p, (int, float)):
            return p
    return _num(xlsx_literal)


# The broker exports are ARCHIVED out of Downloads after a successful pipeline run —
# consume_input_files.py moves the newest of each into ~/Downloads/old_pipeline/ under
# a canonical name, keeping exactly one copy. So 'not in Downloads' does NOT mean 'no
# export exists', and treating it that way is what silently dropped Cash Available back
# to the Wealth Summary's £79 (and Holdings back to the stale workbook column) once the
# 18 Jul export was consumed. Downloads wins when both exist — that's the fresher one.
EXPORT_DIRS = [os.path.join(os.path.expanduser('~'), 'Downloads'),
               os.path.join(os.path.expanduser('~'), 'Downloads', 'old_pipeline')]


def _export_path(name):
    """The newest available copy of a broker export, Downloads before the archive.
    Falls back to the Downloads path when neither exists, so 'missing' messages
    still name the place the user is expected to put it."""
    for d in EXPORT_DIRS:
        p = os.path.join(d, name)
        if os.path.exists(p):
            return p
    return os.path.join(EXPORT_DIRS[0], name)


ACCOUNT_SUMMARY = _export_path('AccountSummary.csv')
# --- AccountSummary.csv 'View all account details' columns (0-indexed) ---
AS_TYPE, AS_NAME, AS_ACCTNO, AS_PRODUCT, AS_HOLDER = 0, 1, 2, 3, 4
AS_PRICE, AS_QTY, AS_VALUE, AS_BOOKCOST, AS_GAIN = 7, 9, 10, 13, 14
AS_CASH = 12                       # 'Cash available' — carried on the ACCOUNT rows
# Broker ticker -> the Investments-sheet ticker the chart/alerts are keyed on.
AS_TICKER_ALIASES = {'AV.': 'AV', 'SPLT': 'PLAT', 'SPDM': 'PALL', 'BT.A': 'BT.A'}
# Fund holdings that belong in the Investments table, not Income Funds (no ticker
# in the broker name, so matched by name prefix) -> Investments ticker.
AS_FUND_AS_INVESTMENT = {'WS GUINNESS GLOBAL ENERGY': 'GUINNESS'}
# The dashboard models the FAMILY's investments only (user decision 2026-07-19).
# The Fidelity export covers accounts held for wider relatives too (Dorothy Wall,
# Olive Elizabeth Sangha, Freda Hibbert) — those are administered here but are not
# part of the family portfolio, and including them inflated every total. Matched on
# the account holder's first name, which is how the Wealth Summary labels them.
FAMILY_HOLDERS = {'PAUL', 'SUSAN', 'LIAM', 'JAYNE'}


def _as_num(s):
    try:
        return float(str(s).replace(',', '').replace('+', '').strip())
    except (TypeError, ValueError):
        return None


def _fidelity_positions(path=ACCOUNT_SUMMARY, funds=False):
    """Open positions from the Fidelity AccountSummary export, one row per
    holding-per-account (the 'View all account details' section — the section
    above it is an aggregate across accounts and would double-count).

    This is the authoritative list of what is actually HELD, and the only source
    of Account / Wrapper / book cost. Returns [] if the export isn't present.

    funds=False returns the equity/ETC/Investments-table holdings; funds=True
    returns the income-fund holdings instead (same section, complementary split).
    """
    if not os.path.exists(path):
        return []
    import csv
    rows = []
    in_detail = False
    with open(path, encoding='utf-8-sig', newline='') as fh:
        for rec in csv.reader(fh):
            if not rec:
                continue
            head = (rec[0] or '').strip()
            if head.lower().startswith('view all account details'):
                in_detail = True
                continue
            if not in_detail or head != 'Asset':
                continue
            name = (rec[AS_NAME] or '').strip()
            if not name or name.lower() == 'cash':
                continue
            ticker = None
            display = name
            if name.endswith(')') and '(' in name:
                ticker = name[name.rfind('(') + 1:-1].strip().upper()
                display = name[:name.rfind('(')].split(',')[0].strip().title()
            else:
                up = name.upper()
                for pref, tk in AS_FUND_AS_INVESTMENT.items():
                    if up.startswith(pref):
                        ticker = tk
                        break
            if funds != (ticker is None):
                continue       # wrong side of the equity / income-fund split
            ticker = AS_TICKER_ALIASES.get(ticker, ticker) if ticker else None
            holder = (rec[AS_HOLDER] or '').strip().split()[0] if rec[AS_HOLDER] else None
            if (holder or '').upper() not in FAMILY_HOLDERS:
                continue   # not a family account — see FAMILY_HOLDERS
            qty = _as_num(rec[AS_QTY])
            cost = _as_num(rec[AS_BOOKCOST])
            rows.append({
                'name': display, 'ticker': ticker,
                'account_no': (rec[AS_ACCTNO] or '').strip() or None,
                'account': holder, 'wrapper': (rec[AS_PRODUCT] or '').strip() or None,
                'quantity': qty,
                'holdings': _as_num(rec[AS_VALUE]),
                'book_cost': cost,
                'buy_price': round(cost / qty * 100.0, 2) if (cost and qty) else None,
                'pl_today': _as_num(rec[AS_GAIN]),
                'broker_price': _as_num(rec[AS_PRICE]),
            })
    return rows


DOWNLOADS = EXPORT_DIRS[0]
# --- TransactionHistory.csv columns (0-indexed, header 'Order date,...') ---
TX_ORDER, TX_COMPLETE, TX_TYPE, TX_WRAPPER, TX_ACCTNO, TX_SOURCE, TX_AMOUNT = 0, 1, 2, 4, 5, 6, 7


def _tx_date(s):
    for fmt in ('%d %b %Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(str(s).strip(), fmt).date()
        except (TypeError, ValueError):
            continue
    return None


def _fund_income_events(downloads=None):
    """(fund name upper, account number) -> the most recent 'Income Received' the
    Fidelity transaction export shows for that holding.

    This is the ONLY source of real dividend dates for the income funds — the
    workbook has none, and the funds aren't in Base Data (no ticker). Fidelity's
    Order date IS the ex-dividend date and Completion date the payment date.
    The export only covers the last ~30 days, so this is 'the latest payment',
    not a history. Nothing is inferred: a fund with no row gets no dates.
    """
    import csv
    events = {}
    # Both the live Downloads copy and the archived one (see EXPORT_DIRS) — the
    # archive is scanned FIRST so a fresher Downloads row overwrites it on the
    # most-recent-payment check below.
    dirs = list(reversed(EXPORT_DIRS)) if downloads is None else [downloads]
    files = [(d, fn) for d in dirs if os.path.isdir(d) for fn in sorted(os.listdir(d))]
    for d, fn in files:
        low = fn.lower()
        if not (low.startswith('transactionhistory') or low.startswith('transactions')):
            continue
        if not low.endswith('.csv'):
            continue
        try:
            with open(os.path.join(d, fn), encoding='utf-8-sig', newline='') as fh:
                for rec in csv.reader(fh):
                    if len(rec) <= TX_AMOUNT or (rec[TX_TYPE] or '').strip() != 'Income Received':
                        continue
                    src = (rec[TX_SOURCE] or '').strip()
                    acct = (rec[TX_ACCTNO] or '').strip()
                    if not src:
                        continue
                    paid = _tx_date(rec[TX_COMPLETE])
                    key = (src.upper(), acct)
                    prev = events.get(key)
                    if prev and prev['paid_date'] and paid and prev['paid_date'] >= paid.isoformat():
                        continue
                    ex = _tx_date(rec[TX_ORDER])
                    events[key] = {
                        'amount': _as_num(rec[TX_AMOUNT]),
                        'ex_div_date': ex.isoformat() if ex else None,
                        'paid_date': paid.isoformat() if paid else None,
                        'wrapper': (rec[TX_WRAPPER] or '').strip() or None,
                    }
        except OSError:
            continue   # a file mid-download / locked by Excel — skip, never fail
    return events


def _fidelity_cash(path=ACCOUNT_SUMMARY):
    """Uninvested cash per family account, from the broker export's 'Cash available'
    column on each ACCOUNT row of the 'View all account details' section.

    The Wealth Summary's cash rows only cover the three standalone Cash Accounts
    (£79 in total) and miss the cash sitting inside every ISA/SIPP, which is the
    bulk of it — that's what made the Cash Available widget wrong (user 2026-07-19).
    """
    if not os.path.exists(path):
        return None, []
    import csv
    rows, total = [], 0.0
    in_detail = False
    with open(path, encoding='utf-8-sig', newline='') as fh:
        for rec in csv.reader(fh):
            if not rec:
                continue
            head = (rec[0] or '').strip()
            if head.lower().startswith('view all account details'):
                in_detail = True
                continue
            if not in_detail or head != 'Account' or len(rec) <= AS_CASH:
                continue
            holder = (rec[AS_HOLDER] or '').strip().split()[0] if rec[AS_HOLDER] else None
            if (holder or '').upper() not in FAMILY_HOLDERS:
                continue
            cash = _as_num(rec[AS_CASH])
            if not cash:
                continue
            total += cash
            rows.append({'account': holder, 'wrapper': (rec[AS_PRODUCT] or '').strip() or None,
                         'cash': round(cash, 2)})
    rows.sort(key=lambda r: -r['cash'])
    return round(total, 2), rows


# Chart symbols priced in USD (per ounce / barrel / MMBtu) rather than UK pence.
# Every LSE equity and ETC the pipeline captures is quoted in pence, so pence is
# the default and only these are the exception — the dashboard labels the unit
# rather than putting a misleading '£' in front of a pence figure.
USD_PRICED_TICKERS = {'PLAT', 'PALL', 'GOLD', 'SILVER', 'SLVR', 'COPP', 'NATGAS',
                      'UKOIL', 'BRENT', 'OIL'}


def _price_unit(ticker):
    tk = str(ticker).strip().upper() if ticker else ''
    return '$' if tk in USD_PRICED_TICKERS else 'p'


def _tv_url(cell):
    """Extract the chart URL from a =HYPERLINK("url","label") formula cell."""
    if isinstance(cell, str) and cell.upper().startswith('=HYPERLINK('):
        try:
            return cell.split('"')[1]
        except IndexError:
            return None
    return None


_LAYOUT_BY_TICKER = None


def _layout_by_ticker():
    """ticker -> the SAVED LAYOUT it was captured in, from the capture manifest —
    the only place that mapping exists. Every chart link in the dashboard uses it
    to drive TradingView DESKTOP (the point is to land on the layout ready to draw
    on, which a browser tab can't do); a ticker with no captured layout falls back
    to its browser chart_url. Read once — the manifest doesn't change mid-run.

    Keyed under BOTH the raw TradingView symbol and its normalized master-sheet
    ticker, because the two sides disagree: the manifest carries TradingView's
    symbol ('PALLADIUM', 'COPPER1!', 'BT.A') while the workbook rows the tables are
    built from carry the sheet ticker ('PALL', 'COPP', 'BT-A'). Keying on the raw
    symbol alone silently drops every commodity and class-suffix row to a browser
    link.

    A ticker can be captured in SEVERAL layouts (GLEN is in both 'FT100 Mining' and
    'FT100 Support Services 2'); the last one in the manifest wins, which is what
    the Activity chips have always done — don't change that to first-wins here or
    existing chips silently jump layout. A raw-symbol match always beats a derived
    master-ticker one.
    """
    global _LAYOUT_BY_TICKER
    if _LAYOUT_BY_TICKER is None:
        by_raw, by_master = {}, {}
        lm_path = os.path.join(SCRIPT_DIR, 'layout_manifest_tmp.json')
        if os.path.exists(lm_path):
            try:
                for e in json.load(open(lm_path, encoding='utf-8')):
                    raw = str(e.get('ticker') or '').strip().upper()
                    if not (raw and e.get('chartId')):
                        continue
                    entry = {'chart_id': e['chartId'], 'layout': e.get('name'),
                             'layout_id': e.get('id')}
                    by_raw[raw] = entry
                    master = _master_key(e.get('ticker'))
                    if master:
                        by_master[master] = entry
            except (ValueError, OSError):
                pass
        _LAYOUT_BY_TICKER = {**by_master, **by_raw}
    return _LAYOUT_BY_TICKER


def _master_key(ticker):
    """A ticker in master-sheet form, dash-normalized (BT.A and BT-A are one key —
    the same equivalence ticker_normalize.master_tickers_match applies)."""
    if not ticker:
        return None
    try:
        info = ticker_normalize.normalize(ticker)
    except Exception:
        info = None
    master = (info or {}).get('master_ticker') or str(ticker)
    return master.strip().upper().replace('.', '-') or None


def _layout_for(ticker):
    """The layout dict for a ticker, or {} — spread into a row so the front end can
    open it in TradingView Desktop."""
    if not ticker:
        return {}
    lookup = _layout_by_ticker()
    return (lookup.get(str(ticker).strip().upper())
            or lookup.get(_master_key(ticker)) or {})


def _read_soi_band(soi, soif, band_substring, base, latest, changes=None):
    """Read one 'Stocks of Interest' section-table band (e.g. 'AT LOWER BOUNDARY')
    into dashboard watchlist row dicts. Returns [] if that band isn't found on the
    sheet (never invents rows) — membership within a band is still hand-curated
    (see refresh_soi_sections.py); this only reads the rows already placed there."""
    rows = []
    band_row = None
    for r in range(1, soi.max_row + 1):
        a = soi.cell(r, SOI_STOCK).value
        if isinstance(a, str) and band_substring in a.upper():
            band_row = r
            break
    if band_row is None:
        return rows
    r = band_row + 2  # skip the band row and the column-header row
    while r <= soi.max_row:
        stock = soi.cell(r, SOI_STOCK).value
        if not stock or (isinstance(stock, str) and stock.strip().upper() == 'STOCK'):
            break  # blank row or the next section's column header -> end of section
        ticker = soi.cell(r, SOI_TICKER).value
        tk = str(ticker).strip().upper() if ticker else None
        low = _num(soi.cell(r, SOI_ALOW).value)
        high = _num(soi.cell(r, SOI_AHIGH).value)
        price = _price_for(ticker, None, latest)
        prox = ((price - low) / low * 100.0) if (price and low) else None
        upside = ((high - low) / low * 100.0) if (high and low) else None
        bd_row = base.get(tk, {})
        updated = soi.cell(r, SOI_UPDATED).value
        chg = (changes or {}).get(tk, {})
        rows.append({
            'stock': stock, 'ticker': ticker,
            'pattern': soi.cell(r, SOI_PATTERN).value,
            'proximity_pct': round(prox, 2) if prox is not None else None,
            'alert_low': low, 'current': price, 'alert_high': high,
            'price_unit': _price_unit(ticker),
            'change': chg.get('change'), 'change_pct': chg.get('change_pct'),
            'upside_pct': round(upside, 2) if upside is not None else None,
            'pe': bd_row.get('pe'), 'div_yield_pct': bd_row.get('div_yield_pct'),
            'chart_note': soi.cell(r, SOI_CHARTNOTE).value,
            'analyst_rating': soi.cell(r, SOI_RATING).value,
            'holdings': _num(soi.cell(r, SOI_HOLDINGS).value),
            'target_value': _num(soi.cell(r, SOI_TARGET).value),
            'notes': soi.cell(r, SOI_NOTES).value,
            'last_updated': updated.strftime('%Y-%m-%d') if isinstance(updated, datetime.datetime) else updated,
            'chart_url': _tv_url(soif.cell(r, SOI_TV).value),
            **_layout_for(ticker),
        })
        r += 1
    return rows


# --- Payslip Summary: one row per pay date, banded by tax year -----------------
# Layout (header row 3, data from row 4): tax-year band rows carry only a label in
# column A ('  Tax Year 2025/26'), and each band ends with a 'TOTAL <year>' row we
# recompute rather than read (those totals are formulas, so they read as None here —
# same constraint as everything else in this file).
PS_SHEET = 'Payslip Summary'
PS_HEADER_ROW = 3
PS_COLS = [
    ('pay_date',      1), ('tax_year',    2), ('gross',       3),
    ('pension_ee',    4), ('pension_er',  5), ('pension_total', 6),
    ('bonus',         7), ('taxable_pay', 8), ('paye',        9),
    ('ni',           10), ('deductions', 11), ('net',        12),
]
PS_MONEY = [k for k, _ in PS_COLS if k not in ('pay_date', 'tax_year')]


def _ps_date(v):
    """Pay date as an ISO string. The sheet mixes real dates and dd/mm/yyyy text."""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v or '').strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# --- Pension annual allowance ------------------------------------------------
# The Payslips screen shows how much annual allowance is left and, from the current
# contribution rate, the month contributions would have to stop to stay inside it.
#
# THIS IS ARITHMETIC ON PAUL'S OWN PAYSLIP FIGURES, NOT TAX ADVICE. Three things it
# cannot know and therefore states as assumptions on the screen:
#   * The TAPER. Adjusted income over £260,000 cuts the allowance by £1 for every £2
#     over, down to £10,000 — and adjusted income covers ALL income, not just this
#     employment, so it can't be derived from payslips.
#   * The MPAA. Flexibly accessing a defined-contribution pension drops the annual
#     allowance to £10,000 and removes carry-forward entirely.
#   * OTHER pension inputs. Only contributions on these payslips are counted; a
#     personal contribution straight into the SIPP, or an AVC paid outside payroll,
#     is an input to the same allowance and would eat into the same figure.
# Both the allowance and the taper are overridable from the screen so the figure can
# be corrected without a code change.
ANNUAL_ALLOWANCE_DEFAULT = 60000        # 2023/24 onward (was £40,000 to 2022/23)
ANNUAL_ALLOWANCE_BY_YEAR = {
    '2020/21': 40000, '2021/22': 40000, '2022/23': 40000,
}
CARRY_FORWARD_YEARS = 3                 # unused allowance survives three tax years


def _allowance_for(tax_year):
    return ANNUAL_ALLOWANCE_BY_YEAR.get(tax_year, ANNUAL_ALLOWANCE_DEFAULT)


def _tax_year_start(tax_year):
    return int(str(tax_year).split('/')[0])


def _tax_year_months(tax_year):
    """The twelve pay months of a UK tax year, April-first, as (year, month)."""
    start = _tax_year_start(tax_year)
    return [((start + (0 if m >= 4 else 1)), m)
            for m in list(range(4, 13)) + list(range(1, 4))]


def pension_allowance(year_totals, rows, today=None):
    """Remaining allowance for the current tax year, and when to stop contributing.

    year_totals: {tax_year: pension input for that year}
    Carry-forward is consumed OLDEST FIRST, after the year's own allowance — HMRC's
    order, and it matters: using the current year first is what leaves the oldest
    unused amount available to expire.
    """
    today = today or datetime.date.today()
    years = sorted(year_totals)
    if not years:
        return None

    # Walk the years oldest -> newest so each year's carry-forward reflects what
    # later years have already consumed.
    unused = {}
    for y in years:
        aa = _allowance_for(y)
        need = year_totals.get(y) or 0
        own = min(need, aa)
        unused[y] = aa - own
        need -= own
        if need > 0:
            start = _tax_year_start(y)
            for back in range(CARRY_FORWARD_YEARS, 0, -1):     # oldest first
                prev = f'{start - back}/{str(start - back + 1)[2:]}'
                if prev not in unused or need <= 0:
                    continue
                take = min(need, unused[prev])
                unused[prev] -= take
                need -= take

    current = _tax_year_of(today.isoformat())
    aa = _allowance_for(current)
    start = _tax_year_start(current)
    # Each prior year is carried with its FULL working — allowance, what was actually
    # contributed, and what survives — so the panel can show how the carry-forward
    # figure was arrived at rather than asserting it (user request 2026-07-19).
    carry = []
    for back in range(CARRY_FORWARD_YEARS, 0, -1):
        prev = f'{start - back}/{str(start - back + 1)[2:]}'
        if prev in unused:
            prev_aa = _allowance_for(prev)
            prev_used = round(year_totals.get(prev) or 0, 2)
            carry.append({'tax_year': prev,
                          'allowance': prev_aa,
                          'contributed': prev_used,
                          'headroom': round(prev_aa - prev_used, 2),
                          'unused': round(unused[prev], 2),
                          'consumed_by_later_years': round(max(0.0, (prev_aa - prev_used) - unused[prev]), 2),
                          'expires_after': f'{start - back + CARRY_FORWARD_YEARS}/'
                                           f'{str(start - back + CARRY_FORWARD_YEARS + 1)[2:]}'})
    carry_total = round(sum(c['unused'] for c in carry), 2)
    used = round(year_totals.get(current) or 0, 2)
    available = round(aa + carry_total, 2)
    remaining = round(available - used, 2)

    # Contribution rate: the most recent payslip in the current tax year, which is
    # what the next one will look like unless the rate is changed.
    cur_rows = [r for r in rows if r.get('tax_year') == current and r.get('pension_total')]
    cur_rows.sort(key=lambda r: r['pay_date'])
    monthly = cur_rows[-1]['pension_total'] if cur_rows else None

    # Which pay months are still to come this tax year, and where the money runs out.
    # Every pay month of this tax year with no payslip recorded yet. A month that has
    # already PASSED but has no payslip still counts — it was almost certainly paid and
    # simply hasn't been uploaded, and dropping it would understate the year's input
    # and overstate what's left.
    paid_months = {tuple(int(x) for x in r['pay_date'].split('-')[:2]) for r in cur_rows}
    this_month = (today.year, today.month)
    upcoming = [(y, m) for (y, m) in _tax_year_months(current) if (y, m) not in paid_months]
    schedule, running, stop_after = [], used, None
    if monthly:
        for (y, m) in upcoming:
            projected = round(running + monthly, 2)
            over = projected > available
            schedule.append({'year': y, 'month': m,
                             'label': datetime.date(y, m, 1).strftime('%b %Y'),
                             'contribution': monthly,
                             'cumulative': projected, 'over': over,
                             'unrecorded': (y, m) < this_month})
            if over and stop_after is None:
                # The LAST affordable month is the one before this — stop after it.
                stop_after = schedule[-2]['label'] if len(schedule) > 1 else 'already over'
            running = projected

    return {
        'tax_year': current,
        'annual_allowance': aa,
        'carry_forward': carry,
        'carry_forward_total': carry_total,
        'available': available,
        'used': used,
        'remaining': remaining,
        'over': remaining < 0,
        'monthly_contribution': monthly,
        'months_affordable': (int(remaining // monthly) if monthly and remaining > 0 else 0),
        'stop_after': stop_after,
        'schedule': schedule,
        'history': [{'tax_year': y, 'allowance': _allowance_for(y),
                     'contributed': round(year_totals.get(y) or 0, 2),
                     'over_allowance': round(max(0.0, (year_totals.get(y) or 0) - _allowance_for(y)), 2)}
                    for y in years],
        'payslips_recorded': len(cur_rows),
        'months_unrecorded': sum(1 for s in schedule if s['unrecorded']),
        'assumptions': [
            f'Standard annual allowance of {aa:,.0f} — NOT tapered. Adjusted income over '
            '260,000 reduces it by 1 for every 2 over, down to 10,000, and adjusted income '
            'covers all income so it cannot be read off a payslip.',
            'No money purchase annual allowance (MPAA). Flexibly accessing a DC pension '
            'cuts the allowance to 10,000 and removes carry-forward.',
            'Only pension input on these payslips is counted (employee + employer). A '
            'personal contribution paid straight into the SIPP, or an AVC outside payroll, '
            'is an input to the same allowance and is not visible here.',
            'Carry-forward is taken oldest first, after the current year own allowance.',
        ],
    }


def build_payslips(workbook=WORKBOOK):
    """Every payslip row, plus per-tax-year totals for the screen's filter.

    Rows come from the workbook's hand-maintained Payslip Summary tab; new ones are
    appended by the dashboard's own upload route, NOT by the pipeline (user decision
    2026-07-19 — payslips are loaded from the Payslips screen).
    """
    rows = []
    if os.path.exists(workbook):
        wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
        if PS_SHEET in wb.sheetnames:
            ws = wb[PS_SHEET]
            for r in ws.iter_rows(min_row=PS_HEADER_ROW + 1, values_only=True):
                cell = lambda i: r[i - 1] if len(r) >= i else None
                date = _ps_date(cell(1))
                if not date:
                    continue      # band label ('  Tax Year 2025/26') or TOTAL row
                row = {k: (_num(cell(i)) if k in PS_MONEY else cell(i))
                       for k, i in PS_COLS}
                row['pay_date'] = date
                row['tax_year'] = str(row['tax_year'] or '').strip() or _tax_year_of(date)
                # Deductions is a formula in the sheet, so derive it
                if not row['deductions']:
                    row['deductions'] = round((row['paye'] or 0) + (row['ni'] or 0)
                                              + (row['pension_ee'] or 0), 2)
                rows.append(row)
        wb.close()

    rows.sort(key=lambda x: x['pay_date'], reverse=True)

    years = {}
    for row in rows:
        acc = years.setdefault(row['tax_year'], {'tax_year': row['tax_year'], 'payslips': 0})
        acc['payslips'] += 1
        for k in PS_MONEY:
            acc[k] = round(acc.get(k, 0) + (row[k] or 0), 2)

    order = sorted(years, reverse=True)
    totals = {y: years[y].get('pension_total') or 0 for y in years}
    return {
        'generated_at': datetime.datetime.now().isoformat(timespec='seconds'),
        'rows': rows,
        'tax_years': [years[y] for y in order],
        'current_tax_year': order[0] if order else None,
        'columns': [k for k, _ in PS_COLS],
        'allowance': pension_allowance(totals, rows),
    }


def _tax_year_of(iso_date):
    """UK tax year containing a date — 6 April to 5 April."""
    d = datetime.date.fromisoformat(iso_date)
    start = d.year if (d.month, d.day) >= (4, 6) else d.year - 1
    return f'{start}/{str(start + 1)[2:]}'


def build(workbook=WORKBOOK):
    wb = openpyxl.load_workbook(workbook, data_only=True)
    wbf = openpyxl.load_workbook(workbook, data_only=False)  # for HYPERLINK formulas
    latest = _latest_prices()
    changes = _price_changes()   # day-over-day price move per ticker
    now = datetime.datetime.now()

    # ---------------- Portfolio (holdings tables) ----------------
    inv = wb['Investments']
    invf = wbf['Investments']
    # Index the Investments sheet by ticker — it owns the alert levels, Type and
    # TradingView link, but NOT what is actually held (its Holdings column goes
    # stale and misses positions, e.g. Centrica). The broker export owns holdings.
    inv_by_ticker = {}
    for r in range(4, inv.max_row + 1):
        tk = inv.cell(r, I_TICKER).value
        if tk:
            inv_by_ticker.setdefault(str(tk).strip().upper(), r)

    def _inv_row(ticker):
        return inv_by_ticker.get(str(ticker).strip().upper()) if ticker else None

    positions = _fidelity_positions()
    if not positions:
        # No broker export in Downloads — fall back to the workbook's own Holdings.
        for r in range(4, inv.max_row + 1):
            hv = _num(inv.cell(r, I_HOLDINGS).value)
            if not inv.cell(r, I_NAME).value or not hv or hv <= 0:
                continue
            positions.append({
                'name': inv.cell(r, I_NAME).value,
                'ticker': inv.cell(r, I_TICKER).value,
                'account': None, 'wrapper': None, 'quantity': None,
                'holdings': hv, 'book_cost': None, 'buy_price': None,
                'pl_today': None, 'broker_price': None,
            })

    investments = []
    inv_value_total = 0.0
    short_term_value = strategic_value = 0.0
    strategic_count = 0
    alert_below = alert_near = alert_above = 0
    for pos in positions:
        ticker = pos['ticker']
        r = _inv_row(ticker)
        holdings = pos['holdings'] or 0.0
        name = inv.cell(r, I_NAME).value if r else pos['name']
        price = (_price_for(ticker, inv.cell(r, I_CURPRICE).value if r else None, latest)
                 or pos['broker_price'])
        low = _num(inv.cell(r, I_ALERT_LOW).value) if r else None
        high = _num(inv.cell(r, I_ALERT_HIGH).value) if r else None
        diff_low = ((price - low) / low * 100.0) if (price and low) else None
        gap_low = ((price - low) / price * 100.0) if (price and low) else None
        inv_value_total += holdings
        itype = inv.cell(r, I_TYPE).value if r else None
        # 'Strategic' replaced 'Long Term' (user request 2026-07-19). The old wording is
        # still matched so a sheet that hasn't been migrated classifies the same way.
        if isinstance(itype, str) and itype.strip().lower() in ('strategic', 'long term'):
            strategic_value += holdings
            strategic_count += 1
        else:
            short_term_value += holdings
        if diff_low is not None:
            if diff_low <= 0:
                alert_below += 1
            elif diff_low <= 5:
                alert_near += 1
            else:
                alert_above += 1
        investments.append({
            'name': name, 'ticker': ticker,
            'type': itype,                       # Short Term / Strategic
            'account': pos['account'], 'wrapper': pos['wrapper'],
            'holdings': round(holdings, 2),
            'quantity': pos['quantity'],
            'buy_price': pos['buy_price'],       # pence, = book cost / quantity
            'book_cost': pos['book_cost'],
            'pl_today': pos['pl_today'],         # £ profit/loss if sold today
            'pl_today_pct': (round(pos['pl_today'] / pos['book_cost'] * 100.0, 2)
                             if (pos['pl_today'] is not None and pos['book_cost']) else None),
            'current_price': price,
            'price_unit': _price_unit(ticker),   # unit of current_price/alert levels
            'gap_to_low_pct': round(gap_low, 2) if gap_low is not None else None,
            'alert_low': low, 'alert_low_source': inv.cell(r, I_ALERT_LOW_SRC).value if r else None,
            'diff_to_low_pct': round(diff_low, 2) if diff_low is not None else None,
            'alert_high': high,
            'chart_url': _tv_url(invf.cell(r, I_TV).value) if r else None,
            **_layout_for(ticker),
        })
    investments.sort(key=lambda x: (-(x['holdings'] or 0)))

    # ---------------- Income Funds ----------------
    inf = wb['Income Funds']
    income_funds = []
    funds_value_total = 0.0
    funds_monthly_income = 0.0
    for r in range(5, inf.max_row + 1):
        fname = inf.cell(r, F_NAME).value
        # The fund list ends at the 'TOTAL (Family)' row; a second per-person
        # sub-table follows it. Stop there so neither is summed into the holdings.
        if isinstance(fname, str) and fname.strip().upper().startswith('TOTAL'):
            break
        curval = _num(inf.cell(r, F_CURVAL).value)
        if not fname or curval is None or curval <= 0:
            continue
        monthly_rev = _num(inf.cell(r, F_MONTHLY_REV).value) or 0.0
        annual_rev = _num(inf.cell(r, F_ANNUAL_REV).value) or 0.0
        yld = _num(inf.cell(r, F_DIVYLD).value)
        funds_value_total += curval
        funds_monthly_income += monthly_rev
        income_funds.append({
            'name': fname, 'account': None, 'wrapper': None,
            'holdings': round(curval, 2),
            'div_yield_pct': round(yld * 100.0, 2) if yld is not None else None,
            'monthly_income': round(monthly_rev, 2),
            'annual_income': round(annual_rev, 2),   # 2026 annual dividend (run-rate)
        })
    funds_annual_income = round(sum(f['annual_income'] for f in income_funds), 2)

    # ---------------- Income fund POSITIONS (per fund, per account) ----------------
    # income_funds above stays as the sheet's family-level roll-up (the Overview
    # metrics, Relevant News and the Historic dividends table all read it). This is
    # the Portfolio screen's per-account view: broker holdings + the real dividend
    # dates from the transaction export.
    income_events = _fund_income_events()
    income_positions = []
    for pos in _fidelity_positions(funds=True):
        ev = income_events.get((pos['name'].upper(), pos['account_no'] or '')) or {}
        income_positions.append({
            'name': pos['name'],
            'account': pos['account'], 'wrapper': pos['wrapper'],
            'holdings': pos['holdings'], 'book_cost': pos['book_cost'],
            'pl_today': pos['pl_today'],
            'pl_today_pct': (round(pos['pl_today'] / pos['book_cost'] * 100.0, 2)
                             if (pos['pl_today'] is not None and pos['book_cost']) else None),
            # Yield is DERIVED from this holding's own last payment (these funds
            # distribute monthly), not matched to the sheet by name — the fund names
            # differ enough between broker and sheet that token matching paired AXA
            # and IFSL with other managers' funds.
            'div_yield_pct': (round(ev['amount'] * 12.0 / pos['holdings'] * 100.0, 2)
                              if (ev.get('amount') and pos['holdings']) else None),
            'last_income': ev.get('amount'),
            'ex_div_date': ev.get('ex_div_date'),
            'payment_date': ev.get('paid_date'),
        })
    income_positions.sort(key=lambda x: -(x['holdings'] or 0))
    income_positions_total = {
        'holdings': round(sum(p['holdings'] or 0 for p in income_positions), 2),
        'book_cost': round(sum(p['book_cost'] or 0 for p in income_positions), 2),
        'pl_today': round(sum(p['pl_today'] or 0 for p in income_positions), 2),
        'last_income': round(sum(p['last_income'] or 0 for p in income_positions), 2),
        'with_dates': sum(1 for p in income_positions if p['payment_date']),
        'count': len(income_positions),
    }

    # ---------------- Wealth Summary (monthly trend, cash, MoM gain) ----------------
    ws = wb['Wealth Summary']
    months = []
    _MON = {m: i for i, m in enumerate(
        ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], start=1)}
    for c in range(4, ws.max_column + 1):
        label = ws.cell(WS_MONTH_ROW, c).value
        if not label:
            continue
        # Only ACTUALS: the Wealth Summary carries forward-projected future months,
        # which would draw a misleading flat line past today. Cap at the current month.
        mkey = None
        parts = str(label).split()
        if len(parts) == 2 and parts[0] in _MON and parts[1].isdigit():
            yr, mo = int(parts[1]), _MON[parts[0]]
            if (yr, mo) > (now.year, now.month):
                continue
            mkey = (yr, mo)
        invest_sum = _num(ws.cell(WS_FIDELITY_TOTAL_ROW, c).value) or 0.0
        if invest_sum <= 0:
            continue  # drop empty / future months with no data
        months.append({'col': c, 'label': str(label), 'value': round(invest_sum, 2), 'key': mkey})
    last_col = months[-1]['col'] if months else 4
    # The chart DROPS the current month: the snapshot lands mid-month, so its Fidelity
    # total is a partial figure that would drag the line down at the right edge. last_col
    # still points at that latest actual month for the cash fallback below.
    chart_months = [m for m in months if m['key'] != (now.year, now.month)]
    value_over_time = {'labels': [m['label'] for m in chart_months],
                       'portfolio': [m['value'] for m in chart_months]}
    # Cash: the broker export is authoritative (it includes the cash held INSIDE each
    # ISA/SIPP). The Wealth Summary's standalone Cash Account rows are the fallback.
    cash_available, cash_rows = _fidelity_cash()
    cash_source = 'Fidelity AccountSummary — cash available across family accounts'
    if cash_available is None:
        cash_available = sum(_num(ws.cell(row, last_col).value) or 0.0 for row in WS_CASH_ROWS)
        cash_rows = []
        cash_source = ('Wealth Summary cash-account rows — standalone Cash Accounts only, '
                       'excludes cash held inside each ISA/SIPP (no broker export in '
                       'Downloads or Downloads/old_pipeline)')
    # Month-over-month change in the investable total (transparent, and consistent
    # with the trend chart). Includes contributions — flagged in the metric caveat.
    #
    # Measured over the last COMPLETE month, not the current one (user request
    # 2026-07-19): the current month is still accruing and, in the Wealth Summary,
    # often just carries the previous figure forward — which showed as a +£193 'gain'
    # for July while June had actually moved £15,882. Same 'last full month' rule the
    # Accounts widget uses, and the month is named on the widget so it is never
    # ambiguous which period the figure covers.
    # 'Has all the data' has to mean more than 'is not the current month'. The Wealth
    # Summary CARRIES THE PREVIOUS FIGURE FORWARD for any account that hasn't been
    # re-imported, so June sat at May's exact total and a June-vs-May reading was £0 —
    # just as misleading as July's +£193. The last month with real data is therefore
    # the last one whose value actually MOVED; everything after it is a copy.
    # Two conditions, both needed: the month must be COMPLETE (the current one is
    # still accruing) and it must have MOVED (anything flat is a carried-forward copy).
    def _is_current(m):
        parts = str(m['label']).split()
        return (len(parts) == 2 and _MON.get(parts[0]) == now.month
                and parts[1] == str(now.year))
    last_complete = len(months) - 1
    while last_complete > 0 and _is_current(months[last_complete]):
        last_complete -= 1
    gain_idx = None
    for i in range(last_complete, 0, -1):
        if months[i]['value'] != months[i - 1]['value']:
            gain_idx = i
            break
    gain_last_month = gain_last_month_pct = gain_month_label = gain_prev_label = None
    if gain_idx is not None:
        cur, prev_m = months[gain_idx], months[gain_idx - 1]
        gain_last_month = round(cur['value'] - prev_m['value'], 2)
        gain_last_month_pct = (round(gain_last_month / prev_m['value'] * 100.0, 2)
                               if prev_m['value'] else None)
        gain_month_label, gain_prev_label = cur['label'], prev_m['label']
    gain_stale_months = (last_complete - gain_idx) if gain_idx is not None else 0

    # Gain SINCE THE START OF THIS YEAR, and the last 6 monthly moves for the
    # sparkline (user request 2026-07-20). Both END on the same month as the
    # headline figure — a card showing two gains measured to different dates
    # invites exactly the misreading the 'name the month' rule exists to stop.
    def _label_ym(label):
        parts = str(label).split()
        if len(parts) == 2 and parts[0] in _MON and parts[1].isdigit():
            return int(parts[1]), _MON[parts[0]]
        return None

    end_i = gain_idx if gain_idx is not None else (last_complete if months else None)
    gain_ytd = gain_ytd_pct = gain_ytd_from = None
    gain_series = []
    if end_i is not None and end_i > 0:
        end_ym = _label_ym(months[end_i]['label'])
        if end_ym:
            # Baseline = the closing value of LAST December when the sheet carries it,
            # otherwise the first month of this year (the series currently starts at
            # Jan, so 'since the start of the year' really means 'since the January
            # figure' — the caveat names both months so it can't be misread).
            base_i = None
            for i, m in enumerate(months):
                ym = _label_ym(m['label'])
                if not ym:
                    continue
                if ym[0] == end_ym[0] - 1 and ym[1] == 12:
                    base_i = i
                elif ym[0] == end_ym[0] and base_i is None:
                    base_i = i
                    break
            if base_i is not None and base_i < end_i:
                base_v = months[base_i]['value']
                gain_ytd = round(months[end_i]['value'] - base_v, 2)
                gain_ytd_pct = round(gain_ytd / base_v * 100.0, 2) if base_v else None
                gain_ytd_from = months[base_i]['label']
        # Month-on-month moves, not the value line — this is the GAIN card, and the
        # value trend already has its own widget.
        for i in range(max(1, end_i - 5), end_i + 1):
            gain_series.append({'label': months[i]['label'],
                                'value': round(months[i]['value'] - months[i - 1]['value'], 2)})

    # Total Portfolio Value = the Wealth Summary's 'Fidelity accounts' block for the
    # latest actual month — the same series the trend chart plots, so the headline and
    # the sparkline can never disagree. (It is NOT holdings + income funds: that summed
    # positions across pension accounts too and came out ~£1M higher than the sheet.)
    portfolio_value = round(months[-1]['value'], 2) if months else round(
        inv_value_total + funds_value_total, 2)

    # ---------------- Fidelity accounts, last FULL month ----------------
    # 'Last full month' = the month before the current one; today's month is still
    # accruing, so comparing it to the previous month understates the increase.
    import re as _re
    accounts, accounts_month, accounts_prev_month = [], None, None
    if len(months) >= 2:
        full = months[-2] if str(months[-1]['label']).startswith(
            now.strftime('%b')) else months[-1]
        idx = months.index(full)
        prev = months[idx - 1] if idx >= 1 else None
        accounts_month = full['label']
        accounts_prev_month = prev['label'] if prev else None
        for row in WS_ACCOUNT_ROWS:
            label = ws.cell(row, 1).value
            if not label:
                continue
            m = _re.match(r'^\s*(.*?)\s*\(([^()]*)\)\s*\(([^()]*)\)\s*$', str(label))
            if m:
                wrapper, holder, acct_no = m.group(1), m.group(2), m.group(3)
            else:
                # The SIPP rows are labelled '  SIPP Savings - Fidelity (2000001606)':
                # one bracket, no holder name — that comes from the pension block header.
                m1 = _re.match(r'^\s*(.*?)\s*\(([^()]*)\)\s*$', str(label))
                wrapper = (m1.group(1) if m1 else str(label)).strip()
                acct_no = m1.group(2) if m1 else None
                holder = WS_SIPP_ROWS.get(row)
            val = _num(ws.cell(row, full['col']).value)
            pval = _num(ws.cell(row, prev['col']).value) if prev else None
            if val is None:
                continue
            accounts.append({
                'account': holder, 'wrapper': wrapper, 'account_no': acct_no,
                'value': round(val, 2),
                'prev_value': round(pval, 2) if pval is not None else None,
                'increase': round(val - pval, 2) if pval is not None else None,
                'increase_pct': (round((val - pval) / pval * 100.0, 2)
                                 if (pval not in (None, 0)) else None),
            })
        accounts.sort(key=lambda a: -(a['value'] or 0))
    # monthly_dividend is completed below, once Base Data dividend yields are read,
    # so it can include SHARE dividends (income + accumulation), not just income funds.

    # ---------------- Historic (completed sales) ----------------
    hist = wb['History']
    sold = []
    # The trading-profit year is DERIVED, never a literal (2026-07-19): it was
    # pinned to 2026, so on 1 Jan the figure would silently have gone to £0 with a
    # widget still captioned "2026". Taken from the calendar year of the most recent
    # completed sale rather than today's date, so an early-January dashboard still
    # reports the year that actually has trades in it instead of an empty one.
    profit_year = None
    profit_ytd = 0.0
    wins = 0
    for r in range(2, hist.max_row + 1):
        sell = hist.cell(r, H_SELL).value
        if not isinstance(sell, datetime.datetime):
            continue  # section dividers / open positions
        cost = _num(hist.cell(r, H_COST).value)
        proceeds = _num(hist.cell(r, H_PROCEEDS).value)
        profit = (proceeds - cost) if (cost is not None and proceeds is not None) else None
        buy = hist.cell(r, H_BUY).value
        days = (sell - buy).days if isinstance(buy, datetime.datetime) else None
        pnl_pct = (profit / cost * 100.0) if (profit is not None and cost) else None
        if profit is not None and profit > 0:
            wins += 1
        sold.append({
            'name': hist.cell(r, H_INV).value,
            'account': hist.cell(r, H_ACCT).value,
            'wrapper': hist.cell(r, H_WRAP).value,
            'buy_date': buy.strftime('%Y-%m-%d') if isinstance(buy, datetime.datetime) else None,
            'sell_date': sell.strftime('%Y-%m-%d'),
            'profit': round(profit, 2) if profit is not None else None,
            'pnl_pct': round(pnl_pct, 2) if pnl_pct is not None else None,
            'days_held': days,
        })
    win_rate = round(wins / len(sold) * 100.0, 1) if sold else None
    if sold:
        profit_year = max(int(s['sell_date'][:4]) for s in sold)
        profit_ytd = sum(s['profit'] for s in sold
                         if s['profit'] is not None and s['sell_date'][:4] == str(profit_year))
    sells_in_year = sum(1 for s in sold if s['sell_date'][:4] == str(profit_year))

    # ---------------- Watchlist (Stocks of Interest, within 5% of alert low) ----------
    # Base Data (literal, from HL.co.uk) gives P/E + Div Yield by ticker — the section
    # table's own P/E/Div-Yield cells are googlefinance/VLOOKUP formulas (empty here).
    bd = wb['Base Data']
    base = {}
    for r in range(3, bd.max_row + 1):
        tkr = bd.cell(r, BD_TICKER).value
        if tkr:
            exdiv = bd.cell(r, BD_EXDIV).value
            base[str(tkr).strip().upper()] = {
                'pe': _num(bd.cell(r, BD_PE).value),
                'div_yield_pct': (round(_num(bd.cell(r, BD_DIVYLD).value) * 100.0, 2)
                                  if _num(bd.cell(r, BD_DIVYLD).value) is not None else None),
                'div_pence': _num(bd.cell(r, BD_DIV_PENCE).value),
                'ex_div': (exdiv.strftime('%Y-%m-%d') if isinstance(exdiv, datetime.datetime) else exdiv)}
    # Monthly SHARE dividends (user request 2026-07-18): add the ACTUAL share income
    # + accumulation dividends that spending_summary computes and exports to
    # data/spending_dividends.json each pipeline run, on top of the income-fund revenue
    # (which stays sourced from the master's Income Funds tab — no double count).
    # Falls back to 0 until that file exists (e.g. a run without the Fidelity exports).
    share_income_monthly = share_accum_monthly = 0.0
    _dv = {}                      # stays {} if the export is missing — see the YTD below
    try:
        with open(os.path.join(REPO, 'data', 'spending_dividends.json'), encoding='utf-8') as _f:
            _dv = json.load(_f)
        share_income_monthly = float(_dv.get('share_income_monthly') or 0)
        share_accum_monthly = float(_dv.get('share_accumulation_monthly') or 0)
    except (OSError, ValueError):
        pass
    # Monthly Dividend = income actually PAID OUT: income funds + stocks-and-shares
    # dividends, each shown as its own line on the card (user request 2026-07-20).
    #
    # The accumulation funds' reinvested income is NOT in here. It used to be, AND was
    # then added again by total_income (= monthly_dividend + accumulative), so Total
    # Income double-counted it: £21,707 + £4,381 = £26,088 against a true £21,707.
    # It has its own 'Accumulative Fund Income' card, so paid-out vs reinvested is the
    # clean split — and it makes the stated relationship true, that the two cards to
    # the left of Total Income add up to it.
    monthly_dividend = round(funds_monthly_income + share_income_monthly, 2)
    total_income_monthly = round(monthly_dividend + share_accum_monthly, 2)
    # Year-to-date total income (user request 2026-07-20). MEASURED where possible:
    # spending_summary exports the month-by-month actuals to spending_dividends.json
    # ('months', each flagged actual/estimated), so sum the actual months of this year
    # rather than multiplying up a rate. Only the months flagged actual count — the
    # rest of the year is projection, and summing those would report a forecast as a
    # result. Falls back to the run-rate when the export predates this field.
    income_ytd_actual = None
    income_ytd_partial = False
    _rows = [m for m in (_dv.get('months') or []) if str(m.get('period', '')).startswith(str(now.year))] \
        if isinstance(_dv, dict) else []
    _actual_rows = [m for m in _rows if m.get('actual')]
    if _actual_rows:
        income_ytd_actual = round(sum(float(m.get('total') or 0) for m in _actual_rows), 2)
        income_ytd_months = len(_actual_rows)
        income_ytd_partial = bool(_actual_rows[-1].get('partial'))
        income_ytd = income_ytd_actual
    else:
        income_ytd_months = now.month
        income_ytd = round(total_income_monthly * income_ytd_months, 2)

    soi = wb['Stocks of Interest']
    soif = wbf['Stocks of Interest']
    # Strictly the 'AT LOWER BOUNDARY — within 5% of alert low' band (user decision
    # 2026-07-17): the Watchlist is the buy-zone list, not the whole Stocks-of-Interest
    # ladder. The other bands stay out.
    watchlist = _read_soi_band(soi, soif, SOI_WATCHLIST_BAND, base, latest, changes)

    # ---------------- Chart statistics ----------------
    # How the OCR/detection run went, from the same channel_results_tmp.json the
    # Activity items read. 'Marked up' mirrors update_master_sheet.marked_up_flag:
    # a detected pattern OR an axis failure means drawings ARE present.
    chart_stats = None
    _cr_path = os.path.join(SCRIPT_DIR, 'channel_results_tmp.json')
    if os.path.exists(_cr_path):
        recs = json.load(open(_cr_path, encoding='utf-8'))
        groups = {'Parallel channel': 0, 'Trend lines only': 0, 'Wedge': 0,
                  'Breakout / breakdown': 0, 'On the line': 0, 'Single boundary': 0,
                  'No lines drawn': 0, 'Not read (axis)': 0}
        levels = axis_fail = undrawn = 0
        for r in recs:
            pat = (r.get('pattern') or '').lower()
            reason = (r.get('reason') or '').lower()
            if r.get('lower') is not None or r.get('upper') is not None:
                levels += 1
            if 'wedge' in pat:
                groups['Wedge'] += 1
            elif 'inside channel' in pat:
                groups['Parallel channel'] += 1
            elif 'broken out' in pat or 'broken down' in pat:
                groups['Breakout / breakdown'] += 1
            elif 'on a drawn line' in pat:
                groups['On the line'] += 1
            elif 'single blue boundary' in pat:
                groups['Single boundary'] += 1
            elif 'trend lines only' in pat:
                groups['Trend lines only'] += 1
            elif 'axis' in reason:
                groups['Not read (axis)'] += 1
                axis_fail += 1
            else:
                groups['No lines drawn'] += 1
                undrawn += 1
        total = len(recs)
        marked = total - undrawn
        chart_stats = {
            'total': total, 'marked_up': marked, 'unmarked': undrawn,
            'marked_up_pct': round(marked / total * 100.0, 1) if total else None,
            'with_levels': levels,
            'with_levels_pct': round(levels / total * 100.0, 1) if total else None,
            'axis_failed': axis_fail,
            'patterns': [{'label': k, 'count': v} for k, v in
                         sorted(groups.items(), key=lambda kv: -kv[1]) if v],
        }

    overview = {
        'generated_at': now.isoformat(timespec='seconds'),
        'currency': 'GBP',
        'metrics': {
            'accumulation_income': {'value': share_accum_monthly,
                                    'caveat': 'Monthly reinvested income from accumulation funds '
                                              '(latest month of the Acc funds’ price-appreciation).'},
            'portfolio_value': {'value': portfolio_value,
                                'caveat': "Wealth Summary 'Fidelity accounts' total for "
                                          + (months[-1]['label'] if months else 'the latest month')
                                          + ' — matches the sheet and the Finance Google Sheet.'},
            'gain_last_month': {'value': gain_last_month, 'pct': gain_last_month_pct,
                                'month': gain_month_label, 'prev_month': gain_prev_label,
                                'stale_months': gain_stale_months,
                                'ytd': gain_ytd, 'ytd_pct': gain_ytd_pct,
                                'ytd_from': gain_ytd_from,
                                'ytd_caveat': (f'Change from {gain_ytd_from} to {gain_month_label}. '
                                               'Same investment-account total as the monthly figure, '
                                               'so it includes contributions as well as market movement.'
                                               if gain_ytd is not None else None),
                                'series': gain_series,
                                'caveat': (f'{gain_month_label} vs {gain_prev_label} — the most recent month '
                                           'with fresh data. Investment-account change, includes contributions.'
                                           + (f' The {gain_stale_months} month(s) since carry the same figure '
                                              'forward in the Wealth Summary.' if gain_stale_months else ''))
                                          if gain_month_label else
                                          'Month-on-month investment-account change (includes contributions).'},
            'trading_profit': {'value': round(profit_ytd, 2), 'year': profit_year,
                               'sells': sells_in_year},
            'monthly_dividend': {'value': monthly_dividend,
                                 'funds': round(funds_monthly_income, 2),
                                 'shares': round(share_income_monthly, 2),
                                 'caveat': (f'Income paid out: income funds £{funds_monthly_income:,.0f}/mo '
                                            f'+ stocks & shares dividends £{share_income_monthly:,.0f}/mo. '
                                            'Reinvested Acc-fund income is the separate Accumulative card, '
                                            'and both are counted in Total Income.')},
            # Total income (user request 2026-07-19): everything the portfolio pays
            # out or reinvests in a month — Monthly Dividend (income funds + share
            # dividends) plus the Accumulative fund income, so the two metrics beside
            # it add up to this one and the split stays visible.
            'total_income': {'value': total_income_monthly,
                             'monthly_dividend': monthly_dividend,
                             'funds': round(funds_monthly_income, 2),
                             'shares': round(share_income_monthly, 2),
                             'accumulative': share_accum_monthly,
                             'annual': round(total_income_monthly * 12, 2),
                             'ytd': income_ytd, 'ytd_months': income_ytd_months,
                             'ytd_measured': income_ytd_actual is not None,
                             'ytd_months_detail': [
                                 {'period': m.get('period'), 'total': m.get('total')}
                                 for m in _actual_rows],
                             'ytd_caveat': (
                                 (f'Measured: {income_ytd_months} actual month(s) of {now.year} from the '
                                  'spending pivots — income funds + share dividends + reinvested Acc income.'
                                  + (' The latest month is still part-way through.'
                                     if income_ytd_partial else ''))
                                 if income_ytd_actual is not None else
                                 (f'{income_ytd_months} month(s) of {now.year} at the current '
                                  f'£{total_income_monthly:,.0f}/mo rate — a run-rate, not '
                                  'measured month-by-month actuals.')),
                             'caveat': (f'Income funds £{funds_monthly_income:,.0f} + shares '
                                        f'£{share_income_monthly:,.0f} + accumulative £{share_accum_monthly:,.0f} '
                                        '— income paid out plus income reinvested inside the Acc funds.')},
            'strategic_value': {'value': round(strategic_value, 2),
                                'count': strategic_count,
                                'pct_of_portfolio': (round(strategic_value / portfolio_value * 100.0, 2)
                                                     if portfolio_value else None),
                                'caveat': (f'{strategic_count} holding(s) on the Investments sheet with '
                                           "Type = 'Strategic', valued from the broker export. "
                                           'Short-term share holdings and income funds are excluded.')},
            'cash_available': {'value': round(cash_available, 2),
                               'pct_of_portfolio': round(cash_available / portfolio_value * 100.0, 2) if portfolio_value else None,
                               'rows': cash_rows, 'caveat': cash_source + '.'},
        },
        'value_over_time': value_over_time,
        'accounts': {'month': accounts_month, 'prev_month': accounts_prev_month,
                     'rows': accounts,
                     'total': round(sum(a['value'] or 0 for a in accounts), 2),
                     'total_increase': round(sum(a['increase'] or 0 for a in accounts), 2)},
        'chart_stats': chart_stats,
        'alert_status': {'below': alert_below, 'near': alert_near,
                         'above': alert_above, 'total': alert_below + alert_near + alert_above},
    }
    # ---------------- Commodities (Portfolio sub-screen, user request 2026-07-20) ----
    # One card per commodity the pipeline captures, read straight off the Investments
    # sheet (which owns the captured price + alert levels) with the day-over-day %
    # move from history.db. history.db keys commodities by the TradingView SYMBOL
    # (GOLD/SILVER/COPPER1!/USOIL/BRENT/PLATINUM/PALLADIUM), NOT the sheet ticker, so
    # each def carries its history key explicitly. Holdings come from the same joined
    # positions the equity holdings use, falling back to the workbook Holdings column;
    # a commodity we don't hold reads None. Order is the user's requested order.
    _hold_by_tk = {}
    for iv in investments:
        k = str(iv['ticker']).strip().upper() if iv['ticker'] else None
        if k:
            _hold_by_tk[k] = (_hold_by_tk.get(k) or 0.0) + (iv['holdings'] or 0.0)
    COMMODITY_DEFS = [
        ('Gold', 'GOLD', 'GOLD'), ('Silver', 'SLVR', 'SILVER'),
        ('Copper', 'COPP', 'COPPER1!'), ('WTI Oil', 'OIL', 'USOIL'),
        ('Brent Oil', 'UKOIL', 'BRENT'), ('Platinum', 'PLAT', 'PLATINUM'),
        ('Palladium', 'PALL', 'PALLADIUM'),
    ]
    commodities = []
    for label, tk, hist_key in COMMODITY_DEFS:
        r = _inv_row(tk)
        price = _price_for(tk, inv.cell(r, I_CURPRICE).value if r else None, latest)
        chg = (changes.get(hist_key) or changes.get(tk.upper()) or {}) if changes else {}
        hold = _hold_by_tk.get(tk.upper())
        if hold is None and r:
            hold = _num(inv.cell(r, I_HOLDINGS).value)
        commodities.append({
            'label': label, 'ticker': tk,
            'price': price, 'price_unit': _price_unit(tk),
            'change': chg.get('change'), 'change_pct': chg.get('change_pct'),
            'holdings': round(hold, 2) if hold else None,
            'alert_low': _num(inv.cell(r, I_ALERT_LOW).value) if r else None,
            'alert_high': _num(inv.cell(r, I_ALERT_HIGH).value) if r else None,
            'chart_url': _tv_url(invf.cell(r, I_TV).value) if r else None,
            **_layout_for(tk),
        })

    portfolio = {'generated_at': now.isoformat(timespec='seconds'),
                 'investments': investments, 'income_funds': income_funds,
                 'commodities': commodities,
                 'income_positions': income_positions,
                 'income_positions_total': income_positions_total,
                 'income_summary': {'year': now.year,
                                    'total_annual_income': funds_annual_income,
                                    # Total income, NOT the paid-out-only Monthly
                                    # Dividend — this read the same figure before
                                    # accumulation was split out of that metric, and
                                    # narrowing it here would silently drop the
                                    # reinvested income from the Portfolio screen.
                                    'total_monthly_income': total_income_monthly,
                                    'fund_count': len(income_funds)}}
    historic = {'generated_at': now.isoformat(timespec='seconds'),
                'sold': sold,
                'summary': {'total_profit': round(profit_ytd, 2), 'year': profit_year,
                            'count': len(sold), 'win_rate': win_rate}}
    watchlist_payload = {'generated_at': now.isoformat(timespec='seconds'),
                         'criterion': 'Within 5% of alert low (Stocks of Interest — at lower boundary)',
                         'rows': watchlist}

    # ---------------- Targets (income + allocation by Type) ----------------
    targets = {
        'generated_at': now.isoformat(timespec='seconds'),
        'income_per_month': total_income_monthly,   # unchanged by the dividend split
        'annual_income': funds_annual_income,
        'allocation': [
            {'label': 'Income Funds', 'value': round(funds_value_total, 2)},
            {'label': 'Short-Term (Shares)', 'value': round(short_term_value, 2)},
            {'label': 'Strategic', 'value': round(strategic_value, 2)},
            {'label': 'Cash', 'value': round(cash_available, 2)},
        ],
    }
    targets['total'] = round(sum(a['value'] for a in targets['allocation']), 2)

    # ---------------- Relevant News (dividend events for what we hold) ------------
    # Each row DESCRIBES the event (user request 2026-07-19) — a date and an amount
    # alone don't say what is happening or what to do about it.
    news = []
    today = now.date()

    def _describe_exdiv(when, pence, yld):
        d = _tx_date(when) or None
        if d is None:
            try:
                d = datetime.date.fromisoformat(str(when)[:10])
            except ValueError:
                d = None
        amt = f'{pence:.2f}p per share' if isinstance(pence, (int, float)) else 'amount not published'
        ytxt = f' ({yld:.2f}% yield)' if isinstance(yld, (int, float)) else ''
        if d is None:
            return 'Ex-dividend date not published.', None
        days = (d - today).days
        if days > 0:
            return (f'Goes ex-dividend in {days} day{"s" if days != 1 else ""} — {amt}{ytxt}. '
                    f'Hold through {d.strftime("%d %b %Y")} to qualify.'), days
        return (f'Last went ex-dividend {d.strftime("%d %b %Y")} ({-days} days ago) — {amt}{ytxt}. '
                f'Next date not yet published in Base Data.'), days

    # One row per TICKER, not per position — the same stock held in three accounts
    # is still one dividend event (Aviva was listed three times).
    by_ticker = {}
    for iv in investments:
        tk = str(iv['ticker']).strip().upper() if iv['ticker'] else None
        if not tk:
            continue
        g = by_ticker.setdefault(tk, {'name': iv['name'], 'ticker': iv['ticker'], 'holdings': 0.0})
        g['holdings'] += iv['holdings'] or 0.0
    for iv in by_ticker.values():              # equity/commodity holdings
        b = base.get(str(iv['ticker']).strip().upper()) if iv['ticker'] else None
        if b and b.get('ex_div'):
            desc, days = _describe_exdiv(b['ex_div'], b.get('div_pence'), b.get('div_yield_pct'))
            news.append({'name': iv['name'], 'ticker': iv['ticker'], 'event': 'Ex-dividend',
                         'description': desc, 'days_until': days, 'past': (days is not None and days < 0),
                         'date': b['ex_div'], 'amount_pence': b.get('div_pence'),
                         'div_yield_pct': b.get('div_yield_pct'), 'holding': iv['holdings']})

    # Income funds: describe the ACTUAL last distribution (from the transaction
    # export), rolled up across the family's accounts in that fund.
    by_fund = {}
    for pos in income_positions:
        g = by_fund.setdefault(pos['name'], {'paid': 0.0, 'n': 0, 'pay': None, 'ex': None,
                                             'holdings': 0.0})
        g['holdings'] += pos['holdings'] or 0.0
        if pos['last_income']:
            g['paid'] += pos['last_income']
            g['n'] += 1
            g['pay'] = max(g['pay'] or '', pos['payment_date'] or '')
            g['ex'] = max(g['ex'] or '', pos['ex_div_date'] or '')
    for fname, g in by_fund.items():
        if g['n'] and g['pay']:
            d = datetime.date.fromisoformat(g['pay'])
            exd = (' (ex-div ' + datetime.date.fromisoformat(g['ex']).strftime('%d %b') + ')') if g['ex'] else ''
            desc = (f'Monthly income of £{g["paid"]:,.2f} paid into cash on '
                    f'{d.strftime("%d %b %Y")}{exd}, across {g["n"]} holding'
                    f'{"s" if g["n"] != 1 else ""}. Next distribution due next month.')
        else:
            desc = ('Accumulation units — income is reinvested into the fund rather than '
                    'paid out, so there is no cash distribution to expect.')
            d = None
        news.append({'name': fname, 'ticker': None, 'event': 'Monthly income',
                     'description': desc, 'past': False,
                     'date': g['pay'], 'amount_pounds': round(g['paid'], 2) if g['n'] else None,
                     'div_yield_pct': (round(g['paid'] * 12 / g['holdings'] * 100.0, 2)
                                       if (g['n'] and g['holdings']) else None),
                     'holding': round(g['holdings'], 2)})
    # Soonest first: upcoming events before past ones, undated last.
    news.sort(key=lambda n: (n['date'] is None, 0 if not n.get('past') else 1,
                             n['date'] or ''))
    # ---------------- Dividend calendar (Overview diary widget, user 2026-07-20) ----
    # A flat list of dated dividend events for THIS holding book, so the diary can
    # group them by month and page through the year. Two kinds: 'ex-div' (the equity
    # ex-dividend date from Base Data — hold through it to qualify — and each income
    # fund's last ex-div date) and 'paid' (each income fund's last cash distribution).
    # Base Data carries one ex-div date per stock (usually the most recent/next), so
    # this is a snapshot of the dates known now, not an exhaustive forward schedule.
    calendar = []
    for iv in by_ticker.values():
        b = base.get(str(iv['ticker']).strip().upper()) if iv['ticker'] else None
        if not (b and b.get('ex_div')):
            continue
        try:
            iso = datetime.date.fromisoformat(str(b['ex_div'])[:10]).isoformat()
        except ValueError:
            continue
        calendar.append({'date': iso, 'kind': 'ex-div', 'name': iv['name'],
                         'ticker': iv['ticker'], 'amount_pence': b.get('div_pence'),
                         'div_yield_pct': b.get('div_yield_pct')})
    for fname, g in by_fund.items():
        if g['n'] and g['ex']:
            calendar.append({'date': g['ex'], 'kind': 'ex-div', 'name': fname,
                             'ticker': None, 'amount_pounds': round(g['paid'], 2)})
        if g['n'] and g['pay']:
            calendar.append({'date': g['pay'], 'kind': 'paid', 'name': fname,
                             'ticker': None, 'amount_pounds': round(g['paid'], 2)})
    calendar.sort(key=lambda e: e['date'])
    news_payload = {'generated_at': now.isoformat(timespec='seconds'), 'rows': news,
                    'calendar': calendar}

    # ---------------- Activity (things to do, with links) ----------------
    SHEET_URL = 'https://docs.google.com/spreadsheets/d/1UjAz_QUuh86_e6yq8QJf2veI8IpkRCyVfWaK6maqiyc/'
    tv_by_ticker = {}
    for r in range(4, inv.max_row + 1):
        tkr = inv.cell(r, I_TICKER).value
        url = _tv_url(invf.cell(r, I_TV).value)
        if tkr and url:
            tv_by_ticker[str(tkr).strip().upper()] = url
    unmarked, inherited = [], []
    cr_path = os.path.join(SCRIPT_DIR, 'channel_results_tmp.json')
    if os.path.exists(cr_path):
        for rr in json.load(open(cr_path, encoding='utf-8')):
            reason = (rr.get('reason') or '').lower()
            t = rr.get('ticker')
            if 'no channel or trend line found near price' in reason:
                key = str(t).strip().upper()
                unmarked.append({'ticker': t, 'chart_url': tv_by_ticker.get(key),
                                 **_layout_for(t)})
            elif 'axis' in reason:
                inherited.append(t)
    below_tickers = [w['ticker'] for w in watchlist if (w.get('proximity_pct') is not None and w['proximity_pct'] <= 0)]

    # Group the unmarked charts by their SAVED LAYOUT (user request 2026-07-20): the
    # user marks a whole layout at a time in TradingView, so the dashboard shows the
    # LAYOUTS that need marking up rather than every individual ticker. Each entry
    # opens that layout in TradingView Desktop (chart_id), carries the count and the
    # tickers within it that need drawing. Tickers with no captured layout become
    # single-ticker entries (layout=None) keeping their browser link.
    _by_layout = {}
    layouts_to_markup = []
    for u in unmarked:
        key = u.get('layout_id') or u.get('layout')
        if key:
            g = _by_layout.get(key)
            if g is None:
                g = {'layout': u.get('layout'), 'layout_id': u.get('layout_id'),
                     'chart_id': u.get('chart_id'), 'chart_url': u.get('chart_url'),
                     'tickers': [], 'count': 0}
                _by_layout[key] = g
                layouts_to_markup.append(g)
            if u.get('chart_id') and not g.get('chart_id'):
                g['chart_id'] = u['chart_id']
            g['count'] += 1
            if u.get('ticker'):
                g['tickers'].append(u['ticker'])
        else:
            layouts_to_markup.append({'layout': None, 'layout_id': None,
                                      'chart_id': u.get('chart_id'), 'chart_url': u.get('chart_url'),
                                      'tickers': [u['ticker']] if u.get('ticker') else [], 'count': 1})
    # Grouped layouts first (most charts to draw first), then the loose no-layout ones.
    layouts_to_markup.sort(key=lambda x: (x['layout'] is None, -x['count'], (x['layout'] or '')))

    def act(category, priority, title, detail, link=None, link_label=None):
        return {'category': category, 'priority': priority, 'title': title,
                'detail': detail, 'link': link, 'link_label': link_label}

    actions = [
        act('Charts', 'high', f'{len(unmarked)} charts have no trend lines drawn',
            'Mark up the channel / trend lines in TradingView so they feed alert levels. See the list below.'),
        act('Charts', 'medium', 'Price extended beyond the trend line / channel',
            'New pattern (added 2026-07-17): the drawn line stops before today. None detected this run — recently refreshed.'),
        act('Charts', 'medium', f'{len(inherited)} charts on inherited levels',
            'Axis read failed this run, so the Alert Low/High are carried over — redraw or re-check these in TradingView.'),
        act('Buys', 'high', f'{len(below_tickers)} stocks at or below alert low',
            'Buy candidates: ' + (', '.join(str(t) for t in below_tickers) or 'none') + '. Review the Watchlist.',
            '#watchlist', 'Open Watchlist'),
        act('Data', 'high', 'Review the figures I changed',
            'Confirm the July Wealth Summary fix (Joint £791,992) and WS Guinness value (£111,655) look right.'),
        act('Data', 'medium', 'Refresh bank exports',
            'Amex (activity.csv) + Barclays (data.csv) are not in Downloads, so spending/wealth tabs go stale past ~6 weeks.'),
        act('Data', 'low', 'Sync the workbook to the Finance Google Sheet', 'Push the latest master workbook to Google Sheets.',
            SHEET_URL, 'Open sheet'),
        act('Code', 'medium', 'Answer any open questions in Claude Code',
            'I may be waiting on a decision — check the Claude Code session.'),
        act('Code', 'low', 'Commit & push pending changes',
            'Lock in the latest dashboard/data changes when you are happy with them.'),
    ]
    activity = {'generated_at': now.isoformat(timespec='seconds'),
                'actions': actions, 'charts_to_markup': unmarked,
                'layouts_to_markup': layouts_to_markup}

    return {'overview': overview, 'portfolio': portfolio, 'historic': historic,
            'watchlist': watchlist_payload, 'targets': targets, 'news': news_payload,
            'activity': activity, 'payslips': build_payslips(workbook)}


def write(out_dir=OUT_DIR, workbook=WORKBOOK):
    os.makedirs(out_dir, exist_ok=True)
    data = build(workbook)
    for name, payload in data.items():
        with open(os.path.join(out_dir, name + '.json'), 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, default=str)
    return data


if __name__ == '__main__':
    d = write()
    ov = d['overview']['metrics']
    print('Wrote dashboard JSON ->', OUT_DIR)
    print(f"  Portfolio value : £{ov['portfolio_value']['value']:,.0f}")
    print(f"  Gain last month : £{(ov['gain_last_month']['value'] or 0):,.0f}")
    print(f"  {ov['trading_profit']['year']} sells profit: £{ov['trading_profit']['value']:,.0f} "
          f"({ov['trading_profit']['sells']} sells)")
    print(f"  Monthly dividend: £{ov['monthly_dividend']['value']:,.0f}")
    print(f"  Cash available  : £{ov['cash_available']['value']:,.0f}")
    print(f"  Holdings: {len(d['portfolio']['investments'])} investments, {len(d['portfolio']['income_funds'])} income funds")
    print(f"  Historic sales  : {d['historic']['summary']['count']} (win rate {d['historic']['summary']['win_rate']}%)")
    print(f"  Alert status    : {d['overview']['alert_status']}")
    ps = d['payslips']
    print(f"  Payslips        : {len(ps['rows'])} across {len(ps['tax_years'])} tax years "
          f"(current {ps['current_tax_year']})")
