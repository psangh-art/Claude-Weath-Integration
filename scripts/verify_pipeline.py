#!/usr/bin/env python3
"""Post-run verification report for the full TradingView -> Excel -> Stocks Buy
Strategy pipeline. Run this after every extract (run_full_pipeline.js does this
automatically as its last step) to get a plain answer to:

  - Were all layouts captured, and did every chart get a real screenshot?
  - Could the upper/lower parallel-channel boundaries be read for each chart?
  - Were all alerts extracted (cross-checked against TradingView live, when reachable)?
  - Did tradingview_layouts.xlsx actually get all this data, sheet by sheet?
  - Was Stocks_Buy_Strategy.xlsx actually updated, and with what?

Never hides a gap to make the run look more complete than it was — an honest
"0 channel reads succeeded (Tesseract not installed)" is the whole point of this
report existing, per the pipeline's own "silence over guessing" principle.

Usage: python verify_pipeline.py [--live-alert-check]
Writes a report to logs/verify_<timestamp>.md and logs/latest-verify.md, and also
prints a summary to stdout.
"""
import os
import sys
import json
from datetime import datetime
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
from config import CFG, downloads_dir
DOWNLOADS = downloads_dir()

CHARTS_MANIFEST = os.path.join(SCRIPT_DIR, 'layout_manifest_tmp.json')
INDICATORS_MANIFEST = os.path.join(SCRIPT_DIR, 'indicator_manifest_tmp.json')
ALERTS_MANIFEST = os.path.join(SCRIPT_DIR, 'alerts_manifest_tmp.json')
CHANNEL_RESULTS = os.path.join(SCRIPT_DIR, 'channel_results_tmp.json')
MASTER_UPDATE_RESULT = os.path.join(SCRIPT_DIR, 'master_update_result_tmp.json')
WORKBOOK_PATH = os.path.join(DOWNLOADS, CFG['layoutsWorkbook'])
MASTER_SHEET_PATH = os.path.join(DOWNLOADS, CFG['masterWorkbook'])
LOGS_DIR = os.path.join(REPO_ROOT, 'logs')


def load_json(path):
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_live_alert_count():
    """Best-effort live cross-check via a short-lived node subprocess. Returns
    None (not a failure) if TradingView isn't reachable right now — this check
    is a nice-to-have, not a hard requirement for the rest of the report."""
    import subprocess
    # src/connection.js keeps an open CDP WebSocket that holds the event loop open
    # forever, so this subprocess must force-exit itself once done rather than
    # relying on the timeout below (which would otherwise fire every single time).
    node_script = (
        "import('../src/core/alerts.js').then(async (m) => {"
        "  try { const r = await m.list(); console.log(JSON.stringify({count: r.alert_count})); }"
        "  catch (e) { console.log(JSON.stringify({error: e.message})); }"
        "  process.exit(0);"
        "});"
    )
    tmp_js = os.path.join(SCRIPT_DIR, '_live_alert_check_tmp.mjs')
    try:
        with open(tmp_js, 'w', encoding='utf-8') as f:
            f.write(node_script)
        result = subprocess.run(['node', tmp_js], cwd=SCRIPT_DIR, capture_output=True, text=True, timeout=20)
        out = json.loads(result.stdout.strip().splitlines()[-1]) if result.stdout.strip() else {}
        return out.get('count')
    except Exception:
        return None
    finally:
        if os.path.exists(tmp_js):
            os.remove(tmp_js)


def section_layouts(charts):
    lines = ["## Layouts & Charts\n"]
    by_layout = defaultdict(list)
    for row in charts:
        by_layout[(row['id'], row['chartId'], row['name'])].append(row)

    total_layouts = len(by_layout)
    total_charts = len(charts)
    ok_charts = sum(1 for r in charts if r.get('screenshot') and not r.get('error'))
    stale_charts = sum(1 for r in charts if r.get('screenshot') and r.get('error') and 'stale' in r['error'])
    failed_charts = sum(1 for r in charts if not r.get('screenshot'))

    lines.append(f"- **{total_layouts} layouts**, **{total_charts} charts** total\n")
    lines.append(f"- ✅ {ok_charts} charts captured cleanly\n")
    if stale_charts:
        lines.append(f"- ⚠️ {stale_charts} charts used a stale screenshot from a previous run (re-run to refresh)\n")
    if failed_charts:
        lines.append(f"- ❌ {failed_charts} charts have NO screenshot at all\n")
    lines.append("\n### Per-layout detail\n\n| Layout | Charts | Status |\n|---|---|---|\n")
    for (lid, cid, name), rows in sorted(by_layout.items(), key=lambda kv: kv[0][2]):
        n = len(rows)
        n_ok = sum(1 for r in rows if r.get('screenshot') and not r.get('error'))
        n_fail = sum(1 for r in rows if not r.get('screenshot'))
        n_stale = sum(1 for r in rows if r.get('screenshot') and r.get('error') and 'stale' in r['error'])
        if n_fail:
            status = f"❌ {n_fail}/{n} missing"
        elif n_stale:
            status = f"⚠️ {n_stale}/{n} stale"
        else:
            status = f"✅ {n_ok}/{n} ok"
        lines.append(f"| {name} | {n} | {status} |\n")
    lines.append("\n")
    return "".join(lines), {'total_layouts': total_layouts, 'total_charts': total_charts,
                              'ok_charts': ok_charts, 'stale_charts': stale_charts, 'failed_charts': failed_charts}


def section_channels(charts, channel_results):
    lines = ["## Channel boundaries (upper/lower parallels)\n\n"]
    if channel_results is None:
        lines.append("⚪ **Not run this session** — channel_results_tmp.json doesn't exist. "
                      "This means either the OCR step (Tesseract) isn't installed, or this "
                      "run only did the chart export without the full pipeline. See "
                      "`channel_detect.py`'s docstring for the one-time install steps.\n\n")
        return "".join(lines), {'attempted': 0, 'succeeded': 0, 'rejected': 0}

    attempted = len(channel_results)
    succeeded = [c for c in channel_results if c.get('kind') in ('parallel', 'single_low', 'single_high')]
    rejected = [c for c in channel_results if c not in succeeded]
    parallel = [c for c in succeeded if c.get('kind') == 'parallel']
    single = [c for c in succeeded if c.get('kind') in ('single_low', 'single_high')]

    lines.append(f"- **{attempted} distinct tickers attempted**\n")
    lines.append(f"- ✅ {len(parallel)} had both upper and lower boundaries identified (parallel channel)\n")
    lines.append(f"- ✅ {len(single)} had a single trendline resolved to Alert Low/High by current price\n")
    lines.append(f"- ❌ {len(rejected)} rejected or not detected\n\n")

    if rejected:
        reason_counts = defaultdict(int)
        for r in rejected:
            reason_counts[r.get('reason') or 'unknown'] += 1
        lines.append("### Rejection reasons\n\n| Reason | Count |\n|---|---|\n")
        for reason, count in sorted(reason_counts.items(), key=lambda kv: -kv[1]):
            lines.append(f"| {reason} | {count} |\n")
        lines.append("\n")

    if parallel:
        lines.append("### Tickers with a parallel-channel read\n\n| Ticker | Lower | Upper |\n|---|---|---|\n")
        for c in parallel:
            lines.append(f"| {c['ticker']} | {c['lower']} | {c['upper']} |\n")
        lines.append("\n")

    if single:
        lines.append("### Tickers with a single-trendline read\n\n| Ticker | Kind | Price |\n|---|---|---|\n")
        for c in single:
            price = c['lower'] if c['kind'] == 'single_low' else c['upper']
            lines.append(f"| {c['ticker']} | {c['kind']} | {price} |\n")
        lines.append("\n")

    return "".join(lines), {'attempted': attempted, 'succeeded': len(succeeded), 'rejected': len(rejected)}


def section_alerts(alerts, live_count):
    lines = ["## Alerts\n\n"]
    n = len(alerts) if alerts is not None else 0
    lines.append(f"- {n} alerts in this export ({sum(1 for a in (alerts or []) if a.get('active'))} active)\n")
    if live_count is not None:
        if live_count == n:
            lines.append(f"- ✅ Matches live TradingView alert count ({live_count})\n\n")
        else:
            lines.append(f"- ⚠️ Live TradingView currently shows {live_count} alerts — export has {n}. "
                          f"If this export is not from just now, that's expected (alerts change over time); "
                          f"if it IS from just now, re-run the export.\n\n")
    else:
        lines.append("- ⚪ Live cross-check skipped (TradingView not reachable right now, or check not requested)\n\n")
    return "".join(lines), {'exported': n, 'live': live_count}


def section_workbook(charts, indicators, alerts):
    lines = ["## tradingview_layouts.xlsx\n\n"]
    if not os.path.exists(WORKBOOK_PATH):
        lines.append(f"❌ **File not found** at `{WORKBOOK_PATH}`\n\n")
        return "".join(lines), {'ok': False}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ok = True
        for sheet_name, expected in [('Charts', len(charts)), ('Indicators', len(indicators or [])), ('Alerts', len(alerts or []))]:
            if sheet_name not in wb.sheetnames:
                lines.append(f"❌ Sheet **{sheet_name}** is missing entirely\n")
                ok = False
                continue
            actual = wb[sheet_name].max_row - 1  # minus header
            if actual == expected:
                lines.append(f"✅ **{sheet_name}**: {actual} rows (matches manifest)\n")
            else:
                lines.append(f"⚠️ **{sheet_name}**: {actual} rows in workbook, {expected} in manifest — mismatch\n")
                ok = False
        lines.append("\n")
        return "".join(lines), {'ok': ok}
    except Exception as e:
        lines.append(f"❌ Could not open workbook: {e}\n\n")
        return "".join(lines), {'ok': False}


def section_master_sheet(master_result):
    lines = ["## Stocks_Buy_Strategy.xlsx\n\n"]
    if not os.path.exists(MASTER_SHEET_PATH):
        lines.append(f"⚪ **Not present** at `{MASTER_SHEET_PATH}` — this step is skipped until the file exists.\n\n")
        return "".join(lines), {'ran': False}
    if master_result is None:
        lines.append("⚪ **Not updated this session** — no `master_update_result_tmp.json` found "
                      "(the OCR step likely didn't run, so there was nothing to apply).\n\n")
        return "".join(lines), {'ran': False}

    lines.append(f"- Last updated: {master_result.get('timestamp')}\n")
    lines.append(f"- ✅ Applied: {len(master_result.get('applied', []))}\n")
    lines.append(f"- ❌ Rejected: {len(master_result.get('rejected', []))}\n")
    lines.append(f"- ⚠️ Manual-protected (skipped): {len(master_result.get('skipped_manual', []))}\n")
    lines.append(f"- ⚪ Noise-skipped (within 3%): {len(master_result.get('skipped_noise', []))}\n")
    lines.append(f"- ⚪ Unmatched (no row / not attempted): {len(master_result.get('unmatched', []))}\n\n")

    if master_result.get('applied'):
        lines.append("### Applied this run\n\n| Ticker | Lower | Upper | Alert Low | Alert High |\n|---|---|---|---|---|\n")
        for a in master_result['applied']:
            cells = [a.get('lower'), a.get('upper'), a.get('alert_low'), a.get('alert_high')]
            lower, upper, alert_low, alert_high = ('—' if c is None else c for c in cells)
            lines.append(f"| {a['ticker']} | {lower} | {upper} | {alert_low} | {alert_high} |\n")
        lines.append("\n")

    # Cross-check the two per-run rewrites (commodity captured prices, below-alert
    # block) against what the saved workbook ACTUALLY contains — this is the last
    # gate before the workbook gets imported into the Finance Google Sheet, so a
    # write that silently didn't land must show up here, not there.
    commodity_prices = master_result.get('commodity_prices')
    below_rows = master_result.get('below_alert_rows')
    if commodity_prices is None and below_rows is None:
        return "".join(lines), {'ran': True}  # result JSON predates these fields

    ok = True
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MASTER_SHEET_PATH, data_only=False)

        lines.append(f"### Commodity captured prices ({len(commodity_prices or [])} this run)\n\n")
        inv = wb['Investments']
        ticker_rows = {}
        for r in range(1, inv.max_row + 1):
            t = inv.cell(row=r, column=3).value  # col C = Ticker
            if isinstance(t, str):
                ticker_rows[t.strip().upper()] = r
        for cp in commodity_prices or []:
            r = ticker_rows.get(cp['ticker'].upper())
            actual = inv.cell(row=r, column=9).value if r else None  # col I = Current Price
            if r and isinstance(actual, (int, float)) and abs(actual - cp['price']) < 0.01:
                lines.append(f"- ✅ {cp['ticker']}: {actual} (value, not formula; checked {cp['checked_at']})\n")
            else:
                lines.append(f"- ❌ {cp['ticker']}: expected {cp['price']} in Investments col I, "
                              f"found {actual!r}\n")
                ok = False

        soi = wb['Stocks of Interest']
        # Count real DATA rows, not 'anything in column A'. The block is no longer a
        # single section: 'On Alert' sits above 'Below Alert Low' (2026-07-15), so a
        # second section band and column-header row now fall INSIDE rows 5-38 and both
        # carry column A text. Counting column A alone read 25 where the run wrote 23,
        # failing the run on a block that was correct. A data row is one with a ticker
        # in column B and a numeric price in column F — bands have no column B, and the
        # header row's column F is the string 'Current'. (2026-07-16: the block widened
        # to the section-table schema, so the live price moved from column C to F;
        # column C now holds the Pattern text.)
        actual_rows = sum(1 for r in range(5, 39)
                          if soi.cell(row=r, column=2).value is not None
                          and isinstance(soi.cell(row=r, column=6).value, (int, float)))
        expected_rows = min(len(below_rows or []), 34)
        if actual_rows == expected_rows:
            lines.append(f"\n✅ **Below-alert block**: {actual_rows} rows at the top of "
                          "'Stocks of Interest' (matches this run)\n")
        else:
            lines.append(f"\n❌ **Below-alert block**: {actual_rows} rows in workbook, "
                          f"{expected_rows} expected from this run\n")
            ok = False
        if 'Below Alert Low' in wb.sheetnames:
            lines.append("❌ Obsolete 'Below Alert Low' sheet has reappeared — it was migrated "
                          "into 'Stocks of Interest' on 2026-07-11 and nothing should recreate it\n")
            ok = False
        lines.append("\n")
    except Exception as e:
        lines.append(f"❌ Could not cross-check workbook contents: {e}\n\n")
        ok = False
    return "".join(lines), {'ran': True, 'ok': ok}


def main():
    live_check = '--live-alert-check' in sys.argv

    charts = load_json(CHARTS_MANIFEST) or []
    indicators = load_json(INDICATORS_MANIFEST) or []
    alerts = load_json(ALERTS_MANIFEST) or []
    channel_results = load_json(CHANNEL_RESULTS)
    master_result = load_json(MASTER_UPDATE_RESULT)

    live_alert_count = get_live_alert_count() if live_check else None

    report = [f"# Pipeline Verification Report — {datetime.now().isoformat()}\n\n"]

    if not charts:
        report.append("⚠️ **No chart manifest found** — has `export-layouts-excel.js` / `run_full_pipeline.js` "
                       "been run yet in this repo? Nothing else in this report can be checked.\n")
        report_text = "".join(report)
        print(report_text)
        return

    layouts_text, layouts_stats = section_layouts(charts)
    channels_text, channels_stats = section_channels(charts, channel_results)
    alerts_text, alerts_stats = section_alerts(alerts, live_alert_count)
    workbook_text, workbook_stats = section_workbook(charts, indicators, alerts)
    master_text, master_stats = section_master_sheet(master_result)

    overall_ok = (
        layouts_stats['failed_charts'] == 0
        and workbook_stats['ok']
        and (master_stats['ran'] or not os.path.exists(MASTER_SHEET_PATH))
        and master_stats.get('ok', True)  # workbook cross-check, when the run recorded one
    )
    report.append(f"**Overall: {'✅ PASS' if overall_ok else '⚠️ SEE DETAILS BELOW'}**\n\n---\n\n")
    report.append(layouts_text)
    report.append(channels_text)
    report.append(alerts_text)
    report.append(workbook_text)
    report.append(master_text)

    report_text = "".join(report)

    os.makedirs(LOGS_DIR, exist_ok=True)
    stamp = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
    with open(os.path.join(LOGS_DIR, f'verify_{stamp}.md'), 'w', encoding='utf-8') as f:
        f.write(report_text)
    with open(os.path.join(LOGS_DIR, 'latest-verify.md'), 'w', encoding='utf-8') as f:
        f.write(report_text)

    try:
        print(report_text)
    except UnicodeEncodeError:
        # Windows consoles often default to a cp1252-family codepage that can't
        # encode the checkmark/warning emoji used in the report. The file was
        # already written as UTF-8 above; fall back to an ASCII-safe console echo.
        sys.stdout.buffer.write(report_text.encode('utf-8', errors='replace'))
        sys.stdout.buffer.write(b'\n')
    print(f"\nReport saved to logs/verify_{stamp}.md and logs/latest-verify.md")


if __name__ == '__main__':
    main()
