#!/usr/bin/env python3
"""Apply channel-detection results from a TradingView chart export into the master
'Stocks Buy Strategy.xlsx' sheet, following the rules in
Claude_Code_Handoff_Instructions.md sections 3-4-5-9. Never writes a value it isn't
confident in — silence (leaving a row untouched) is always preferable to a wrong
number in a live trading sheet.

Usage:
  python update_master_sheet.py <master_in.xlsx> <charts_manifest.json> \
      <channel_results.json> <master_out.xlsx> <feedback_md_path>

charts_manifest.json: Charts-sheet rows from export-layouts-excel.js
  [{"id", "chartId", "name", "ticker", "description", "screenshot", "error"}, ...]
channel_results.json: output of `channel_detect.py --batch`
  [{"ticker", "screenshot", "lower", "upper", "x_frac", "reason"}, ...]
"""
import sys
import json
import re
from datetime import date
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ticker_normalize import normalize, master_tickers_match
from refresh_soi_sections import pattern_label

SHEET_NAME = 'Investments'
# A new 'Marked Up' column was inserted at B (2026-07-16), shifting every data
# column one to the right — see insert_marked_up_col_2026-07-16.py. All constants
# below moved +1 except COL_CHART (col A, unchanged).
COL_CHART = 1
COL_MARKED_UP = 2          # 'Yes'/'No' — has the user drawn channel/trend lines on the chart
COL_SHARE_NAME = 3
COL_TICKER = 4
COL_HOLDINGS = 5
COL_TARGET_VALUE = 7
COL_CURRENT_PRICE = 10
COL_ALERT_LOW = 13
COL_ALERT_LOW_SOURCE = 14
COL_ALERT_HIGH = 16
COL_CLAUDE_NOTES = 32


def marked_up_flag(detection):
    """'Yes' if the user has drawn channel/trend lines on this chart, else 'No'
    (Investments column B, user request 2026-07-16). A detected pattern OR an axis
    read that failed both mean drawings ARE present — a failed axis read doesn't
    remove the user's lines. 'no channel or trend line found near price', a
    macro/reference instrument, or no capture at all are the only No cases."""
    if not detection:
        return 'No'
    if detection.get('kind'):
        return 'Yes'
    reason = (detection.get('reason') or '').lower()
    if 'macro/reference' in reason:
        return 'No'
    if 'axis' in reason:            # drawings present; the axis just couldn't be OCR'd
        return 'Yes'
    return 'No'
HEADER_ROW = 2
LAST_CHECKED_HEADER = 'Chart Last Checked'

CHART_YES_FONT = 'FF276221'
CHART_YES_FILL = 'FFC6EFCE'
CHART_NO_FONT = 'FF666666'
CHART_NO_FILL = 'FFF2F2F2'

REFRESH_NOISE_THRESHOLD = 0.03  # don't rewrite if new Alert Low is within 3% of existing

ALERT_LOW_BUFFER = 1.05  # Alert Low sits 5% above the support line, as an early warning


def is_noise_refresh(ws, row, existing_source, alert_low, on_alert):
    """True when a re-read is close enough to what's already there to leave alone.

    The threshold exists to stop harmless churn, but it used to gate the whole row
    update — so a row already holding a BROKEN pair (Alert Low >= Alert High, as 8
    rows did after the unclamped buffer shipped) could never be repaired: the new
    Alert Low landed within 3% of the bad one, the row was skipped, and the stale
    inversion survived every subsequent run. A row is only 'noise' if what's already
    in the sheet is coherent; an inverted pair, or a level price has now reached, is
    always rewritten."""
    if on_alert:
        return False
    if existing_source != 'Auto' or not isinstance(alert_low, (int, float)):
        return False
    existing_low = ws.cell(row=row, column=COL_ALERT_LOW).value
    if not isinstance(existing_low, (int, float)) or not existing_low:
        return False
    existing_high = ws.cell(row=row, column=COL_ALERT_HIGH).value
    if isinstance(existing_high, (int, float)) and existing_high and existing_low >= existing_high:
        return False        # currently inverted — repair it rather than call it noise
    return abs(alert_low - existing_low) / existing_low <= REFRESH_NOISE_THRESHOLD


def buffered_alert_low(lower, upper, on_alert=False, price=None):
    """Alert Low IS the drawn support line — no buffer (user decision 2026-07-16).

    The ×1.05 early-warning buffer was RETIRED. Reviewing the live sheet the user
    corrected GLEN/NG/ADM to sit AT the drawn line, not 5% above it (GLEN 544->518,
    NG 1278->1183, ADM 3687->3510) and asked for the same everywhere: Alert Low is
    the line itself. `on_alert`/`price` are no longer used (kept in the signature so
    the call sites are untouched).

    The CLAMP guard stays: Alert Low must never reach or cross Alert High. A fresh
    low meeting a stale high once shipped inverted pairs (ICG 1874.35 vs a stale
    1874.2), so if the level would touch/cross Alert High it sits just under it.
    """
    if lower is None:
        return None
    alert_low = lower
    if upper is not None and alert_low >= upper:
        # Keep a usable band: sit just under Alert High rather than crossing it.
        alert_low = min(alert_low, upper * 0.999)
    return round(alert_low, 2)


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


_HEADER_NAVY = 'FF1F3864'
_THIN_SIDE = Side(style='thin')
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def get_or_create_last_checked_col(ws):
    """Find the 'Chart Last Checked' column by header text, or append a new one
    at the end of the sheet if it doesn't exist yet. Stable across runs once
    created — always looked up by header, never assumed to be a fixed index,
    since this column was added after the sheet's original layout.

    The header cell is (re)styled every run to match the sheet's other headers
    (Arial 9 bold white on navy, centered, wrapped, bordered) — the column was
    originally created bare (AK2 flagged by the user 2026-07-12), and doing it
    idempotently here means a future recreation can never regress it."""
    max_col = ws.max_column
    col = None
    for c in range(1, max_col + 1):
        if ws.cell(row=HEADER_ROW, column=c).value == LAST_CHECKED_HEADER:
            col = c
            break
    if col is None:
        col = max_col + 1
        ws.cell(row=HEADER_ROW, column=col, value=LAST_CHECKED_HEADER)
    header = ws.cell(row=HEADER_ROW, column=col)
    header.font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
    header.fill = PatternFill(fill_type='solid', fgColor=_HEADER_NAVY)
    header.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    header.border = _CELL_BORDER
    return col


def stamp_last_checked(ws, row, col, today):
    """Write the 'Chart Last Checked' date with the sheet's body styling (Arial 9,
    bordered, centered) so the column reads as part of the table, not an add-on."""
    cell = ws.cell(row=row, column=col, value=today)
    cell.font = Font(name='Arial', size=9)
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = _CELL_BORDER


def master_index_key(ticker):
    """The index key for a ticker as written in the sheet.

    This MUST agree with ticker_normalize.normalize()'s master_ticker, or the row can
    never be matched. It used to be a blunt .replace('.', '-'), which is right for a
    BT.A-style CLASS SUFFIX ('BT.A' -> 'BT-A') but wrong for the bare trailing dot
    TradingView puts on some LSE symbols: normalize() STRIPS that ('AT.' -> 'AT') while
    the index produced 'AT-', so the two never met.

    It hid because 'AT.' (Ashtead Technology, Investments row 111) is the only ticker in
    the sheet stored WITH the trailing dot — its siblings SN, AV, NG, UU and QQ are all
    stored bare. The row silently fell out of matching every run: reported 'no existing
    row in master sheet', its Alert Low left at a stale 429.87 while the chart read
    400.27, and — the tell — no 'Chart Last Checked' stamp at all, unlike every other
    processed row. Found by the test-analyst audit 2026-07-15. Of 355 indexed rows this
    changes exactly one key, AT- -> AT."""
    key = str(ticker).strip().upper()
    if key.endswith('.') and len(key) > 1:
        return key[:-1]
    return key.replace('.', '-')


def build_master_index(ws):
    """Map normalized master ticker -> row number, skipping section-header rows
    (rows with no Chart Yes/No value in column A aren't real ticker rows)."""
    index = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        chart_val = ws.cell(row=r, column=COL_CHART).value
        ticker = ws.cell(row=r, column=COL_TICKER).value
        if chart_val not in ('Yes', 'No') or not ticker:
            continue
        index[master_index_key(ticker)] = r
    return index


def set_chart_flag(ws, row, is_yes):
    cell = ws.cell(row=row, column=COL_CHART)
    cell.value = 'Yes' if is_yes else 'No'
    font = copy(cell.font)
    font.color = openpyxl.styles.colors.Color(rgb=CHART_YES_FONT if is_yes else CHART_NO_FONT)
    cell.font = font
    fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=CHART_YES_FILL if is_yes else CHART_NO_FILL)
    cell.fill = fill


def append_note(ws, row, note):
    cell = ws.cell(row=row, column=COL_CLAUDE_NOTES)
    existing = cell.value
    cell.value = f"{existing} | {note}" if existing else note


def process(master_ws, charts, channel_by_ticker):
    """Returns (applied, rejected, skipped_manual, skipped_noise, unmatched, matches)
    — the first five are lists of dicts describing what happened for each charted
    ticker, for the feedback log; `matches` records every charted ticker that found
    a master row, with its captured live price, so main() can rebuild the
    below-alert table after all Alert Low/High updates have landed.
    Every row that was actually looked at this run (applied, rejected, manual-skipped,
    or noise-skipped) gets a 'Chart Last Checked' timestamp — 'unmatched' rows are
    NOT stamped, since those were never actually attempted against this ticker."""
    index = build_master_index(master_ws)
    col_last_checked = get_or_create_last_checked_col(master_ws)
    today = date.today().isoformat()

    applied, rejected, skipped_manual, skipped_noise, unmatched = [], [], [], [], []
    matches = []
    seen_tickers = set()

    for row in charts:
        raw_ticker = row.get('ticker')
        if not raw_ticker:
            continue
        norm = normalize(raw_ticker, row.get('description'))
        if norm is None or norm['kind'] == 'macro_excluded' or norm['master_ticker'] is None:
            continue
        master_ticker = norm['master_ticker']
        if master_ticker in seen_tickers:
            continue  # same ticker may appear in multiple layouts; only process once per run
        seen_tickers.add(master_ticker)

        company = row.get('description') or raw_ticker
        detection = channel_by_ticker.get(raw_ticker) or channel_by_ticker.get(master_ticker)

        master_row = None
        for key, r in index.items():
            if master_tickers_match(key, master_ticker):
                master_row = r
                break

        if master_row is None:
            unmatched.append({'ticker': master_ticker, 'company': company, 'reason': 'no existing row in master sheet'})
            continue

        # Carry on_alert on the match: the detection is keyed by the CHART ticker
        # (e.g. 'SILVER') while matches are keyed by the MASTER ticker ('SLVR'), so
        # re-looking the detection up downstream by master ticker silently misses.
        matches.append({'ticker': master_ticker, 'company': company, 'row': master_row,
                        'price': row.get('price'), 'checked_at': row.get('priceCheckedAt'),
                        'chart_id': row.get('chartId'),
                        'on_alert': bool((detection or {}).get('on_alert')),
                        'detection': detection})
        # Column B: has the user marked up this chart? Maintained every run.
        master_ws.cell(row=master_row, column=COL_MARKED_UP, value=marked_up_flag(detection))

        # Commodities can't be priced by GOOGLEFINANCE at all any more (verified
        # 2026-07-11: TVC: and CURRENCY:XAU/XAG/XPT/XPD all return #N/A), so their
        # Current Price cell gets the TradingView-captured live price written as a
        # VALUE — replacing the dead formula — with 'Chart Last Checked' as its
        # freshness stamp. Equity/index rows keep their working formulas.
        if norm['kind'] == 'commodity' and isinstance(row.get('price'), (int, float)):
            master_ws.cell(row=master_row, column=COL_CURRENT_PRICE, value=round(row['price'], 2))
            stamp_last_checked(master_ws, master_row, col_last_checked, today)
            matches[-1]['commodity_price_written'] = True

        if detection is None:
            unmatched.append({'ticker': master_ticker, 'company': company, 'reason': 'not yet attempted (no channel-detection result for this run)'})
            continue

        if detection.get('reason'):
            rejected.append({'ticker': master_ticker, 'company': company, 'reason': detection['reason']})
            stamp_last_checked(master_ws, master_row, col_last_checked, today)
            continue

        existing_source = master_ws.cell(row=master_row, column=COL_ALERT_LOW_SOURCE).value
        if existing_source == 'Manual':
            skipped_manual.append({'ticker': master_ticker, 'company': company})
            stamp_last_checked(master_ws, master_row, col_last_checked, today)
            continue

        kind = detection.get('kind', 'parallel')
        lower, upper = detection['lower'], detection['upper']
        on_alert = bool(detection.get('on_alert'))

        if kind == 'parallel':
            alert_low = buffered_alert_low(lower, upper, on_alert, row.get('price'))
            alert_high = upper

            if is_noise_refresh(master_ws, master_row, existing_source, alert_low, on_alert):
                skipped_noise.append({'ticker': master_ticker, 'company': company,
                                      'existing_low': master_ws.cell(row=master_row, column=COL_ALERT_LOW).value,
                                      'new_low': alert_low})
                stamp_last_checked(master_ws, master_row, col_last_checked, today)
                continue

            master_ws.cell(row=master_row, column=COL_ALERT_LOW, value=alert_low)
            master_ws.cell(row=master_row, column=COL_ALERT_LOW_SOURCE, value='Auto')
            master_ws.cell(row=master_row, column=COL_ALERT_HIGH, value=alert_high)
            set_chart_flag(master_ws, master_row, True)
            append_note(master_ws, master_row,
                        f"Auto: Alert Low {alert_low}, Alert High {alert_high} ({today})")
            applied.append({'ticker': master_ticker, 'company': company, 'lower': lower, 'upper': upper,
                             'alert_low': alert_low, 'alert_high': alert_high})

        elif kind == 'single_low':
            # No parallel channel — a single trendline sat below the current price,
            # so it's used as Alert Low only. Alert High is left completely
            # untouched (not cleared, not guessed).
            # Clamp against whatever Alert High is ALREADY in the sheet: this run
            # only re-read the support side, but the buffer can still push Alert Low
            # past a previously-written Alert High (HIK, DCC, SSE all inverted this
            # way — a fresh low meeting a stale high).
            existing_high = master_ws.cell(row=master_row, column=COL_ALERT_HIGH).value
            alert_low = buffered_alert_low(
                lower, existing_high if isinstance(existing_high, (int, float)) and existing_high else None,
                on_alert, row.get('price'))

            if is_noise_refresh(master_ws, master_row, existing_source, alert_low, on_alert):
                skipped_noise.append({'ticker': master_ticker, 'company': company,
                                      'existing_low': master_ws.cell(row=master_row, column=COL_ALERT_LOW).value,
                                      'new_low': alert_low})
                stamp_last_checked(master_ws, master_row, col_last_checked, today)
                continue

            master_ws.cell(row=master_row, column=COL_ALERT_LOW, value=alert_low)
            master_ws.cell(row=master_row, column=COL_ALERT_LOW_SOURCE, value='Auto')
            set_chart_flag(master_ws, master_row, True)
            append_note(master_ws, master_row,
                        f"Auto: Alert Low {alert_low} from single trendline below price ({today})")
            applied.append({'ticker': master_ticker, 'company': company, 'lower': lower, 'upper': None,
                             'alert_low': alert_low, 'alert_high': None})

        elif kind == 'single_high':
            # Single trendline above the current price -> Alert High only; Alert
            # Low is left completely untouched.
            alert_high = upper
            master_ws.cell(row=master_row, column=COL_ALERT_HIGH, value=alert_high)
            set_chart_flag(master_ws, master_row, True)
            append_note(master_ws, master_row,
                        f"Auto: Alert High {alert_high} from single trendline above price ({today})")
            applied.append({'ticker': master_ticker, 'company': company, 'lower': None, 'upper': upper,
                             'alert_low': None, 'alert_high': alert_high})

        else:
            rejected.append({'ticker': master_ticker, 'company': company, 'reason': f'unrecognized detection kind: {kind!r}'})
            stamp_last_checked(master_ws, master_row, col_last_checked, today)
            continue

        stamp_last_checked(master_ws, master_row, col_last_checked, today)

    return applied, rejected, skipped_manual, skipped_noise, unmatched, matches


def build_below_alert_rows(master_ws, matches):
    """Every matched ticker whose captured live price sits at or below its (possibly
    just-updated) Alert Low, worst gap first — the input for the alert table at the
    top of 'Stocks of Interest'. Reads the master sheet AFTER process() so freshly
    applied Alert Lows are what get compared.

    Rows carry on_alert (price sitting ON a drawn line, per channel_detect's
    ON_ALERT_TOL) so refresh_block can split them into their own section — those
    have reached the level rather than fallen through it."""
    rows = []
    for m in matches:
        price = m['price']
        if not isinstance(price, (int, float)):
            continue
        alert_low = master_ws.cell(row=m['row'], column=COL_ALERT_LOW).value
        if not isinstance(alert_low, (int, float)) or not alert_low or price >= alert_low:
            continue
        alert_high = master_ws.cell(row=m['row'], column=COL_ALERT_HIGH).value
        # Holdings/Target in Investments are sometimes FORMULAS (e.g. =ROUND(...));
        # with data_only=False .value is the formula string, which is meaningless
        # once written into the Stocks of Interest sheet. Only carry real numbers
        # over — anything else becomes a blank cell rather than a broken formula.
        holdings = master_ws.cell(row=m['row'], column=COL_HOLDINGS).value
        target_value = master_ws.cell(row=m['row'], column=COL_TARGET_VALUE).value
        rows.append({
            'ticker': m['ticker'],
            'share_name': master_ws.cell(row=m['row'], column=COL_SHARE_NAME).value or m['company'],
            'price': price,
            'alert_low': alert_low,
            'alert_high': alert_high if isinstance(alert_high, (int, float)) else None,
            'gap_pct': (price - alert_low) / alert_low * 100,
            'holdings': holdings if isinstance(holdings, (int, float)) else None,
            'target_value': target_value if isinstance(target_value, (int, float)) else None,
            'checked_at': m['checked_at'],
            'chart_id': m.get('chart_id'),
            'on_alert': bool(m.get('on_alert')),
            # Pattern label for the block's new Pattern column — the SAME labelling
            # (and inherited-vs-fresh distinction) the section tables use, so a row
            # whose axis read failed this run reads 'Not read this run — level
            # inherited' rather than the misleading 'No lines drawn'.
            'pattern': pattern_label(m.get('detection')),
        })
    rows.sort(key=lambda r: r['gap_pct'])
    return rows


_TRACKER_HEADING_RE = re.compile(r'^##.*Coverage Tracker.*$', re.MULTILINE)
_TABLE_ROW_RE = re.compile(r'^\|(.+)\|\s*$')


def _parse_tracker_table(block):
    """Parse the '| Ticker | Company | Status | Why | Last checked |' table in the
    Coverage Tracker block into a dict keyed by ticker. Returns (rows_dict, header_lines)
    or (None, None) if the table isn't in the expected shape."""
    lines = block.splitlines()
    table_start = None
    for i, line in enumerate(lines):
        if _TABLE_ROW_RE.match(line) and 'Ticker' in line:
            table_start = i
            break
    if table_start is None:
        return None, None
    header_lines = lines[table_start:table_start + 2]  # header + separator
    rows = {}
    row_order = []
    i = table_start + 2
    while i < len(lines) and _TABLE_ROW_RE.match(lines[i]):
        cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
        if len(cells) >= 5:
            ticker = cells[0]
            rows[ticker] = {'company': cells[1], 'status': cells[2], 'why': cells[3], 'last_checked': cells[4]}
            row_order.append(ticker)
        i += 1
    return {'rows': rows, 'order': row_order, 'header': header_lines, 'start': table_start,
            'end': i, 'lines': lines}, None


def _update_coverage_tracker(content, applied, rejected, unmatched, today):
    """Best-effort update of the Coverage Tracker table: remove resolved tickers (move
    them to a 'Resolved since last update' bullet list) and add/refresh unresolved ones.
    Returns the updated content, or the original content unchanged if the table isn't
    in a shape this can safely parse (never guess at document structure)."""
    heading_match = _TRACKER_HEADING_RE.search(content)
    if not heading_match:
        return content, False

    # Coverage Tracker block runs from its heading to the next '---' after it.
    rest = content[heading_match.start():]
    end_match = re.search(r'\n---\n', rest)
    if not end_match:
        return content, False
    block = rest[:end_match.start()]

    parsed, _ = _parse_tracker_table(block)
    if parsed is None:
        return content, False

    rows, order, lines = parsed['rows'], parsed['order'], parsed['lines']

    resolved_bullets = []
    for a in applied:
        if a['ticker'] in rows:
            del rows[a['ticker']]
            order.remove(a['ticker'])
            alert_low = a['alert_low'] if a['alert_low'] is not None else 'unchanged'
            alert_high = a['alert_high'] if a['alert_high'] is not None else 'unchanged'
            resolved_bullets.append(f"- **{a['ticker']}** ✅ — Applied: Alert Low {alert_low}, Alert High {alert_high}.")

    for item, reason_key in [(r, 'reason') for r in rejected] + [(u, 'reason') for u in unmatched]:
        ticker = item['ticker']
        why = item[reason_key]
        if ticker in rows:
            rows[ticker]['why'] = why
            rows[ticker]['last_checked'] = today
        else:
            rows[ticker] = {'company': item.get('company', ''), 'status': '⚪ Not yet attempted', 'why': why, 'last_checked': today}
            order.append(ticker)

    new_table_lines = list(parsed['header'])
    for ticker in order:
        r = rows[ticker]
        new_table_lines.append(f"| {ticker} | {r['company']} | {r['status']} | {r['why']} | {r['last_checked']} |")

    updated_lines = lines[:parsed['start']] + new_table_lines + lines[parsed['end']:]

    # Merge resolved bullets into the existing "Resolved since last update:" list, if any.
    updated_block = "\n".join(updated_lines)
    if resolved_bullets:
        resolved_marker = re.search(r'\*\*Resolved since last update:\*\*\n((?:- .*\n?)*)', updated_block)
        if resolved_marker:
            insert_at = resolved_marker.end(1)
            addition = "\n".join(resolved_bullets) + "\n"
            updated_block = updated_block[:insert_at] + addition + updated_block[insert_at:]
        else:
            updated_block = updated_block.rstrip('\n') + "\n\n**Resolved since last update:**\n" + "\n".join(resolved_bullets) + "\n"

    new_content = (content[:heading_match.start()] + updated_block.rstrip('\n') + "\n\n"
                   + rest[end_match.start():].lstrip('\n'))
    return new_content, True


def update_feedback_md(feedback_path, applied, rejected, skipped_manual, skipped_noise, unmatched):
    today = date.today().isoformat()
    lines_new = []
    lines_new.append(f"## {today} — Automated batch import ({len(applied) + len(rejected) + len(unmatched)} tickers)\n")
    lines_new.append("\n**Source:** `tradingview_layouts.xlsx`, built + processed automatically by the tradingview-mcp export pipeline.\n")
    lines_new.append("**Method:** OCR axis read + colour-line channel detection (see `channel_detect.py`), applied via `update_master_sheet.py`.\n\n")

    if applied:
        plural = '' if len(applied) == 1 else 's'
        lines_new.append(f"### ✅ Applied — {len(applied)} ticker{plural} updated with fresh Alert Low/High\n\n")
        lines_new.append("| Ticker | Lower | Upper | Alert Low | Alert High |\n|---|---|---|---|---|\n")
        for a in applied:
            cells = [a['lower'], a['upper'], a['alert_low'], a['alert_high']]
            lower, upper, alert_low, alert_high = ('—' if c is None else c for c in cells)
            lines_new.append(f"| {a['ticker']} | {lower} | {upper} | {alert_low} | {alert_high} |\n")
        lines_new.append("\n")

    if skipped_noise:
        lines_new.append(f"{len(skipped_noise)} re-read within {int(REFRESH_NOISE_THRESHOLD*100)}% of already-current values — left untouched to avoid noise: "
                          + ", ".join(s['ticker'] for s in skipped_noise) + "\n\n")

    if rejected:
        lines_new.append(f"### ❌ Rejected — nothing written, for cause\n\n")
        lines_new.append("| Ticker | Reason |\n|---|---|\n")
        for r in rejected:
            lines_new.append(f"| {r['ticker']} | {r['reason']} |\n")
        lines_new.append("\n")

    if skipped_manual:
        lines_new.append("### ⚠️ Skipped — Alert Low Source is \"Manual\", not overwritten\n\n")
        lines_new.append(", ".join(s['ticker'] for s in skipped_manual) + "\n\n")

    if unmatched:
        lines_new.append("### ⚠️ Unmatched / not attempted\n\n")
        lines_new.append("| Ticker | Company | Reason |\n|---|---|---|\n")
        for u in unmatched:
            lines_new.append(f"| {u['ticker']} | {u['company']} | {u['reason']} |\n")
        lines_new.append("\n")

    lines_new.append("---\n\n")
    new_section = "".join(lines_new)

    try:
        with open(feedback_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        content = "# Chart Import Findings\n\n---\n\n## \U0001F4CB Coverage Tracker\n\n---\n\n"

    content, tracker_updated = _update_coverage_tracker(content, applied, rejected, unmatched, today)
    if not tracker_updated:
        print("WARNING: could not parse the Coverage Tracker table in the expected shape — "
              "left it untouched. Update it manually for this session.", file=sys.stderr)

    # Insert the new dated entry after the Coverage Tracker block (its own trailing
    # '---'), i.e. right before the previously-newest dated entry — never above the
    # tracker, which must stay pinned at the top of the file.
    inserted = False
    heading_match = _TRACKER_HEADING_RE.search(content)
    if heading_match:
        end_match = re.search(r'\n---\n', content[heading_match.start():])
        if end_match:
            idx = heading_match.start() + end_match.end()
            content = content[:idx] + "\n" + new_section + content[idx:]
            inserted = True
    if not inserted:
        marker = re.search(r'\n---\n', content)
        if marker:
            idx = marker.end()
            content = content[:idx] + "\n" + new_section + content[idx:]
            inserted = True

    if not inserted:
        content = content + "\n" + new_section

    with open(feedback_path, 'w', encoding='utf-8') as f:
        f.write(content)

    return feedback_path


def main():
    master_in, charts_path, channel_path, master_out, feedback_path = sys.argv[1:6]

    charts = load_json(charts_path)
    channel_results = load_json(channel_path)
    channel_by_ticker = {c['ticker']: c for c in channel_results if c.get('ticker')}

    wb = openpyxl.load_workbook(master_in, data_only=False)
    ws = wb[SHEET_NAME]

    applied, rejected, skipped_manual, skipped_noise, unmatched, matches = process(ws, charts, channel_by_ticker)

    # Rebuild the below-alert table at the top of 'Stocks of Interest' from this
    # run's captured prices vs the (post-update) Alert Lows. Same workbook save.
    below_rows = build_below_alert_rows(ws, matches)
    from add_below_alert_sheet import refresh_block, SHEET_NAME as SOI_SHEET
    refresh_block(wb[SOI_SHEET], below_rows)
    n_on_alert = sum(1 for r in below_rows if r.get('on_alert'))
    print(f"Alert table: {len(below_rows)} ticker(s) at/under Alert Low "
          f"({n_on_alert} on alert) -> top of '{SOI_SHEET}'")

    wb.save(master_out)
    update_feedback_md(feedback_path, applied, rejected, skipped_manual, skipped_noise, unmatched)

    # Machine-readable summary alongside the human-readable feedback log, so
    # verify_pipeline.py can report on this run without re-parsing markdown.
    import os
    result_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'master_update_result_tmp.json')
    with open(result_path, 'w', encoding='utf-8') as f:
        json.dump({
            'timestamp': date.today().isoformat(),
            'master_out': master_out,
            'applied': applied,
            'rejected': rejected,
            'skipped_manual': skipped_manual,
            'skipped_noise': skipped_noise,
            'unmatched': unmatched,
            'below_alert_rows': below_rows,
            'commodity_prices': [
                {'ticker': m['ticker'], 'price': round(m['price'], 2), 'checked_at': m['checked_at']}
                for m in matches if m.get('commodity_price_written')
            ],
        }, f, indent=2)

    print(f"Applied: {len(applied)}, Rejected: {len(rejected)}, Manual-skipped: {len(skipped_manual)}, "
          f"Noise-skipped: {len(skipped_noise)}, Unmatched: {len(unmatched)}")
    print(f"Saved -> {master_out}")
    print(f"Feedback log updated -> {feedback_path}")
    print(f"Summary written -> {result_path}")


if __name__ == '__main__':
    main()
