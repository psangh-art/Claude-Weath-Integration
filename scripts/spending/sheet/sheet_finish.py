"""Post-processing on the finished sheet: column widths, splitting the risk-metrics
and Targets tables onto their own tab, the note row at the top, and the save.

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# scripts/ is on sys.path (it is the entry script's directory), same as the
# rest of this repo's cross-module imports.
from xlsx_sheet_copy import copy_cell_style, copy_sheet_into



def preserve_manual_sheets(new_wb, output_path):
    """This builder regenerates the whole workbook every run — but the user
    keeps hand-maintained tabs (e.g. 'Payslip Summary', 'Retirement Income
    Plan') alongside the generated ones, and a rebuild used to silently drop
    them (lost tabs reported 2026-07-12, restored from the 2026-07-02 copy).
    Carry over every sheet in the existing file whose name this run didn't
    generate, so manual tabs survive rebuilds."""
    if not os.path.exists(output_path):
        return
    try:
        old_wb = load_workbook(output_path)
    except Exception as e:
        print(f"WARNING: could not read existing {output_path} to preserve "
              f"manual tabs ({e}) — generated tabs only this run.", file=sys.stderr)
        return
    generated = set(new_wb.sheetnames)
    for name in old_wb.sheetnames:
        if name in generated:
            continue
        copy_sheet_into(old_wb[name], new_wb.create_sheet(name))
        print(f"Preserved manual tab: {name}")


def finish_workbook(ctx):
    # --- shared state in (generated) ---
    hist_months = ctx.hist_months
    output_path = ctx.output_path
    sum_months = ctx.sum_months
    wb = ctx.wb
    ws = ctx.ws

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    # Historic month cols: narrower
    for i, m in enumerate(sum_months):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 9 if m in hist_months else 11

    # Freeze panes: col A–C (labels) and row 1–2 (title + month headers) always visible
    ws.freeze_panes = "D3"

    # ── Move 'Investment Risk Metrics' and '2026 Targets' to a separate sheet ──
    section_rows = {}
    for row in ws.iter_rows():
        v = str(row[0].value or '').strip()
        if v in ('Investment Risk Metrics', '2026 Targets'):
            section_rows[v] = row[0].row

    if 'Investment Risk Metrics' in section_rows:
        move_start = section_rows['Investment Risk Metrics']
        move_end = ws.max_row

        ws_targets = wb.create_sheet("Targets")
        ws_targets.sheet_properties.tabColor = "E67E22"

        dst_row = 1
        for sr in range(move_start, move_end + 1):
            for sc in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=sr, column=sc)
                dst_cell = ws_targets.cell(row=dst_row, column=sc)
                dst_cell.value = src_cell.value
                if src_cell.data_type == 'f':
                    dst_cell.data_type = 'f'
                copy_cell_style(src_cell, dst_cell)
            if sr in ws.row_dimensions:
                ws_targets.row_dimensions[dst_row].height = ws.row_dimensions[sr].height
            dst_row += 1

        # Copy column widths
        for col_letter, col_dim in ws.column_dimensions.items():
            ws_targets.column_dimensions[col_letter].width = col_dim.width

        # Copy merged cells within the moved range
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= move_start and merged.max_row <= move_end:
                offset = move_start - 1
                try:
                    ws_targets.merge_cells(
                        start_row=merged.min_row - offset, start_column=merged.min_col,
                        end_row=merged.max_row - offset,   end_column=merged.max_col)
                except Exception:
                    pass

        ws_targets.freeze_panes = "D3"

        # Remove the moved rows from the main sheet (delete from bottom up not needed —
        # delete_rows handles the range in one call)
        ws.delete_rows(move_start, move_end - move_start + 1)

    # Force Excel to recalculate all formulas when the file is opened
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Post-process: ensure every cell with a string starting "=" is stored as a formula
    # (openpyxl sometimes stores formula strings as text data_type="s" instead of "f")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.data_type = "f"

    # ── Insert "This file is..." note row at top of Wealth Summary ─────────────
    import re as _re
    for sheet in wb.worksheets:
        # Capture merge ranges BEFORE insert (insert_rows does not shift these)
        old_merges = [str(m) for m in sheet.merged_cells.ranges]
        for m in list(sheet.merged_cells.ranges):
            try:
                sheet.unmerge_cells(str(m))
            except KeyError:
                sheet.merged_cells.ranges.discard(m)

        sheet.insert_rows(1)

        # Formula text needs manual row-number adjustment (insert_rows shifts
        # cell positions but not the formula strings themselves)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    def _bump(m):
                        return m.group(1) + str(int(m.group(2)) + 1)
                    cell.value = _re.sub(r"([A-Z]{1,3})(\d+)", _bump, cell.value)
                    cell.data_type = "f"

        # Re-apply merges, shifted down by 1 row
        for m_str in old_merges:
            # m_str like "A47:O47"
            match = _re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", m_str)
            if match:
                c1, r1, c2, r2 = match.groups()
                from openpyxl.utils import column_index_from_string
                sheet.merge_cells(start_row=int(r1)+1, start_column=column_index_from_string(c1),
                                  end_row=int(r2)+1, end_column=column_index_from_string(c2))

        # Shift freeze panes down by 1 row (D3 -> D4)
        if sheet.freeze_panes:
            fp = sheet.freeze_panes
            col_part = ''.join(c for c in fp if c.isalpha())
            row_part = ''.join(c for c in fp if c.isdigit())
            if row_part:
                sheet.freeze_panes = f"{col_part}{int(row_part)+1}"

    # Write the note row
    note_text = "This file is 'Spending Summary · XLSX' from downloads"
    for sheet in wb.worksheets:
        c = sheet.cell(row=1, column=1, value=note_text)
        c.font = Font(italic=True, color="7F4000", size=10, name="Arial")
        c.fill = PatternFill("solid", fgColor="FFF3CD")
        c.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 16
        # Extend yellow shading across all columns and merge
        for col in range(2, sheet.max_column + 1):
            sheet.cell(row=1, column=col).fill = PatternFill("solid", fgColor="FFF3CD")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

    preserve_manual_sheets(wb, output_path)
    wb.save(output_path)
