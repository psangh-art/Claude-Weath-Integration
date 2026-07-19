"""Section 1 of the Wealth Summary: monthly spend by category, the TOTAL row and
the expense-reimbursements table beneath it.

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from ..constants import CATEGORIES
from .style import (ALT_FILL, BOLD_FONT, C_NAVY, EST_FONT, EST_HDR_FILL, F, FOOT_FILL,
                    HDR_FILL, HDR_FONT, NUM_FMT, P, TOT_FILL)


def write_spending(ctx):
    # --- shared state in (generated) ---
    actual_set = ctx.actual_set
    all_months = ctx.all_months
    calc_cur = ctx.calc_cur
    col_fill = ctx.col_fill
    reimbursements = ctx.reimbursements
    spend_month_labels = ctx.spend_month_labels
    spend_months = ctx.spend_months
    spend_pivot = ctx.spend_pivot
    val_font = ctx.val_font
    ws = ctx.ws

    # ── Section 1: Spending Summary ────────────────────────────────────────────
    n_spend_cols = 3 + len(spend_months)
    R = calc_cur + 0  # no blank spacer

    # Section title
    ws.merge_cells(start_row=R+1, start_column=1, end_row=R+1, end_column=n_spend_cols)
    title = ws.cell(row=R+1, column=1, value="Amex & Barclays Spend")
    title.font = F(bold=True, color="FFFFFF", size=12)
    title.fill = P(C_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[R+1].height = 24

    # Column headers — col1=Category, col2=blank, col3=Total, col4+=months (aligned with income table)
    headers = ["Category", "", "Total"] + spend_month_labels
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=R+2, column=col, value=h)
        if col <= 3:
            c.font = HDR_FONT
            c.fill = HDR_FILL
        else:
            m = spend_months[col - 4]
            c.font = HDR_FONT if m in actual_set else EST_FONT
            c.fill = HDR_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[R+2].height = 18

    # Data rows
    for row_idx, cat in enumerate(CATEGORIES, R+3):
        is_alt = (row_idx % 2 == 0)
        font  = BOLD_FONT

        c = ws.cell(row=row_idx, column=1, value=cat)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="D9E1F2"))

        # col 2 blank
        ws.cell(row=row_idx, column=2).border = Border(bottom=Side(style="thin", color="D9E1F2"))

        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(spend_months))
        tc = ws.cell(row=row_idx, column=3,
                     value=f"=SUM({fc}{row_idx}:{lc}{row_idx})")
        tc.number_format = NUM_FMT
        tc.font = font
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.fill = TOT_FILL
        tc.border = Border(bottom=Side(style="thin", color="D9E1F2"))

        for col_idx, m in enumerate(spend_months, 4):
            val = spend_pivot.loc[cat, m] if cat in spend_pivot.index else 0
            fill = col_fill(m, is_alt, ALT_FILL, ALT_FILL)
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=int(round(val)) or None)
            cell.number_format = NUM_FMT
            cell.font = val_font(m)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="D9E1F2"))
            if fill: cell.fill = fill

    # Footer: total spend
    foot_row = R + 3 + len(CATEGORIES)
    exp_rows = [R + 3 + i for i in range(len(CATEGORIES))]
    ws.cell(row=foot_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=foot_row, column=1).fill = FOOT_FILL
    ws.cell(row=foot_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=foot_row, column=2).fill = FOOT_FILL  # blank col
    for col_idx in range(3, 4 + len(spend_months)):
        cl = get_column_letter(col_idx)
        refs = "+".join(f"{cl}{r}" for r in exp_rows)
        cell = ws.cell(row=foot_row, column=col_idx, value=f"={refs}")
        cell.number_format = NUM_FMT
        cell.font = BOLD_FONT
        cell.fill = FOOT_FILL
        cell.alignment = Alignment(horizontal="right", vertical="center")


    # ── Note row below TOTAL ────────────────────────────────────────────────────
    note_row = foot_row + 1
    note_cell = ws.cell(row=note_row, column=1,
                        value="Susan's lease car additional miles at 30.3p.  Contract to End July 2027")
    note_cell.font = Font(italic=True, color="808080", size=9)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 14
    # Merge across all spend columns so it reads as one line
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=3 + len(all_months))

    # ── Expense Reimbursements table (underneath Amex & Barclays Spend) ────────
    REIMB_START_ROW = note_row + 2  # 1 blank row gap below the lease-car note
    fc_r = get_column_letter(4)
    lc_r = get_column_letter(3 + len(all_months))

    # Title row — spans full width like the spend table
    ws.merge_cells(start_row=REIMB_START_ROW, start_column=1,
                   end_row=REIMB_START_ROW, end_column=3 + len(all_months))
    rt = ws.cell(row=REIMB_START_ROW, column=1, value="Expense Reimbursements")
    rt.font = F(bold=True, color="FFFFFF", size=12)
    rt.fill = P("8E44AD")  # purple — distinct from navy spend table
    rt.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[REIMB_START_ROW].height = 22

    # Header row — Category | (blank) | Total | Jan..Dec, matching spend table
    header_row = REIMB_START_ROW + 1
    hc = ws.cell(row=header_row, column=1, value="Source")
    hc.font = F(bold=True, color="FFFFFF", size=10)
    hc.fill = P("8E44AD")
    hc.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=header_row, column=2).fill = P("8E44AD")
    htot = ws.cell(row=header_row, column=3, value="Total")
    htot.font = F(bold=True, color="FFFFFF", size=10)
    htot.fill = P("8E44AD")
    htot.alignment = Alignment(horizontal="right", vertical="center")
    for i, m in enumerate(all_months):
        col = 4 + i
        c = ws.cell(row=header_row, column=col, value=m.strftime("%b %Y"))
        c.font = F(bold=True, color="FFFFFF", size=10)
        c.fill = P("8E44AD")
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[header_row].height = 20

    # Build month -> amount map per source
    reimb_list = reimbursements or []
    by_source = {"Royal Mail": {}, "Expleo": {}}
    for r in reimb_list:
        memo_upper = r["memo"].upper()
        source = "Royal Mail" if "ROYAL MAIL" in memo_upper else "Expleo"
        m = r["month"]
        by_source[source][m] = by_source[source].get(m, 0) + r["amount"]

    # Two data rows: Royal Mail, Expleo
    for ri, source in enumerate(["Royal Mail", "Expleo"]):
        row_n = header_row + 1 + ri
        bg = "F4ECF7" if ri % 2 == 1 else "FFFFFF"
        c = ws.cell(row=row_n, column=1, value=source)
        c.font = F(size=10)
        c.fill = P(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="E8DAEF"))
        ws.cell(row=row_n, column=2).fill = P(bg)
        ws.cell(row=row_n, column=2).border = Border(bottom=Side(style="thin", color="E8DAEF"))

        # Month columns
        for i, m in enumerate(all_months):
            col = 4 + i
            val = by_source[source].get(m, 0)
            cell = ws.cell(row=row_n, column=col)
            if val:
                cell.value = val
                cell.number_format = NUM_FMT
            cell.font = F(size=10)
            cell.fill = P(bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="E8DAEF"))

        # Total column
        tc = ws.cell(row=row_n, column=3, value=f"=SUM({fc_r}{row_n}:{lc_r}{row_n})")
        tc.number_format = NUM_FMT
        tc.font = F(bold=True, size=10)
        tc.fill = P(bg)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border = Border(bottom=Side(style="thin", color="E8DAEF"))

        ws.row_dimensions[row_n].height = 16

    # TOTAL row
    reimb_row = header_row + 3
    ws.cell(row=reimb_row, column=1, value="TOTAL").font = F(bold=True, size=10)
    ws.cell(row=reimb_row, column=1).fill = P("D2B4DE")
    ws.cell(row=reimb_row, column=2).fill = P("D2B4DE")
    tot_total = ws.cell(row=reimb_row, column=3,
                         value=f"=SUM(C{header_row+1}:C{header_row+2})")
    tot_total.number_format = NUM_FMT
    tot_total.font = F(bold=True, size=10)
    tot_total.fill = P("D2B4DE")
    tot_total.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(len(all_months)):
        col = 4 + i
        col_l = get_column_letter(col)
        cm = ws.cell(row=reimb_row, column=col,
                      value=f"=SUM({col_l}{header_row+1}:{col_l}{header_row+2})")
        cm.number_format = NUM_FMT
        cm.font = F(bold=True, size=10)
        cm.fill = P("D2B4DE")
        cm.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[reimb_row].height = 18

    # Note
    reimb_note_row = reimb_row + 1
    ws.merge_cells(start_row=reimb_note_row, start_column=1,
                   end_row=reimb_note_row, end_column=3 + len(all_months))
    rn = ws.cell(row=reimb_note_row, column=1,
                 value="Reimbursements from Royal Mail or Expleo for expenses incurred on the family's "
                       "behalf — recovered from employers, excluded from family spending totals.")
    rn.font = Font(italic=True, color="808080", size=8)
    rn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[reimb_note_row].height = 28

    # --- shared state out (generated) ---
    ctx.reimb_note_row = reimb_note_row
