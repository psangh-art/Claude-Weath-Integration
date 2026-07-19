"""Cross-section totals that can only be written once the row numbers above are
known: the 'Total Income and Accumulations' formulas, the Total Income + Salary
row, and the data feeding the risk-metrics table.

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .style import BODY_FONT, C_SUM_T, EST_BODY, EST_FONT, EST_HDR_FILL, F, NUM_FMT, P


def write_totals(ctx):
    # --- shared state in (generated) ---
    acc_cur_row = ctx.acc_cur_row
    actual_set = ctx.actual_set
    all_months = ctx.all_months
    spend_month_labels = ctx.spend_month_labels
    spend_months = ctx.spend_months
    spend_pivot = ctx.spend_pivot
    summary_data = ctx.summary_data
    tot_acc_r = ctx.tot_acc_r
    tot_r = ctx.tot_r
    total_inc_acc_row = ctx.total_inc_acc_row
    ws = ctx.ws

    # ── Update Total Income and Accumulations with Excel formulas now that row numbers are known ──
    # = Total Income (tot_r) + TOTAL Accumulative (tot_acc_r)
    col_letters = [get_column_letter(4 + i) for i in range(len(all_months))]
    for col_idx, m in enumerate(all_months, 4):
        cl = get_column_letter(col_idx)
        cell = ws.cell(row=total_inc_acc_row, column=col_idx,
                       value=f"={cl}{tot_r}+{cl}{tot_acc_r}")
        cell.number_format = NUM_FMT
        cell.font = F(bold=True, color="042C53")
        cell.fill = P(C_SUM_T)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    # Col C = annual total (sum of all 12 months)
    sum_range = "+".join(f"{cl}{total_inc_acc_row}" for cl in col_letters)
    c_total = ws.cell(row=total_inc_acc_row, column=3, value=f"={sum_range}")
    c_total.number_format = NUM_FMT
    c_total.font = F(bold=True, color="042C53")
    c_total.fill = P(C_SUM_T)
    c_total.alignment = Alignment(horizontal="right", vertical="center")
    fid3_start = acc_cur_row + 1

    C_FID3_H = "4A235B"   # deep purple header
    C_FID3_A = "F5EEF8"   # light purple alt rows
    C_FID3_T = "D2B4DE"   # purple total

    FID3_FILL  = P(C_FID3_H)
    FID3A_FILL = P(C_FID3_A)
    FID3T_FILL = P(C_FID3_T)
    FID3_HFONT = F(bold=True, color="FFFFFF")
    FID3_TFONT = F(bold=True, color="4A235B")

    n_fid3_cols = 3 + len(spend_months)
    ws.merge_cells(start_row=fid3_start, start_column=1,
                   end_row=fid3_start, end_column=n_fid3_cols)
    t3 = ws.cell(row=fid3_start, column=1, value="Fidelity")
    t3.font = F(bold=True, color="FFFFFF", size=12)
    t3.fill = FID3_FILL
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fid3_start].height = 24

    fid3_hdr = fid3_start + 1
    for col, h in enumerate(["Category", "", "Total"] + spend_month_labels, 1):
        c = ws.cell(row=fid3_hdr, column=col, value=h)
        if col <= 3:
            c.font = FID3_HFONT
            c.fill = FID3_FILL
        else:
            m = spend_months[col - 4]
            c.font = FID3_HFONT if m in actual_set else EST_FONT
            c.fill = FID3_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[fid3_hdr].height = 18

    fid3_data_row = fid3_start + 2
    fill = FID3A_FILL
    c = ws.cell(row=fid3_data_row, column=1, value="Fidelity card payments")
    c.font = BODY_FONT; c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="thin", color="D2B4DE"))
    ws.cell(row=fid3_data_row, column=2).fill = fill
    ws.cell(row=fid3_data_row, column=2).border = Border(bottom=Side(style="thin", color="D2B4DE"))

    fc = get_column_letter(4)
    lc = get_column_letter(3 + len(spend_months))
    tc = ws.cell(row=fid3_data_row, column=3,
                 value=f"=SUM({fc}{fid3_data_row}:{lc}{fid3_data_row})")
    tc.number_format = NUM_FMT; tc.font = FID3_TFONT; tc.fill = FID3T_FILL
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = Border(bottom=Side(style="thin", color="D2B4DE"))

    for col_idx, m in enumerate(spend_months, 4):
        val = spend_pivot.loc["Fidelity", m] if "Fidelity" in spend_pivot.index else 0
        is_est = m not in actual_set
        cell = ws.cell(row=fid3_data_row, column=col_idx,
                       value=int(round(val)) or None)
        cell.number_format = NUM_FMT
        cell.font = BODY_FONT if not is_est else EST_BODY
        cell.fill = FID3A_FILL if not is_est else P("ECE8F4")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="D2B4DE"))

    # ── Total Income + Salary row ──────────────────────────────────────────────
    combo_row = fid3_data_row + 1
    c = ws.cell(row=combo_row, column=1, value="Total Income & Salary")
    c.font = F(bold=True, color="FFFFFF"); c.fill = FID3_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=combo_row, column=2).fill = FID3_FILL  # blank

    # Income table: col1=label, col2=units, col3=Total, col4+=months (same as spend now)
    # Reference tot_r directly — all cols now aligned
    for col_idx in range(3, 4 + len(spend_months)):
        cl = get_column_letter(col_idx)
        cell = ws.cell(row=combo_row, column=col_idx, value=f"={cl}{tot_r}")
        cell.number_format = NUM_FMT
        cell.font = F(bold=True, color="FFFFFF"); cell.fill = FID3_FILL
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Investment Risk Metrics Table ──────────────────────────────────────────
    # Classify all Fidelity holdings into: Shares, Income Funds, Non-Income Funds
    # Columns: Group | Paul SIPP | Paul ISA | Joint Acct | Susan SIPP | Susan ISA | Jayne JISA | Liam ISA | Liam Inv | TOTAL
    # Rows: Shares, Income Funds, Non-Income Funds, TOTAL

    METRIC_ACCS_ORDER = [
        ('2000001606', 'Paul SIPP'),
        ('SANX002282', 'Paul ISA'),
        ('SANQ000468', 'Joint Acct'),
        ('2000001604', 'Susan SIPP'),
        ('SANX002617', 'Susan ISA'),
        ('SANX002936', 'Jayne JISA'),
        ('AS10303823', 'Liam ISA'),
        ('AG10131710', 'Liam Inv Acct'),
    ]

    HOLDING_GROUPS = {
        'Aegon High Yield Bond B Inc':                   'Income Funds',
        'Aegon High Yield Bond B Acc':                   'Income Funds',
        'Schroder High Yield Opportunities Fund Z Inc':  'Income Funds',
        'Schroder High Yield Opportunities Fund Z Acc':  'Income Funds',
        'Man High Yield Opportunities Fund Prof D Inc':  'Income Funds',
        'Man High Yield Opportunities Fund Prof D Acc':  'Income Funds',
        'WS Guinness Global Energy Fund I Acc':          'Non-Income Funds',
        'AUTOTRADER GROUP PLC,ORD GBP0.01(AUTO)':       'Shares',
        'AVIVA,ORD GBP0.328947368(AV.)':                 'Shares',
        'RELX PLC,ORD GBP0.1444(REL)':                  'Shares',
        'THE SAGE GROUP PLC,GBP0.01051948(SGE)':         'Shares',
        'WEIR GROUP,ORD GBP0.125(WEIR)':                 'Shares',
        'LEGAL & GENERAL GROUP,ORD GBP0.025(LGEN)':      'Shares',
    }
    GROUPS_ORDER = ['Shares', 'Income Funds', 'Non-Income Funds']

    # Build data: group → account → value
    metric_data = {g: {a: 0.0 for a, _ in METRIC_ACCS_ORDER} for g in GROUPS_ORDER}
    metric_data['_cash'] = {a: 0.0 for a, _ in METRIC_ACCS_ORDER}

    import csv as _csv2, io as _io2, os as _os2
    # Find AccountSummary path — passed in via summary_data
    _acct_path = summary_data.get('_account_summary_path', '')
    if not _acct_path or not _os2.path.exists(_acct_path):
        for _p in ['AccountSummary.csv', '/home/claude/AccountSummary.csv']:
            if _os2.path.exists(_p):
                _acct_path = _p
                break
    with open(_acct_path, encoding='utf-8-sig') as _f:
        _content = _f.read()
    _lines = _content.replace('\r','').split('\n')
    _hi = max(i for i, l in enumerate(_lines) if l.startswith('Type,Holdings,Account number'))
    for _row in _csv2.DictReader(_io2.StringIO('\n'.join(_lines[_hi:]))):
        _acc = _row.get('Account number','').strip()
        if _acc not in {a for a,_ in METRIC_ACCS_ORDER}: continue
        _t = _row.get('Type','').strip()
        _val = float(_row.get('Value (£)','0').replace(',','') or 0)
        if _t == 'Asset':
            _fund = _row.get('Holdings','').strip()
            _grp = HOLDING_GROUPS.get(_fund)
            if _grp:
                metric_data[_grp][_acc] += _val
        elif _t == 'Account':
            # Total account value — used to derive cash (account total - sum of assets)
            metric_data['_cash'][_acc] = _val

    # Cash = account total - sum of classified assets
    metric_cash = {}
    for acc, _ in METRIC_ACCS_ORDER:
        asset_sum = sum(metric_data[g][acc] for g in GROUPS_ORDER)
        metric_cash[acc] = max(0, metric_data['_cash'][acc] - asset_sum)

    # --- shared state out (generated) ---
    ctx.GROUPS_ORDER = GROUPS_ORDER
    ctx.METRIC_ACCS_ORDER = METRIC_ACCS_ORDER
    ctx.combo_row = combo_row
    ctx.metric_cash = metric_cash
    ctx.metric_data = metric_data
