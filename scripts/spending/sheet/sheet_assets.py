"""Wealth Summary rows: the summary table, the Fidelity accounts, both sets of
pensions, the assets section, the growth/total rows and the calculations block.

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..constants import ESTIMATES_AS_OF
from .style import C_SUM_T, F, HIST_MAP, P, P_SUM_A, P_SUM_H, P_SUM_S


def write_asset_rows(ctx):
    # --- shared state in (generated) ---
    acc_holdings = ctx.acc_holdings
    actual_set = ctx.actual_set
    all_months = ctx.all_months
    anchors = ctx.anchors
    fid_pivot = ctx.fid_pivot
    history = ctx.history
    is_history_month = ctx.is_history_month
    live_and_hist_safe = ctx.live_and_hist_safe
    projection_with_hist_override = ctx.projection_with_hist_override
    summary_data = ctx.summary_data
    ws = ctx.ws

    hist_periods_set = set()
    for series in history.values():
        hist_periods_set.update(series.keys())
    hist_months = sorted(p for p in hist_periods_set if p not in set(all_months))

    # Display only 2026 months — all tables share same columns
    sum_months = list(all_months)   # Jan–Dec 2026 only
    n_sum_cols = 3 + len(sum_months)

    # Helper: look up a value for a given month — history first for hist months,
    # live dict for live months. live_dict keys are pd.Period.
    def sum_val(live_dict, m):
        """Return value for month m: from live_dict if available, else history."""
        if m in live_dict and live_dict[m]:
            return live_dict[m]
        return None

    def write_total_row(r, label, monthly_dict, fill_col, font_col="042C53"):
        c = ws.cell(row=r, column=1, value=label)
        c.font = F(bold=True, color=font_col); c.fill = P(fill_col)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).fill = P(fill_col)
        ws.cell(row=r, column=3).fill = P(fill_col)
        for col_idx, m in enumerate(sum_months, 4):
            v = monthly_dict.get(m, 0)
            cell = ws.cell(row=r, column=col_idx, value=int(round(v)) if v else None)
            cell.number_format = '#,##0'; cell.font = F(bold=True, color=font_col)
            cell.fill = P(fill_col); cell.alignment = Alignment(horizontal="right", vertical="center")
        return r + 1

    def sum_row(r, label, values_dict, alt=False, bold=False, header_fill=None,
                label_indent="", font_color="000000", total_fill=None, hist_key=None):
        """Write one row: col1=Label, col2=blank, col3=blank, col4+=months.
        hist_key: if provided, uses history[hist_key] for historic months."""
        fill = header_fill or (P_SUM_A if alt else None)
        font = F(bold=bold, color=font_color)
        c = ws.cell(row=r, column=1, value=label_indent + label)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="D5D8DC"))
        if fill: c.fill = fill

        for col in (2, 3):
            bc = ws.cell(row=r, column=col)
            bc.border = Border(bottom=Side(style="thin", color="D5D8DC"))
            if fill: bc.fill = fill

        hist_row_name = HIST_MAP.get(hist_key, hist_key) if hist_key else None
        hist_series = history.get(hist_row_name, {}) if hist_row_name else {}

        for col_idx, m in enumerate(sum_months, 4):
            is_hist = is_history_month(m)
            if is_hist and hist_key and m in hist_series:
                # Previous Data owns this month and has a value — use it
                val = hist_series[m]
            elif is_hist and hist_key and m not in hist_series:
                # Previous Data owns this month but no entry — blank
                val = None
            else:
                # Either not a history month, or no hist_key — use values_dict
                val = values_dict.get(m, 0) if values_dict else 0
            is_est = (not is_hist) and (m not in actual_set)
            # Write value — allow zero (don't filter with truthiness)
            if val is not None:
                cell = ws.cell(row=r, column=col_idx, value=int(round(val)))
                cell.number_format = '#,##0'
            else:
                cell = ws.cell(row=r, column=col_idx, value=None)
                cell.number_format = '#,##0'
            if is_hist:
                cell.font = F(bold=bold, color="7F8C8D")
                cell.fill = P("E8EAEB") if alt else P("F2F3F4")
            elif is_est:
                cell.font = F(bold=bold, color="555577")
                if fill: cell.fill = fill
            else:
                cell.font = font
                if fill: cell.fill = fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="D5D8DC"))

        return r + 1

    def sum_total_row(r, label, monthly_dict, fill_col, font_col="042C53"):
        """Write a total/subtotal row across all sum_months."""
        font = F(bold=True, color=font_col)
        p_fill = P(fill_col)
        c = ws.cell(row=r, column=1, value=label)
        c.font = font; c.fill = p_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).fill = p_fill
        ws.cell(row=r, column=3).fill = p_fill
        for col_idx, m in enumerate(sum_months, 4):
            v = monthly_dict.get(m, 0)
            cell = ws.cell(row=r, column=col_idx, value=int(round(v)) if v else None)
            cell.number_format = '#,##0'; cell.font = font; cell.fill = p_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
        return r + 1

    # Section title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_sum_cols)
    t = ws.cell(row=1, column=1, value="Family Wealth Summary")
    t.font = F(bold=True, color="FFFFFF", size=12)
    t.fill = P_SUM_H
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row: Label | blank | blank | hist months (grey) | live months
    ws.cell(row=2, column=1, value="").fill = P_SUM_H
    ws.cell(row=2, column=2, value="").fill = P_SUM_H
    ws.cell(row=2, column=3, value="").fill = P_SUM_H
    for col_idx, m in enumerate(sum_months, 4):
        is_hist = m in hist_months
        is_est = (not is_hist) and (m not in actual_set)
        label = m.strftime("%b %Y")
        c = ws.cell(row=2, column=col_idx, value=label)
        if is_hist:
            c.font = F(bold=True, color="AAAAAA")
            c.fill = P("BFC9CA")
        elif is_est:
            c.font = F(bold=True, color="CCCCDD")
            c.fill = P("4A4A6A")
        else:
            c.font = F(bold=True, color="FFFFFF")
            c.fill = P_SUM_H
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    cur_row = 3

    # ── 1. Fidelity accounts ──────────────────────────────────────────────────
    # Sub-header
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Fidelity accounts")
    sh.font = F(bold=True, color="FFFFFF", size=9)
    sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    # Mapping: summary row → history row name
    def hist(key):
        """Return history series for this key."""
        return history.get(HIST_MAP.get(key, ""), {})

    def live_and_hist(live_dict, hist_key_str):
        """Alias for live_and_hist_safe — history always wins for pre-cutoff months."""
        return live_and_hist_safe(live_dict, hist_key_str)

    # Each family account — exclude SIPP accounts (shown in pensions instead)
    fid_total_val = summary_data["fidelity_total"]
    fid_by_acc = summary_data["fidelity_by_acc"]
    SIPP_ACCS = {"2000001606", "2000001604"}  # moved to pensions sections
    FIDELITY_ACC_LABELS = {
        "AW10032966": "Cash Account (Paul)",
        "SANX002282": "Investment ISA (Paul)",
        "SANQ000468": "Investment Account (Joint)",   # removed "(Paul)"
        "SANX002936": "Junior ISA (Jayne)",
        "AW10261123": "Cash Account (Susan)",
        "SANX002617": "Investment ISA (Susan)",
        "AW10580794": "Cash Account (Liam)",
        "AS10303823": "Investment ISA (Liam)",
        "AG10131710": "Investment Account (Liam)",
    }

    fid_non_sipp_accs = [(acc, val) for acc, val in sorted(fid_by_acc.items(), key=lambda x: x[1], reverse=True)
                         if acc not in SIPP_ACCS]

    # Compute monthly growth for SANQ000468 from accumulated fund values
    sanq_monthly = {}
    for fund_h, fd_h in acc_holdings.get("SANQ000468", {}).items():
        mv = fd_h.get("monthly_values", {})
        for m, v in mv.items():
            sanq_monthly[m] = sanq_monthly.get(m, 0) + v
    # Add LGEN and any non-Acc holdings (flat) to get full account value
    sanq_non_acc_val = fid_by_acc.get("SANQ000468", 0) - sum(
        fd_h["value"] for fd_h in acc_holdings.get("SANQ000468", {}).values()
    )
    for m in all_months:
        sanq_monthly[m] = sanq_monthly.get(m, 0) + sanq_non_acc_val

    # The AccountSummary total is the ACTUAL current value of the account. The
    # forward growth-projection past the build anchor (anchors.data_month in
    # build_acc_holdings)
    # is unreliable — it collapsed the current month (was hardcoded to override only
    # JUNE, so when "now" advanced to July the July value dropped ~£145k to a broken
    # projection: bug #20, 2026-07-18). Hold the current data month and every month
    # after it flat at the actual total — exactly as every OTHER Fidelity account row
    # already is (see the `else` branch below) — keeping only the reliable historical
    # backward ramp for earlier months. hold_from is the first month past the
    # build anchor, so it now tracks the snapshot automatically.
    sanq_actual = fid_by_acc.get("SANQ000468", 0)
    HOLD_FROM = anchors.hold_from
    if sanq_actual:
        for m in all_months:
            if m >= HOLD_FROM:
                sanq_monthly[m] = sanq_actual

    fid_rows_start = cur_row  # track start row for Fidelity total formula

    for i, (acc, val) in enumerate(fid_non_sipp_accs):
        # Show the account number alongside the friendly name (user request
        # 2026-07-18) — e.g. "Investment Account (Joint) (SANQ000468)".
        _base = FIDELITY_ACC_LABELS.get(acc)
        label = f"{_base} ({acc})" if _base else acc
        if acc == "SANQ000468" and sanq_monthly:
            monthly_vals = {m: sanq_monthly.get(m, val) for m in all_months}
        else:
            monthly_vals = {m: val for m in all_months}
        # Fidelity account rows: no per-account history, live values only
        cur_row = sum_row(cur_row, label, monthly_vals, alt=(i % 2 == 0),
                          label_indent="  ", font_color="042C53")
    fid_rows_end = cur_row

    # ── 2. Paul's pensions ────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Paul's pensions")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    paul_sipp_val = fid_by_acc.get("2000001606", 0)
    paul_sipp_vals_flat = {m: paul_sipp_val for m in all_months}
    cur_row = sum_row(cur_row, "SIPP Savings – Fidelity (2000001606)", paul_sipp_vals_flat,
                      alt=True, label_indent="  ", font_color="1A5276")

    arriva_merged = projection_with_hist_override(summary_data["arriva"], "arriva")
    cur_row = sum_row(cur_row, "PS Arriva (Defined Benefit)", arriva_merged,
                      alt=False, label_indent="  ", font_color="1A5276")  # pre-merged

    expleo_merged = projection_with_hist_override(summary_data["expleo_sw"], "expleo")
    cur_row = sum_row(cur_row, "Expleo Scottish Widows", expleo_merged,  # pre-merged
                      alt=True, label_indent="  ", font_color="1A5276")

    # Paul total placeholder — updated below once paul_sipp_growth_vals is computed
    paul_total_placeholder = {m: paul_sipp_val + summary_data["arriva"].get(m, 0) + summary_data["expleo_sw"].get(m, 0)
                              for m in all_months}
    cur_row = write_total_row(cur_row, "Total Paul's pensions", paul_total_placeholder, C_SUM_T)

    # ── 3. Susan's pensions ───────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Susan's pensions")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    SUSAN_HIST_KEYS = {
        "Capita (RMSPS) – to 2018 (65 yrs)": "capita",
        "RMPP 2012–2023 (65 years)":          "rmpp",
        "Collective Pension (2023+)":         "collective",
        "Cash Balance":                        "cash_balance",
        "AVC Bonus Plan (Scottish Widows)":    "avc",
    }
    susan_total_by_month = {m: 0 for m in all_months}
    for i, (pen_label, pen_vals) in enumerate(summary_data["susan_pensions"].items()):
        hk = SUSAN_HIST_KEYS.get(pen_label)
        merged = projection_with_hist_override(pen_vals, hk) if hk else pen_vals
        cur_row = sum_row(cur_row, pen_label, merged, alt=(i % 2 == 0),
                          label_indent="  ", font_color="1A5276")  # hist_key omitted — pre-merged
        # Accumulate from merged (history-corrected) not raw pen_vals
        for m in all_months:
            susan_total_by_month[m] += merged.get(m, 0)

    sipp_vals = summary_data["susan_fidelity_sipp"]
    for m in all_months:
        if m not in sipp_vals:
            sipp_vals[m] = list(sipp_vals.values())[0] if sipp_vals else 0
    susan_fid_merged = live_and_hist(sipp_vals, "susan_fidelity")
    cur_row = sum_row(cur_row, "SIPP Savings – Fidelity (2000001604)", susan_fid_merged,
                      alt=True, label_indent="  ", font_color="1A5276",
                      hist_key="susan_fidelity")
    for m in all_months:
        susan_total_by_month[m] = susan_total_by_month.get(m, 0) + sipp_vals.get(m, 0)

    cur_row = write_total_row(cur_row, "Total Susan's pensions", susan_total_by_month, C_SUM_T)
    susan_excl_sipp = {m: susan_total_by_month.get(m, 0) - sipp_vals.get(m, 0) for m in all_months}

    # ── 4. Assets section ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Other Assets")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    mf_proj = {m: 190000 for m in all_months}
    mf_merged = projection_with_hist_override(mf_proj, "mf")
    mf_vals = {m: mf_merged.get(m, 190000) for m in all_months}
    cur_row = sum_row(cur_row, "MF", mf_merged, alt=True, label_indent="  ",
                      font_color="1A5276")  # pre-merged

    bar_data = summary_data.get("barclays_by_month", {})
    cur_row = sum_row(cur_row, "Cash (Barclays)", bar_data, alt=False, label_indent="  ",
                      font_color="1A5276")

    house_at_est = 450450   # as at ESTIMATES_AS_OF
    house_vals = {}
    est_idx_h = next((i for i, m in enumerate(all_months) if m == ESTIMATES_AS_OF), 0)
    for i, m in enumerate(all_months):
        offset = i - est_idx_h
        house_vals[m] = round(house_at_est * ((1.05) ** (offset / 12)))
    house_merged = projection_with_hist_override(house_vals, "house")
    cur_row = sum_row(cur_row, "House", house_merged, alt=True, label_indent="  ",
                      font_color="1A5276")  # pre-merged

    # Liam ISA (AS10303823) and Investment Account (AG10131710) already in Fidelity accounts — not duplicated here

    car_at_est = 42074  # actual value from history at ESTIMATES_AS_OF (Cars Paul)
    car_vals = {}
    for i, m in enumerate(all_months):
        offset = i - est_idx_h
        car_vals[m] = round(car_at_est * ((1 - 0.05) ** (offset / 12)))
    car_merged = projection_with_hist_override(car_vals, "cars")
    cur_row = sum_row(cur_row, "Cars Paul (Mercedes AMG GTS 2016)", car_merged, alt=False,
                      label_indent="  ", font_color="1A5276")  # pre-merged

    assets_dicts = [mf_vals, bar_data, house_vals, car_vals]
    assets_by_month = {m: sum(d.get(m, 0) for d in assets_dicts) for m in all_months}
    cur_row = write_total_row(cur_row, "Total Other Assets", assets_by_month, C_SUM_T)

    # ── Compute Fidelity growth ───────────────────────────────────────────────
    fid_growth = {all_months[0]: round(fid_total_val)}
    for i in range(1, len(all_months)):
        m = all_months[i]
        monthly_income = fid_pivot[m].sum() if m in fid_pivot.columns else 0
        fid_growth[m] = round(fid_growth[all_months[i-1]] + monthly_income)

    # ── SIPP growth: pinned history early, AccountSummary anchors the snapshot
    # month, income grows the months after it. The snapshot is a statement of
    # what the accounts are worth ON ITS OWN DATE — anchoring it to a fixed month
    # made every later export overwrite the wrong column.
    paul_sipp_val  = fid_by_acc.get("2000001606", 0)   # value at anchors.data_month
    susan_sipp_val = fid_by_acc.get("2000001604", 0)
    paul_hist_series  = hist("paul_pension")
    paul_sipp_growth_vals = {}
    susan_sipp_growth_vals = {}
    sipp_vals_from_summary = summary_data.get("susan_fidelity_sipp", {})

    ps_running = paul_hist_series.get(all_months[0], paul_sipp_val)
    ss_running = sipp_vals_from_summary.get(all_months[0], susan_sipp_val)

    for i, m in enumerate(all_months):
        if m == anchors.data_month:
            # AccountSummary is ground truth for this month — override any projection
            ps_running = paul_sipp_val
            ss_running = susan_sipp_val
            paul_sipp_growth_vals[m] = paul_sipp_val
            susan_sipp_growth_vals[m] = susan_sipp_val
        elif is_history_month(m):
            # Pre-cutoff: use history if available
            if m in paul_hist_series:
                paul_sipp_growth_vals[m] = paul_hist_series[m]
                ps_running = paul_hist_series[m]
            else:
                paul_sipp_growth_vals[m] = ps_running
            ss_val = sipp_vals_from_summary.get(m, ss_running)
            susan_sipp_growth_vals[m] = ss_val
            ss_running = ss_val
        else:
            # Post-cutoff (Jun+): grow from previous month using income
            ps_inc = fid_pivot.loc["2000001606", m] if "2000001606" in fid_pivot.index and m in fid_pivot.columns else 0
            ss_inc = fid_pivot.loc["2000001604", m] if "2000001604" in fid_pivot.index and m in fid_pivot.columns else 0
            ps_running = round(ps_running + ps_inc)
            ss_running = round(ss_running + ss_inc)
            paul_sipp_growth_vals[m] = ps_running
            susan_sipp_growth_vals[m] = ss_running

    fid_non_sipp_total = sum(val for acc, val in fid_by_acc.items() if acc not in SIPP_ACCS)
    sanq_flat = fid_by_acc.get("SANQ000468", 0)
    sanq_growth = {m: sanq_monthly.get(m, sanq_flat) for m in all_months}
    fid_non_sipp_with_growth = {m: fid_non_sipp_total - sanq_flat + sanq_growth[m] for m in all_months}
    total_fidelity_growth = {m: fid_non_sipp_with_growth[m] + paul_sipp_growth_vals[m] + susan_sipp_growth_vals[m]
                             for m in all_months}

    # ── 5. Fidelity accounts total (all including SIPPs) ─────────────────────
    fid_total_row_num = cur_row
    fid_all_merged = {**hist("fidelity_all"), **total_fidelity_growth}
    cur_row = write_total_row(cur_row, "Fidelity accounts", fid_all_merged, C_SUM_T)

    # Update Paul SIPP row values with growth
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "  SIPP Savings – Fidelity (2000001606)":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(paul_sipp_growth_vals[m]))
            break
    # Update Susan SIPP row values with growth
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "  SIPP Savings – Fidelity (2000001604)":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(susan_sipp_growth_vals[m]))
            break
    # Update paul_total using history-corrected arriva and expleo values
    arriva_merged = projection_with_hist_override(summary_data["arriva"], "arriva")
    expleo_merged_upd = projection_with_hist_override(summary_data["expleo_sw"], "expleo")
    paul_total = {m: paul_sipp_growth_vals[m] + arriva_merged.get(m, 0) + expleo_merged_upd.get(m, 0)
                  for m in all_months}
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "Total Paul's pensions":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(paul_total[m]))
            break

    # Update Susan total using history-corrected pension + SIPP growth values
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "Total Susan's pensions":
            for col_idx, m in enumerate(all_months, 4):
                sp_merged = sum(
                    projection_with_hist_override(summary_data["susan_pensions"].get(k, {}),
                                  SUSAN_HIST_KEYS.get(k)).get(m, 0)
                    for k in summary_data["susan_pensions"]
                )
                ws.cell(row=_r, column=col_idx).value = int(round(sp_merged + susan_sipp_growth_vals[m]))
            break

    # ── 8. TOTAL FAMILY WEALTH — at the very bottom ───────────────────────────
    grand_row = cur_row
    cur_row += 1
    paul_non_sipp = {m: summary_data["arriva"][m] + summary_data["expleo_sw"][m] for m in all_months}
    grand_by_month = {m: total_fidelity_growth[m] + paul_non_sipp.get(m, 0)
                      + susan_excl_sipp.get(m, 0) + assets_by_month.get(m, 0)
                      for m in all_months}
    # Build historic grand total from history rows — covers ALL historic periods
    # (not just display months) so Yearly Increase can reference prior year values
    hist_grand = {}
    all_hist_periods = sorted(set().union(*[set(s.keys()) for s in history.values()]))
    for m in all_hist_periods:
        fid_h    = hist("fidelity_all").get(m, 0)
        arriva_h = hist("arriva").get(m, 0)
        expleo_h = hist("expleo").get(m, 0)
        capita_h = hist("capita").get(m, 0)
        rmpp_h   = hist("rmpp").get(m, 0)
        coll_h   = hist("collective").get(m, 0)
        cb_h     = hist("cash_balance").get(m, 0)
        avc_h    = hist("avc").get(m, 0)
        mf_h     = hist("mf").get(m, 0)
        house_h  = hist("house").get(m, 0)
        # Liam ISA & Investment Account is inside fid_h (All Fidelity accounts) — not added separately
        # Jayne ISA is also inside fid_h — not added separately
        susan_non_fid_h = capita_h + rmpp_h + coll_h + cb_h + avc_h
        hist_grand[m] = (fid_h + arriva_h + expleo_h + susan_non_fid_h +
                         mf_h + house_h)

    all_grand = {**hist_grand, **grand_by_month}

    c = ws.cell(row=grand_row, column=1, value="TOTAL FAMILY WEALTH")
    c.font = F(bold=True, color="FFFFFF", size=11); c.fill = P_SUM_H
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[grand_row].height = 22
    for col in range(1, 4):
        ws.cell(row=grand_row, column=col).fill = P_SUM_H
    for col_idx, m in enumerate(sum_months, 4):
        v = all_grand.get(m, 0)
        cell = ws.cell(row=grand_row, column=col_idx, value=int(round(v)) if v else None)
        cell.number_format = '#,##0'; cell.font = F(bold=True, color="FFFFFF")
        cell.fill = P_SUM_H; cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── 9. Total Income and Accumulations — below TOTAL FAMILY WEALTH ───────────────────────
    # Values written now as placeholders; updated with Excel formulas once
    # Income (tot_r) and Accumulative (tot_acc_r) row numbers are known.
    fid_income_monthly = {m: int(round(fid_pivot[m].sum())) if (not fid_pivot.empty and m in fid_pivot.columns) else 0
                          for m in all_months}
    acc_monthly_increase = {m: 0 for m in all_months}
    for acc_hh, funds_hh in acc_holdings.items():
        for fund_hh, fd_hh in funds_hh.items():
            pa = fd_hh.get("price_appreciation", {})
            for mk, inc in pa.items():
                if mk in all_months:
                    acc_monthly_increase[mk] = acc_monthly_increase.get(mk, 0) + inc
    total_inc_acc = {m: fid_income_monthly.get(m, 0) + acc_monthly_increase.get(m, 0)
                     for m in all_months}
    total_inc_acc_row = cur_row  # remember for later formula update
    cur_row = write_total_row(cur_row, "Total Income and Accumulations", total_inc_acc, C_SUM_T)

    # ── Calculations section (placed AFTER all assets rows) ───────────────────
    calc_start = cur_row + 1  # one blank row after Total Income and Accumulations

    C_CALC_H = "17202A"
    P_CALC_H = P(C_CALC_H)
    P_CALC_A = P("EAECF0")

    ws.merge_cells(start_row=calc_start, start_column=1, end_row=calc_start, end_column=n_sum_cols)
    ch = ws.cell(row=calc_start, column=1, value="Summary Calculations")
    ch.font = F(bold=True, color="FFFFFF", size=12); ch.fill = P_CALC_H
    ch.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[calc_start].height = 24

    calc_hdr = calc_start + 1
    for col_idx, m in enumerate(sum_months, 4):
        is_hist = m in hist_months
        c = ws.cell(row=calc_hdr, column=col_idx,
                    value=m.strftime("%b %Y") if not (col_idx <= 3) else None)
        c.font = F(bold=True, color="AAAAAA" if is_hist else "FFFFFF")
        c.fill = P("BFC9CA") if is_hist else P_CALC_H
        c.alignment = Alignment(horizontal="right", vertical="center")
    for col in (1, 2, 3):
        ws.cell(row=calc_hdr, column=col).fill = P_CALC_H
    ws.row_dimensions[calc_hdr].height = 16

    def mcol(m): return get_column_letter(4 + sum_months.index(m)) if m in sum_months else None

    def calc_row_fn(r, label, formulas_by_month, alt=False, bold=False, font_color="000000"):
        fill = P_CALC_A if alt else None
        font = F(bold=bold, color=font_color)
        c = ws.cell(row=r, column=1, value=label)
        c.font = font; c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="CCCCCC"))
        if fill: c.fill = fill
        for col in (2, 3):
            bc = ws.cell(row=r, column=col)
            bc.border = Border(bottom=Side(style="thin", color="CCCCCC"))
            if fill: bc.fill = fill
        for col_idx, m in enumerate(sum_months, 4):
            formula = formulas_by_month.get(m)
            cell = ws.cell(row=r, column=col_idx)
            if formula is None:
                cell.value = None
            elif isinstance(formula, str) and formula.startswith("="):
                # Write as proper Excel formula — never as text
                cell.data_type = "f"
                cell.value = formula
                cell.number_format = '#,##0;(#,##0);"-"'
            elif isinstance(formula, (int, float)):
                cell.value = int(round(formula))
                cell.number_format = '#,##0;(#,##0);"-"'
            else:
                cell.value = formula
            cell.font = font
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="CCCCCC"))
            if fill: cell.fill = fill
        return r + 1

    # Scan summary rows to find key row numbers
    row_refs = {}
    for _r in range(3, calc_start):
        v = ws.cell(row=_r, column=1).value
        if v: row_refs[str(v).strip()] = _r

    gw_row      = row_refs.get("TOTAL FAMILY WEALTH", grand_row)
    fid_row_sum = row_refs.get("Total Fidelity")
    paul_row_s  = row_refs.get("Total Paul's pensions")
    susan_row_s = row_refs.get("Total Susan's pensions")
    assets_row  = row_refs.get("Total Other Assets")
    house_row   = next((r for k, r in row_refs.items() if k.startswith("House")), None)
    car_row     = next((r for k, r in row_refs.items() if "Cars" in k), None)
    arriva_row  = next((r for k, r in row_refs.items() if "Arriva" in k), None)
    expleo_row  = next((r for k, r in row_refs.items() if "Expleo" in k), None)

    calc_cur = calc_hdr + 1

    # 1. Total — all assets incl cash and property (= TOTAL FAMILY WEALTH)
    calc_cur = calc_row_fn(calc_cur, "Total",
        {m: f"={mcol(m)}{gw_row}" for m in all_months},
        alt=True, bold=True, font_color="17202A")

    # 2. Investments & Cash — Total minus House and Car
    def inv_formula(m):
        parts = [f"{mcol(m)}{gw_row}"]
        if house_row: parts.append(f"-{mcol(m)}{house_row}")
        if car_row:   parts.append(f"-{mcol(m)}{car_row}")
        return "=" + "".join(parts)
    calc_cur = calc_row_fn(calc_cur, "Investments & Cash",
        {m: inv_formula(m) for m in sum_months}, alt=False)

    # 3. Yearly Increase — this month's TOTAL FAMILY WEALTH minus same month last year
    # Prior year values come from all_grand dict (which includes historic data)
    yearly = {}
    for i, m in enumerate(sum_months):
        m_minus_12 = m - 12
        prior_val = all_grand.get(m_minus_12, 0)
        if prior_val:
            # Use direct cell reference for current month, subtract hardcoded prior value
            yearly[m] = f"={mcol(m)}{gw_row}-{int(prior_val)}"
    calc_cur = calc_row_fn(calc_cur, "Yearly Increase", yearly, alt=True)

    # 4. PS Pension — Paul: Fidelity SIPP + Arriva + Expleo SW
    paul_fid_r = next((r for k, r in row_refs.items() if "SIPP Savings (Paul)" in k or "2000001606" in k), None)
    def ps_formula(m):
        parts = []
        if paul_fid_r:  parts.append(f"{mcol(m)}{paul_fid_r}")
        if arriva_row:  parts.append(f"{mcol(m)}{arriva_row}")
        if expleo_row:  parts.append(f"{mcol(m)}{expleo_row}")
        return ("=" + "+".join(parts)) if parts else None
    calc_cur = calc_row_fn(calc_cur, "PS Pension (Paul)", {m: ps_formula(m) for m in sum_months}, alt=False)

    # 5. SS Pension — Susan: total pensions row (incl SIPP)
    calc_cur = calc_row_fn(calc_cur, "SS Pension (Susan)",
        {m: f"={mcol(m)}{susan_row_s}" for m in sum_months} if susan_row_s else {},
        alt=True)

    # 6. Monthly Change — vs previous month
    mc_formulas = {}
    for i, m in enumerate(sum_months):
        if i > 0:
            mc_formulas[m] = f"={mcol(m)}{gw_row}-{mcol(sum_months[i-1])}{gw_row}"
    calc_cur = calc_row_fn(calc_cur, "Monthly Change", mc_formulas, alt=False)

    # 7. Av Monthly Change — average monthly change since start of all data
    av_formulas = {}
    for i, m in enumerate(sum_months):
        if i >= 2:
            av_formulas[m] = f"=ROUND(({mcol(m)}{gw_row}-{mcol(sum_months[0])}{gw_row})/{i},0)"
    calc_cur = calc_row_fn(calc_cur, "Av Monthly Change", av_formulas, alt=True)

    # 8. Monthly Investment Increase — income generated from investments (fid_pivot)
    monthly_inv = {}
    for m in sum_months:
        inc = round(fid_pivot[m].sum()) if m in fid_pivot.columns else 0
        if inc: monthly_inv[m] = inc
    calc_cur = calc_row_fn(calc_cur, "Monthly Investment Increase", monthly_inv,
                           alt=False, font_color="145A32")

    # --- shared state out (generated) ---
    ctx.calc_cur = calc_cur
    ctx.hist_months = hist_months
    ctx.sum_months = sum_months
    ctx.total_inc_acc_row = total_inc_acc_row
