"""The Investment Risk Metrics table and the Targets table (both later moved onto
their own sheet by the finish phase).

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import json
import os

import pandas as pd
from openpyxl.styles import Alignment, Font

from ..constants import EQUITY_DIVIDENDS_ANNUAL
from ..loaders import load_income_history
from .style import F, P


def _resolve_cell_num(ws, cell):
    """Numeric value of a cell, resolving a simple '=SUM(range)' formula by summing
    the referenced cells (recursively). The retirement-plan section reads section
    totals back off the sheet, and those totals are written as SUM formulas — a
    plain float() on them raises 'could not convert string to float: =SUM(...)'.
    Returns 0.0 for anything non-numeric it can't resolve."""
    v = cell.value
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.strip().upper().startswith('=SUM(') and v.strip().endswith(')'):
        rng = v.strip()[5:-1]
        total = 0.0
        try:
            for row in ws[rng]:
                cells = row if isinstance(row, tuple) else (row,)
                for c in cells:
                    total += _resolve_cell_num(ws, c)
        except (ValueError, KeyError):
            return 0.0
        return total
    try:
        return float(v or 0)
    except (ValueError, TypeError):
        return 0.0


def write_targets(ctx):
    # --- shared state in (generated) ---
    GROUPS_ORDER = ctx.GROUPS_ORDER
    METRIC_ACCS_ORDER = ctx.METRIC_ACCS_ORDER
    acc_holdings = ctx.acc_holdings
    all_months = ctx.all_months
    anchors = ctx.anchors
    combo_row = ctx.combo_row
    fid_pivot = ctx.fid_pivot
    metric_cash = ctx.metric_cash
    metric_data = ctx.metric_data
    spend_pivot = ctx.spend_pivot
    ws = ctx.ws

    # ── Write table ─────────────────────────────────────────────────────────────
    C_MET_H = "2C3E50"    # dark slate header
    C_MET_A = "F2F3F4"    # light grey alt
    C_MET_T = "D5D8DC"    # grey total
    C_MET_S = "E8F4F8"    # light blue for Shares
    C_MET_I = "EAF5EA"    # light green for Income
    C_MET_N = "FEF9E7"    # light yellow for Non-Income

    GROUP_COLOURS = {
        'Shares':           C_MET_S,
        'Income Funds':     C_MET_I,
        'Non-Income Funds': C_MET_N,
    }

    met_start = combo_row + 3 if 'combo_row' in dir() else ws.max_row + 3
    n_acc_cols = len(METRIC_ACCS_ORDER)
    n_met_cols = 1 + n_acc_cols + 1  # Group label + accounts + Total

    # Section title
    ws.merge_cells(start_row=met_start, start_column=1,
                   end_row=met_start, end_column=n_met_cols)
    t = ws.cell(row=met_start, column=1, value="Investment Risk Metrics")
    t.font = F(bold=True, color="FFFFFF", size=12)
    t.fill = P(C_MET_H)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[met_start].height = 24

    # Column headers
    hdr_row = met_start + 1
    ws.cell(row=hdr_row, column=1, value="Group").font = F(bold=True, color="FFFFFF", size=10)
    ws.cell(row=hdr_row, column=1).fill = P(C_MET_H)
    ws.cell(row=hdr_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for ci, (acc, label) in enumerate(METRIC_ACCS_ORDER, 2):
        c = ws.cell(row=hdr_row, column=ci, value=label)
        c.font = F(bold=True, color="FFFFFF", size=9)
        c.fill = P(C_MET_H)
        c.alignment = Alignment(horizontal="right", vertical="center")
    tot_hdr = ws.cell(row=hdr_row, column=2 + n_acc_cols, value="TOTAL")
    tot_hdr.font = F(bold=True, color="FFFFFF", size=10)
    tot_hdr.fill = P(C_MET_H)
    tot_hdr.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[hdr_row].height = 18

    # Data rows — value rows
    val_rows = {}
    cur_met = hdr_row + 1
    grand_met_total = sum(
        sum(metric_data[g][acc] for acc, _ in METRIC_ACCS_ORDER)
        for g in GROUPS_ORDER
    ) + sum(metric_cash.values())

    all_groups = GROUPS_ORDER + ['Cash']
    for gi, grp in enumerate(all_groups):
        row_fill = GROUP_COLOURS.get(grp, C_MET_A)
        r = cur_met
        val_rows[grp] = r
        # Label
        lbl = ws.cell(row=r, column=1, value=grp)
        lbl.font = F(bold=False, color="000000", size=10)
        lbl.fill = P(row_fill)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        # Account values
        row_total = 0
        for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
            val = metric_data[grp][acc] if grp in metric_data else metric_cash.get(acc, 0)
            if grp == 'Cash':
                val = metric_cash.get(acc, 0)
            row_total += val
            cell = ws.cell(row=r, column=ci)
            if val > 0:
                cell.value = round(val)
                cell.number_format = '#,##0'
            cell.fill = P(row_fill)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = F(size=10)
        # Row total
        tc = ws.cell(row=r, column=2 + n_acc_cols)
        if row_total > 0:
            tc.value = round(row_total)
            tc.number_format = '#,##0'
        tc.fill = P(row_fill)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.font = F(bold=True, size=10)
        ws.row_dimensions[r].height = 16
        cur_met += 1

    # % of total rows
    pct_start = cur_met
    cur_met += 1  # blank row
    ws.row_dimensions[cur_met - 1].height = 6

    for gi, grp in enumerate(all_groups):
        row_fill = GROUP_COLOURS.get(grp, C_MET_A)
        r = pct_start + gi
        lbl = ws.cell(row=r, column=1, value=f"{grp} %")
        lbl.font = F(italic=True, color="555555", size=9)
        lbl.fill = P(row_fill)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        row_total = sum(metric_data[grp][acc] for acc, _ in METRIC_ACCS_ORDER) if grp in metric_data else sum(metric_cash.values())
        for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
            val = metric_data[grp][acc] if grp in metric_data else metric_cash.get(acc, 0)
            if grp == 'Cash':
                val = metric_cash.get(acc, 0)
            acc_total_val = metric_data['_cash'].get(acc, 0)
            if acc_total_val > 0:
                pct = val / acc_total_val * 100
                cell = ws.cell(row=r, column=ci)
                if pct > 0:
                    cell.value = round(pct, 1)
                    cell.number_format = '0.0"%"'
                cell.fill = P(row_fill)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = F(italic=True, size=9, color="555555")
        # Overall % of portfolio
        overall_pct = row_total / grand_met_total * 100 if grand_met_total > 0 else 0
        tc = ws.cell(row=r, column=2 + n_acc_cols)
        if overall_pct > 0:
            tc.value = round(overall_pct, 1)
            tc.number_format = '0.0"%"'
        tc.fill = P(row_fill)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.font = F(bold=True, italic=True, size=9, color="555555")
        ws.row_dimensions[r].height = 14
        cur_met += 1

    # Grand total row
    cur_met += 1
    gt_row = cur_met
    ws.cell(row=gt_row, column=1, value="TOTAL INVESTMENTS").font = F(bold=True, color="FFFFFF", size=10)
    ws.cell(row=gt_row, column=1).fill = P(C_MET_H)
    ws.cell(row=gt_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
        acc_total_val = metric_data['_cash'].get(acc, 0)
        cell = ws.cell(row=gt_row, column=ci)
        if acc_total_val > 0:
            cell.value = round(acc_total_val)
            cell.number_format = '#,##0'
        cell.fill = P(C_MET_H)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = F(bold=True, color="FFFFFF", size=10)
    tc = ws.cell(row=gt_row, column=2 + n_acc_cols)
    fid_total_val = sum(metric_data['_cash'].get(acc, 0) for acc, _ in METRIC_ACCS_ORDER)
    if fid_total_val > 0:
        tc.value = round(fid_total_val)
        tc.number_format = '#,##0'
    tc.fill = P(C_MET_H)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.font = F(bold=True, color="FFFFFF", size=10)
    ws.row_dimensions[gt_row].height = 18

    # Note row
    note_met = gt_row + 1
    nc = ws.cell(row=note_met, column=1,
                 value="Values as at AccountSummary export date. Cash = account total minus classified assets. Totals match Fidelity accounts section above.")
    nc.font = Font(italic=True, color="888888", size=8)
    nc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=note_met, start_column=1,
                   end_row=note_met, end_column=n_met_cols)
    ws.row_dimensions[note_met].height = 12

    # ── Targets Table ───────────────────────────────────────────────────────────
    C_TGT_H  = "1A3A5C"   # dark navy header
    C_TGT_ON = "E9F7EF"   # green — on/above target
    C_TGT_OF = "FDEDEC"   # red   — below target
    C_TGT_NR = "FEF9E7"   # amber — within 10% of target
    C_TGT_LB = "EBF5FB"   # blue  — label rows

    tgt_start = note_met + 3
    # Columns: Metric | Actual | Target | Status | Notes
    TGT_COLS = ["Metric", "Actual", "Target", "Status", "Notes"]
    TGT_WIDTHS = [38, 16, 16, 12, 40]

    # Title
    ws.merge_cells(start_row=tgt_start, start_column=1, end_row=tgt_start, end_column=5)
    tt = ws.cell(row=tgt_start, column=1, value="2026 Targets")
    tt.font = F(bold=True, color="FFFFFF", size=12)
    tt.fill = P(C_TGT_H)
    tt.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tgt_start].height = 24

    # Headers
    hdr_tgt = tgt_start + 1
    for ci, (h, w) in enumerate(zip(TGT_COLS, TGT_WIDTHS), 1):
        c = ws.cell(row=hdr_tgt, column=ci, value=h)
        c.font = F(bold=True, color="FFFFFF", size=10)
        c.fill = P(C_TGT_H)
        c.alignment = Alignment(horizontal="left" if ci == 1 else "right", vertical="center")
    ws.row_dimensions[hdr_tgt].height = 18

    # ── Compute actuals ────────────────────────────────────────────────────────
    # Both pivots already carry the whole year: pinned history for the early
    # months, transaction data for the covered ones and estimates for the rest
    # (see estimate_future_months). So the annual figures are just their row
    # sums — no month literals, and nothing to re-tune as the calendar moves.
    inc_hist_data = load_income_history()
    fid_annual = 0
    if not fid_pivot.empty:
        cols = [m for m in all_months if m in fid_pivot.columns]
        fid_annual = float(fid_pivot[cols].sum().sum())
        # Any pinned history month the pivot has no column for
        for month_vals in inc_hist_data.values():
            for period, val in month_vals.items():
                if period not in fid_pivot.columns:
                    fid_annual += val

    salary_annual = 0
    if 'Salary' in spend_pivot.index:
        cols = [m for m in all_months if m in spend_pivot.columns]
        salary_annual = float(spend_pivot.loc['Salary', cols].sum())

    equity_annual = EQUITY_DIVIDENDS_ANNUAL
    total_annual_income = fid_annual + salary_annual + equity_annual
    income_avg_pm = round(total_annual_income / 12)

    # Emit the monthly SHARE-dividend figures for the Investment Dashboard's Monthly
    # Dividend metric (user request 2026-07-18: add share income + accumulation
    # dividends). share_income = equity dividends run at equity_annual/12; share
    # accumulation = the Acc funds' reinvested income (inc_ppu × units), already
    # computed per fund/month in build_acc_holdings as 'price_appreciation' — take the
    # latest month's total across all Acc funds (same figure as the Wealth Summary's
    # 'TOTAL Accumulative (price appreciation)' row). Written to data/ for the dashboard;
    # income-fund revenue stays sourced from the master's Income Funds tab (no dup).
    try:
        acc_appr = {}
        for _funds in (acc_holdings or {}).values():
            for _fd in _funds.values():
                for _per, _v in (_fd.get('price_appreciation') or {}).items():
                    acc_appr[_per] = acc_appr.get(_per, 0) + _v
        _dividends = {
            'generated_at': pd.Timestamp.now().isoformat(timespec='seconds'),
            'share_income_monthly': round(equity_annual / 12.0, 2),
            'share_accumulation_monthly': round(acc_appr[max(acc_appr)], 2) if acc_appr else 0.0,
        }
        # Repo root. This file moved from scripts/spending_summary.py to
        # scripts/spending/sheet/sheet_targets.py on 2026-07-19, so it is now FOUR
        # levels down, not two — dashboard_data.py reads this JSON for the Monthly
        # Dividend figure, and a wrong path here fails silently.
        _repo = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
        _dp = os.path.join(_repo, 'data', 'spending_dividends.json')
        os.makedirs(os.path.dirname(_dp), exist_ok=True)
        with open(_dp, 'w', encoding='utf-8') as _f:
            json.dump(_dividends, _f, indent=2)
        print(f"  Share dividends -> income £{_dividends['share_income_monthly']:,.0f}/mo, "
              f"accumulation £{_dividends['share_accumulation_monthly']:,.0f}/mo")
    except Exception as _e:
        print(f"  (share-dividend export skipped: {_e})")

    # These targets are all "what is it worth NOW" figures, so they read the
    # snapshot month's column — month columns start at column 4, and _r is a
    # 0-indexed tuple. This used to be a hardcoded _r[7] (May) and quietly went on
    # reporting a stale month once the export moved on.
    _CUR = 3 + all_months.index(anchors.data_month)

    # 2. Paul SIPP 25% drawdown
    paul_sipp_val_tgt = 0
    for _r in ws.iter_rows():
        if '2000001606' in str(_r[0].value or '') and 'SIPP' in str(_r[0].value or ''):
            paul_sipp_val_tgt = _resolve_cell_num(ws, _r[_CUR])
            break
    paul_drawdown = round(paul_sipp_val_tgt * 0.25)

    # 3. Susan SIPP 25% drawdown
    susan_sipp_val_tgt = 0
    for _r in ws.iter_rows():
        if '2000001604' in str(_r[0].value or '') and 'SIPP' in str(_r[0].value or ''):
            susan_sipp_val_tgt = _resolve_cell_num(ws, _r[_CUR])
            break
    susan_drawdown = round(susan_sipp_val_tgt * 0.25)

    # 4. ISA values combined
    paul_isa_val = susan_isa_val = 0
    for _r in ws.iter_rows():
        if str(_r[0].value or '').strip() == 'Investment ISA (Paul)':
            paul_isa_val = _resolve_cell_num(ws, _r[_CUR])
        if str(_r[0].value or '').strip() == 'Investment ISA (Susan)':
            susan_isa_val = _resolve_cell_num(ws, _r[_CUR])
    isa_combined = round(paul_isa_val + susan_isa_val)

    # 5. Growth % (Shares + Non-Income Funds) and Income % of total invested
    shares_val = inc_fund_val = non_inc_val = 0
    for _r in ws.iter_rows():
        v = str(_r[0].value or '').strip()
        if v == 'Shares':
            shares_val = _resolve_cell_num(ws, _r[9])
        elif v == 'Income Funds':
            inc_fund_val = _resolve_cell_num(ws, _r[9])
        elif v == 'Non-Income Funds':
            non_inc_val = _resolve_cell_num(ws, _r[9])
    total_invested = shares_val + inc_fund_val + non_inc_val
    growth_pct  = round((shares_val + non_inc_val) / total_invested * 100, 1) if total_invested else 0
    income_pct  = round(inc_fund_val / total_invested * 100, 1) if total_invested else 0

    # 6. Fidelity service fees — sum from live TransactionHistory
    import csv as _csv3, io as _io3
    import sys as _sys3
    _th_path = next((a for a in _sys3.argv[1:] if 'TransactionHistory' in a or 'transaction' in a.lower()), 'TransactionHistory.csv')
    with open(_th_path) as _f3:
        _lines3 = _f3.read().replace('\r','').split('\n')
    _start3 = next(i for i,l in enumerate(_lines3) if l.startswith('Order date'))
    _rows3 = list(_csv3.DictReader(_io3.StringIO('\n'.join(_lines3[_start3:]))))
    svc_fees_live = sum(abs(float(r.get('Amount','0') or 0))
                        for r in _rows3
                        if r.get('Transaction type','').strip() == 'Service Fee'
                        and r.get('Status','').strip() == 'Completed')
    # Annualise from Jan-May actual (5 months)
    svc_fees_annual = round(svc_fees_live / 5 * 12)

    def status_cell(ws, r, c, actual, target, higher_is_better=True, is_pct=False):
        """Write RAG status cell."""
        if target == 0:
            ws.cell(row=r, column=c, value="—")
            return "—"
        pct_of_tgt = actual / target
        if higher_is_better:
            if pct_of_tgt >= 1.0:   rag, txt = C_TGT_ON, "✓ On Target"
            elif pct_of_tgt >= 0.9: rag, txt = C_TGT_NR, "⚠ Near Target"
            else:                    rag, txt = C_TGT_OF, "✗ Below Target"
        else:  # lower is better (e.g. costs)
            if pct_of_tgt <= 1.0:   rag, txt = C_TGT_ON, "✓ Within Budget"
            elif pct_of_tgt <= 1.1: rag, txt = C_TGT_NR, "⚠ Slightly Over"
            else:                    rag, txt = C_TGT_OF, "✗ Over Budget"
        cell = ws.cell(row=r, column=c, value=txt)
        cell.fill = P(rag)
        cell.font = F(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return rag

    def tgt_row(ws, r, metric, actual, target, status_rag, note,
                actual_fmt="£{:,.0f}", target_fmt="£{:,.0f}", alt=False):
        bg = status_rag if status_rag not in (True, False) else (C_TGT_LB if alt else "FFFFFF")
        # Metric label
        c = ws.cell(row=r, column=1, value=metric)
        c.font = F(size=10); c.fill = P(C_TGT_LB)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Actual
        act_cell = ws.cell(row=r, column=2, value=actual_fmt.format(actual) if actual_fmt else actual)
        act_cell.font = F(bold=True, size=11)
        act_cell.fill = P(status_rag if isinstance(status_rag, str) and len(status_rag)==6 else "FFFFFF")
        act_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Target
        tgt_cell = ws.cell(row=r, column=3, value=target_fmt.format(target) if target_fmt else target)
        tgt_cell.font = F(size=10, color="444444"); tgt_cell.fill = P("F8F9FA")
        tgt_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Note
        nc = ws.cell(row=r, column=5, value=note)
        nc.font = F(italic=True, size=9, color="666666"); nc.fill = P(C_TGT_LB)
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22

    # ── Write target rows ──────────────────────────────────────────────────────
    tgt_data_start = hdr_tgt + 1

    # Row 1: Income per month
    r = tgt_data_start
    rag = status_cell(ws, r, 4, income_avg_pm, 10000, higher_is_better=True)
    tgt_row(ws, r,
            "Income per month (avg over 12 months)",
            income_avg_pm, 10000, rag,
            f"Annual total £{int(total_annual_income):,} ÷ 12. Includes salary, Fidelity fund income & equity dividends.")

    # Row 2: Paul SIPP 25% drawdown
    r += 1
    PAUL_DRAWDOWN_TAKEN = 0
    rag2 = status_cell(ws, r, 4, PAUL_DRAWDOWN_TAKEN + paul_drawdown, 269000, higher_is_better=True)
    tgt_row(ws, r,
            "Paul Pension — 25% tax-free drawdown available",
            paul_drawdown, 269000, rag2,
            f"SIPP value £{int(paul_sipp_val_tgt):,} × 25% = £{paul_drawdown:,}. Taken to date: £0.")

    # Row 3: Susan SIPP 25% drawdown
    r += 1
    SUSAN_DRAWDOWN_TAKEN = 0
    rag3 = status_cell(ws, r, 4, SUSAN_DRAWDOWN_TAKEN + susan_drawdown, 269000, higher_is_better=True)
    tgt_row(ws, r,
            "Susan Pension — 25% tax-free drawdown available",
            susan_drawdown, 269000, rag3,
            f"SIPP value £{int(susan_sipp_val_tgt):,} × 25% = £{susan_drawdown:,}. Taken to date: £0.")

    # Row 4: ISA values
    r += 1
    rag4 = status_cell(ws, r, 4, isa_combined, 1000000, higher_is_better=True)
    tgt_row(ws, r,
            "ISA Values (Paul ISA + Susan ISA combined)",
            isa_combined, 1000000, rag4,
            f"Paul ISA £{int(paul_isa_val):,} + Susan ISA £{int(susan_isa_val):,}.")

    # Row 5: Growth funds % — target is to REACH 60%, so higher is better
    r += 1
    rag5 = status_cell(ws, r, 4, growth_pct, 60, higher_is_better=True)
    tgt_row(ws, r,
            "Growth funds % of total invested (Shares + Non-Income Funds)",
            growth_pct, 60, rag5,
            f"Shares £{int(shares_val):,} + Non-Income £{int(non_inc_val):,} = £{int(shares_val+non_inc_val):,} of £{int(total_invested):,}.",
            actual_fmt="{:.1f}%", target_fmt="{:.0f}%")

    # Row 6: Income fund % — target is to stay AT/BELOW 40%, lower is better
    r += 1
    rag6 = status_cell(ws, r, 4, income_pct, 40, higher_is_better=False)
    tgt_row(ws, r,
            "Income funds % of total invested",
            income_pct, 40, rag6,
            f"Income Funds £{int(inc_fund_val):,} of £{int(total_invested):,} total invested. Currently {income_pct:.1f}% — target is to reduce to 40%.",
            actual_fmt="{:.1f}%", target_fmt="{:.0f}%")

    # Row 7: Fidelity service costs
    r += 1
    rag7 = status_cell(ws, r, 4, svc_fees_annual, 3000, higher_is_better=False)
    tgt_row(ws, r,
            "Fidelity annual service costs",
            svc_fees_annual, 3000, rag7,
            f"Jan–May actual £{svc_fees_live:,.2f} × 12/5 = £{svc_fees_annual:,} annualised estimate.")

    # Single comprehensive sheet
    ws.title = "Wealth Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    # --- shared state out (generated) ---
    ctx.ws = ws
