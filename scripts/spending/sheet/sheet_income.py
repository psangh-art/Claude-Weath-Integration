"""Sections 2 and 3: income by account (with the salary row) and the accumulative
holdings block.

A phase of the Wealth Summary build, split out of write_excel on 2026-07-19 (that
function was 2,115 lines). The body below is the ORIGINAL code, moved verbatim;
`ctx` carries the state the phases hand between each other, and the unpack/repack
lines around it were generated from a read/write analysis rather than written by
hand — keep them in step with any new shared variable.
"""
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from ..constants import ACCOUNT_LABELS, ACCOUNT_OWNER, EQUITY_DIVIDENDS_INLINE, FAMILY_ORDER
from .style import (BODY_FONT, EST_BODY, EST_FONT, EST_HDR_FILL, F, FIDA_FILL, FIDT_FILL,
                    FID_FILL, FID_HFONT, NUM_FMT, P, SAL_FILL, SAL_FONT)


def write_income(ctx):
    # --- shared state in (generated) ---
    acc_fund_map = ctx.acc_fund_map
    acc_holdings = ctx.acc_holdings
    actual_months = ctx.actual_months
    actual_set = ctx.actual_set
    fid_month_labels = ctx.fid_month_labels
    fid_months = ctx.fid_months
    fid_pivot = ctx.fid_pivot
    holdings = ctx.holdings
    reimb_note_row = ctx.reimb_note_row
    spend_pivot = ctx.spend_pivot
    ws = ctx.ws

    # ── Spacer rows ────────────────────────────────────────────────────────────
    fid_start_row = reimb_note_row + 2
    tot_r = fid_start_row  # fallback; overwritten when income section is written



    # ── Section 2: Income ─────────────────────────────────────────────────────
    if not fid_pivot.empty:
        # Build combined income rows: Salary first, then Fidelity accounts
        fid_accounts = list(fid_pivot.index)
        # Account + Units + Total + months
        n_fid_cols = 3 + len(fid_months)

        # Section title
        ws.merge_cells(start_row=fid_start_row, start_column=1,
                       end_row=fid_start_row, end_column=n_fid_cols)
        ft = ws.cell(row=fid_start_row, column=1, value="Income")
        ft.font = F(bold=True, color="FFFFFF", size=12)
        ft.fill = FID_FILL
        ft.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[fid_start_row].height = 24

        # Column headers
        fid_hdr_row = fid_start_row + 1
        # Income table: col1=Account, col2=Units, col3=Total, col4+=months
        fid_headers = ["Account", "Units", "Total"] + fid_month_labels
        for col, h in enumerate(fid_headers, 1):
            c = ws.cell(row=fid_hdr_row, column=col, value=h)
            if col <= 3:
                c.font = FID_HFONT
                c.fill = FID_FILL
            else:
                m = fid_months[col - 4]
                c.font = FID_HFONT if m in actual_set else EST_FONT
                c.fill = FID_FILL if m in actual_set else EST_HDR_FILL
            c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
        ws.row_dimensions[fid_hdr_row].height = 18

        # ── Salary row ────────────────────────────────────────────────────────
        sal_row = fid_start_row + 2
        sal_label = ws.cell(row=sal_row, column=1, value="Salary")
        sal_label.font = SAL_FONT
        sal_label.fill = SAL_FILL
        sal_label.alignment = Alignment(horizontal="left", vertical="center")
        sal_label.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        # Units col (col 2) — blank for salary
        su = ws.cell(row=sal_row, column=2, value=None)
        su.fill = SAL_FILL
        su.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(fid_months))
        sal_tot = ws.cell(row=sal_row, column=3,
                          value=f"=SUM({fc}{sal_row}:{lc}{sal_row})")
        sal_tot.number_format = NUM_FMT
        sal_tot.font = SAL_FONT
        sal_tot.fill = SAL_FILL
        sal_tot.alignment = Alignment(horizontal="right", vertical="center")
        sal_tot.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        for col_idx, m in enumerate(fid_months, 4):
            val = spend_pivot.loc["Salary", m] if "Salary" in spend_pivot.index else 0
            is_est = m not in actual_set
            cell = ws.cell(row=sal_row, column=col_idx,
                           value=int(round(val)) or None)
            cell.number_format = NUM_FMT
            cell.font = SAL_FONT if not is_est else F(bold=True, color="778866")
            cell.fill = SAL_FILL if not is_est else P("EEF5E8")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        # ── Per-person sections ───────────────────────────────────────────────
        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(fid_months))
        current_row = sal_row + 1
        account_total_rows = []
        equity_data_rows = []

        for person, person_accs in acc_fund_map.items():
            # Person sub-header — spans all cols
            ph = ws.cell(row=current_row, column=1, value=person)
            ph.font = F(bold=True, color="FFFFFF", size=9)
            ph.fill = P("1E8449")
            ph.alignment = Alignment(horizontal="left", vertical="center")
            for col in range(2, 4 + len(fid_months)):
                ws.cell(row=current_row, column=col).fill = P("1E8449")
            current_row += 1

            for acc, fund_df in person_accs.items():
                label = ACCOUNT_LABELS.get(acc, acc)
                funds_in_acc = list(fund_df.index)
                n_fund_rows = len(funds_in_acc)

                acc_r = current_row
                account_total_rows.append(acc_r)

                # Col 1: account label
                c = ws.cell(row=acc_r, column=1, value=label)
                c.font = F(bold=True, color="145A32")
                c.fill = FIDT_FILL
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                # Col 2: units — blank for account header row
                uc = ws.cell(row=acc_r, column=2, value=None)
                uc.fill = FIDT_FILL
                uc.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                # Col 3: total
                tc = ws.cell(row=acc_r, column=3,
                             value=f"=SUM({fc}{acc_r+1}:{lc}{acc_r+n_fund_rows})")
                tc.number_format = NUM_FMT
                tc.font = F(bold=True, color="145A32")
                tc.fill = FIDT_FILL
                tc.alignment = Alignment(horizontal="right", vertical="center")
                tc.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                for col_idx, m in enumerate(fid_months, 4):
                    val = fid_pivot.loc[acc, m] if acc in fid_pivot.index else 0
                    is_est = m not in actual_set
                    cell = ws.cell(row=acc_r, column=col_idx,
                                   value=int(round(val)) or None)
                    cell.number_format = NUM_FMT
                    cell.font = F(bold=True, color="145A32") if not is_est else F(bold=True, color="558855")
                    cell.fill = FIDT_FILL if not is_est else P("C8E6C9")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                current_row += 1

                # Fund rows — col1=indented name, col2=units, col3=total, col4+=months
                for fund_offset, fund in enumerate(funds_in_acc):
                    fr = current_row
                    fill = FIDA_FILL if (fund_offset % 2 == 0) else None

                    # Col 1: fund name indented
                    c = ws.cell(row=fr, column=1, value=f"  {fund}")
                    c.font = BODY_FONT
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: c.fill = fill

                    # Col 2: units held — try exact match then normalised match
                    units = holdings.get((acc, fund), 0)
                    if not units:
                        # Strip all spaces for comparison
                        def norm(s): return s.replace(" ", "").upper()
                        units = next(
                            (v for (a, f), v in holdings.items()
                             if a == acc and norm(f) == norm(fund)),
                            0
                        )
                    uc = ws.cell(row=fr, column=2,
                                 value=round(units, 2) if units > 0 else None)
                    uc.font = BODY_FONT
                    uc.number_format = '#,##0.##'
                    uc.alignment = Alignment(horizontal="right", vertical="center")
                    uc.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: uc.fill = fill

                    # Col 3: total
                    tc = ws.cell(row=fr, column=3,
                                 value=f"=SUM({fc}{fr}:{lc}{fr})")
                    tc.number_format = NUM_FMT
                    tc.font = BODY_FONT
                    tc.alignment = Alignment(horizontal="right", vertical="center")
                    tc.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: tc.fill = fill

                    # Col 4+: month values
                    for col_idx, m in enumerate(fid_months, 4):
                        val = fund_df.loc[fund, m] if fund in fund_df.index else 0
                        is_est = m not in actual_set
                        fill2 = (FIDA_FILL if (fund_offset % 2 == 0) else None) if not is_est else (P("DCF0DC") if (fund_offset % 2 == 0) else P("EBF5EB"))
                        cell = ws.cell(row=fr, column=col_idx,
                                       value=int(round(val)) or None)
                        cell.number_format = NUM_FMT
                        cell.font = BODY_FONT if not is_est else EST_BODY
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill2: cell.fill = fill2

                    current_row += 1

                # ── Equity dividends — write inline under Paul SIPP (2000001606) ──
                if acc == "2000001606":
                    for eq_offset, (eq_label, eq_divs) in enumerate(EQUITY_DIVIDENDS_INLINE):
                        eq_row = current_row
                        equity_data_rows.append(eq_row)
                        fill_eq = FIDA_FILL if (eq_offset % 2 == 0) else None
                        c = ws.cell(row=eq_row, column=1, value=eq_label)
                        c.font = F(color="145A32", italic=True, size=10)
                        c.alignment = Alignment(horizontal="left", vertical="center")
                        c.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: c.fill = fill_eq
                        ws.cell(row=eq_row, column=2).border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: ws.cell(row=eq_row, column=2).fill = fill_eq

                        # Total column (col 3)
                        tc_eq = ws.cell(row=eq_row, column=3, value=f"=SUM({fc}{eq_row}:{lc}{eq_row})")
                        tc_eq.number_format = NUM_FMT
                        tc_eq.font = F(color="145A32", italic=True, size=10)
                        tc_eq.alignment = Alignment(horizontal="right", vertical="center")
                        tc_eq.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: tc_eq.fill = fill_eq
                        for col_idx, m in enumerate(fid_months, 4):
                            val = eq_divs.get(str(m), 0)
                            cell = ws.cell(row=eq_row, column=col_idx)
                            if val:
                                cell.value = val
                                cell.number_format = NUM_FMT
                            cell.font = F(color="145A32", italic=True, size=10)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                            if fill_eq: cell.fill = fill_eq
                        ws.row_dimensions[eq_row].height = 15
                        current_row += 1

        # Note row
        note_eq = current_row
        nc = ws.cell(row=note_eq, column=1,
                     value="  * Sep/Nov equity dividend amounts are estimates based on prior year interim payments")
        nc.font = Font(italic=True, color="999999", size=8)
        nc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=note_eq, start_column=1,
                       end_row=note_eq, end_column=3 + len(fid_months))
        ws.row_dimensions[note_eq].height = 12
        current_row += 1

        # June provisional note
        note_jun = current_row
        nj = ws.cell(row=note_jun, column=1,
                     value="  * June 2026 figures are PROVISIONAL — based on partial transaction data (60-day export, 10 Jun) "
                           "and May-value estimates for fund income/salary. Will be revised once full June data is provided.")
        nj.font = Font(italic=True, bold=True, color="B7791F", size=8)
        nj.fill = P("FFF8E1")
        nj.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=note_jun, start_column=1,
                       end_row=note_jun, end_column=3 + len(fid_months))
        ws.row_dimensions[note_jun].height = 14
        current_row += 1

        # Grand total — Salary + all fund accounts + equity dividends
        tot_r = current_row
        c = ws.cell(row=tot_r, column=1, value="Total Income")
        c.font = F(bold=True, color="145A32")
        c.fill = FIDT_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=tot_r, column=2).fill = FIDT_FILL
        all_income_rows = [sal_row] + account_total_rows + equity_data_rows
        for col_idx in range(3, 4 + len(fid_months)):
            cl = get_column_letter(col_idx)
            refs = "+".join(f"{cl}{rr}" for rr in all_income_rows)
            cell = ws.cell(row=tot_r, column=col_idx, value=f"={refs}")
            cell.number_format = NUM_FMT
            cell.font = F(bold=True, color="145A32")
            cell.fill = FIDT_FILL
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Section 3: Accumulative Holdings ─────────────────────────────────────
    # Estimate monthly growth for Acc funds using the income yield from Inc equivalents
    # Yield proxy: monthly income per account / Inc fund value → apply to Acc fund value

    # Build yield map from Inc fund income data
    # Key: normalised fund base name → monthly income rate
    ACC_INC_EQUIV = {
        "Aegon High Yield Bond B Acc":               "Aegon High Yield Bond B Inc",
        "Schroder High Yield Opportunities Fund Z Acc": "Schroder High Yield Opportunities Fund Z Inc",
        "WS Guinness Global Energy Fund I Acc":       None,  # no Inc equivalent — use 6% annual
        "Man High Yield Opportunities Fund Prof D Acc": "Man High Yield Opportunities Fund Prof D Inc",
    }

    # Estimate annual yield rates from actual income / May 26 value
    def est_annual_yield(inc_fund_name, acc_value):
        """Estimate annual yield for an Acc fund from its Inc equivalent's income rate."""
        if inc_fund_name is None:
            return 0.06  # 6% default for unknown
        # Sum total 2026 income across all accounts for this fund from fid_pivot
        total_inc = 0
        if inc_fund_name in fid_pivot.index:
            total_inc = fid_pivot.loc[inc_fund_name, actual_months].sum()
        # Annualise: divide by months of data, multiply by 12
        months_of_data = len(actual_months)
        if months_of_data > 0 and total_inc > 0:
            # Find total Inc fund value from AccountSummary
            inc_total_val = sum(
                v for (a, f), v_dict in {}.items()
            )
            # Fall back: use the Acc fund value as proxy
            monthly_rate = total_inc / acc_value if acc_value > 0 else 0
            return monthly_rate * 12
        return 0.07  # 7% default

    C_ACC_H = "0E4D6B"   # teal-navy header
    C_ACC_A = "E8F4F8"   # light teal alt
    C_ACC_T = "A8D5E2"   # teal total

    ACC_FILL  = P(C_ACC_H)
    ACCA_FILL = P(C_ACC_A)
    ACCT_FILL = P(C_ACC_T)
    ACC_HFONT = F(bold=True, color="FFFFFF")
    ACC_TFONT = F(bold=True, color="0E4D6B")

    acc3_start = tot_r + 2

    n_acc_cols = 3 + len(fid_months)   # Label + Units + Total + months (same as income table)
    ws.merge_cells(start_row=acc3_start, start_column=1,
                   end_row=acc3_start, end_column=n_acc_cols)
    t_acc = ws.cell(row=acc3_start, column=1, value="Accumulative Holdings")
    t_acc.font = F(bold=True, color="FFFFFF", size=12)
    t_acc.fill = ACC_FILL
    t_acc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[acc3_start].height = 24

    acc_hdr_row = acc3_start + 1
    for col, h in enumerate(["Account / Fund", "Units", "Total"] + fid_month_labels, 1):
        c = ws.cell(row=acc_hdr_row, column=col, value=h)
        if col <= 3:
            c.font = ACC_HFONT; c.fill = ACC_FILL
        else:
            m = fid_months[col - 4]
            c.font = ACC_HFONT if m in actual_set else EST_FONT
            c.fill = ACC_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[acc_hdr_row].height = 18

    acc_cur_row = acc_hdr_row + 1
    acc_account_total_rows = []

    for person in FAMILY_ORDER:
        person_accs = [acc for acc, owner in ACCOUNT_OWNER.items() if owner == person]
        for acc in person_accs:
            if acc not in acc_holdings:
                continue
            funds = acc_holdings[acc]
            label = f"{ACCOUNT_LABELS.get(acc, acc)} ({person})"
            n_fund_rows = len(funds)

            # Account header row
            acc_ar = acc_cur_row
            acc_account_total_rows.append(acc_ar)

            c = ws.cell(row=acc_ar, column=1, value=label)
            c.font = ACC_TFONT; c.fill = ACCT_FILL
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            ws.cell(row=acc_ar, column=2).fill = ACCT_FILL  # units blank for account
            ws.cell(row=acc_ar, column=2).border = Border(bottom=Side(style="thin", color=C_ACC_T))

            fc_a = get_column_letter(4)
            lc_a = get_column_letter(3 + len(fid_months))
            tc = ws.cell(row=acc_ar, column=3,
                         value=f"=SUM({fc_a}{acc_ar+1}:{lc_a}{acc_ar+n_fund_rows})")
            tc.number_format = NUM_FMT; tc.font = ACC_TFONT; tc.fill = ACCT_FILL
            tc.alignment = Alignment(horizontal="right", vertical="center")
            tc.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            # Sum of account values for each month (from fund rows below)
            for col_idx, m in enumerate(fid_months, 4):
                cell = ws.cell(row=acc_ar, column=col_idx,
                               value=f"=SUM({get_column_letter(col_idx)}{acc_ar+1}:{get_column_letter(col_idx)}{acc_ar+n_fund_rows})")
                cell.number_format = NUM_FMT; cell.font = ACC_TFONT; cell.fill = ACCT_FILL
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            acc_cur_row += 1

            # Fund rows — use actual price×units from transaction history
            for fund_offset, (fund, fund_data) in enumerate(sorted(funds.items())):
                fr = acc_cur_row
                is_alt = (fund_offset % 2 == 0)
                fill = ACCA_FILL if is_alt else None
                units = fund_data["units"]
                # Use price_appreciation (ignores new unit purchases, just price-driven gain)
                month_vals = fund_data.get("price_appreciation", fund_data.get("monthly_values", {}))
                may_val = fund_data["value"]

                # For future months not in price history, extrapolate using known annual yield
                # Annual yields from Fidelity (guaranteed non-negative)
                FUND_YIELDS = {
                    "Aegon High Yield Bond B Acc":                  0.0732,
                    "Schroder High Yield Opportunities Fund Z Acc": 0.0769,
                    "WS Guinness Global Energy Fund I Acc":         0.0235,
                }
                annual_yield = FUND_YIELDS.get(fund, 0.05)
                monthly_growth = (1 + annual_yield) ** (1/12) - 1

                if month_vals:
                    last_known = max(month_vals.keys())
                    last_val = month_vals[last_known]
                    for m in fid_months:
                        if m not in month_vals:
                            offset = (m - last_known).n
                            month_vals[m] = round(last_val * ((1 + monthly_growth) ** offset))

                c = ws.cell(row=fr, column=1, value=f"  {fund}")
                c.font = BODY_FONT; c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: c.fill = fill

                uc = ws.cell(row=fr, column=2, value=round(units, 2) if units else None)
                uc.number_format = '#,##0.##'; uc.font = BODY_FONT
                uc.alignment = Alignment(horizontal="right", vertical="center")
                uc.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: uc.fill = fill

                tc2 = ws.cell(row=fr, column=3,
                              value=f"=SUM({fc_a}{fr}:{lc_a}{fr})")
                tc2.number_format = NUM_FMT; tc2.font = BODY_FONT
                tc2.alignment = Alignment(horizontal="right", vertical="center")
                tc2.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: tc2.fill = fill

                for col_idx, m in enumerate(fid_months, 4):
                    val = month_vals.get(m, 0)
                    is_est = m not in actual_set
                    cell = ws.cell(row=fr, column=col_idx, value=int(round(val)) if val else None)
                    cell.number_format = NUM_FMT
                    cell.font = BODY_FONT if not is_est else EST_BODY
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                    if fill: cell.fill = fill

                acc_cur_row += 1

    # Grand total row — shows INCREASE each month (new units × price), not total value
    tot_acc_r = acc_cur_row
    c = ws.cell(row=tot_acc_r, column=1, value="TOTAL Accumulative (price appreciation)")
    c.font = F(bold=True, color="FFFFFF"); c.fill = ACC_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=tot_acc_r, column=2).fill = ACC_FILL
    ws.cell(row=tot_acc_r, column=3).fill = ACC_FILL

    for col_idx, m in enumerate(fid_months, 4):
        monthly_increase = sum(
            fd_h.get("price_appreciation", {}).get(m, 0)
            for funds_h in acc_holdings.values()
            for fd_h in funds_h.values()
        )
        cell = ws.cell(row=tot_acc_r, column=col_idx,
                       value=int(round(monthly_increase)) if monthly_increase else None)
        cell.number_format = NUM_FMT; cell.font = F(bold=True, color="FFFFFF")
        cell.fill = ACC_FILL; cell.alignment = Alignment(horizontal="right", vertical="center")
    acc_cur_row += 1

    # --- shared state out (generated) ---
    ctx.acc_cur_row = acc_cur_row
    ctx.tot_acc_r = tot_acc_r
    ctx.tot_r = tot_r
