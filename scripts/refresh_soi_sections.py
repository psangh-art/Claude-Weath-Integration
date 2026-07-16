#!/usr/bin/env python3
"""Rebuild the 'Stocks of Interest' SECTION TABLES from the current pipeline data.

These tables (below the auto-built below-alert block) were hand-maintained, and drifted
badly: audited 2026-07-15, 17 of 25 rows sat in the wrong section and 21 of 25 carried
an Alert Low that no longer matched the pipeline, all stamped 'Last Updated 2026-07-06'.
Worse, the drift ran the wrong way — BEZ (+15.5%) and AUTO (+11.4%) were still listed as
'AT LOWER BOUNDARY — highest priority' while nine stocks genuinely at their buy point
(DGE, AZN, GSK, GOLD, HSBA, CRDA, PLAT, PRU, ALW) sat in lower-priority sections. The
user asked for it to be pipeline-maintained instead (2026-07-15).

WHAT IS OWNED BY WHOM
  Pipeline-owned, rewritten every run:
    C Pattern       - from channel_detect's pattern string (the seven agreed patterns)
    E Alert Low     - Investments col L
    G Alert High    - Investments col O
    P Last Updated  - today
    section placement + within-section ordering (nearest to its alert low first)
  Regenerated every run (row-relative formulas — they MUST be rewritten, not copied,
  or a row that moves keeps pointing at its old row):
    D Proximity, F Current, H Upside %, I P/E, J Div Yield, Q TradingView link
  Hand-written, CARRIED per ticker and never invented:
    A Stock, K Chart Note, L Analyst Rating, M Holdings, N Target Value, O Notes

MEMBERSHIP IS NOT AUTOMATIC. The tables are a curated watchlist: this script re-sections
and refreshes the stocks already listed, and never adds or drops one. Add a stock by
hand and the next run will place and maintain it. (Contrast the below-alert block above,
which IS a full auto-generated list.)

Usage:
  python refresh_soi_sections.py [workbook.xlsx] [--apply]
    default workbook ~/Downloads/Stocks_Buy_Strategy.xlsx; without --apply it reports
    the plan and writes nothing.
"""
import os
import re
import sys
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from ticker_normalize import normalize

SHEET = 'Stocks of Interest'
INV = 'Investments'

# The region this script owns. Row 41/42 (title + strategy) and row 87 (summary) are
# rewritten in place; 43-85 is the rebuildable band. Row 90 onward (the FTSE promotion
# tables) must NEVER be touched — hence a hard cap rather than inserting rows.
# Band enlarged 43-79 -> 43-85 and the reference block shifted 84->90 on 2026-07-16
# (shift_soi_reference_down_2026-07-16.py) when the curated watchlist outgrew 37 rows.
TITLE_ROW, STRATEGY_ROW = 41, 42
REGION_START, REGION_END = 43, 85
SUMMARY_ROW = 87
LAST_COL = 17

# +1 vs the pre-2026-07-16 layout: a 'Marked Up' column was inserted at Investments!B.
INV_TICKER_COL, INV_LOW_COL, INV_HIGH_COL = 4, 13, 16

FONT_NAME = 'Arial'
WHITE = 'FFFFFFFF'
SUBHEAD_FILL = 'FF2E5077'
_THIN = Side(style='thin')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADER = ['Stock', 'Ticker', 'Pattern', 'Proximity', 'Alert Low', 'Current', 'Alert High',
          'Upside %', 'P/E', 'Div Yield', 'Chart Note', 'Analyst Rating', 'Holdings (£)',
          'Target Value (£)', 'Notes', 'Last Updated', 'TradingView']

# (key, band text, band fill, data tint) — same palette as the existing tables.
SECTIONS = [
    ('at_lower',  '🟢  AT LOWER BOUNDARY — within 5% of alert low (Highest priority)',
     'FF1A5733', 'FFF0FFF4'),
    ('near',      '🟡  NEAR LOWER BOUNDARY — 5–15% above alert low (Watch closely)',
     'FF9C6500', 'FFFFFDF0'),
    ('watch',     '⬜  WATCHLIST — 15–30% above alert low (Monitor)',
     'FF44546A', 'FFF7F9FC'),
    ('breakout',  '⬆   BREAKOUTS — Broken above resistance, momentum',
     'FF1F4E79', 'FFF2F7FC'),
    ('beyond',    '⬛  BEYOND 30% — well above alert low (Reference only)',
     'FF595959', 'FFF7F7F7'),
]
SECTION_TITLE = {k: t for k, t, _, _ in SECTIONS}

# channel_detect pattern string -> the short label the sheet shows.
PATTERN_LABELS = [
    ('wedge', 'price broke ABOVE', 'Wedge — broken above'),
    ('wedge', None, 'Wedge'),
    ('price ON a drawn line', None, 'On the line'),
    ('price INSIDE channel', None, 'Parallel channel'),
    ('price ABOVE channel', None, 'Breakout above channel'),
    ('price BELOW channel', None, 'Below channel — band governs'),
    ('trend lines only', None, 'Trend lines'),
    ('no lines read', None, 'No lines drawn'),
    ('blue channel (no price', None, 'Parallel channel'),
]


def emit(text='', err=False):
    """print() that survives Windows' cp1252 console. The section bands carry emoji, and
    a bare print of them raises UnicodeEncodeError — the same crash verify_pipeline.py
    hit in 2026-07-10 while printing a report it had already written correctly."""
    stream = sys.stderr if err else sys.stdout
    try:
        print(text, file=stream)
    except UnicodeEncodeError:
        stream.flush()
        stream.buffer.write((text + '\n').encode('utf-8'))
        stream.buffer.flush()


def pattern_label(rec):
    """The sheet's Pattern label for one channel_detect record.

    'No lines drawn' and 'not read this run' must not be conflated. A chart whose axis
    OCR failed, or whose read the price-bracket guard withheld, still HAS the user's
    drawings on it — labelling it 'No lines drawn' would say his trendlines had
    vanished. It also matters because such a row keeps the Alert Low from an earlier
    run: the label is the only signal that the level is inherited rather than fresh.
    (2026-07-15: 7 of 353 charts — AUTO, NWG, TW., SMT, BA, ENT, III — read cleanly at
    09:22 but not in the 16:00 capture, six of them withheld by the bracket guard.)"""
    if not rec:
        return 'No chart'
    pat = rec.get('pattern')
    if not rec.get('kind'):
        # channel_detect sets pattern='no lines read' for BOTH causes — an undrawn chart
        # and a failed axis read — so the pattern string cannot tell them apart. Only
        # `reason` can, and every axis-failure reason names the axis: 'no OCR-readable
        # axis labels', 'fewer than 3 clean axis labels...', 'OCR axis labels [x-y] do
        # not bracket known price'. An undrawn chart says 'no channel or trend line
        # found near price' (UKW).
        reason = (rec.get('reason') or '').lower()
        if 'macro/reference' in reason:
            return 'Not a buy-list chart'
        if 'axis' in reason:
            return 'Not read this run — level inherited'
        return 'No lines drawn'
    if not pat:
        return ''
    for head, extra, label in PATTERN_LABELS:
        if pat.startswith(head) and (extra is None or extra in pat):
            return label
    return ''


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def harvest(ws):
    """Per-ticker records for every stock currently listed in the section tables.

    Keyed by ticker so a stock keeps its hand-written columns wherever it lands. GSK was
    listed twice (rows 58 and 77, 'Near Lower' and 'Breakouts'); duplicates collapse to
    one record, preferring whichever row carries more hand-written content."""
    def weight(rec):
        return sum(1 for k in ('chart_note', 'rating', 'notes', 'holdings', 'target')
                   if rec.get(k) not in (None, ''))
    found = {}
    for r in range(REGION_START, REGION_END + 1):
        t = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=1).value
        if not isinstance(t, str) or not t.strip() or t.strip() == 'Ticker':
            continue
        if not isinstance(name, str) or not name.strip():
            continue
        rec = {'ticker': t.strip(), 'name': name.strip(),
               'chart_note': ws.cell(row=r, column=11).value,
               'rating': ws.cell(row=r, column=12).value,
               'holdings': ws.cell(row=r, column=13).value,
               'target': ws.cell(row=r, column=14).value,
               'notes': ws.cell(row=r, column=15).value}
        prev = found.get(rec['ticker'])
        if prev is None or weight(rec) > weight(prev):
            found[rec['ticker']] = rec
    return list(found.values())


def classify(gap, price, high):
    if isinstance(high, (int, float)) and isinstance(price, (int, float)) and high and price > high:
        return 'breakout'
    if gap is None:
        return None
    if gap < 0.05:
        return 'at_lower'
    if gap < 0.15:
        return 'near'
    if gap < 0.30:
        return 'watch'
    return 'beyond'


def build_plan(wb, prices, chart_ids, patterns):
    ws, inv = wb[SHEET], wb[INV]
    levels = {}
    for r in range(3, inv.max_row + 1):
        t = inv.cell(row=r, column=INV_TICKER_COL).value
        if isinstance(t, str) and t.strip():
            levels[t.strip().upper()] = (inv.cell(row=r, column=INV_LOW_COL).value,
                                         inv.cell(row=r, column=INV_HIGH_COL).value)

    rows, unplaced = [], []
    for rec in harvest(ws):
        key = rec['ticker'].upper()
        low, high = levels.get(key, (None, None))
        price = prices.get(key)
        gap = ((price - low) / low) if (isinstance(price, (int, float))
                                        and isinstance(low, (int, float)) and low) else None
        sec = classify(gap, price, high)
        if sec is None:
            # No live price or no Alert Low — the row cannot be placed by proximity.
            # Keep it (never silently drop a curated stock) and say so.
            unplaced.append(rec['ticker'])
            sec = 'beyond'
        rec.update({'low': low, 'high': high, 'price': price, 'gap': gap, 'section': sec,
                    'pattern': pattern_label(patterns.get(key)),
                    'chart_id': chart_ids.get(key)})
        rows.append(rec)

    # Nearest to its alert low first — the sheet's whole purpose is 'what is close?'.
    rows.sort(key=lambda x: (x['gap'] if x['gap'] is not None else 99))
    return rows, unplaced


def needed_rows(rows):
    n = 0
    for key, _, _, _ in SECTIONS:
        members = [x for x in rows if x['section'] == key]
        if members:
            n += 2 + len(members) + 1   # band + header + data + trailing blank
    return n


def write_plan(wb, rows, today):
    ws = wb[SHEET]
    # Unmerge BEFORE clearing: the section bands are merged (A43:O43 and friends), and a
    # MergedCell's .value is read-only — writing to one raises AttributeError.
    for m in [str(m) for m in ws.merged_cells.ranges
              if REGION_START <= m.min_row <= REGION_END]:
        ws.unmerge_cells(m)
    # Clear the region so a shorter rebuild leaves no styled ghosts behind.
    for r in range(REGION_START, REGION_END + 1):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(name=FONT_NAME, size=9)

    r = REGION_START
    for key, band_text, band_fill, tint in SECTIONS:
        members = [x for x in rows if x['section'] == key]
        if not members:
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        b = ws.cell(row=r, column=1, value=band_text)
        b.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        b.fill = PatternFill(fill_type='solid', fgColor=band_fill)
        b.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 15
        r += 1

        for c, h in enumerate(HEADER, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(name=FONT_NAME, size=8, bold=True, color=WHITE)
            cell.fill = PatternFill(fill_type='solid', fgColor=SUBHEAD_FILL)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 15
        r += 1

        for m in members:
            gf = normalize(m['ticker']) or {}
            gt = gf.get('google_finance_ticker')
            vals = {
                1: m['name'], 2: m['ticker'], 3: m['pattern'],
                4: f'=IFERROR(TEXT((F{r}-E{r})/E{r},"0.0%")&" above low","")',
                5: m['low'],
                6: f"=IFERROR(VLOOKUP(B{r},'Investments'!$D:$J,7,FALSE()),\"\")",
                7: m['high'],
                8: f'=IFERROR((G{r}-E{r})/E{r},"")',
                9: (f'=IFERROR(googlefinance("{gt}","pe"),"")' if gt else None),
                10: f"=IFERROR(VLOOKUP(B{r},'Base Data'!$A:$Q,9,FALSE()),\"\")",
                11: m['chart_note'], 12: m['rating'], 13: m['holdings'],
                14: m['target'], 15: m['notes'], 16: today,
                17: (f'=HYPERLINK("https://www.tradingview.com/chart/{m["chart_id"]}/","📊 Layout")'
                     if m.get('chart_id') else None),
            }
            for c in range(1, LAST_COL + 1):
                cell = ws.cell(row=r, column=c, value=vals.get(c))
                cell.font = Font(name=FONT_NAME, size=9)
                cell.fill = PatternFill(fill_type='solid', fgColor=tint)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal='left', vertical='top')
            ws.cell(row=r, column=16).number_format = 'yyyy-mm-dd'
            r += 1
        r += 1
    return r - 1


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    apply = '--apply' in sys.argv
    path = args[0] if args else os.path.expanduser('~/Downloads/Stocks_Buy_Strategy.xlsx')

    manifest = load_json(os.path.join(SCRIPT_DIR, 'layout_manifest_tmp.json'))
    results = load_json(os.path.join(SCRIPT_DIR, 'channel_results_tmp.json'))
    prices, chart_ids = {}, {}
    for row in manifest:
        t = row.get('ticker')
        if not t:
            continue
        keys = {t.strip().upper()}
        mt = (normalize(t) or {}).get('master_ticker')
        if mt:
            keys.add(mt.upper())
        for k in keys:
            if isinstance(row.get('price'), (int, float)):
                prices[k] = row['price']
            if row.get('chartId'):
                chart_ids[k] = row['chartId']
    patterns = {}
    for rec in results:
        t = (rec.get('ticker') or '').strip().upper()
        if not t:
            continue
        patterns[t] = rec
        mt = (normalize(t) or {}).get('master_ticker')
        if mt:
            patterns.setdefault(mt.upper(), rec)

    wb = openpyxl.load_workbook(path)
    rows, unplaced = build_plan(wb, prices, chart_ids, patterns)

    need, avail = needed_rows(rows), REGION_END - REGION_START + 1
    emit(f'{len(rows)} curated stocks -> {need} rows needed, {avail} available '
         f'(rows {REGION_START}-{REGION_END})')
    for key, _, _, _ in SECTIONS:
        members = [x for x in rows if x['section'] == key]
        if members:
            emit(f'  {SECTION_TITLE[key][:44]:46} {len(members):2d}  '
                 + ', '.join(m['ticker'] for m in members))
    if unplaced:
        emit('  no live price or alert low (left in reference section): ' + ', '.join(unplaced))
    if need > avail:
        print(f'\nABORT: needs {need} rows but only {avail} are reserved before the FTSE '
              f'promotion tables at row 90. Remove a stock, or move those tables down '
              f'(shift_soi_reference_down_*.py + bump REGION_END/SUMMARY_ROW).',
              file=sys.stderr)
        return 2
    if not apply:
        print('\nDry run — nothing written. Re-run with --apply.')
        return 0

    today = datetime.now().date()
    last = write_plan(wb, rows, today)
    counts = {k: len([x for x in rows if x['section'] == k]) for k, _, _, _ in SECTIONS}
    ws = wb[SHEET]
    ws.cell(row=SUMMARY_ROW, column=1,
            value=('Summary: %d at boundary | %d near | %d watchlist | %d breakout | %d beyond'
                   '   — rebuilt by the pipeline %s'
                   % (counts['at_lower'], counts['near'], counts['watch'],
                      counts['breakout'], counts['beyond'], today.isoformat())))
    ws.cell(row=STRATEGY_ROW, column=1,
            value=('Strategy: Long only | Buy at lower parallel/trendline, sell at upper | '
                   'Sections and levels rebuilt each pipeline run — add or remove a stock by '
                   'hand and it will be placed and maintained from the next run.'))
    wb.save(path)
    print(f'\nWritten -> {path} (region ends row {last}; FTSE tables at 90 untouched)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
