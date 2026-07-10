#!/usr/bin/env python3
"""Build tradingview_layouts.xlsx from manifest JSON files produced by
export-layouts-excel.js. One row per individual chart (not per layout) on the
Charts sheet — a layout with 6 panes produces 6 rows, each with its own cropped
screenshot rather than one squeezed multi-pane grid image. Optional Indicators
and Alerts manifests add two more sheets to the same workbook, so chart images,
current indicator values, and price alerts all live in one file.

Charts manifest: [{"id": int, "chartId": str, "name": str, "ticker": str|None,
"description": str|None, "screenshot": str|None, "error": str|None}, ...]
Indicators manifest: [{"layoutId": int, "chartId": str, "layoutName": str,
"ticker": str|None, "company": str|None, "indicator": str, "field": str, "value": any}, ...]
Alerts manifest: [{"alertId": int, "symbol": str, "message": str, "conditionType": str|None,
"targetPrice": number|None, "resolution": str, "active": bool, "created": str,
"lastFired": str|None, "expiration": str|None}, ...]

Usage: python build_layout_excel.py <charts.json> <output.xlsx> [<indicators.json>] [<alerts.json>]
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from ticker_normalize import normalize

MAX_DISPLAY_WIDTH = 700

def add_charts_sheet(wb, rows):
    ws = wb.active
    ws.title = 'Charts'
    ws.append(['Layout ID', 'Chart ID', 'Layout Name', 'Symbol', 'Company',
               'Google Finance Ticker', 'Google Finance Formula', 'Screenshot'])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    widths = [15, 15, 32, 16, 32, 20, 40, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for i, row in enumerate(rows):
        r = i + 2
        if row.get('screenshot'):
            with PILImage.open(row['screenshot']) as im:
                w, h = im.size
            scale = min(1.0, MAX_DISPLAY_WIDTH / w)
            disp_w, disp_h = w * scale, h * scale

            img = XLImage(row['screenshot'])
            img.width, img.height = disp_w, disp_h
            ws.add_image(img, f"H{r}")
            screenshot_value = row['screenshot'].split('\\')[-1].split('/')[-1]
            ws.row_dimensions[r].height = disp_h * 0.75
        else:
            screenshot_value = f"FAILED - {row.get('error', 'unknown error')}"

        norm = normalize(row.get('ticker'))

        ws.cell(row=r, column=1, value=row['id'])
        ws.cell(row=r, column=2, value=row['chartId'])
        c3 = ws.cell(row=r, column=3, value=row['name'])
        c3.font = c3.font.copy(bold=True)
        ws.cell(row=r, column=4, value=row.get('ticker'))
        ws.cell(row=r, column=5, value=row.get('description'))
        ws.cell(row=r, column=6, value=norm['google_finance_ticker'] if norm else None)
        ws.cell(row=r, column=7, value=norm['google_finance_formula'] if norm else None)
        ws.cell(row=r, column=8, value=screenshot_value)
    return len(rows)

def add_simple_sheet(wb, title, header, rows, keys, bold_cols=()):
    ws = wb.create_sheet(title)
    ws.append(header)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for i, row in enumerate(rows):
        r = i + 2
        for ci, key in enumerate(keys):
            cell = ws.cell(row=r, column=ci + 1, value=row.get(key))
            if ci in bold_cols:
                cell.font = cell.font.copy(bold=True)
    return len(rows)

def main():
    charts_path, out_path = sys.argv[1], sys.argv[2]
    indicators_path = sys.argv[3] if len(sys.argv) > 3 else None
    alerts_path = sys.argv[4] if len(sys.argv) > 4 else None

    with open(charts_path, 'r', encoding='utf-8') as f:
        chart_rows = json.load(f)

    wb = Workbook()
    chart_count = add_charts_sheet(wb, chart_rows)
    print(f"Charts sheet: {chart_count} rows")

    if indicators_path:
        with open(indicators_path, 'r', encoding='utf-8') as f:
            indicator_rows = json.load(f)
        count = add_simple_sheet(
            wb, 'Indicators',
            ['Layout ID', 'Chart ID', 'Layout Name', 'Symbol', 'Company', 'Indicator', 'Field', 'Value'],
            indicator_rows,
            ['layoutId', 'chartId', 'layoutName', 'ticker', 'company', 'indicator', 'field', 'value'],
            bold_cols=(3,),
        )
        print(f"Indicators sheet: {count} rows")

    if alerts_path:
        with open(alerts_path, 'r', encoding='utf-8') as f:
            alert_rows = json.load(f)
        count = add_simple_sheet(
            wb, 'Alerts',
            ['Alert ID', 'Symbol', 'Message', 'Condition', 'Target Price', 'Resolution', 'Active', 'Created', 'Last Fired', 'Expiration'],
            alert_rows,
            ['alertId', 'symbol', 'message', 'conditionType', 'targetPrice', 'resolution', 'active', 'created', 'lastFired', 'expiration'],
            bold_cols=(1,),
        )
        print(f"Alerts sheet: {count} rows")

    wb.save(out_path)
    print(f"Saved workbook to {out_path}")

if __name__ == '__main__':
    main()
