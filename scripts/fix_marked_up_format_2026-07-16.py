#!/usr/bin/env python3
"""One-off: restyle Investments column B ('Marked Up') to MATCH column A ('Chart').

The 'Marked Up' column was inserted 2026-07-16 with its VALUES set but no styling,
so its data cells sat on openpyxl's Calibri/Arial-10, no-fill, centred defaults —
visibly out of step with the green/grey Yes/No 'Chart' flag right beside it. This
applies the same per-cell style the pipeline now writes (update_master_sheet.
set_marked_up_flag) to every existing data row, preserving each cell's current
Yes/No value, and fixes the B2 header alignment to match A2. Idempotent — re-running
is a no-op on already-correct cells. Safe to re-run.
"""
import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment

from update_master_sheet import (
    SHEET_NAME, HEADER_ROW, COL_CHART, COL_MARKED_UP, COL_TICKER,
    set_marked_up_flag,
)

PATH = os.path.expanduser('~/Downloads/Stocks_Buy_Strategy.xlsx')


def main():
    backup = f"{PATH}.bak-{datetime.now():%Y%m%d-%H%M%S}"
    shutil.copyfile(PATH, backup)
    print(f"Backup -> {backup}")

    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET_NAME]

    # Header B2: match A2 (left/top). Font/fill already match (Arial 9 bold white
    # on navy from the insert); only the alignment was centred.
    ws.cell(row=HEADER_ROW, column=COL_MARKED_UP).alignment = Alignment(
        horizontal='left', vertical='top')

    styled = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        chart_val = ws.cell(row=r, column=COL_CHART).value
        ticker = ws.cell(row=r, column=COL_TICKER).value
        if chart_val not in ('Yes', 'No') or not ticker:
            continue  # section-header / non-ticker row — skip, same rule as build_master_index
        cur = ws.cell(row=r, column=COL_MARKED_UP).value
        set_marked_up_flag(ws, r, str(cur).strip().lower() == 'yes')
        styled += 1

    wb.save(PATH)
    print(f"Styled {styled} data rows in column B; header alignment fixed. Saved -> {PATH}")


if __name__ == '__main__':
    main()
