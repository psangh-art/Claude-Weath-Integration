#!/usr/bin/env python3
"""One-off: give the 'Stocks of Interest' curated watchlist band more room.

The pipeline-rebuilt band (rows 43-79, 37 rows) has run out of space — the curated
watchlist grew to 24 stocks needing 39 rows, and refresh_soi_sections aborts rather
than overwrite the static FTSE promotion / demotion / dividend-key reference tables
that begin at row 84. This shifts that whole static block DOWN by 6 rows (84-135 ->
90-141), enlarging the reserved band to rows 43-85 (43 rows) with the summary moving
to row 87.

Safe because (verified before writing): the block below row 84 is entirely static
content — its only formulas are self-contained =HYPERLINK("literal url") links (no
relative refs to break), NOTHING elsewhere in the workbook references those rows, and
the sheet has no conditional formatting. Values, styles, merged ranges and row heights
are all carried to +6. Backs the workbook up first. Idempotent guard: refuses to run
twice by checking the FTSE title has already moved.

Run this ONCE, then the matching constant bump in refresh_soi_sections.py
(REGION_END 79->85, SUMMARY_ROW 81->87) must already be in place.
"""
import os
import shutil
from copy import copy
from datetime import datetime

import openpyxl

PATH = os.path.expanduser('~/Downloads/Stocks_Buy_Strategy.xlsx')
SHEET = 'Stocks of Interest'
SRC_START, SRC_END = 84, 135   # the static reference block, inclusive
SHIFT = 6                       # 84 -> 90; band grows 43-79 -> 43-85, summary 81 -> 87
LAST_COL = 17


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET]

    # Idempotence guard: FTSE title lives at A84 before the shift, A90 after.
    if ws.cell(row=SRC_START, column=1).value in (None, '') and \
       isinstance(ws.cell(row=SRC_START + SHIFT, column=1).value, str) and \
       'FTSE' in str(ws.cell(row=SRC_START + SHIFT, column=1).value):
        print('Already shifted (FTSE block sits at row 90) — nothing to do.')
        return

    backup = f"{PATH}.bak-{datetime.now():%Y%m%d-%H%M%S}"
    shutil.copyfile(PATH, backup)
    print(f"Backup -> {backup}")

    # 1. Capture every source cell's value + style, and the row heights.
    buf = {}
    heights = {}
    for r in range(SRC_START, SRC_END + 1):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            buf[(r, c)] = (cell.value, copy(cell._style))
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            heights[r] = rd.height

    # 2. Capture merged ranges fully inside the block, then unmerge them.
    merges = [(m.min_row, m.min_col, m.max_row, m.max_col)
              for m in ws.merged_cells.ranges
              if SRC_START <= m.min_row and m.max_row <= SRC_END + 0]
    for mn_r, mn_c, mx_r, mx_c in merges:
        ws.unmerge_cells(start_row=mn_r, start_column=mn_c, end_row=mx_r, end_column=mx_c)

    # 3. Clear the source block AND the destination tail (rows 84 .. 141), so the
    #    six vacated rows (84-89) end up truly empty and nothing stale lingers.
    for r in range(SRC_START, SRC_END + SHIFT + 1):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
            cell.border = openpyxl.styles.Border()
            cell.font = openpyxl.styles.Font()
            cell.alignment = openpyxl.styles.Alignment()
            cell.number_format = 'General'

    # 4. Write the captured content back at +SHIFT, styles and heights included.
    for (r, c), (val, style) in buf.items():
        dst = ws.cell(row=r + SHIFT, column=c)
        dst.value = val
        dst._style = copy(style)
    for r, h in heights.items():
        ws.row_dimensions[r + SHIFT].height = h

    # 5. Re-merge at the new position.
    for mn_r, mn_c, mx_r, mx_c in merges:
        ws.merge_cells(start_row=mn_r + SHIFT, start_column=mn_c,
                       end_row=mx_r + SHIFT, end_column=mx_c)

    wb.save(PATH)
    ftse = ws.cell(row=SRC_START + SHIFT, column=1).value
    print(f"Shifted rows {SRC_START}-{SRC_END} down {SHIFT} -> "
          f"{SRC_START + SHIFT}-{SRC_END + SHIFT}. FTSE block now at row "
          f"{SRC_START + SHIFT}: {str(ftse)[:40]!r}")
    print(f"Saved -> {PATH}")


if __name__ == '__main__':
    main()
