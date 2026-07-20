#!/usr/bin/env python3
"""Payslip PDF -> a row on the workbook's 'Payslip Summary' tab.

Owned by the Investment Dashboard's Payslips screen, NOT the pipeline (user decision
2026-07-19: "we don't need to include payslip uploads in the pipeline — we'll load it
via this new payslips screen"). The pipeline never calls this; the dashboard's
POST /api/payslips/upload does.

Two stages, deliberately separate so a payslip whose layout we don't recognise fails
loudly instead of writing a wrong number into a live financial sheet:

  extract_fields(pdf)  — pull the figures out of the PDF text
  append_row(fields)   — write them onto the tab, in the right tax-year band

Usage:
    python payslip_ingest.py <payslip.pdf> [--dry-run]
    python payslip_ingest.py <payslip.pdf> --dump      # print the raw text and stop
"""
import os
import re
import sys
import json
import datetime
import shutil

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKBOOK = os.path.join(os.path.expanduser('~'), 'Downloads', 'Stocks_Buy_Strategy.xlsx')
SHEET = 'Payslip Summary'
HEADER_ROW = 3
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARCHIVE = os.path.join(REPO, 'data', 'payslips')

# Column -> the labels a payslip might use for it, tried in order. Kept as data so a
# new payroll layout is a one-line addition rather than a code change.
FIELD_LABELS = {
    'gross':       ['gross pay', 'total gross pay', 'gross salary', 'total payments',
                    'salary'],
    'pension_ee':  ['pension ee', 'employee pension', 'pension (employee)',
                    'salary sacrifice', 'pension sacrifice', 'scottish widows ee',
                    'ee %'],
    'pension_er':  ['pension er', 'employer pension', 'pension (employer)',
                    'employers pension', "employer's pension"],
    'bonus':       ['bonus', 'employee referral', 'referral'],
    'taxable_pay': ['taxable pay', 'taxable gross'],
    'paye':        ['paye', 'paye tax', 'income tax', 'tax paid'],
    'ni':          ['national insurance', 'ni contribution', 'nic', 'employee ni'],
    'net':         ['net pay', 'total net pay', 'payment amount'],
}
MONEY = re.compile(r'-?[\d,]+\.\d{2}')
DATE_PATTERNS = [
    (re.compile(r'\b(\d{2})[/-](\d{2})[/-](\d{4})\b'), '%d/%m/%Y'),
    (re.compile(r'\b(\d{4})-(\d{2})-(\d{2})\b'), '%Y-%m-%d'),
]


class PayslipParseError(Exception):
    """The PDF didn't yield the fields we need — never guess, ask the human."""


def pdf_text(path):
    try:
        import pdfplumber
    except ImportError as e:
        raise PayslipParseError(
            'pdfplumber is not installed — run: python -m pip install pdfplumber') from e
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            out.append(page.extract_text() or '')
    text = '\n'.join(out)
    if not text.strip():
        raise PayslipParseError(
            'No text in this PDF — it looks like a scan/image. A text payslip is needed '
            '(or the file has to go through OCR first).')
    return text


def _money_after(line, label):
    """The first money-looking number on the line, after the label."""
    idx = line.lower().find(label)
    if idx < 0:
        return None
    m = MONEY.search(line, idx + len(label))
    return float(m.group(0).replace(',', '')) if m else None


def _period_lines(text):
    """Only the THIS-PERIOD lines. Everything under a 'CUMULATIVE YEAR TO DATE' header,
    and any 'YTD ...' line, carries the year-to-date totals — reading those as the
    period figure is exactly how gross came out as the YTD 'Total Gross Payments'
    instead of this month's salary. Drop them so the period scan can't reach them."""
    out = []
    for l in text.splitlines():
        s = l.strip()
        if not s:
            continue
        if 'cumulative year to date' in s.lower():
            break
        if s.lower().startswith('ytd'):
            continue
        out.append(s)
    return out


def _columnar_summary(lines):
    """Some payslips print the pay summary as a header ROW of column names with the
    values on the NEXT row (so _money_after finds nothing beside the label — this is
    why 'Net Pay' couldn't be read). Detect the header carrying both 'gross pay' and
    'net pay' and take the money tokens off the following line by position:
    Gross | PAYE | NIC | Others | Net. Gross is deliberately NOT taken here — it's the
    post-sacrifice figure; the sheet's Gross Salary comes from the 'Salary' pay line."""
    for i in range(len(lines) - 1):
        low = lines[i].lower()
        if 'gross pay' in low and 'net pay' in low and 'paye' in low:
            nums = MONEY.findall(lines[i + 1])
            if len(nums) == 5:
                v = [float(n.replace(',', '')) for n in nums]
                return {'paye': v[1], 'ni': v[2], 'net': v[4]}
            break
    return {}


def extract_fields(path):
    """Best-effort field extraction. Returns (fields, text, missing)."""
    text = pdf_text(path)
    lines = _period_lines(text)

    fields = {}
    # The two-line summary block is the only place Net Pay appears with its value.
    fields.update(_columnar_summary(lines))
    for key, labels in FIELD_LABELS.items():
        if key in fields:
            continue
        for label in labels:
            for line in lines:
                if label in line.lower():
                    val = _money_after(line, label)
                    if val is not None:
                        fields[key] = val
                        break
            if key in fields:
                break

    # Salary-sacrifice pension is quoted negative in the Payments column; the sheet
    # records the magnitude.
    if fields.get('pension_ee'):
        fields['pension_ee'] = abs(fields['pension_ee'])

    # Pay date: the latest date on the page that isn't in the future
    today = datetime.date.today()
    dates = []
    for pat, fmt in DATE_PATTERNS:
        for m in pat.finditer(text):
            try:
                d = datetime.datetime.strptime(m.group(0), fmt).date()
            except ValueError:
                continue
            if datetime.date(2000, 1, 1) <= d <= today:
                dates.append(d)
    if dates:
        fields['pay_date'] = max(dates).isoformat()

    if fields.get('pension_ee') and fields.get('pension_er'):
        fields['pension_total'] = round(fields['pension_ee'] + fields['pension_er'], 2)
    # Taxable Pay and Total Deductions are written as FORMULAS by append_row (to match
    # the neighbour rows), so they are deliberately NOT computed as values here.
    if fields.get('pay_date'):
        fields['tax_year'] = tax_year_of(fields['pay_date'])

    missing = [k for k in ('pay_date', 'gross', 'net') if not fields.get(k)]
    return fields, text, missing


def tax_year_of(iso_date):
    """UK tax year containing a date — 6 April to 5 April."""
    d = datetime.date.fromisoformat(iso_date)
    start = d.year if (d.month, d.day) >= (4, 6) else d.year - 1
    return f'{start}/{str(start + 1)[2:]}'


COLS = ['pay_date', 'tax_year', 'gross', 'pension_ee', 'pension_er', 'pension_total',
        'bonus', 'taxable_pay', 'paye', 'ni', 'deductions', 'net']


def _row_style(ws, row, template_row):
    """Copy the styling of an existing payslip row — a new row must match the format
    of the ones beside it, never openpyxl's defaults (CLAUDE.md rule, 2026-07-16)."""
    for col in range(1, len(COLS) + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=row, column=col)
        dst.font = Font(name=src.font.name, size=src.font.size, bold=src.font.bold,
                        color=src.font.color)
        dst.alignment = Alignment(horizontal=src.alignment.horizontal,
                                  vertical=src.alignment.vertical)
        dst.number_format = src.number_format
        if src.fill and src.fill.fgColor and src.fill.fgColor.rgb:
            dst.fill = src.fill.copy()
        dst.border = Border(bottom=Side(style='thin', color='D9E1F2'))


def append_row(fields, workbook=WORKBOOK, dry_run=False):
    """Insert the payslip into its tax-year band, keeping pay dates in order."""
    if not os.path.exists(workbook):
        raise PayslipParseError(f'Workbook not found: {workbook}')
    wb = openpyxl.load_workbook(workbook)
    if SHEET not in wb.sheetnames:
        raise PayslipParseError(f"'{SHEET}' tab not found in {os.path.basename(workbook)}")
    ws = wb[SHEET]

    pay_date = datetime.date.fromisoformat(fields['pay_date'])
    tax_year = fields.get('tax_year') or tax_year_of(fields['pay_date'])

    # Walk the sheet: find this tax year's band and where the date belongs in it.
    band_start = band_end = None
    in_band = False
    last_data_row = None
    insert_at = None
    for r in range(HEADER_ROW + 1, ws.max_row + 2):
        label = str(ws.cell(row=r, column=1).value or '').strip()
        if label.lower().startswith('tax year'):
            if in_band:
                break
            in_band = tax_year in label
            if in_band:
                band_start = r
            continue
        if not in_band:
            continue
        if label.upper().startswith('TOTAL'):
            band_end = r
            break
        existing = ws.cell(row=r, column=1).value
        d = existing.date() if isinstance(existing, datetime.datetime) else None
        if d is None:
            try:
                d = datetime.datetime.strptime(str(existing).strip(), '%d/%m/%Y').date()
            except (ValueError, TypeError):
                continue
        last_data_row = r
        if d == pay_date:
            raise PayslipParseError(
                f'A payslip dated {pay_date:%d/%m/%Y} is already on the sheet (row {r}).')
        if d > pay_date and insert_at is None:
            insert_at = r

    if band_start is None:
        raise PayslipParseError(
            f"No 'Tax Year {tax_year}' band on the {SHEET} tab — add the band (and its "
            f"TOTAL row) by hand once, then re-upload.")
    target = insert_at or band_end or (last_data_row + 1 if last_data_row else band_start + 1)
    template = last_data_row or band_start + 1

    if dry_run:
        wb.close()
        return {'row': target, 'tax_year': tax_year, 'written': False}

    ws.insert_rows(target)
    ws.cell(row=target, column=1, value=pay_date.strftime('%d/%m/%Y'))
    ws.cell(row=target, column=2, value=tax_year)
    for i, key in enumerate(COLS[2:], start=3):
        if key in ('taxable_pay', 'deductions'):
            continue  # written as formulas below, to match the neighbour rows
        val = fields.get(key)
        if val is not None:
            ws.cell(row=target, column=i, value=val)
    # Taxable Pay = Gross − Pension(EE); Total Deductions = PAYE + NI (col 8 and 11).
    ws.cell(row=target, column=8, value=f'=C{target}-D{target}')
    ws.cell(row=target, column=11, value=f'=I{target}+J{target}')
    _row_style(ws, target, template + (1 if template >= target else 0))

    # Extend the band's TOTAL SUM ranges to cover the inserted row. openpyxl does NOT
    # adjust formulas on insert, so the tax-year TOTAL would otherwise keep summing the
    # old range and silently leave the new payslip out. The TOTAL sat at band_end and
    # was pushed down one row by the insert. (Payslips always land in the latest — last
    # — band, so bands below aren't a concern.)
    if band_end is not None:
        total_row = band_end + 1
        first, last = band_start + 1, total_row - 1
        for c in range(3, len(COLS) + 1):
            cell = ws.cell(row=total_row, column=c)
            if isinstance(cell.value, str) and cell.value.upper().startswith('=SUM('):
                lc = get_column_letter(c)
                cell.value = f'=SUM({lc}{first}:{lc}{last})'

    backup = workbook + '.bak-' + datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    shutil.copyfile(workbook, backup)
    wb.save(workbook)
    wb.close()
    return {'row': target, 'tax_year': tax_year, 'written': True, 'backup': backup}


def archive_pdf(path, fields):
    """Keep the source PDF — the sheet row is a summary, the payslip is the record."""
    os.makedirs(ARCHIVE, exist_ok=True)
    stamp = fields.get('pay_date') or datetime.date.today().isoformat()
    dst = os.path.join(ARCHIVE, f'payslip_{stamp}.pdf')
    shutil.copyfile(path, dst)
    return dst


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    flags = {a for a in sys.argv[1:] if a.startswith('--')}
    if not args:
        print(__doc__)
        return 2
    path = args[0]

    try:
        if '--dump' in flags:
            print(pdf_text(path))
            return 0
        fields, text, missing = extract_fields(path)
        if missing:
            raise PayslipParseError(
                'Could not read ' + ', '.join(missing) + ' from this payslip. '
                'Run with --dump to see the text, then add the right labels to '
                'FIELD_LABELS in payslip_ingest.py.')
        result = append_row(fields, dry_run='--dry-run' in flags)
        if not result['written']:
            result['archived'] = None
        else:
            result['archived'] = archive_pdf(path, fields)
        print(json.dumps({'ok': True, 'fields': fields, **result}, indent=2))
        return 0
    except PayslipParseError as e:
        print(json.dumps({'ok': False, 'error': str(e)}, indent=2), file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
