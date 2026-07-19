#!/usr/bin/env python3
"""
Spending Summary Generator
--------------------------
Usage:
    python spending_summary.py                                    # uses defaults
    python spending_summary.py amex.csv barclays.csv fidelity.csv
    python spending_summary.py amex.csv barclays.csv fidelity.csv output.xlsx

Reads Amex (activity.csv), Barclays (data.csv) and Fidelity
(TransactionHistory.csv) and produces a categorised monthly spending summary
plus a Fidelity income-by-account section as a formatted Excel file.
"""

import sys
import os
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xlsx_sheet_copy import copy_sheet_into, copy_cell_style

# The reading/categorising/pivoting half of this report lives in the `spending`
# package (split out 2026-07-19 — this file was 4,020 lines). What stays here is
# the Excel writer and the CLI. Imported by name rather than `import *` so every
# cross-module use is greppable.
from spending.constants import (ACCOUNT_LABELS, ACCOUNT_OWNER, CATEGORIES,
                                EQUITY_DIVIDENDS_ANNUAL, EQUITY_DIVIDENDS_INLINE,
                                ESTIMATES_AS_OF, FAMILY_ORDER)
from spending.anchors import resolve_anchors
from spending.loaders import (load_amex, load_barclays, load_fidelity_income,
                              load_history, load_income_history)
from spending.holdings import build_acc_holdings, build_holdings
from spending.pivots import (build_account_fund_pivot, build_fidelity_pivot,
                             build_spending_pivot, build_summary_data,
                             estimate_future_months)

# Windows' default console codepage (cp1252) can't encode characters like
# "→" used in console-only status prints below — reconfigure to UTF-8 so a
# cosmetic console print can't crash the script after the workbook is
# already written (same class of bug fixed in verify_pipeline.py).
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass











# ── Excel styles helper ────────────────────────────────────────────────────────
NUM_FMT = '#,##0;(#,##0);"-"'






def _resolve_cell_num(ws, cell):
    """Numeric value of a cell, resolving a simple '=SUM(range)' formula by summing
    the referenced cells (recursively). The retirement-plan section reads section
    totals back off the sheet, and those totals are written as SUM formulas — a
    plain float() on them raises 'could not convert string to float: =SUM(...)'.
    Returns 0.0 for anything non-numeric it can't resolve."""
    v = cell.value
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.strip().upper().startswith('=SUM(') and v.strip().endswith(')'):
        rng = v.strip()[5:-1]
        total = 0.0
        try:
            for row in ws[rng]:
                cells = row if isinstance(row, tuple) else (row,)
                for c in cells:
                    total += _resolve_cell_num(ws, c)
        except (ValueError, KeyError):
            return 0.0
        return total
    try:
        return float(v or 0)
    except (ValueError, TypeError):
        return 0.0


# ── Write Excel ────────────────────────────────────────────────────────────────
def write_excel(spend_pivot, actual_months, future_months, fid_pivot,
                acc_fund_map, holdings, summary_data, acc_holdings, anchors,
                output_path, reimbursements=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wealth Summary"

    # Single comprehensive sheet — no separate tabs

    def fml(cell, formula):
        """Write an Excel formula to a cell, ensuring it is stored as a formula not text."""
        cell.data_type = "f"
        cell.value = formula
        cell.number_format = '#,##0;(#,##0);"-"'

    all_months = actual_months + future_months
    spend_months = all_months
    fid_months   = all_months

    spend_month_labels = [m.strftime("%b %Y") for m in all_months]
    fid_month_labels   = [m.strftime("%b %Y") for m in all_months]
    actual_set = set(actual_months)  # months with real CSV data (Jan–Apr)

    # ── History boundary ───────────────────────────────────────────────────────
    # Previous Data owns everything BEFORE this month; live files own this month
    # and everything after. It sits one month past the end of load_history(), so
    # it moves by itself when that pinned series is extended.
    HISTORY_CUTOFF = anchors.wealth_cutoff

    def is_history_month(m):
        """True if this month belongs to Previous Data (read-only)."""
        return m < HISTORY_CUTOFF

    def live_and_hist_safe(live_dict, hist_key_str):
        """Merge values with strict boundary:
        - Months BEFORE HISTORY_CUTOFF: Previous Data only (never overwritten by live)
        - Months FROM HISTORY_CUTOFF: live CSV files only
        """
        merged = {}
        hist_series = history.get(HIST_MAP.get(hist_key_str, ""), {})
        for m in all_months:
            if is_history_month(m):
                # History owns this month — ignore live value
                if m in hist_series:
                    merged[m] = hist_series[m]
                # If no history value exists, leave blank (don't backfill from live)
            else:
                # Live owns this month
                v = live_dict.get(m)
                if v:
                    merged[m] = v
        return merged

    def projection_with_hist_override(proj_dict, hist_key_str):
        """For formula-projected series (Expleo SW, Susan pensions, Cars):
        - Use hardcoded history value if it exists for a month (actuals take priority)
        - Otherwise use the live projection (covers all months incl. June+)
        """
        merged = {}
        hist_series = history.get(HIST_MAP.get(hist_key_str, ""), {})
        for m in all_months:
            if m in hist_series:
                merged[m] = hist_series[m]
            elif m in proj_dict:
                merged[m] = proj_dict[m]
        return merged


    C_NAVY   = "1F4E79"
    C_TEAL   = "1A5276"
    C_ALT    = "F2F7FB"
    C_SAL    = "E2EFDA"
    C_TOT    = "D9E1F2"
    C_FOOT   = "BDD7EE"
    C_FID_H  = "145A32"
    C_FID_A  = "EAF5EA"
    C_FID_T  = "A9DFBF"
    C_EST_H  = "4A4A6A"   # muted navy for estimated month headers
    C_EST_V  = "F7F7FC"   # very light lavender for estimated cells
    C_EST_A  = "EDEDF5"   # alt row estimated

    P = lambda c: PatternFill("solid", start_color=c, end_color=c)
    HDR_FILL  = P(C_NAVY)
    EST_HDR_FILL = P(C_EST_H)
    ALT_FILL  = P(C_ALT)
    EST_FILL  = P(C_EST_V)
    EST_ALT   = P(C_EST_A)
    SAL_FILL  = P(C_SAL)
    TOT_FILL  = P(C_TOT)
    FOOT_FILL = P(C_FOOT)
    FID_FILL  = P(C_FID_H)
    FIDA_FILL = P(C_FID_A)
    FIDT_FILL = P(C_FID_T)

    def F(bold=False, color="000000", size=10, name="Arial", italic=False):
        return Font(bold=bold, color=color, name=name, size=size, italic=italic)

    HDR_FONT  = F(bold=True,  color="FFFFFF")
    EST_FONT  = F(bold=True,  color="CCCCDD")   # dimmed for estimated headers
    BODY_FONT = F()
    EST_BODY  = F(color="555577")               # muted text for estimated values
    BOLD_FONT = F(bold=True)
    SAL_FONT  = F(bold=True,  color="375623")
    FID_HFONT = F(bold=True,  color="FFFFFF")
    FID_TFONT = F(bold=True,  color="145A32")

    def col_fill(m, is_alt, base_fill, base_alt):
        """Return appropriate fill for actual vs estimated column."""
        if m in actual_set:
            return base_alt if is_alt else None
        else:
            return EST_ALT if is_alt else EST_FILL

    def val_font(m, bold=False, sal=False):
        if sal: return SAL_FONT
        if m not in actual_set: return EST_BODY
        return BOLD_FONT if bold else BODY_FONT

    # ── Summary Table ──────────────────────────────────────────────────────────
    C_SUM_H  = "2C3E50"   # dark slate header
    C_SUM_A  = "F4F6F7"   # light alt row
    C_SUM_T  = "D5D8DC"   # total / subtotal
    C_SUM_S  = "1A5276"   # section subheader (dark blue)
    C_SUM_SA = "EAF2FF"   # section alt

    P_SUM_H  = P(C_SUM_H)
    P_SUM_A  = P(C_SUM_A)
    P_SUM_T  = P(C_SUM_T)
    P_SUM_S  = P(C_SUM_S)
    P_SUM_SA = P(C_SUM_SA)

    n_sum_cols = 3 + len(all_months)  # Label | blank | blank | month cols (live)
    # All tables: col1=Label, col2=Units/blank, col3=blank, col4+=months

    HIST_MAP = {
        "fidelity_all":      "All Fidelity accounts",
        "arriva":            "PS Arriva pension (Def Bene)",
        "expleo":            "Paul Explo Scottish Widows",
        "susan_fidelity":    "Susan Fidelity Pension",
        "capita":            "Capita (RMSPS) - to 2018 (65 yrs)",
        "rmpp":              "RMPP 2012 - 2023 (65 years)",
        "collective":        "Collective Pension (2023 +)",
        "cash_balance":      "Cash Balance",
        "avc":               "AVC Bonus Plan (Scottish widows)",
        "mf":                "MF",
        "house":             "House",
        "liam":              "Liam ISA",
        "jayne":             "Jaynes ISA",
        "cars":              "Cars Paul",
        "paul_pension":      "Paul Pension",
        "susan_pension":     "Susan Pension",
    }

    # ── Historic months (from Previous_Data CSV — used for calcs, NOT displayed) ─
    history = load_history()  # hardcoded — no file needed
    hist_periods_set = set()
    for series in history.values():
        hist_periods_set.update(series.keys())
    hist_months = sorted(p for p in hist_periods_set if p not in set(all_months))

    # Display only 2026 months — all tables share same columns
    sum_months = list(all_months)   # Jan–Dec 2026 only
    n_sum_cols = 3 + len(sum_months)

    # Helper: look up a value for a given month — history first for hist months,
    # live dict for live months. live_dict keys are pd.Period.
    def sum_val(live_dict, m):
        """Return value for month m: from live_dict if available, else history."""
        if m in live_dict and live_dict[m]:
            return live_dict[m]
        return None

    def write_total_row(r, label, monthly_dict, fill_col, font_col="042C53"):
        c = ws.cell(row=r, column=1, value=label)
        c.font = F(bold=True, color=font_col); c.fill = P(fill_col)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).fill = P(fill_col)
        ws.cell(row=r, column=3).fill = P(fill_col)
        for col_idx, m in enumerate(sum_months, 4):
            v = monthly_dict.get(m, 0)
            cell = ws.cell(row=r, column=col_idx, value=int(round(v)) if v else None)
            cell.number_format = '#,##0'; cell.font = F(bold=True, color=font_col)
            cell.fill = P(fill_col); cell.alignment = Alignment(horizontal="right", vertical="center")
        return r + 1

    def sum_row(r, label, values_dict, alt=False, bold=False, header_fill=None,
                label_indent="", font_color="000000", total_fill=None, hist_key=None):
        """Write one row: col1=Label, col2=blank, col3=blank, col4+=months.
        hist_key: if provided, uses history[hist_key] for historic months."""
        fill = header_fill or (P_SUM_A if alt else None)
        font = F(bold=bold, color=font_color)
        c = ws.cell(row=r, column=1, value=label_indent + label)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="D5D8DC"))
        if fill: c.fill = fill

        for col in (2, 3):
            bc = ws.cell(row=r, column=col)
            bc.border = Border(bottom=Side(style="thin", color="D5D8DC"))
            if fill: bc.fill = fill

        hist_row_name = HIST_MAP.get(hist_key, hist_key) if hist_key else None
        hist_series = history.get(hist_row_name, {}) if hist_row_name else {}

        for col_idx, m in enumerate(sum_months, 4):
            is_hist = is_history_month(m)
            if is_hist and hist_key and m in hist_series:
                # Previous Data owns this month and has a value — use it
                val = hist_series[m]
            elif is_hist and hist_key and m not in hist_series:
                # Previous Data owns this month but no entry — blank
                val = None
            else:
                # Either not a history month, or no hist_key — use values_dict
                val = values_dict.get(m, 0) if values_dict else 0
            is_est = (not is_hist) and (m not in actual_set)
            # Write value — allow zero (don't filter with truthiness)
            if val is not None:
                cell = ws.cell(row=r, column=col_idx, value=int(round(val)))
                cell.number_format = '#,##0'
            else:
                cell = ws.cell(row=r, column=col_idx, value=None)
                cell.number_format = '#,##0'
            if is_hist:
                cell.font = F(bold=bold, color="7F8C8D")
                cell.fill = P("E8EAEB") if alt else P("F2F3F4")
            elif is_est:
                cell.font = F(bold=bold, color="555577")
                if fill: cell.fill = fill
            else:
                cell.font = font
                if fill: cell.fill = fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="D5D8DC"))

        return r + 1

    def sum_total_row(r, label, monthly_dict, fill_col, font_col="042C53"):
        """Write a total/subtotal row across all sum_months."""
        font = F(bold=True, color=font_col)
        p_fill = P(fill_col)
        c = ws.cell(row=r, column=1, value=label)
        c.font = font; c.fill = p_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).fill = p_fill
        ws.cell(row=r, column=3).fill = p_fill
        for col_idx, m in enumerate(sum_months, 4):
            v = monthly_dict.get(m, 0)
            cell = ws.cell(row=r, column=col_idx, value=int(round(v)) if v else None)
            cell.number_format = '#,##0'; cell.font = font; cell.fill = p_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
        return r + 1

    # Section title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_sum_cols)
    t = ws.cell(row=1, column=1, value="Family Wealth Summary")
    t.font = F(bold=True, color="FFFFFF", size=12)
    t.fill = P_SUM_H
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row: Label | blank | blank | hist months (grey) | live months
    ws.cell(row=2, column=1, value="").fill = P_SUM_H
    ws.cell(row=2, column=2, value="").fill = P_SUM_H
    ws.cell(row=2, column=3, value="").fill = P_SUM_H
    for col_idx, m in enumerate(sum_months, 4):
        is_hist = m in hist_months
        is_est = (not is_hist) and (m not in actual_set)
        label = m.strftime("%b %Y")
        c = ws.cell(row=2, column=col_idx, value=label)
        if is_hist:
            c.font = F(bold=True, color="AAAAAA")
            c.fill = P("BFC9CA")
        elif is_est:
            c.font = F(bold=True, color="CCCCDD")
            c.fill = P("4A4A6A")
        else:
            c.font = F(bold=True, color="FFFFFF")
            c.fill = P_SUM_H
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    cur_row = 3

    # ── 1. Fidelity accounts ──────────────────────────────────────────────────
    # Sub-header
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Fidelity accounts")
    sh.font = F(bold=True, color="FFFFFF", size=9)
    sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    # Mapping: summary row → history row name
    def hist(key):
        """Return history series for this key."""
        return history.get(HIST_MAP.get(key, ""), {})

    def live_and_hist(live_dict, hist_key_str):
        """Alias for live_and_hist_safe — history always wins for pre-cutoff months."""
        return live_and_hist_safe(live_dict, hist_key_str)

    # Each family account — exclude SIPP accounts (shown in pensions instead)
    fid_total_val = summary_data["fidelity_total"]
    fid_by_acc = summary_data["fidelity_by_acc"]
    SIPP_ACCS = {"2000001606", "2000001604"}  # moved to pensions sections
    FIDELITY_ACC_LABELS = {
        "AW10032966": "Cash Account (Paul)",
        "SANX002282": "Investment ISA (Paul)",
        "SANQ000468": "Investment Account (Joint)",   # removed "(Paul)"
        "SANX002936": "Junior ISA (Jayne)",
        "AW10261123": "Cash Account (Susan)",
        "SANX002617": "Investment ISA (Susan)",
        "AW10580794": "Cash Account (Liam)",
        "AS10303823": "Investment ISA (Liam)",
        "AG10131710": "Investment Account (Liam)",
    }

    fid_non_sipp_accs = [(acc, val) for acc, val in sorted(fid_by_acc.items(), key=lambda x: x[1], reverse=True)
                         if acc not in SIPP_ACCS]

    # Compute monthly growth for SANQ000468 from accumulated fund values
    sanq_monthly = {}
    for fund_h, fd_h in acc_holdings.get("SANQ000468", {}).items():
        mv = fd_h.get("monthly_values", {})
        for m, v in mv.items():
            sanq_monthly[m] = sanq_monthly.get(m, 0) + v
    # Add LGEN and any non-Acc holdings (flat) to get full account value
    sanq_non_acc_val = fid_by_acc.get("SANQ000468", 0) - sum(
        fd_h["value"] for fd_h in acc_holdings.get("SANQ000468", {}).values()
    )
    for m in all_months:
        sanq_monthly[m] = sanq_monthly.get(m, 0) + sanq_non_acc_val

    # The AccountSummary total is the ACTUAL current value of the account. The
    # forward growth-projection past the build anchor (anchors.data_month in
    # build_acc_holdings)
    # is unreliable — it collapsed the current month (was hardcoded to override only
    # JUNE, so when "now" advanced to July the July value dropped ~£145k to a broken
    # projection: bug #20, 2026-07-18). Hold the current data month and every month
    # after it flat at the actual total — exactly as every OTHER Fidelity account row
    # already is (see the `else` branch below) — keeping only the reliable historical
    # backward ramp for earlier months. hold_from is the first month past the
    # build anchor, so it now tracks the snapshot automatically.
    sanq_actual = fid_by_acc.get("SANQ000468", 0)
    HOLD_FROM = anchors.hold_from
    if sanq_actual:
        for m in all_months:
            if m >= HOLD_FROM:
                sanq_monthly[m] = sanq_actual

    fid_rows_start = cur_row  # track start row for Fidelity total formula

    for i, (acc, val) in enumerate(fid_non_sipp_accs):
        # Show the account number alongside the friendly name (user request
        # 2026-07-18) — e.g. "Investment Account (Joint) (SANQ000468)".
        _base = FIDELITY_ACC_LABELS.get(acc)
        label = f"{_base} ({acc})" if _base else acc
        if acc == "SANQ000468" and sanq_monthly:
            monthly_vals = {m: sanq_monthly.get(m, val) for m in all_months}
        else:
            monthly_vals = {m: val for m in all_months}
        # Fidelity account rows: no per-account history, live values only
        cur_row = sum_row(cur_row, label, monthly_vals, alt=(i % 2 == 0),
                          label_indent="  ", font_color="042C53")
    fid_rows_end = cur_row

    # ── 2. Paul's pensions ────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Paul's pensions")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    paul_sipp_val = fid_by_acc.get("2000001606", 0)
    paul_sipp_vals_flat = {m: paul_sipp_val for m in all_months}
    cur_row = sum_row(cur_row, "SIPP Savings – Fidelity (2000001606)", paul_sipp_vals_flat,
                      alt=True, label_indent="  ", font_color="1A5276")

    arriva_merged = projection_with_hist_override(summary_data["arriva"], "arriva")
    cur_row = sum_row(cur_row, "PS Arriva (Defined Benefit)", arriva_merged,
                      alt=False, label_indent="  ", font_color="1A5276")  # pre-merged

    expleo_merged = projection_with_hist_override(summary_data["expleo_sw"], "expleo")
    cur_row = sum_row(cur_row, "Expleo Scottish Widows", expleo_merged,  # pre-merged
                      alt=True, label_indent="  ", font_color="1A5276")

    # Paul total placeholder — updated below once paul_sipp_growth_vals is computed
    paul_total_placeholder = {m: paul_sipp_val + summary_data["arriva"].get(m, 0) + summary_data["expleo_sw"].get(m, 0)
                              for m in all_months}
    cur_row = write_total_row(cur_row, "Total Paul's pensions", paul_total_placeholder, C_SUM_T)

    # ── 3. Susan's pensions ───────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Susan's pensions")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    SUSAN_HIST_KEYS = {
        "Capita (RMSPS) – to 2018 (65 yrs)": "capita",
        "RMPP 2012–2023 (65 years)":          "rmpp",
        "Collective Pension (2023+)":         "collective",
        "Cash Balance":                        "cash_balance",
        "AVC Bonus Plan (Scottish Widows)":    "avc",
    }
    susan_total_by_month = {m: 0 for m in all_months}
    for i, (pen_label, pen_vals) in enumerate(summary_data["susan_pensions"].items()):
        hk = SUSAN_HIST_KEYS.get(pen_label)
        merged = projection_with_hist_override(pen_vals, hk) if hk else pen_vals
        cur_row = sum_row(cur_row, pen_label, merged, alt=(i % 2 == 0),
                          label_indent="  ", font_color="1A5276")  # hist_key omitted — pre-merged
        # Accumulate from merged (history-corrected) not raw pen_vals
        for m in all_months:
            susan_total_by_month[m] += merged.get(m, 0)

    sipp_vals = summary_data["susan_fidelity_sipp"]
    for m in all_months:
        if m not in sipp_vals:
            sipp_vals[m] = list(sipp_vals.values())[0] if sipp_vals else 0
    susan_fid_merged = live_and_hist(sipp_vals, "susan_fidelity")
    cur_row = sum_row(cur_row, "SIPP Savings – Fidelity (2000001604)", susan_fid_merged,
                      alt=True, label_indent="  ", font_color="1A5276",
                      hist_key="susan_fidelity")
    for m in all_months:
        susan_total_by_month[m] = susan_total_by_month.get(m, 0) + sipp_vals.get(m, 0)

    cur_row = write_total_row(cur_row, "Total Susan's pensions", susan_total_by_month, C_SUM_T)
    susan_excl_sipp = {m: susan_total_by_month.get(m, 0) - sipp_vals.get(m, 0) for m in all_months}

    # ── 4. Assets section ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_sum_cols)
    sh = ws.cell(row=cur_row, column=1, value="Other Assets")
    sh.font = F(bold=True, color="FFFFFF", size=9); sh.fill = P_SUM_S
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cur_row += 1

    mf_proj = {m: 190000 for m in all_months}
    mf_merged = projection_with_hist_override(mf_proj, "mf")
    mf_vals = {m: mf_merged.get(m, 190000) for m in all_months}
    cur_row = sum_row(cur_row, "MF", mf_merged, alt=True, label_indent="  ",
                      font_color="1A5276")  # pre-merged

    bar_data = summary_data.get("barclays_by_month", {})
    cur_row = sum_row(cur_row, "Cash (Barclays)", bar_data, alt=False, label_indent="  ",
                      font_color="1A5276")

    house_at_est = 450450   # as at ESTIMATES_AS_OF
    house_vals = {}
    est_idx_h = next((i for i, m in enumerate(all_months) if m == ESTIMATES_AS_OF), 0)
    for i, m in enumerate(all_months):
        offset = i - est_idx_h
        house_vals[m] = round(house_at_est * ((1.05) ** (offset / 12)))
    house_merged = projection_with_hist_override(house_vals, "house")
    cur_row = sum_row(cur_row, "House", house_merged, alt=True, label_indent="  ",
                      font_color="1A5276")  # pre-merged

    # Liam ISA (AS10303823) and Investment Account (AG10131710) already in Fidelity accounts — not duplicated here

    car_at_est = 42074  # actual value from history at ESTIMATES_AS_OF (Cars Paul)
    car_vals = {}
    for i, m in enumerate(all_months):
        offset = i - est_idx_h
        car_vals[m] = round(car_at_est * ((1 - 0.05) ** (offset / 12)))
    car_merged = projection_with_hist_override(car_vals, "cars")
    cur_row = sum_row(cur_row, "Cars Paul (Mercedes AMG GTS 2016)", car_merged, alt=False,
                      label_indent="  ", font_color="1A5276")  # pre-merged

    assets_dicts = [mf_vals, bar_data, house_vals, car_vals]
    assets_by_month = {m: sum(d.get(m, 0) for d in assets_dicts) for m in all_months}
    cur_row = write_total_row(cur_row, "Total Other Assets", assets_by_month, C_SUM_T)

    # ── Compute Fidelity growth ───────────────────────────────────────────────
    fid_growth = {all_months[0]: round(fid_total_val)}
    for i in range(1, len(all_months)):
        m = all_months[i]
        monthly_income = fid_pivot[m].sum() if m in fid_pivot.columns else 0
        fid_growth[m] = round(fid_growth[all_months[i-1]] + monthly_income)

    # ── SIPP growth: pinned history early, AccountSummary anchors the snapshot
    # month, income grows the months after it. The snapshot is a statement of
    # what the accounts are worth ON ITS OWN DATE — anchoring it to a fixed month
    # made every later export overwrite the wrong column.
    paul_sipp_val  = fid_by_acc.get("2000001606", 0)   # value at anchors.data_month
    susan_sipp_val = fid_by_acc.get("2000001604", 0)
    paul_hist_series  = hist("paul_pension")
    paul_sipp_growth_vals = {}
    susan_sipp_growth_vals = {}
    sipp_vals_from_summary = summary_data.get("susan_fidelity_sipp", {})

    ps_running = paul_hist_series.get(all_months[0], paul_sipp_val)
    ss_running = sipp_vals_from_summary.get(all_months[0], susan_sipp_val)

    for i, m in enumerate(all_months):
        if m == anchors.data_month:
            # AccountSummary is ground truth for this month — override any projection
            ps_running = paul_sipp_val
            ss_running = susan_sipp_val
            paul_sipp_growth_vals[m] = paul_sipp_val
            susan_sipp_growth_vals[m] = susan_sipp_val
        elif is_history_month(m):
            # Pre-cutoff: use history if available
            if m in paul_hist_series:
                paul_sipp_growth_vals[m] = paul_hist_series[m]
                ps_running = paul_hist_series[m]
            else:
                paul_sipp_growth_vals[m] = ps_running
            ss_val = sipp_vals_from_summary.get(m, ss_running)
            susan_sipp_growth_vals[m] = ss_val
            ss_running = ss_val
        else:
            # Post-cutoff (Jun+): grow from previous month using income
            ps_inc = fid_pivot.loc["2000001606", m] if "2000001606" in fid_pivot.index and m in fid_pivot.columns else 0
            ss_inc = fid_pivot.loc["2000001604", m] if "2000001604" in fid_pivot.index and m in fid_pivot.columns else 0
            ps_running = round(ps_running + ps_inc)
            ss_running = round(ss_running + ss_inc)
            paul_sipp_growth_vals[m] = ps_running
            susan_sipp_growth_vals[m] = ss_running

    fid_non_sipp_total = sum(val for acc, val in fid_by_acc.items() if acc not in SIPP_ACCS)
    sanq_flat = fid_by_acc.get("SANQ000468", 0)
    sanq_growth = {m: sanq_monthly.get(m, sanq_flat) for m in all_months}
    fid_non_sipp_with_growth = {m: fid_non_sipp_total - sanq_flat + sanq_growth[m] for m in all_months}
    total_fidelity_growth = {m: fid_non_sipp_with_growth[m] + paul_sipp_growth_vals[m] + susan_sipp_growth_vals[m]
                             for m in all_months}

    # ── 5. Fidelity accounts total (all including SIPPs) ─────────────────────
    fid_total_row_num = cur_row
    fid_all_merged = {**hist("fidelity_all"), **total_fidelity_growth}
    cur_row = write_total_row(cur_row, "Fidelity accounts", fid_all_merged, C_SUM_T)

    # Update Paul SIPP row values with growth
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "  SIPP Savings – Fidelity (2000001606)":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(paul_sipp_growth_vals[m]))
            break
    # Update Susan SIPP row values with growth
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "  SIPP Savings – Fidelity (2000001604)":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(susan_sipp_growth_vals[m]))
            break
    # Update paul_total using history-corrected arriva and expleo values
    arriva_merged = projection_with_hist_override(summary_data["arriva"], "arriva")
    expleo_merged_upd = projection_with_hist_override(summary_data["expleo_sw"], "expleo")
    paul_total = {m: paul_sipp_growth_vals[m] + arriva_merged.get(m, 0) + expleo_merged_upd.get(m, 0)
                  for m in all_months}
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "Total Paul's pensions":
            for col_idx, m in enumerate(all_months, 4):
                ws.cell(row=_r, column=col_idx).value = int(round(paul_total[m]))
            break

    # Update Susan total using history-corrected pension + SIPP growth values
    for _r in range(3, fid_total_row_num):
        if ws.cell(row=_r, column=1).value == "Total Susan's pensions":
            for col_idx, m in enumerate(all_months, 4):
                sp_merged = sum(
                    projection_with_hist_override(summary_data["susan_pensions"].get(k, {}),
                                  SUSAN_HIST_KEYS.get(k)).get(m, 0)
                    for k in summary_data["susan_pensions"]
                )
                ws.cell(row=_r, column=col_idx).value = int(round(sp_merged + susan_sipp_growth_vals[m]))
            break

    # ── 8. TOTAL FAMILY WEALTH — at the very bottom ───────────────────────────
    grand_row = cur_row
    cur_row += 1
    paul_non_sipp = {m: summary_data["arriva"][m] + summary_data["expleo_sw"][m] for m in all_months}
    grand_by_month = {m: total_fidelity_growth[m] + paul_non_sipp.get(m, 0)
                      + susan_excl_sipp.get(m, 0) + assets_by_month.get(m, 0)
                      for m in all_months}
    # Build historic grand total from history rows — covers ALL historic periods
    # (not just display months) so Yearly Increase can reference prior year values
    hist_grand = {}
    all_hist_periods = sorted(set().union(*[set(s.keys()) for s in history.values()]))
    for m in all_hist_periods:
        fid_h    = hist("fidelity_all").get(m, 0)
        arriva_h = hist("arriva").get(m, 0)
        expleo_h = hist("expleo").get(m, 0)
        capita_h = hist("capita").get(m, 0)
        rmpp_h   = hist("rmpp").get(m, 0)
        coll_h   = hist("collective").get(m, 0)
        cb_h     = hist("cash_balance").get(m, 0)
        avc_h    = hist("avc").get(m, 0)
        mf_h     = hist("mf").get(m, 0)
        house_h  = hist("house").get(m, 0)
        # Liam ISA & Investment Account is inside fid_h (All Fidelity accounts) — not added separately
        # Jayne ISA is also inside fid_h — not added separately
        susan_non_fid_h = capita_h + rmpp_h + coll_h + cb_h + avc_h
        hist_grand[m] = (fid_h + arriva_h + expleo_h + susan_non_fid_h +
                         mf_h + house_h)

    all_grand = {**hist_grand, **grand_by_month}

    c = ws.cell(row=grand_row, column=1, value="TOTAL FAMILY WEALTH")
    c.font = F(bold=True, color="FFFFFF", size=11); c.fill = P_SUM_H
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[grand_row].height = 22
    for col in range(1, 4):
        ws.cell(row=grand_row, column=col).fill = P_SUM_H
    for col_idx, m in enumerate(sum_months, 4):
        v = all_grand.get(m, 0)
        cell = ws.cell(row=grand_row, column=col_idx, value=int(round(v)) if v else None)
        cell.number_format = '#,##0'; cell.font = F(bold=True, color="FFFFFF")
        cell.fill = P_SUM_H; cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── 9. Total Income and Accumulations — below TOTAL FAMILY WEALTH ───────────────────────
    # Values written now as placeholders; updated with Excel formulas once
    # Income (tot_r) and Accumulative (tot_acc_r) row numbers are known.
    fid_income_monthly = {m: int(round(fid_pivot[m].sum())) if (not fid_pivot.empty and m in fid_pivot.columns) else 0
                          for m in all_months}
    acc_monthly_increase = {m: 0 for m in all_months}
    for acc_hh, funds_hh in acc_holdings.items():
        for fund_hh, fd_hh in funds_hh.items():
            pa = fd_hh.get("price_appreciation", {})
            for mk, inc in pa.items():
                if mk in all_months:
                    acc_monthly_increase[mk] = acc_monthly_increase.get(mk, 0) + inc
    total_inc_acc = {m: fid_income_monthly.get(m, 0) + acc_monthly_increase.get(m, 0)
                     for m in all_months}
    total_inc_acc_row = cur_row  # remember for later formula update
    cur_row = write_total_row(cur_row, "Total Income and Accumulations", total_inc_acc, C_SUM_T)

    # ── Calculations section (placed AFTER all assets rows) ───────────────────
    calc_start = cur_row + 1  # one blank row after Total Income and Accumulations

    C_CALC_H = "17202A"
    P_CALC_H = P(C_CALC_H)
    P_CALC_A = P("EAECF0")

    ws.merge_cells(start_row=calc_start, start_column=1, end_row=calc_start, end_column=n_sum_cols)
    ch = ws.cell(row=calc_start, column=1, value="Summary Calculations")
    ch.font = F(bold=True, color="FFFFFF", size=12); ch.fill = P_CALC_H
    ch.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[calc_start].height = 24

    calc_hdr = calc_start + 1
    for col_idx, m in enumerate(sum_months, 4):
        is_hist = m in hist_months
        c = ws.cell(row=calc_hdr, column=col_idx,
                    value=m.strftime("%b %Y") if not (col_idx <= 3) else None)
        c.font = F(bold=True, color="AAAAAA" if is_hist else "FFFFFF")
        c.fill = P("BFC9CA") if is_hist else P_CALC_H
        c.alignment = Alignment(horizontal="right", vertical="center")
    for col in (1, 2, 3):
        ws.cell(row=calc_hdr, column=col).fill = P_CALC_H
    ws.row_dimensions[calc_hdr].height = 16

    def mcol(m): return get_column_letter(4 + sum_months.index(m)) if m in sum_months else None

    def calc_row_fn(r, label, formulas_by_month, alt=False, bold=False, font_color="000000"):
        fill = P_CALC_A if alt else None
        font = F(bold=bold, color=font_color)
        c = ws.cell(row=r, column=1, value=label)
        c.font = font; c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="CCCCCC"))
        if fill: c.fill = fill
        for col in (2, 3):
            bc = ws.cell(row=r, column=col)
            bc.border = Border(bottom=Side(style="thin", color="CCCCCC"))
            if fill: bc.fill = fill
        for col_idx, m in enumerate(sum_months, 4):
            formula = formulas_by_month.get(m)
            cell = ws.cell(row=r, column=col_idx)
            if formula is None:
                cell.value = None
            elif isinstance(formula, str) and formula.startswith("="):
                # Write as proper Excel formula — never as text
                cell.data_type = "f"
                cell.value = formula
                cell.number_format = '#,##0;(#,##0);"-"'
            elif isinstance(formula, (int, float)):
                cell.value = int(round(formula))
                cell.number_format = '#,##0;(#,##0);"-"'
            else:
                cell.value = formula
            cell.font = font
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="CCCCCC"))
            if fill: cell.fill = fill
        return r + 1

    # Scan summary rows to find key row numbers
    row_refs = {}
    for _r in range(3, calc_start):
        v = ws.cell(row=_r, column=1).value
        if v: row_refs[str(v).strip()] = _r

    gw_row      = row_refs.get("TOTAL FAMILY WEALTH", grand_row)
    fid_row_sum = row_refs.get("Total Fidelity")
    paul_row_s  = row_refs.get("Total Paul's pensions")
    susan_row_s = row_refs.get("Total Susan's pensions")
    assets_row  = row_refs.get("Total Other Assets")
    house_row   = next((r for k, r in row_refs.items() if k.startswith("House")), None)
    car_row     = next((r for k, r in row_refs.items() if "Cars" in k), None)
    arriva_row  = next((r for k, r in row_refs.items() if "Arriva" in k), None)
    expleo_row  = next((r for k, r in row_refs.items() if "Expleo" in k), None)

    calc_cur = calc_hdr + 1

    # 1. Total — all assets incl cash and property (= TOTAL FAMILY WEALTH)
    calc_cur = calc_row_fn(calc_cur, "Total",
        {m: f"={mcol(m)}{gw_row}" for m in all_months},
        alt=True, bold=True, font_color="17202A")

    # 2. Investments & Cash — Total minus House and Car
    def inv_formula(m):
        parts = [f"{mcol(m)}{gw_row}"]
        if house_row: parts.append(f"-{mcol(m)}{house_row}")
        if car_row:   parts.append(f"-{mcol(m)}{car_row}")
        return "=" + "".join(parts)
    calc_cur = calc_row_fn(calc_cur, "Investments & Cash",
        {m: inv_formula(m) for m in sum_months}, alt=False)

    # 3. Yearly Increase — this month's TOTAL FAMILY WEALTH minus same month last year
    # Prior year values come from all_grand dict (which includes historic data)
    yearly = {}
    for i, m in enumerate(sum_months):
        m_minus_12 = m - 12
        prior_val = all_grand.get(m_minus_12, 0)
        if prior_val:
            # Use direct cell reference for current month, subtract hardcoded prior value
            yearly[m] = f"={mcol(m)}{gw_row}-{int(prior_val)}"
    calc_cur = calc_row_fn(calc_cur, "Yearly Increase", yearly, alt=True)

    # 4. PS Pension — Paul: Fidelity SIPP + Arriva + Expleo SW
    paul_fid_r = next((r for k, r in row_refs.items() if "SIPP Savings (Paul)" in k or "2000001606" in k), None)
    def ps_formula(m):
        parts = []
        if paul_fid_r:  parts.append(f"{mcol(m)}{paul_fid_r}")
        if arriva_row:  parts.append(f"{mcol(m)}{arriva_row}")
        if expleo_row:  parts.append(f"{mcol(m)}{expleo_row}")
        return ("=" + "+".join(parts)) if parts else None
    calc_cur = calc_row_fn(calc_cur, "PS Pension (Paul)", {m: ps_formula(m) for m in sum_months}, alt=False)

    # 5. SS Pension — Susan: total pensions row (incl SIPP)
    calc_cur = calc_row_fn(calc_cur, "SS Pension (Susan)",
        {m: f"={mcol(m)}{susan_row_s}" for m in sum_months} if susan_row_s else {},
        alt=True)

    # 6. Monthly Change — vs previous month
    mc_formulas = {}
    for i, m in enumerate(sum_months):
        if i > 0:
            mc_formulas[m] = f"={mcol(m)}{gw_row}-{mcol(sum_months[i-1])}{gw_row}"
    calc_cur = calc_row_fn(calc_cur, "Monthly Change", mc_formulas, alt=False)

    # 7. Av Monthly Change — average monthly change since start of all data
    av_formulas = {}
    for i, m in enumerate(sum_months):
        if i >= 2:
            av_formulas[m] = f"=ROUND(({mcol(m)}{gw_row}-{mcol(sum_months[0])}{gw_row})/{i},0)"
    calc_cur = calc_row_fn(calc_cur, "Av Monthly Change", av_formulas, alt=True)

    # 8. Monthly Investment Increase — income generated from investments (fid_pivot)
    monthly_inv = {}
    for m in sum_months:
        inc = round(fid_pivot[m].sum()) if m in fid_pivot.columns else 0
        if inc: monthly_inv[m] = inc
    calc_cur = calc_row_fn(calc_cur, "Monthly Investment Increase", monthly_inv,
                           alt=False, font_color="145A32")

    # ── Section 1: Spending Summary ────────────────────────────────────────────
    n_spend_cols = 3 + len(spend_months)
    R = calc_cur + 0  # no blank spacer

    # Section title
    ws.merge_cells(start_row=R+1, start_column=1, end_row=R+1, end_column=n_spend_cols)
    title = ws.cell(row=R+1, column=1, value="Amex & Barclays Spend")
    title.font = F(bold=True, color="FFFFFF", size=12)
    title.fill = P(C_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[R+1].height = 24

    # Column headers — col1=Category, col2=blank, col3=Total, col4+=months (aligned with income table)
    headers = ["Category", "", "Total"] + spend_month_labels
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=R+2, column=col, value=h)
        if col <= 3:
            c.font = HDR_FONT
            c.fill = HDR_FILL
        else:
            m = spend_months[col - 4]
            c.font = HDR_FONT if m in actual_set else EST_FONT
            c.fill = HDR_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[R+2].height = 18

    # Data rows
    for row_idx, cat in enumerate(CATEGORIES, R+3):
        is_alt = (row_idx % 2 == 0)
        font  = BOLD_FONT

        c = ws.cell(row=row_idx, column=1, value=cat)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="D9E1F2"))

        # col 2 blank
        ws.cell(row=row_idx, column=2).border = Border(bottom=Side(style="thin", color="D9E1F2"))

        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(spend_months))
        tc = ws.cell(row=row_idx, column=3,
                     value=f"=SUM({fc}{row_idx}:{lc}{row_idx})")
        tc.number_format = NUM_FMT
        tc.font = font
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.fill = TOT_FILL
        tc.border = Border(bottom=Side(style="thin", color="D9E1F2"))

        for col_idx, m in enumerate(spend_months, 4):
            val = spend_pivot.loc[cat, m] if cat in spend_pivot.index else 0
            fill = col_fill(m, is_alt, ALT_FILL, ALT_FILL)
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=int(round(val)) or None)
            cell.number_format = NUM_FMT
            cell.font = val_font(m)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="D9E1F2"))
            if fill: cell.fill = fill

    # Footer: total spend
    foot_row = R + 3 + len(CATEGORIES)
    exp_rows = [R + 3 + i for i in range(len(CATEGORIES))]
    ws.cell(row=foot_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=foot_row, column=1).fill = FOOT_FILL
    ws.cell(row=foot_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=foot_row, column=2).fill = FOOT_FILL  # blank col
    for col_idx in range(3, 4 + len(spend_months)):
        cl = get_column_letter(col_idx)
        refs = "+".join(f"{cl}{r}" for r in exp_rows)
        cell = ws.cell(row=foot_row, column=col_idx, value=f"={refs}")
        cell.number_format = NUM_FMT
        cell.font = BOLD_FONT
        cell.fill = FOOT_FILL
        cell.alignment = Alignment(horizontal="right", vertical="center")


    # ── Note row below TOTAL ────────────────────────────────────────────────────
    note_row = foot_row + 1
    note_cell = ws.cell(row=note_row, column=1,
                        value="Susan's lease car additional miles at 30.3p.  Contract to End July 2027")
    note_cell.font = Font(italic=True, color="808080", size=9)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 14
    # Merge across all spend columns so it reads as one line
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=3 + len(all_months))

    # ── Expense Reimbursements table (underneath Amex & Barclays Spend) ────────
    REIMB_START_ROW = note_row + 2  # 1 blank row gap below the lease-car note
    fc_r = get_column_letter(4)
    lc_r = get_column_letter(3 + len(all_months))

    # Title row — spans full width like the spend table
    ws.merge_cells(start_row=REIMB_START_ROW, start_column=1,
                   end_row=REIMB_START_ROW, end_column=3 + len(all_months))
    rt = ws.cell(row=REIMB_START_ROW, column=1, value="Expense Reimbursements")
    rt.font = F(bold=True, color="FFFFFF", size=12)
    rt.fill = P("8E44AD")  # purple — distinct from navy spend table
    rt.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[REIMB_START_ROW].height = 22

    # Header row — Category | (blank) | Total | Jan..Dec, matching spend table
    header_row = REIMB_START_ROW + 1
    hc = ws.cell(row=header_row, column=1, value="Source")
    hc.font = F(bold=True, color="FFFFFF", size=10)
    hc.fill = P("8E44AD")
    hc.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=header_row, column=2).fill = P("8E44AD")
    htot = ws.cell(row=header_row, column=3, value="Total")
    htot.font = F(bold=True, color="FFFFFF", size=10)
    htot.fill = P("8E44AD")
    htot.alignment = Alignment(horizontal="right", vertical="center")
    for i, m in enumerate(all_months):
        col = 4 + i
        c = ws.cell(row=header_row, column=col, value=m.strftime("%b %Y"))
        c.font = F(bold=True, color="FFFFFF", size=10)
        c.fill = P("8E44AD")
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[header_row].height = 20

    # Build month -> amount map per source
    reimb_list = reimbursements or []
    by_source = {"Royal Mail": {}, "Expleo": {}}
    for r in reimb_list:
        memo_upper = r["memo"].upper()
        source = "Royal Mail" if "ROYAL MAIL" in memo_upper else "Expleo"
        m = r["month"]
        by_source[source][m] = by_source[source].get(m, 0) + r["amount"]

    # Two data rows: Royal Mail, Expleo
    for ri, source in enumerate(["Royal Mail", "Expleo"]):
        row_n = header_row + 1 + ri
        bg = "F4ECF7" if ri % 2 == 1 else "FFFFFF"
        c = ws.cell(row=row_n, column=1, value=source)
        c.font = F(size=10)
        c.fill = P(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="E8DAEF"))
        ws.cell(row=row_n, column=2).fill = P(bg)
        ws.cell(row=row_n, column=2).border = Border(bottom=Side(style="thin", color="E8DAEF"))

        # Month columns
        for i, m in enumerate(all_months):
            col = 4 + i
            val = by_source[source].get(m, 0)
            cell = ws.cell(row=row_n, column=col)
            if val:
                cell.value = val
                cell.number_format = NUM_FMT
            cell.font = F(size=10)
            cell.fill = P(bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="E8DAEF"))

        # Total column
        tc = ws.cell(row=row_n, column=3, value=f"=SUM({fc_r}{row_n}:{lc_r}{row_n})")
        tc.number_format = NUM_FMT
        tc.font = F(bold=True, size=10)
        tc.fill = P(bg)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border = Border(bottom=Side(style="thin", color="E8DAEF"))

        ws.row_dimensions[row_n].height = 16

    # TOTAL row
    reimb_row = header_row + 3
    ws.cell(row=reimb_row, column=1, value="TOTAL").font = F(bold=True, size=10)
    ws.cell(row=reimb_row, column=1).fill = P("D2B4DE")
    ws.cell(row=reimb_row, column=2).fill = P("D2B4DE")
    tot_total = ws.cell(row=reimb_row, column=3,
                         value=f"=SUM(C{header_row+1}:C{header_row+2})")
    tot_total.number_format = NUM_FMT
    tot_total.font = F(bold=True, size=10)
    tot_total.fill = P("D2B4DE")
    tot_total.alignment = Alignment(horizontal="right", vertical="center")
    for i in range(len(all_months)):
        col = 4 + i
        col_l = get_column_letter(col)
        cm = ws.cell(row=reimb_row, column=col,
                      value=f"=SUM({col_l}{header_row+1}:{col_l}{header_row+2})")
        cm.number_format = NUM_FMT
        cm.font = F(bold=True, size=10)
        cm.fill = P("D2B4DE")
        cm.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[reimb_row].height = 18

    # Note
    reimb_note_row = reimb_row + 1
    ws.merge_cells(start_row=reimb_note_row, start_column=1,
                   end_row=reimb_note_row, end_column=3 + len(all_months))
    rn = ws.cell(row=reimb_note_row, column=1,
                 value="Reimbursements from Royal Mail or Expleo for expenses incurred on the family's "
                       "behalf — recovered from employers, excluded from family spending totals.")
    rn.font = Font(italic=True, color="808080", size=8)
    rn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[reimb_note_row].height = 28

    # ── Spacer rows ────────────────────────────────────────────────────────────
    fid_start_row = reimb_note_row + 2
    tot_r = fid_start_row  # fallback; overwritten when income section is written



    # ── Section 2: Income ─────────────────────────────────────────────────────
    if not fid_pivot.empty:
        # Build combined income rows: Salary first, then Fidelity accounts
        fid_accounts = list(fid_pivot.index)
        # Account + Units + Total + months
        n_fid_cols = 3 + len(fid_months)

        # Section title
        ws.merge_cells(start_row=fid_start_row, start_column=1,
                       end_row=fid_start_row, end_column=n_fid_cols)
        ft = ws.cell(row=fid_start_row, column=1, value="Income")
        ft.font = F(bold=True, color="FFFFFF", size=12)
        ft.fill = FID_FILL
        ft.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[fid_start_row].height = 24

        # Column headers
        fid_hdr_row = fid_start_row + 1
        # Income table: col1=Account, col2=Units, col3=Total, col4+=months
        fid_headers = ["Account", "Units", "Total"] + fid_month_labels
        for col, h in enumerate(fid_headers, 1):
            c = ws.cell(row=fid_hdr_row, column=col, value=h)
            if col <= 3:
                c.font = FID_HFONT
                c.fill = FID_FILL
            else:
                m = fid_months[col - 4]
                c.font = FID_HFONT if m in actual_set else EST_FONT
                c.fill = FID_FILL if m in actual_set else EST_HDR_FILL
            c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
        ws.row_dimensions[fid_hdr_row].height = 18

        # ── Salary row ────────────────────────────────────────────────────────
        sal_row = fid_start_row + 2
        sal_label = ws.cell(row=sal_row, column=1, value="Salary")
        sal_label.font = SAL_FONT
        sal_label.fill = SAL_FILL
        sal_label.alignment = Alignment(horizontal="left", vertical="center")
        sal_label.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        # Units col (col 2) — blank for salary
        su = ws.cell(row=sal_row, column=2, value=None)
        su.fill = SAL_FILL
        su.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(fid_months))
        sal_tot = ws.cell(row=sal_row, column=3,
                          value=f"=SUM({fc}{sal_row}:{lc}{sal_row})")
        sal_tot.number_format = NUM_FMT
        sal_tot.font = SAL_FONT
        sal_tot.fill = SAL_FILL
        sal_tot.alignment = Alignment(horizontal="right", vertical="center")
        sal_tot.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        for col_idx, m in enumerate(fid_months, 4):
            val = spend_pivot.loc["Salary", m] if "Salary" in spend_pivot.index else 0
            is_est = m not in actual_set
            cell = ws.cell(row=sal_row, column=col_idx,
                           value=int(round(val)) or None)
            cell.number_format = NUM_FMT
            cell.font = SAL_FONT if not is_est else F(bold=True, color="778866")
            cell.fill = SAL_FILL if not is_est else P("EEF5E8")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="A9DFBF"))

        # ── Per-person sections ───────────────────────────────────────────────
        fc = get_column_letter(4)
        lc = get_column_letter(3 + len(fid_months))
        current_row = sal_row + 1
        account_total_rows = []
        equity_data_rows = []

        for person, person_accs in acc_fund_map.items():
            # Person sub-header — spans all cols
            ph = ws.cell(row=current_row, column=1, value=person)
            ph.font = F(bold=True, color="FFFFFF", size=9)
            ph.fill = P("1E8449")
            ph.alignment = Alignment(horizontal="left", vertical="center")
            for col in range(2, 4 + len(fid_months)):
                ws.cell(row=current_row, column=col).fill = P("1E8449")
            current_row += 1

            for acc, fund_df in person_accs.items():
                label = ACCOUNT_LABELS.get(acc, acc)
                funds_in_acc = list(fund_df.index)
                n_fund_rows = len(funds_in_acc)

                acc_r = current_row
                account_total_rows.append(acc_r)

                # Col 1: account label
                c = ws.cell(row=acc_r, column=1, value=label)
                c.font = F(bold=True, color="145A32")
                c.fill = FIDT_FILL
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                # Col 2: units — blank for account header row
                uc = ws.cell(row=acc_r, column=2, value=None)
                uc.fill = FIDT_FILL
                uc.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                # Col 3: total
                tc = ws.cell(row=acc_r, column=3,
                             value=f"=SUM({fc}{acc_r+1}:{lc}{acc_r+n_fund_rows})")
                tc.number_format = NUM_FMT
                tc.font = F(bold=True, color="145A32")
                tc.fill = FIDT_FILL
                tc.alignment = Alignment(horizontal="right", vertical="center")
                tc.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                for col_idx, m in enumerate(fid_months, 4):
                    val = fid_pivot.loc[acc, m] if acc in fid_pivot.index else 0
                    is_est = m not in actual_set
                    cell = ws.cell(row=acc_r, column=col_idx,
                                   value=int(round(val)) or None)
                    cell.number_format = NUM_FMT
                    cell.font = F(bold=True, color="145A32") if not is_est else F(bold=True, color="558855")
                    cell.fill = FIDT_FILL if not is_est else P("C8E6C9")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=Side(style="thin", color="A9DFBF"))

                current_row += 1

                # Fund rows — col1=indented name, col2=units, col3=total, col4+=months
                for fund_offset, fund in enumerate(funds_in_acc):
                    fr = current_row
                    fill = FIDA_FILL if (fund_offset % 2 == 0) else None

                    # Col 1: fund name indented
                    c = ws.cell(row=fr, column=1, value=f"  {fund}")
                    c.font = BODY_FONT
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: c.fill = fill

                    # Col 2: units held — try exact match then normalised match
                    units = holdings.get((acc, fund), 0)
                    if not units:
                        # Strip all spaces for comparison
                        def norm(s): return s.replace(" ", "").upper()
                        units = next(
                            (v for (a, f), v in holdings.items()
                             if a == acc and norm(f) == norm(fund)),
                            0
                        )
                    uc = ws.cell(row=fr, column=2,
                                 value=round(units, 2) if units > 0 else None)
                    uc.font = BODY_FONT
                    uc.number_format = '#,##0.##'
                    uc.alignment = Alignment(horizontal="right", vertical="center")
                    uc.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: uc.fill = fill

                    # Col 3: total
                    tc = ws.cell(row=fr, column=3,
                                 value=f"=SUM({fc}{fr}:{lc}{fr})")
                    tc.number_format = NUM_FMT
                    tc.font = BODY_FONT
                    tc.alignment = Alignment(horizontal="right", vertical="center")
                    tc.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                    if fill: tc.fill = fill

                    # Col 4+: month values
                    for col_idx, m in enumerate(fid_months, 4):
                        val = fund_df.loc[fund, m] if fund in fund_df.index else 0
                        is_est = m not in actual_set
                        fill2 = (FIDA_FILL if (fund_offset % 2 == 0) else None) if not is_est else (P("DCF0DC") if (fund_offset % 2 == 0) else P("EBF5EB"))
                        cell = ws.cell(row=fr, column=col_idx,
                                       value=int(round(val)) or None)
                        cell.number_format = NUM_FMT
                        cell.font = BODY_FONT if not is_est else EST_BODY
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill2: cell.fill = fill2

                    current_row += 1

                # ── Equity dividends — write inline under Paul SIPP (2000001606) ──
                if acc == "2000001606":
                    for eq_offset, (eq_label, eq_divs) in enumerate(EQUITY_DIVIDENDS_INLINE):
                        eq_row = current_row
                        equity_data_rows.append(eq_row)
                        fill_eq = FIDA_FILL if (eq_offset % 2 == 0) else None
                        c = ws.cell(row=eq_row, column=1, value=eq_label)
                        c.font = F(color="145A32", italic=True, size=10)
                        c.alignment = Alignment(horizontal="left", vertical="center")
                        c.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: c.fill = fill_eq
                        ws.cell(row=eq_row, column=2).border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: ws.cell(row=eq_row, column=2).fill = fill_eq

                        # Total column (col 3)
                        tc_eq = ws.cell(row=eq_row, column=3, value=f"=SUM({fc}{eq_row}:{lc}{eq_row})")
                        tc_eq.number_format = NUM_FMT
                        tc_eq.font = F(color="145A32", italic=True, size=10)
                        tc_eq.alignment = Alignment(horizontal="right", vertical="center")
                        tc_eq.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                        if fill_eq: tc_eq.fill = fill_eq
                        for col_idx, m in enumerate(fid_months, 4):
                            val = eq_divs.get(str(m), 0)
                            cell = ws.cell(row=eq_row, column=col_idx)
                            if val:
                                cell.value = val
                                cell.number_format = NUM_FMT
                            cell.font = F(color="145A32", italic=True, size=10)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.border = Border(bottom=Side(style="thin", color="D5F5E3"))
                            if fill_eq: cell.fill = fill_eq
                        ws.row_dimensions[eq_row].height = 15
                        current_row += 1

        # Note row
        note_eq = current_row
        nc = ws.cell(row=note_eq, column=1,
                     value="  * Sep/Nov equity dividend amounts are estimates based on prior year interim payments")
        nc.font = Font(italic=True, color="999999", size=8)
        nc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=note_eq, start_column=1,
                       end_row=note_eq, end_column=3 + len(fid_months))
        ws.row_dimensions[note_eq].height = 12
        current_row += 1

        # June provisional note
        note_jun = current_row
        nj = ws.cell(row=note_jun, column=1,
                     value="  * June 2026 figures are PROVISIONAL — based on partial transaction data (60-day export, 10 Jun) "
                           "and May-value estimates for fund income/salary. Will be revised once full June data is provided.")
        nj.font = Font(italic=True, bold=True, color="B7791F", size=8)
        nj.fill = P("FFF8E1")
        nj.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=note_jun, start_column=1,
                       end_row=note_jun, end_column=3 + len(fid_months))
        ws.row_dimensions[note_jun].height = 14
        current_row += 1

        # Grand total — Salary + all fund accounts + equity dividends
        tot_r = current_row
        c = ws.cell(row=tot_r, column=1, value="Total Income")
        c.font = F(bold=True, color="145A32")
        c.fill = FIDT_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=tot_r, column=2).fill = FIDT_FILL
        all_income_rows = [sal_row] + account_total_rows + equity_data_rows
        for col_idx in range(3, 4 + len(fid_months)):
            cl = get_column_letter(col_idx)
            refs = "+".join(f"{cl}{rr}" for rr in all_income_rows)
            cell = ws.cell(row=tot_r, column=col_idx, value=f"={refs}")
            cell.number_format = NUM_FMT
            cell.font = F(bold=True, color="145A32")
            cell.fill = FIDT_FILL
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Section 3: Accumulative Holdings ─────────────────────────────────────
    # Estimate monthly growth for Acc funds using the income yield from Inc equivalents
    # Yield proxy: monthly income per account / Inc fund value → apply to Acc fund value

    # Build yield map from Inc fund income data
    # Key: normalised fund base name → monthly income rate
    ACC_INC_EQUIV = {
        "Aegon High Yield Bond B Acc":               "Aegon High Yield Bond B Inc",
        "Schroder High Yield Opportunities Fund Z Acc": "Schroder High Yield Opportunities Fund Z Inc",
        "WS Guinness Global Energy Fund I Acc":       None,  # no Inc equivalent — use 6% annual
        "Man High Yield Opportunities Fund Prof D Acc": "Man High Yield Opportunities Fund Prof D Inc",
    }

    # Estimate annual yield rates from actual income / May 26 value
    def est_annual_yield(inc_fund_name, acc_value):
        """Estimate annual yield for an Acc fund from its Inc equivalent's income rate."""
        if inc_fund_name is None:
            return 0.06  # 6% default for unknown
        # Sum total 2026 income across all accounts for this fund from fid_pivot
        total_inc = 0
        if inc_fund_name in fid_pivot.index:
            total_inc = fid_pivot.loc[inc_fund_name, actual_months].sum()
        # Annualise: divide by months of data, multiply by 12
        months_of_data = len(actual_months)
        if months_of_data > 0 and total_inc > 0:
            # Find total Inc fund value from AccountSummary
            inc_total_val = sum(
                v for (a, f), v_dict in {}.items()
            )
            # Fall back: use the Acc fund value as proxy
            monthly_rate = total_inc / acc_value if acc_value > 0 else 0
            return monthly_rate * 12
        return 0.07  # 7% default

    C_ACC_H = "0E4D6B"   # teal-navy header
    C_ACC_A = "E8F4F8"   # light teal alt
    C_ACC_T = "A8D5E2"   # teal total

    ACC_FILL  = P(C_ACC_H)
    ACCA_FILL = P(C_ACC_A)
    ACCT_FILL = P(C_ACC_T)
    ACC_HFONT = F(bold=True, color="FFFFFF")
    ACC_TFONT = F(bold=True, color="0E4D6B")

    acc3_start = tot_r + 2

    n_acc_cols = 3 + len(fid_months)   # Label + Units + Total + months (same as income table)
    ws.merge_cells(start_row=acc3_start, start_column=1,
                   end_row=acc3_start, end_column=n_acc_cols)
    t_acc = ws.cell(row=acc3_start, column=1, value="Accumulative Holdings")
    t_acc.font = F(bold=True, color="FFFFFF", size=12)
    t_acc.fill = ACC_FILL
    t_acc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[acc3_start].height = 24

    acc_hdr_row = acc3_start + 1
    for col, h in enumerate(["Account / Fund", "Units", "Total"] + fid_month_labels, 1):
        c = ws.cell(row=acc_hdr_row, column=col, value=h)
        if col <= 3:
            c.font = ACC_HFONT; c.fill = ACC_FILL
        else:
            m = fid_months[col - 4]
            c.font = ACC_HFONT if m in actual_set else EST_FONT
            c.fill = ACC_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[acc_hdr_row].height = 18

    acc_cur_row = acc_hdr_row + 1
    acc_account_total_rows = []

    for person in FAMILY_ORDER:
        person_accs = [acc for acc, owner in ACCOUNT_OWNER.items() if owner == person]
        for acc in person_accs:
            if acc not in acc_holdings:
                continue
            funds = acc_holdings[acc]
            label = f"{ACCOUNT_LABELS.get(acc, acc)} ({person})"
            n_fund_rows = len(funds)

            # Account header row
            acc_ar = acc_cur_row
            acc_account_total_rows.append(acc_ar)

            c = ws.cell(row=acc_ar, column=1, value=label)
            c.font = ACC_TFONT; c.fill = ACCT_FILL
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            ws.cell(row=acc_ar, column=2).fill = ACCT_FILL  # units blank for account
            ws.cell(row=acc_ar, column=2).border = Border(bottom=Side(style="thin", color=C_ACC_T))

            fc_a = get_column_letter(4)
            lc_a = get_column_letter(3 + len(fid_months))
            tc = ws.cell(row=acc_ar, column=3,
                         value=f"=SUM({fc_a}{acc_ar+1}:{lc_a}{acc_ar+n_fund_rows})")
            tc.number_format = NUM_FMT; tc.font = ACC_TFONT; tc.fill = ACCT_FILL
            tc.alignment = Alignment(horizontal="right", vertical="center")
            tc.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            # Sum of account values for each month (from fund rows below)
            for col_idx, m in enumerate(fid_months, 4):
                cell = ws.cell(row=acc_ar, column=col_idx,
                               value=f"=SUM({get_column_letter(col_idx)}{acc_ar+1}:{get_column_letter(col_idx)}{acc_ar+n_fund_rows})")
                cell.number_format = NUM_FMT; cell.font = ACC_TFONT; cell.fill = ACCT_FILL
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Border(bottom=Side(style="thin", color=C_ACC_T))

            acc_cur_row += 1

            # Fund rows — use actual price×units from transaction history
            for fund_offset, (fund, fund_data) in enumerate(sorted(funds.items())):
                fr = acc_cur_row
                is_alt = (fund_offset % 2 == 0)
                fill = ACCA_FILL if is_alt else None
                units = fund_data["units"]
                # Use price_appreciation (ignores new unit purchases, just price-driven gain)
                month_vals = fund_data.get("price_appreciation", fund_data.get("monthly_values", {}))
                may_val = fund_data["value"]

                # For future months not in price history, extrapolate using known annual yield
                # Annual yields from Fidelity (guaranteed non-negative)
                FUND_YIELDS = {
                    "Aegon High Yield Bond B Acc":                  0.0732,
                    "Schroder High Yield Opportunities Fund Z Acc": 0.0769,
                    "WS Guinness Global Energy Fund I Acc":         0.0235,
                }
                annual_yield = FUND_YIELDS.get(fund, 0.05)
                monthly_growth = (1 + annual_yield) ** (1/12) - 1

                if month_vals:
                    last_known = max(month_vals.keys())
                    last_val = month_vals[last_known]
                    for m in fid_months:
                        if m not in month_vals:
                            offset = (m - last_known).n
                            month_vals[m] = round(last_val * ((1 + monthly_growth) ** offset))

                c = ws.cell(row=fr, column=1, value=f"  {fund}")
                c.font = BODY_FONT; c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: c.fill = fill

                uc = ws.cell(row=fr, column=2, value=round(units, 2) if units else None)
                uc.number_format = '#,##0.##'; uc.font = BODY_FONT
                uc.alignment = Alignment(horizontal="right", vertical="center")
                uc.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: uc.fill = fill

                tc2 = ws.cell(row=fr, column=3,
                              value=f"=SUM({fc_a}{fr}:{lc_a}{fr})")
                tc2.number_format = NUM_FMT; tc2.font = BODY_FONT
                tc2.alignment = Alignment(horizontal="right", vertical="center")
                tc2.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                if fill: tc2.fill = fill

                for col_idx, m in enumerate(fid_months, 4):
                    val = month_vals.get(m, 0)
                    is_est = m not in actual_set
                    cell = ws.cell(row=fr, column=col_idx, value=int(round(val)) if val else None)
                    cell.number_format = NUM_FMT
                    cell.font = BODY_FONT if not is_est else EST_BODY
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=Side(style="thin", color="B2E0EC"))
                    if fill: cell.fill = fill

                acc_cur_row += 1

    # Grand total row — shows INCREASE each month (new units × price), not total value
    tot_acc_r = acc_cur_row
    c = ws.cell(row=tot_acc_r, column=1, value="TOTAL Accumulative (price appreciation)")
    c.font = F(bold=True, color="FFFFFF"); c.fill = ACC_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=tot_acc_r, column=2).fill = ACC_FILL
    ws.cell(row=tot_acc_r, column=3).fill = ACC_FILL

    for col_idx, m in enumerate(fid_months, 4):
        monthly_increase = sum(
            fd_h.get("price_appreciation", {}).get(m, 0)
            for funds_h in acc_holdings.values()
            for fd_h in funds_h.values()
        )
        cell = ws.cell(row=tot_acc_r, column=col_idx,
                       value=int(round(monthly_increase)) if monthly_increase else None)
        cell.number_format = NUM_FMT; cell.font = F(bold=True, color="FFFFFF")
        cell.fill = ACC_FILL; cell.alignment = Alignment(horizontal="right", vertical="center")
    acc_cur_row += 1

    # ── Update Total Income and Accumulations with Excel formulas now that row numbers are known ──
    # = Total Income (tot_r) + TOTAL Accumulative (tot_acc_r)
    col_letters = [get_column_letter(4 + i) for i in range(len(all_months))]
    for col_idx, m in enumerate(all_months, 4):
        cl = get_column_letter(col_idx)
        cell = ws.cell(row=total_inc_acc_row, column=col_idx,
                       value=f"={cl}{tot_r}+{cl}{tot_acc_r}")
        cell.number_format = NUM_FMT
        cell.font = F(bold=True, color="042C53")
        cell.fill = P(C_SUM_T)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    # Col C = annual total (sum of all 12 months)
    sum_range = "+".join(f"{cl}{total_inc_acc_row}" for cl in col_letters)
    c_total = ws.cell(row=total_inc_acc_row, column=3, value=f"={sum_range}")
    c_total.number_format = NUM_FMT
    c_total.font = F(bold=True, color="042C53")
    c_total.fill = P(C_SUM_T)
    c_total.alignment = Alignment(horizontal="right", vertical="center")
    fid3_start = acc_cur_row + 1

    C_FID3_H = "4A235B"   # deep purple header
    C_FID3_A = "F5EEF8"   # light purple alt rows
    C_FID3_T = "D2B4DE"   # purple total

    FID3_FILL  = P(C_FID3_H)
    FID3A_FILL = P(C_FID3_A)
    FID3T_FILL = P(C_FID3_T)
    FID3_HFONT = F(bold=True, color="FFFFFF")
    FID3_TFONT = F(bold=True, color="4A235B")

    n_fid3_cols = 3 + len(spend_months)
    ws.merge_cells(start_row=fid3_start, start_column=1,
                   end_row=fid3_start, end_column=n_fid3_cols)
    t3 = ws.cell(row=fid3_start, column=1, value="Fidelity")
    t3.font = F(bold=True, color="FFFFFF", size=12)
    t3.fill = FID3_FILL
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fid3_start].height = 24

    fid3_hdr = fid3_start + 1
    for col, h in enumerate(["Category", "", "Total"] + spend_month_labels, 1):
        c = ws.cell(row=fid3_hdr, column=col, value=h)
        if col <= 3:
            c.font = FID3_HFONT
            c.fill = FID3_FILL
        else:
            m = spend_months[col - 4]
            c.font = FID3_HFONT if m in actual_set else EST_FONT
            c.fill = FID3_FILL if m in actual_set else EST_HDR_FILL
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    ws.row_dimensions[fid3_hdr].height = 18

    fid3_data_row = fid3_start + 2
    fill = FID3A_FILL
    c = ws.cell(row=fid3_data_row, column=1, value="Fidelity card payments")
    c.font = BODY_FONT; c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="thin", color="D2B4DE"))
    ws.cell(row=fid3_data_row, column=2).fill = fill
    ws.cell(row=fid3_data_row, column=2).border = Border(bottom=Side(style="thin", color="D2B4DE"))

    fc = get_column_letter(4)
    lc = get_column_letter(3 + len(spend_months))
    tc = ws.cell(row=fid3_data_row, column=3,
                 value=f"=SUM({fc}{fid3_data_row}:{lc}{fid3_data_row})")
    tc.number_format = NUM_FMT; tc.font = FID3_TFONT; tc.fill = FID3T_FILL
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = Border(bottom=Side(style="thin", color="D2B4DE"))

    for col_idx, m in enumerate(spend_months, 4):
        val = spend_pivot.loc["Fidelity", m] if "Fidelity" in spend_pivot.index else 0
        is_est = m not in actual_set
        cell = ws.cell(row=fid3_data_row, column=col_idx,
                       value=int(round(val)) or None)
        cell.number_format = NUM_FMT
        cell.font = BODY_FONT if not is_est else EST_BODY
        cell.fill = FID3A_FILL if not is_est else P("ECE8F4")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="D2B4DE"))

    # ── Total Income + Salary row ──────────────────────────────────────────────
    combo_row = fid3_data_row + 1
    c = ws.cell(row=combo_row, column=1, value="Total Income & Salary")
    c.font = F(bold=True, color="FFFFFF"); c.fill = FID3_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=combo_row, column=2).fill = FID3_FILL  # blank

    # Income table: col1=label, col2=units, col3=Total, col4+=months (same as spend now)
    # Reference tot_r directly — all cols now aligned
    for col_idx in range(3, 4 + len(spend_months)):
        cl = get_column_letter(col_idx)
        cell = ws.cell(row=combo_row, column=col_idx, value=f"={cl}{tot_r}")
        cell.number_format = NUM_FMT
        cell.font = F(bold=True, color="FFFFFF"); cell.fill = FID3_FILL
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Investment Risk Metrics Table ──────────────────────────────────────────
    # Classify all Fidelity holdings into: Shares, Income Funds, Non-Income Funds
    # Columns: Group | Paul SIPP | Paul ISA | Joint Acct | Susan SIPP | Susan ISA | Jayne JISA | Liam ISA | Liam Inv | TOTAL
    # Rows: Shares, Income Funds, Non-Income Funds, TOTAL

    METRIC_ACCS_ORDER = [
        ('2000001606', 'Paul SIPP'),
        ('SANX002282', 'Paul ISA'),
        ('SANQ000468', 'Joint Acct'),
        ('2000001604', 'Susan SIPP'),
        ('SANX002617', 'Susan ISA'),
        ('SANX002936', 'Jayne JISA'),
        ('AS10303823', 'Liam ISA'),
        ('AG10131710', 'Liam Inv Acct'),
    ]

    HOLDING_GROUPS = {
        'Aegon High Yield Bond B Inc':                   'Income Funds',
        'Aegon High Yield Bond B Acc':                   'Income Funds',
        'Schroder High Yield Opportunities Fund Z Inc':  'Income Funds',
        'Schroder High Yield Opportunities Fund Z Acc':  'Income Funds',
        'Man High Yield Opportunities Fund Prof D Inc':  'Income Funds',
        'Man High Yield Opportunities Fund Prof D Acc':  'Income Funds',
        'WS Guinness Global Energy Fund I Acc':          'Non-Income Funds',
        'AUTOTRADER GROUP PLC,ORD GBP0.01(AUTO)':       'Shares',
        'AVIVA,ORD GBP0.328947368(AV.)':                 'Shares',
        'RELX PLC,ORD GBP0.1444(REL)':                  'Shares',
        'THE SAGE GROUP PLC,GBP0.01051948(SGE)':         'Shares',
        'WEIR GROUP,ORD GBP0.125(WEIR)':                 'Shares',
        'LEGAL & GENERAL GROUP,ORD GBP0.025(LGEN)':      'Shares',
    }
    GROUPS_ORDER = ['Shares', 'Income Funds', 'Non-Income Funds']

    # Build data: group → account → value
    metric_data = {g: {a: 0.0 for a, _ in METRIC_ACCS_ORDER} for g in GROUPS_ORDER}
    metric_data['_cash'] = {a: 0.0 for a, _ in METRIC_ACCS_ORDER}

    import csv as _csv2, io as _io2, os as _os2
    # Find AccountSummary path — passed in via summary_data
    _acct_path = summary_data.get('_account_summary_path', '')
    if not _acct_path or not _os2.path.exists(_acct_path):
        for _p in ['AccountSummary.csv', '/home/claude/AccountSummary.csv']:
            if _os2.path.exists(_p):
                _acct_path = _p
                break
    with open(_acct_path, encoding='utf-8-sig') as _f:
        _content = _f.read()
    _lines = _content.replace('\r','').split('\n')
    _hi = max(i for i, l in enumerate(_lines) if l.startswith('Type,Holdings,Account number'))
    for _row in _csv2.DictReader(_io2.StringIO('\n'.join(_lines[_hi:]))):
        _acc = _row.get('Account number','').strip()
        if _acc not in {a for a,_ in METRIC_ACCS_ORDER}: continue
        _t = _row.get('Type','').strip()
        _val = float(_row.get('Value (£)','0').replace(',','') or 0)
        if _t == 'Asset':
            _fund = _row.get('Holdings','').strip()
            _grp = HOLDING_GROUPS.get(_fund)
            if _grp:
                metric_data[_grp][_acc] += _val
        elif _t == 'Account':
            # Total account value — used to derive cash (account total - sum of assets)
            metric_data['_cash'][_acc] = _val

    # Cash = account total - sum of classified assets
    metric_cash = {}
    for acc, _ in METRIC_ACCS_ORDER:
        asset_sum = sum(metric_data[g][acc] for g in GROUPS_ORDER)
        metric_cash[acc] = max(0, metric_data['_cash'][acc] - asset_sum)

    # ── Write table ─────────────────────────────────────────────────────────────
    C_MET_H = "2C3E50"    # dark slate header
    C_MET_A = "F2F3F4"    # light grey alt
    C_MET_T = "D5D8DC"    # grey total
    C_MET_S = "E8F4F8"    # light blue for Shares
    C_MET_I = "EAF5EA"    # light green for Income
    C_MET_N = "FEF9E7"    # light yellow for Non-Income

    GROUP_COLOURS = {
        'Shares':           C_MET_S,
        'Income Funds':     C_MET_I,
        'Non-Income Funds': C_MET_N,
    }

    met_start = combo_row + 3 if 'combo_row' in dir() else ws.max_row + 3
    n_acc_cols = len(METRIC_ACCS_ORDER)
    n_met_cols = 1 + n_acc_cols + 1  # Group label + accounts + Total

    # Section title
    ws.merge_cells(start_row=met_start, start_column=1,
                   end_row=met_start, end_column=n_met_cols)
    t = ws.cell(row=met_start, column=1, value="Investment Risk Metrics")
    t.font = F(bold=True, color="FFFFFF", size=12)
    t.fill = P(C_MET_H)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[met_start].height = 24

    # Column headers
    hdr_row = met_start + 1
    ws.cell(row=hdr_row, column=1, value="Group").font = F(bold=True, color="FFFFFF", size=10)
    ws.cell(row=hdr_row, column=1).fill = P(C_MET_H)
    ws.cell(row=hdr_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for ci, (acc, label) in enumerate(METRIC_ACCS_ORDER, 2):
        c = ws.cell(row=hdr_row, column=ci, value=label)
        c.font = F(bold=True, color="FFFFFF", size=9)
        c.fill = P(C_MET_H)
        c.alignment = Alignment(horizontal="right", vertical="center")
    tot_hdr = ws.cell(row=hdr_row, column=2 + n_acc_cols, value="TOTAL")
    tot_hdr.font = F(bold=True, color="FFFFFF", size=10)
    tot_hdr.fill = P(C_MET_H)
    tot_hdr.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[hdr_row].height = 18

    # Data rows — value rows
    val_rows = {}
    cur_met = hdr_row + 1
    grand_met_total = sum(
        sum(metric_data[g][acc] for acc, _ in METRIC_ACCS_ORDER)
        for g in GROUPS_ORDER
    ) + sum(metric_cash.values())

    all_groups = GROUPS_ORDER + ['Cash']
    for gi, grp in enumerate(all_groups):
        row_fill = GROUP_COLOURS.get(grp, C_MET_A)
        r = cur_met
        val_rows[grp] = r
        # Label
        lbl = ws.cell(row=r, column=1, value=grp)
        lbl.font = F(bold=False, color="000000", size=10)
        lbl.fill = P(row_fill)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        # Account values
        row_total = 0
        for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
            val = metric_data[grp][acc] if grp in metric_data else metric_cash.get(acc, 0)
            if grp == 'Cash':
                val = metric_cash.get(acc, 0)
            row_total += val
            cell = ws.cell(row=r, column=ci)
            if val > 0:
                cell.value = round(val)
                cell.number_format = '#,##0'
            cell.fill = P(row_fill)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = F(size=10)
        # Row total
        tc = ws.cell(row=r, column=2 + n_acc_cols)
        if row_total > 0:
            tc.value = round(row_total)
            tc.number_format = '#,##0'
        tc.fill = P(row_fill)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.font = F(bold=True, size=10)
        ws.row_dimensions[r].height = 16
        cur_met += 1

    # % of total rows
    pct_start = cur_met
    cur_met += 1  # blank row
    ws.row_dimensions[cur_met - 1].height = 6

    for gi, grp in enumerate(all_groups):
        row_fill = GROUP_COLOURS.get(grp, C_MET_A)
        r = pct_start + gi
        lbl = ws.cell(row=r, column=1, value=f"{grp} %")
        lbl.font = F(italic=True, color="555555", size=9)
        lbl.fill = P(row_fill)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        row_total = sum(metric_data[grp][acc] for acc, _ in METRIC_ACCS_ORDER) if grp in metric_data else sum(metric_cash.values())
        for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
            val = metric_data[grp][acc] if grp in metric_data else metric_cash.get(acc, 0)
            if grp == 'Cash':
                val = metric_cash.get(acc, 0)
            acc_total_val = metric_data['_cash'].get(acc, 0)
            if acc_total_val > 0:
                pct = val / acc_total_val * 100
                cell = ws.cell(row=r, column=ci)
                if pct > 0:
                    cell.value = round(pct, 1)
                    cell.number_format = '0.0"%"'
                cell.fill = P(row_fill)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = F(italic=True, size=9, color="555555")
        # Overall % of portfolio
        overall_pct = row_total / grand_met_total * 100 if grand_met_total > 0 else 0
        tc = ws.cell(row=r, column=2 + n_acc_cols)
        if overall_pct > 0:
            tc.value = round(overall_pct, 1)
            tc.number_format = '0.0"%"'
        tc.fill = P(row_fill)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.font = F(bold=True, italic=True, size=9, color="555555")
        ws.row_dimensions[r].height = 14
        cur_met += 1

    # Grand total row
    cur_met += 1
    gt_row = cur_met
    ws.cell(row=gt_row, column=1, value="TOTAL INVESTMENTS").font = F(bold=True, color="FFFFFF", size=10)
    ws.cell(row=gt_row, column=1).fill = P(C_MET_H)
    ws.cell(row=gt_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for ci, (acc, _) in enumerate(METRIC_ACCS_ORDER, 2):
        acc_total_val = metric_data['_cash'].get(acc, 0)
        cell = ws.cell(row=gt_row, column=ci)
        if acc_total_val > 0:
            cell.value = round(acc_total_val)
            cell.number_format = '#,##0'
        cell.fill = P(C_MET_H)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = F(bold=True, color="FFFFFF", size=10)
    tc = ws.cell(row=gt_row, column=2 + n_acc_cols)
    fid_total_val = sum(metric_data['_cash'].get(acc, 0) for acc, _ in METRIC_ACCS_ORDER)
    if fid_total_val > 0:
        tc.value = round(fid_total_val)
        tc.number_format = '#,##0'
    tc.fill = P(C_MET_H)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.font = F(bold=True, color="FFFFFF", size=10)
    ws.row_dimensions[gt_row].height = 18

    # Note row
    note_met = gt_row + 1
    nc = ws.cell(row=note_met, column=1,
                 value="Values as at AccountSummary export date. Cash = account total minus classified assets. Totals match Fidelity accounts section above.")
    nc.font = Font(italic=True, color="888888", size=8)
    nc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=note_met, start_column=1,
                   end_row=note_met, end_column=n_met_cols)
    ws.row_dimensions[note_met].height = 12

    # ── Targets Table ───────────────────────────────────────────────────────────
    C_TGT_H  = "1A3A5C"   # dark navy header
    C_TGT_ON = "E9F7EF"   # green — on/above target
    C_TGT_OF = "FDEDEC"   # red   — below target
    C_TGT_NR = "FEF9E7"   # amber — within 10% of target
    C_TGT_LB = "EBF5FB"   # blue  — label rows

    tgt_start = note_met + 3
    # Columns: Metric | Actual | Target | Status | Notes
    TGT_COLS = ["Metric", "Actual", "Target", "Status", "Notes"]
    TGT_WIDTHS = [38, 16, 16, 12, 40]

    # Title
    ws.merge_cells(start_row=tgt_start, start_column=1, end_row=tgt_start, end_column=5)
    tt = ws.cell(row=tgt_start, column=1, value="2026 Targets")
    tt.font = F(bold=True, color="FFFFFF", size=12)
    tt.fill = P(C_TGT_H)
    tt.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tgt_start].height = 24

    # Headers
    hdr_tgt = tgt_start + 1
    for ci, (h, w) in enumerate(zip(TGT_COLS, TGT_WIDTHS), 1):
        c = ws.cell(row=hdr_tgt, column=ci, value=h)
        c.font = F(bold=True, color="FFFFFF", size=10)
        c.fill = P(C_TGT_H)
        c.alignment = Alignment(horizontal="left" if ci == 1 else "right", vertical="center")
    ws.row_dimensions[hdr_tgt].height = 18

    # ── Compute actuals ────────────────────────────────────────────────────────
    # Both pivots already carry the whole year: pinned history for the early
    # months, transaction data for the covered ones and estimates for the rest
    # (see estimate_future_months). So the annual figures are just their row
    # sums — no month literals, and nothing to re-tune as the calendar moves.
    inc_hist_data = load_income_history()
    fid_annual = 0
    if not fid_pivot.empty:
        cols = [m for m in all_months if m in fid_pivot.columns]
        fid_annual = float(fid_pivot[cols].sum().sum())
        # Any pinned history month the pivot has no column for
        for month_vals in inc_hist_data.values():
            for period, val in month_vals.items():
                if period not in fid_pivot.columns:
                    fid_annual += val

    salary_annual = 0
    if 'Salary' in spend_pivot.index:
        cols = [m for m in all_months if m in spend_pivot.columns]
        salary_annual = float(spend_pivot.loc['Salary', cols].sum())

    equity_annual = EQUITY_DIVIDENDS_ANNUAL
    total_annual_income = fid_annual + salary_annual + equity_annual
    income_avg_pm = round(total_annual_income / 12)

    # Emit the monthly SHARE-dividend figures for the Investment Dashboard's Monthly
    # Dividend metric (user request 2026-07-18: add share income + accumulation
    # dividends). share_income = equity dividends run at equity_annual/12; share
    # accumulation = the Acc funds' reinvested income (inc_ppu × units), already
    # computed per fund/month in build_acc_holdings as 'price_appreciation' — take the
    # latest month's total across all Acc funds (same figure as the Wealth Summary's
    # 'TOTAL Accumulative (price appreciation)' row). Written to data/ for the dashboard;
    # income-fund revenue stays sourced from the master's Income Funds tab (no dup).
    try:
        acc_appr = {}
        for _funds in (acc_holdings or {}).values():
            for _fd in _funds.values():
                for _per, _v in (_fd.get('price_appreciation') or {}).items():
                    acc_appr[_per] = acc_appr.get(_per, 0) + _v
        _dividends = {
            'generated_at': pd.Timestamp.now().isoformat(timespec='seconds'),
            'share_income_monthly': round(equity_annual / 12.0, 2),
            'share_accumulation_monthly': round(acc_appr[max(acc_appr)], 2) if acc_appr else 0.0,
        }
        _repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        _dp = os.path.join(_repo, 'data', 'spending_dividends.json')
        os.makedirs(os.path.dirname(_dp), exist_ok=True)
        with open(_dp, 'w', encoding='utf-8') as _f:
            json.dump(_dividends, _f, indent=2)
        print(f"  Share dividends -> income £{_dividends['share_income_monthly']:,.0f}/mo, "
              f"accumulation £{_dividends['share_accumulation_monthly']:,.0f}/mo")
    except Exception as _e:
        print(f"  (share-dividend export skipped: {_e})")

    # These targets are all "what is it worth NOW" figures, so they read the
    # snapshot month's column — month columns start at column 4, and _r is a
    # 0-indexed tuple. This used to be a hardcoded _r[7] (May) and quietly went on
    # reporting a stale month once the export moved on.
    _CUR = 3 + all_months.index(anchors.data_month)

    # 2. Paul SIPP 25% drawdown
    paul_sipp_val_tgt = 0
    for _r in ws.iter_rows():
        if '2000001606' in str(_r[0].value or '') and 'SIPP' in str(_r[0].value or ''):
            paul_sipp_val_tgt = _resolve_cell_num(ws, _r[_CUR])
            break
    paul_drawdown = round(paul_sipp_val_tgt * 0.25)

    # 3. Susan SIPP 25% drawdown
    susan_sipp_val_tgt = 0
    for _r in ws.iter_rows():
        if '2000001604' in str(_r[0].value or '') and 'SIPP' in str(_r[0].value or ''):
            susan_sipp_val_tgt = _resolve_cell_num(ws, _r[_CUR])
            break
    susan_drawdown = round(susan_sipp_val_tgt * 0.25)

    # 4. ISA values combined
    paul_isa_val = susan_isa_val = 0
    for _r in ws.iter_rows():
        if str(_r[0].value or '').strip() == 'Investment ISA (Paul)':
            paul_isa_val = _resolve_cell_num(ws, _r[_CUR])
        if str(_r[0].value or '').strip() == 'Investment ISA (Susan)':
            susan_isa_val = _resolve_cell_num(ws, _r[_CUR])
    isa_combined = round(paul_isa_val + susan_isa_val)

    # 5. Growth % (Shares + Non-Income Funds) and Income % of total invested
    shares_val = inc_fund_val = non_inc_val = 0
    for _r in ws.iter_rows():
        v = str(_r[0].value or '').strip()
        if v == 'Shares':
            shares_val = _resolve_cell_num(ws, _r[9])
        elif v == 'Income Funds':
            inc_fund_val = _resolve_cell_num(ws, _r[9])
        elif v == 'Non-Income Funds':
            non_inc_val = _resolve_cell_num(ws, _r[9])
    total_invested = shares_val + inc_fund_val + non_inc_val
    growth_pct  = round((shares_val + non_inc_val) / total_invested * 100, 1) if total_invested else 0
    income_pct  = round(inc_fund_val / total_invested * 100, 1) if total_invested else 0

    # 6. Fidelity service fees — sum from live TransactionHistory
    import csv as _csv3, io as _io3
    import sys as _sys3
    _th_path = next((a for a in _sys3.argv[1:] if 'TransactionHistory' in a or 'transaction' in a.lower()), 'TransactionHistory.csv')
    with open(_th_path) as _f3:
        _lines3 = _f3.read().replace('\r','').split('\n')
    _start3 = next(i for i,l in enumerate(_lines3) if l.startswith('Order date'))
    _rows3 = list(_csv3.DictReader(_io3.StringIO('\n'.join(_lines3[_start3:]))))
    svc_fees_live = sum(abs(float(r.get('Amount','0') or 0))
                        for r in _rows3
                        if r.get('Transaction type','').strip() == 'Service Fee'
                        and r.get('Status','').strip() == 'Completed')
    # Annualise from Jan-May actual (5 months)
    svc_fees_annual = round(svc_fees_live / 5 * 12)

    def status_cell(ws, r, c, actual, target, higher_is_better=True, is_pct=False):
        """Write RAG status cell."""
        if target == 0:
            ws.cell(row=r, column=c, value="—")
            return "—"
        pct_of_tgt = actual / target
        if higher_is_better:
            if pct_of_tgt >= 1.0:   rag, txt = C_TGT_ON, "✓ On Target"
            elif pct_of_tgt >= 0.9: rag, txt = C_TGT_NR, "⚠ Near Target"
            else:                    rag, txt = C_TGT_OF, "✗ Below Target"
        else:  # lower is better (e.g. costs)
            if pct_of_tgt <= 1.0:   rag, txt = C_TGT_ON, "✓ Within Budget"
            elif pct_of_tgt <= 1.1: rag, txt = C_TGT_NR, "⚠ Slightly Over"
            else:                    rag, txt = C_TGT_OF, "✗ Over Budget"
        cell = ws.cell(row=r, column=c, value=txt)
        cell.fill = P(rag)
        cell.font = F(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return rag

    def tgt_row(ws, r, metric, actual, target, status_rag, note,
                actual_fmt="£{:,.0f}", target_fmt="£{:,.0f}", alt=False):
        bg = status_rag if status_rag not in (True, False) else (C_TGT_LB if alt else "FFFFFF")
        # Metric label
        c = ws.cell(row=r, column=1, value=metric)
        c.font = F(size=10); c.fill = P(C_TGT_LB)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Actual
        act_cell = ws.cell(row=r, column=2, value=actual_fmt.format(actual) if actual_fmt else actual)
        act_cell.font = F(bold=True, size=11)
        act_cell.fill = P(status_rag if isinstance(status_rag, str) and len(status_rag)==6 else "FFFFFF")
        act_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Target
        tgt_cell = ws.cell(row=r, column=3, value=target_fmt.format(target) if target_fmt else target)
        tgt_cell.font = F(size=10, color="444444"); tgt_cell.fill = P("F8F9FA")
        tgt_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Note
        nc = ws.cell(row=r, column=5, value=note)
        nc.font = F(italic=True, size=9, color="666666"); nc.fill = P(C_TGT_LB)
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22

    # ── Write target rows ──────────────────────────────────────────────────────
    tgt_data_start = hdr_tgt + 1

    # Row 1: Income per month
    r = tgt_data_start
    rag = status_cell(ws, r, 4, income_avg_pm, 10000, higher_is_better=True)
    tgt_row(ws, r,
            "Income per month (avg over 12 months)",
            income_avg_pm, 10000, rag,
            f"Annual total £{int(total_annual_income):,} ÷ 12. Includes salary, Fidelity fund income & equity dividends.")

    # Row 2: Paul SIPP 25% drawdown
    r += 1
    PAUL_DRAWDOWN_TAKEN = 0
    rag2 = status_cell(ws, r, 4, PAUL_DRAWDOWN_TAKEN + paul_drawdown, 269000, higher_is_better=True)
    tgt_row(ws, r,
            "Paul Pension — 25% tax-free drawdown available",
            paul_drawdown, 269000, rag2,
            f"SIPP value £{int(paul_sipp_val_tgt):,} × 25% = £{paul_drawdown:,}. Taken to date: £0.")

    # Row 3: Susan SIPP 25% drawdown
    r += 1
    SUSAN_DRAWDOWN_TAKEN = 0
    rag3 = status_cell(ws, r, 4, SUSAN_DRAWDOWN_TAKEN + susan_drawdown, 269000, higher_is_better=True)
    tgt_row(ws, r,
            "Susan Pension — 25% tax-free drawdown available",
            susan_drawdown, 269000, rag3,
            f"SIPP value £{int(susan_sipp_val_tgt):,} × 25% = £{susan_drawdown:,}. Taken to date: £0.")

    # Row 4: ISA values
    r += 1
    rag4 = status_cell(ws, r, 4, isa_combined, 1000000, higher_is_better=True)
    tgt_row(ws, r,
            "ISA Values (Paul ISA + Susan ISA combined)",
            isa_combined, 1000000, rag4,
            f"Paul ISA £{int(paul_isa_val):,} + Susan ISA £{int(susan_isa_val):,}.")

    # Row 5: Growth funds % — target is to REACH 60%, so higher is better
    r += 1
    rag5 = status_cell(ws, r, 4, growth_pct, 60, higher_is_better=True)
    tgt_row(ws, r,
            "Growth funds % of total invested (Shares + Non-Income Funds)",
            growth_pct, 60, rag5,
            f"Shares £{int(shares_val):,} + Non-Income £{int(non_inc_val):,} = £{int(shares_val+non_inc_val):,} of £{int(total_invested):,}.",
            actual_fmt="{:.1f}%", target_fmt="{:.0f}%")

    # Row 6: Income fund % — target is to stay AT/BELOW 40%, lower is better
    r += 1
    rag6 = status_cell(ws, r, 4, income_pct, 40, higher_is_better=False)
    tgt_row(ws, r,
            "Income funds % of total invested",
            income_pct, 40, rag6,
            f"Income Funds £{int(inc_fund_val):,} of £{int(total_invested):,} total invested. Currently {income_pct:.1f}% — target is to reduce to 40%.",
            actual_fmt="{:.1f}%", target_fmt="{:.0f}%")

    # Row 7: Fidelity service costs
    r += 1
    rag7 = status_cell(ws, r, 4, svc_fees_annual, 3000, higher_is_better=False)
    tgt_row(ws, r,
            "Fidelity annual service costs",
            svc_fees_annual, 3000, rag7,
            f"Jan–May actual £{svc_fees_live:,.2f} × 12/5 = £{svc_fees_annual:,} annualised estimate.")

    # Single comprehensive sheet
    ws.title = "Wealth Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    # Historic month cols: narrower
    for i, m in enumerate(sum_months):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 9 if m in hist_months else 11

    # Freeze panes: col A–C (labels) and row 1–2 (title + month headers) always visible
    ws.freeze_panes = "D3"

    # ── Move 'Investment Risk Metrics' and '2026 Targets' to a separate sheet ──
    section_rows = {}
    for row in ws.iter_rows():
        v = str(row[0].value or '').strip()
        if v in ('Investment Risk Metrics', '2026 Targets'):
            section_rows[v] = row[0].row

    if 'Investment Risk Metrics' in section_rows:
        move_start = section_rows['Investment Risk Metrics']
        move_end = ws.max_row

        ws_targets = wb.create_sheet("Targets")
        ws_targets.sheet_properties.tabColor = "E67E22"

        dst_row = 1
        for sr in range(move_start, move_end + 1):
            for sc in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=sr, column=sc)
                dst_cell = ws_targets.cell(row=dst_row, column=sc)
                dst_cell.value = src_cell.value
                if src_cell.data_type == 'f':
                    dst_cell.data_type = 'f'
                copy_cell_style(src_cell, dst_cell)
            if sr in ws.row_dimensions:
                ws_targets.row_dimensions[dst_row].height = ws.row_dimensions[sr].height
            dst_row += 1

        # Copy column widths
        for col_letter, col_dim in ws.column_dimensions.items():
            ws_targets.column_dimensions[col_letter].width = col_dim.width

        # Copy merged cells within the moved range
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= move_start and merged.max_row <= move_end:
                offset = move_start - 1
                try:
                    ws_targets.merge_cells(
                        start_row=merged.min_row - offset, start_column=merged.min_col,
                        end_row=merged.max_row - offset,   end_column=merged.max_col)
                except Exception:
                    pass

        ws_targets.freeze_panes = "D3"

        # Remove the moved rows from the main sheet (delete from bottom up not needed —
        # delete_rows handles the range in one call)
        ws.delete_rows(move_start, move_end - move_start + 1)

    # Force Excel to recalculate all formulas when the file is opened
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Post-process: ensure every cell with a string starting "=" is stored as a formula
    # (openpyxl sometimes stores formula strings as text data_type="s" instead of "f")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.data_type = "f"

    # ── Insert "This file is..." note row at top of Wealth Summary ─────────────
    import re as _re
    for sheet in wb.worksheets:
        # Capture merge ranges BEFORE insert (insert_rows does not shift these)
        old_merges = [str(m) for m in sheet.merged_cells.ranges]
        for m in list(sheet.merged_cells.ranges):
            try:
                sheet.unmerge_cells(str(m))
            except KeyError:
                sheet.merged_cells.ranges.discard(m)

        sheet.insert_rows(1)

        # Formula text needs manual row-number adjustment (insert_rows shifts
        # cell positions but not the formula strings themselves)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    def _bump(m):
                        return m.group(1) + str(int(m.group(2)) + 1)
                    cell.value = _re.sub(r"([A-Z]{1,3})(\d+)", _bump, cell.value)
                    cell.data_type = "f"

        # Re-apply merges, shifted down by 1 row
        for m_str in old_merges:
            # m_str like "A47:O47"
            match = _re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", m_str)
            if match:
                c1, r1, c2, r2 = match.groups()
                from openpyxl.utils import column_index_from_string
                sheet.merge_cells(start_row=int(r1)+1, start_column=column_index_from_string(c1),
                                  end_row=int(r2)+1, end_column=column_index_from_string(c2))

        # Shift freeze panes down by 1 row (D3 -> D4)
        if sheet.freeze_panes:
            fp = sheet.freeze_panes
            col_part = ''.join(c for c in fp if c.isalpha())
            row_part = ''.join(c for c in fp if c.isdigit())
            if row_part:
                sheet.freeze_panes = f"{col_part}{int(row_part)+1}"

    # Write the note row
    note_text = "This file is 'Spending Summary · XLSX' from downloads"
    for sheet in wb.worksheets:
        c = sheet.cell(row=1, column=1, value=note_text)
        c.font = Font(italic=True, color="7F4000", size=10, name="Arial")
        c.fill = PatternFill("solid", fgColor="FFF3CD")
        c.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 16
        # Extend yellow shading across all columns and merge
        for col in range(2, sheet.max_column + 1):
            sheet.cell(row=1, column=col).fill = PatternFill("solid", fgColor="FFF3CD")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

    preserve_manual_sheets(wb, output_path)
    wb.save(output_path)


def preserve_manual_sheets(new_wb, output_path):
    """This builder regenerates the whole workbook every run — but the user
    keeps hand-maintained tabs (e.g. 'Payslip Summary', 'Retirement Income
    Plan') alongside the generated ones, and a rebuild used to silently drop
    them (lost tabs reported 2026-07-12, restored from the 2026-07-02 copy).
    Carry over every sheet in the existing file whose name this run didn't
    generate, so manual tabs survive rebuilds."""
    if not os.path.exists(output_path):
        return
    try:
        old_wb = load_workbook(output_path)
    except Exception as e:
        print(f"WARNING: could not read existing {output_path} to preserve "
              f"manual tabs ({e}) — generated tabs only this run.", file=sys.stderr)
        return
    generated = set(new_wb.sheetnames)
    for name in old_wb.sheetnames:
        if name in generated:
            continue
        copy_sheet_into(old_wb[name], new_wb.create_sheet(name))
        print(f"Preserved manual tab: {name}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    amex_path    = args[0] if len(args) > 0 else "activity.csv"
    bar_path     = args[1] if len(args) > 1 else "data.csv"
    fid_path     = args[2] if len(args) > 2 else "TransactionHistory.csv"
    summary_path = args[3] if len(args) > 3 else "AccountSummary.csv"
    output_path  = args[4] if len(args) > 4 else "spending_summary.xlsx"
    # Optional: a pending-orders export (every row has Completion date ==
    # 'Pending') — net units from pending Buy/Sell orders are layered on top
    # of the settled holdings figure. Not required; omit to skip.
    pending_path = args[5] if len(args) > 5 else None

    # Partial inputs are allowed (user request 2026-07-18): build with whatever
    # sources are present. Only bail if NONE of the three are — nothing to do.
    def _have(p):
        return bool(p) and os.path.exists(p)
    if not any(_have(p) for p in (amex_path, bar_path, fid_path)):
        print("Error: none of the spending sources found (Amex / Barclays / Fidelity).")
        sys.exit(1)
    missing = [name for name, p in (('Amex', amex_path), ('Barclays', bar_path),
                                    ('Fidelity', fid_path)) if not _have(p)]
    if missing:
        print(f"  Partial inputs — building without: {', '.join(missing)}")

    print(f"  Amex:     {amex_path if _have(amex_path) else '(not provided)'}")
    print(f"  Barclays: {bar_path if _have(bar_path) else '(not provided)'}")
    print(f"  Fidelity: {fid_path if _have(fid_path) else '(not provided)'}")
    if pending_path:
        print(f"  Pending:  {pending_path}")
    print(f"  Output:   {output_path}\n")

    # Every month boundary this run depends on, derived from the inputs.
    anchors = resolve_anchors(summary_path)
    print(f"Anchors: {anchors.describe()}")
    for _w in anchors.warnings():
        print(f"  WARNING: {_w}")
    print()

    print("Loading Amex...")
    amex_df = load_amex(amex_path)
    print(f"  {len(amex_df)} transactions\n")

    print("Loading Barclays...")
    bar_df = load_barclays(bar_path)
    print(f"  {len(bar_df)} transactions\n")

    print("Loading Fidelity income...")
    fid_df = load_fidelity_income(fid_path, anchors.year)
    print(f"  {len(fid_df)} income entries\n")

    print("Building holdings...")
    settled_holdings = build_holdings(fid_path, summary_path)
    holdings = build_holdings(fid_path, summary_path, pending_path)
    src = "AccountSummary.csv" if os.path.exists(summary_path) else "transaction history"
    print(f"  {len(holdings)} positions (from {src})")
    if pending_path and os.path.exists(pending_path):
        changed = sorted(k for k in set(settled_holdings) | set(holdings)
                          if round(settled_holdings.get(k, 0), 2) != round(holdings.get(k, 0), 2))
        if changed:
            print(f"  {len(changed)} position(s) adjusted by pending orders:")
            for acc, fund in changed:
                before = settled_holdings.get((acc, fund), 0)
                after = holdings.get((acc, fund), 0)
                print(f"    {acc} — {fund}: {before:,.2f} -> {after:,.2f}")
        else:
            print("  No pending Buy/Sell orders affected current holdings")
    print()

    print("Building pivots...")
    spend_pivot, spend_months       = build_spending_pivot(amex_df, bar_df, anchors)
    fid_pivot, fid_months           = build_fidelity_pivot(fid_df, anchors)
    acc_fund_map, acc_fund_months   = build_account_fund_pivot(fid_df)

    # Always show the full reporting year — the transaction files only cover a
    # 60-day window, the pinned history covers the early months, and everything
    # in between or beyond is estimated.
    all_months = list(anchors.months)

    # Split the year into complete actuals vs months that must be estimated.
    # Done PER PIVOT, from that pivot's own coverage: the spending sources and the
    # Fidelity export cover different months, and a bank export that is missing
    # entirely must not make its months read as actual zeroes (that once left the
    # salary row blank for a month and dragged the median down for every other).
    spend_actual, spend_future = anchors.split_months(spend_months)
    fid_actual, fid_future = anchors.split_months(fid_months)
    actual_months, future_months = anchors.split_months(
        sorted(set(spend_months) | set(fid_months)))

    spend_months = fid_months = acc_fund_months = all_months
    spend_pivot = spend_pivot.reindex(columns=all_months + ["Total"], fill_value=0)
    if not fid_pivot.empty:
        fid_pivot = fid_pivot.reindex(columns=all_months + ["Total"], fill_value=0)
        fid_pivot["Total"] = fid_pivot[all_months].sum(axis=1)
    for person in acc_fund_map:
        for acc in acc_fund_map[person]:
            df = acc_fund_map[person][acc]
            df = df.reindex(columns=all_months + ["Total"], fill_value=0)
            df["Total"] = df[all_months].sum(axis=1)
            acc_fund_map[person][acc] = df

    def _months(ms):
        return ", ".join(str(m) for m in ms) if ms else "none"
    print(f"  Anchors:  {anchors.describe()}")
    print(f"  Spend   — actual: {_months(spend_actual)} | estimated: {_months(spend_future)}")
    print(f"  Income  — actual: {_months(fid_actual)} | estimated: {_months(fid_future)}")

    # Full year Jan–Dec (actuals + future, always 12 months)
    full_months = all_months

    # ── Inject missing stocks BEFORE estimation ───────────────────────────────
    # Stocks held (from AccountSummary) but with no 2026 income transactions
    # Uses confirmed dividend pence-per-share × units for accurate estimates
    STOCK_INJECT = {
        "LEGAL & GENERAL GROUP,ORD GBP0.025(LGEN)": [
            (6, 15.36),   # Final: pay Jun 4 (ex Apr 23)
            (9, 6.12),    # Interim: pay Sep 25 (ex Aug 20)
        ],
    }
    FUND_NAME_MAP = {
        "LEGAL & GENERAL GROUP,ORD GBP0.025(LGEN)": "LEGAL & GENERAL GROUP, ORD GBP0.025 (LGEN)",
    }
    # Track injected fund names so estimate_future_months can skip them
    injected_funds = set()

    for person in acc_fund_map:
        for acc in list(acc_fund_map[person].keys()):
            for summary_name, payments in STOCK_INJECT.items():
                income_name = FUND_NAME_MAP.get(summary_name, summary_name)
                fund_df = acc_fund_map[person][acc]
                if income_name in fund_df.index and fund_df.loc[income_name].sum() > 0:
                    continue
                units = holdings.get((acc, summary_name), 0)
                if units <= 0:
                    continue
                new_row = {m: 0 for m in full_months + ["Total"]}
                for pay_month, ppm in payments:
                    for m in full_months:
                        if m.month == pay_month and m.year == anchors.year:
                            new_row[m] = round(units * ppm / 100)
                new_row["Total"] = sum(new_row[m] for m in full_months)
                new_df = pd.DataFrame([new_row], index=[income_name])
                new_df.index.name = "fund"
                fund_df = fund_df.reindex(columns=full_months + ["Total"], fill_value=0)
                acc_fund_map[person][acc] = pd.concat([fund_df, new_df]).sort_index()
                injected_funds.add(income_name)
                print(f"  Injected {income_name} → {acc} ({units:,.0f} units, "
                      f"Jun=£{round(units*payments[0][1]/100):,}, "
                      f"Sep=£{round(units*payments[1][1]/100):,})")

    # Apply estimates to all pivots (injected funds are skipped)
    spend_pivot    = estimate_future_months(spend_pivot, spend_actual, spend_future, anchors)
    if not fid_pivot.empty:
        fid_pivot  = estimate_future_months(fid_pivot, fid_actual, fid_future, anchors)
    for person in acc_fund_map:
        for acc in acc_fund_map[person]:
            acc_fund_map[person][acc] = estimate_future_months(
                acc_fund_map[person][acc], fid_actual, fid_future, anchors,
                skip_funds=injected_funds)

    # ── Partial-month provisional fallback ────────────────────────────────────
    # The transaction export only covers the last 60 days, and the snapshot is
    # usually taken mid-month, so a monthly distribution due later in the partial
    # month isn't in the file yet. Carry the previous month's figure as a
    # placeholder — it is replaced by the real one on the next export.
    partial = anchors.partial_month
    prev = (partial - 1) if partial else None
    provisional_funds = set()
    if partial:
        for person in acc_fund_map:
            for acc, fund_df in acc_fund_map[person].items():
                if partial in fund_df.columns and prev in fund_df.columns:
                    for fund in fund_df.index:
                        if fund in injected_funds:
                            continue
                        cur_val = fund_df.loc[fund, partial]
                        prev_val = fund_df.loc[fund, prev]
                        if (pd.isna(cur_val) or cur_val == 0) and prev_val and prev_val > 0:
                            fund_df.loc[fund, partial] = prev_val
                            provisional_funds.add((acc, fund))
    if provisional_funds:
        print(f"  {partial}: provisional estimates on {len(provisional_funds)} fund rows "
              f"(carried from {prev})")

    # Rebuild fid_pivot account totals after estimation + injection
    for person in acc_fund_map:
        for acc in acc_fund_map[person]:
            fund_df = acc_fund_map[person][acc]
            acc_total = fund_df[full_months].sum()
            if acc in fid_pivot.index:
                for m in full_months:
                    # Only overwrite months the pinned income history doesn't own
                    if m >= anchors.hist_cutoff:
                        fid_pivot.loc[acc, m] = acc_total[m]
                # Always update Total
                fid_pivot.loc[acc, "Total"] = fid_pivot.loc[acc, full_months].sum()

    all_months = full_months
    spend_months = fid_months = acc_fund_months = all_months

    print(f"  Fidelity: {len(fid_pivot)} accounts\n")

    print("Writing Excel...")
    # The sheet shades months as actual vs estimated. Everything up to and
    # including the snapshot month is shown as actual (the partial month has real
    # data in it, just not all of it); everything after is projection.
    actual_for_excel = [m for m in anchors.months if m <= anchors.data_month]
    future_for_excel = [m for m in anchors.months if m > anchors.data_month]
    full_months = actual_for_excel + future_for_excel  # the whole year, in order

    summary_data = build_summary_data(summary_path, full_months, anchors)
    acc_holdings = build_acc_holdings(summary_path, anchors, fid_path, inc_income_df=fid_df)
    # Expense Reimbursements: Royal Mail or Expleo credits under £1,000 (excl. salary)
    REIMBURSEMENT_KEYWORDS = ["ROYAL MAIL", "EXPLEO"]

    # One-time backfill from expenses.csv (Jan-Jun 2026 full year feed) — these
    # months are no longer covered by the live data.csv (60-day window), so
    # hardcode them here. Going forward, data.csv provides ongoing updates.
    REIMBURSEMENT_BACKFILL = [
        {"date": pd.Timestamp("2026-02-04"), "amount": 36.80,
         "memo": "ROYAL MAIL            \t1612 2000284071 K BGC"},
        {"date": pd.Timestamp("2026-03-06"), "amount": 68.24,
         "memo": "EXPLEO UK LIMITED     \t1160304622 BGC\t"},
        {"date": pd.Timestamp("2026-03-20"), "amount": 163.03,
         "memo": "EXPLEO UK LIMITED     \t1161123617 BGC\t"},
        {"date": pd.Timestamp("2026-04-07"), "amount": 342.97,
         "memo": "EXPLEO UK LIMITED     \t1162170846 BGC\t"},
    ]

    reimbursements = []
    seen_keys = set()
    for entry in REIMBURSEMENT_BACKFILL:
        key = (entry["date"].date(), round(entry["amount"], 2))
        if key not in seen_keys:
            seen_keys.add(key)
            reimbursements.append({
                "date": entry["date"],
                "month": entry["date"].to_period("M"),
                "amount": entry["amount"],
                "memo": entry["memo"].strip()
            })

    # Live extraction from data.csv — covers ongoing/future updates
    try:
        bar_df_r = pd.read_csv("data.csv")
        bar_df_r["Amount"] = pd.to_numeric(bar_df_r["Amount"], errors="coerce").fillna(0)
        bar_df_r["Date"] = pd.to_datetime(bar_df_r["Date"], dayfirst=True, errors="coerce")
        bar_df_r = bar_df_r.dropna(subset=["Date"])
        for _, r in bar_df_r.iterrows():
            memo = str(r.get("Memo", "")).upper()
            amt = r["Amount"]
            if amt > 0 and amt < 1000 and any(k in memo for k in REIMBURSEMENT_KEYWORDS):
                if "EUKPT" in memo:
                    continue
                key = (r["Date"].date(), round(amt, 2))
                if key not in seen_keys:
                    seen_keys.add(key)
                    reimbursements.append({
                        "date": r["Date"],
                        "month": r["Date"].to_period("M"),
                        "amount": amt,
                        "memo": str(r.get("Memo", "")).strip()
                    })
    except Exception:
        pass

    write_excel(spend_pivot, actual_for_excel, future_for_excel, fid_pivot,
                acc_fund_map, holdings, summary_data, acc_holdings, anchors,
                output_path, reimbursements=reimbursements)
    print(f"Done → {output_path}\n")

    # Console preview of Fidelity section
    if not fid_pivot.empty:
        col_w, num_w = 32, 10
        hdr = f"{'Account':<{col_w}} {'Total':>{num_w}}" + "".join(
            f"  {m.strftime('%b %y'):>{num_w}}" for m in fid_months)
        print("Fidelity Income by Account")
        print("-" * len(hdr))
        print(hdr)
        print("-" * len(hdr))
        for acc in fid_pivot.index:
            label = ACCOUNT_LABELS.get(acc, acc)
            total = int(round(fid_pivot.loc[acc, "Total"]))
            monthly = [int(round(fid_pivot.loc[acc, m])) for m in fid_months]
            vals = f"{total:>{num_w},}" + "".join(f"  {v:>{num_w},}" for v in monthly)
            print(f"{label:<{col_w}} {vals}")
        gtotal = int(round(fid_pivot["Total"].sum()))
        gmonthly = [int(round(sum(fid_pivot[m]))) for m in fid_months]
        print("-" * len(hdr))
        gvals = f"{gtotal:>{num_w},}" + "".join(f"  {v:>{num_w},}" for v in gmonthly)
        print(f"{'TOTAL':<{col_w}} {gvals}")


if __name__ == "__main__":
    main()
