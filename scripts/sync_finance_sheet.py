"""Sync Stocks_Buy_Strategy.xlsx into the Finance Google Sheet over the Sheets API.

Replaces the manual browser route (delete every data tab, then File -> Import),
which cannot be automated — Chrome will not open a native file picker for a tab
the Claude extension drives in the background — and which deletes the tabs BEFORE
importing, so a failure leaves the live sheet EMPTY. That happened on 2026-07-20.

This writes each tab IN PLACE: no deletion, no picker, no window where the sheet
has no data, and it is re-runnable unattended.

    python scripts/sync_finance_sheet.py              # sync
    python scripts/sync_finance_sheet.py --dry-run    # report, write nothing
    python scripts/sync_finance_sheet.py --tabs Investments,Wealth Summary

Auth: the service account whose key sits at ~/.secrets/finance-sheets-sync.json
(see check_sheets_auth.py, and CLAUDE.md for the one-time setup). The Finance
sheet must be shared with that account as Editor.

── Three decisions worth understanding before changing anything ──────────────

1. VALUES + NUMBER FORMATS. Fills, borders, merges and column widths are NOT
   pushed — they survive untouched because clear() wipes values only, so the look
   of the sheet is whatever the last full xlsx import established. Import the xlsx
   by hand ONCE to set the appearance, then let this keep the numbers current.
   NUMBER formats are the exception and must be pushed: writing a formula with
   USER_ENTERED clears the cell's format (see sheets_number_format), which turned
   '-0.4%' into '-0.003685720404' on the first run. They are taken from the
   workbook, so the sync can also repair a sheet whose formats were already lost.
   The cost of the values-only model is that a STRUCTURAL change (inserting a
   column, moving a section) misaligns the surviving formatting and needs a fresh
   manual import to re-establish it. That is the trade-off; it is not a bug.

2. FORMULAS ARE PUSHED AS FORMULAS, evaluated by Google. The workbook is written
   by openpyxl, which does not compute — every formula cell reads as None locally
   (see the note at the top of dashboard_data.py). So pushing "values" would push
   thousands of blanks. Reading with data_only=False gives the formula text, and
   USER_ENTERED makes Google evaluate it. This is also what makes the
   GOOGLEFINANCE formulas in column J work at all — they only exist Google-side.

3. TWO TABS ARE NEVER TOUCHED:
   - ClaudeCode — the permanent run log, and Google requires >= 1 sheet to exist.
   - Stats — it contains CHARTS and no data (1x1 in the xlsx). The API cannot
     recreate charts from openpyxl, and clearing it would silently destroy the
     ones the last import created.
"""
import argparse
import datetime as dt
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import CFG, downloads_file  # noqa: E402
from ssl_certs import ensure_ca_bundle  # noqa: E402

ensure_ca_bundle()   # Norton re-signs TLS; see ssl_certs.py. Must precede any request.

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

DEFAULT_KEY = os.path.join(os.path.expanduser("~"), ".secrets",
                           "finance-sheets-sync.json")
KEY_PATH = os.path.expanduser(CFG.get("sheetsServiceAccountKey") or DEFAULT_KEY)
SHEET_ID = CFG["financeSheetId"]
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

LOG_TAB = "ClaudeCode"
# Never written by this script. See decision 3 in the module docstring.
PROTECTED = {LOG_TAB, "Stats"}


# ── reading the workbook ─────────────────────────────────────────────────────
def cell_value(v):
    """One xlsx cell -> something the Sheets API will accept.

    Dates go across as ISO text: USER_ENTERED makes Google parse that back into a
    real date, whereas a datetime object is not JSON-serialisable. Everything
    unrecognised is str()'d rather than dropped — a value the user can see and
    question beats a silent blank in a financial sheet.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S") if (v.hour or v.minute or v.second) \
            else v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.time):
        return v.strftime("%H:%M:%S")
    return str(v)


def read_workbook(path, wanted=None):
    """[(title, rows, parse_cells)] from the xlsx.

    `parse_cells` is {(row, col): value} for the cells that must be PARSED by
    Google rather than stored verbatim — formulas and real date/time cells.

    It is built from the openpyxl CELL TYPE, never by sniffing the rendered
    string, and that distinction is load-bearing. A first version tested whether
    the text looked like an ISO date, which promptly converted Investments'
    'Chart Last Checked' column — 347 cells that are plain TEXT in the workbook —
    into real dates, i.e. exactly the silent coercion the two-pass design exists
    to prevent. Only the source type knows the difference.

    data_only=False so formulas come through as formula text (openpyxl never
    computes, so the cached values are None locally).
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    out = []
    for ws in wb.worksheets:
        if ws.title in PROTECTED:
            continue
        if wanted and ws.title not in wanted:
            continue
        rows, parse_cells = [], {}
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                              min_col=1, max_col=ws.max_column), start=1):
            vals = []
            for ci, c in enumerate(row, start=1):
                v = c.value
                rendered = cell_value(v)
                if isinstance(v, str) and v.startswith("="):
                    parse_cells[(ri, ci)] = (rendered, c.number_format)
                elif isinstance(v, (dt.datetime, dt.date, dt.time)):
                    parse_cells[(ri, ci)] = (rendered, c.number_format)
                vals.append(rendered)
            rows.append(vals)
        # Trim wholly-empty trailing rows — openpyxl's max_row over-reports on
        # sheets that have had rows deleted, and writing them just pads the tab.
        while rows and all(c == "" for c in rows[-1]):
            parse_cells = {k: v for k, v in parse_cells.items() if k[0] != len(rows)}
            rows.pop()
        out.append((ws.title, rows, parse_cells))
    return out


# ── the sync ─────────────────────────────────────────────────────────────────
def open_sheet():
    import gspread
    from google.oauth2.service_account import Credentials
    if not os.path.exists(KEY_PATH):
        raise SystemExit(f"No service-account key at {KEY_PATH}.\n"
                         "Run scripts/check_sheets_auth.py for the setup check.")
    with open(KEY_PATH, encoding="utf-8") as f:
        info = json.load(f)
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return gspread.authorize(creds).open_by_key(SHEET_ID), info.get("client_email")


def _a1_col(n):
    """1 -> A, 27 -> AA."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _parsed_ranges(title, cells):
    """Group {(row,col): value} into contiguous per-column A1 ranges.

    The parsed pass covers ~6,400 cells, nearly all of them whole columns of
    formulas — grouping turns that into a handful of ranges instead of 6,400
    single-cell writes.
    """
    by_col = {}
    for (r, c), v in cells.items():
        by_col.setdefault(c, []).append((r, v))
    out = []
    for c, items in sorted(by_col.items()):
        items.sort()
        run_start, run_vals, prev = None, [], None
        for r, v in items:
            if prev is not None and r == prev + 1:
                run_vals.append([v])
            else:
                if run_start is not None:
                    out.append((f"'{title}'!{_a1_col(c)}{run_start}:"
                                f"{_a1_col(c)}{run_start + len(run_vals) - 1}", run_vals))
                run_start, run_vals = r, [[v]]
            prev = r
        if run_start is not None:
            out.append((f"'{title}'!{_a1_col(c)}{run_start}:"
                        f"{_a1_col(c)}{run_start + len(run_vals) - 1}", run_vals))
    return out


def sheets_number_format(excel_fmt):
    """An openpyxl number_format -> a Sheets numberFormat, or None to leave alone.

    Writing a formula with USER_ENTERED CLEARS the cell's number format. Measured,
    not assumed: after the first full sync every formula cell came back
    `numberFormat: None`, so Investments' '-0.4%' rendered as '-0.003685720404'
    while literal cells (written RAW) kept their '#,##0.00'.

    The formats are taken from the WORKBOOK rather than read back off the sheet,
    which makes the sync self-sufficient — it no longer depends on a prior manual
    import having established them, and it can repair a sheet whose formats have
    already been lost. Excel and Sheets share pattern syntax for everything this
    workbook uses (#,##0.00, 0.0%, £#,##0, dd/mm/yyyy); `type` is only a hint, so
    a rough classification is enough.
    """
    if not excel_fmt or excel_fmt == "General":
        return None
    f = excel_fmt
    # Strip Excel's colour/section decorations, which Sheets does not accept.
    for tok in ("[Red]", "[Black]", "[Blue]", "[Green]", "[$-409]", "[$-809]"):
        f = f.replace(tok, "")
    f = f.split(";")[0].strip()          # positive section only
    if not f:
        return None
    low = f.lower()
    if "%" in f:
        kind = "PERCENT"
    elif any(sym in f for sym in ("£", "$", "€")):
        kind = "CURRENCY"
    elif any(ch in low for ch in ("y", "d")) and "m" in low:
        kind = "DATE"
    elif "h" in low and ":" in f:
        kind = "TIME"
    else:
        kind = "NUMBER"
    return {"type": kind, "pattern": f}


def _restore_formats(sh, ws, wanted):
    """Apply {(row,col): numberFormat}, grouped into contiguous column runs."""
    if not wanted:
        return 0
    by_col = {}
    for (r, c), nf in wanted.items():
        by_col.setdefault(c, []).append((r, nf))

    requests = []
    for c, items in sorted(by_col.items()):
        items.sort()
        run_start = run_prev = None
        run_fmt = None
        def flush():
            if run_start is None:
                return
            requests.append({"repeatCell": {
                "range": {"sheetId": ws.id,
                          "startRowIndex": run_start - 1, "endRowIndex": run_prev,
                          "startColumnIndex": c - 1, "endColumnIndex": c},
                "cell": {"userEnteredFormat": {"numberFormat": run_fmt}},
                "fields": "userEnteredFormat.numberFormat"}})
        for r, nf in items:
            if run_start is not None and r == run_prev + 1 and nf == run_fmt:
                run_prev = r
                continue
            flush()
            run_start, run_prev, run_fmt = r, r, nf
        flush()

    # Chunked: a tab can need a few hundred runs and the API caps request size.
    for i in range(0, len(requests), 200):
        sh.batch_update({"requests": requests[i:i + 200]})
    return len(wanted)


def sync_tab(sh, title, rows, parse_cells, dry_run=False):
    """Write one tab in place. Returns a short status string.

    TWO PASSES, because a single write can only carry one valueInputOption and
    neither option is right for the whole grid:

      RAW           — everything except formulas and dates. USER_ENTERED would
                      COERCE these: a free-text Chart Note reading '1-2' becomes a
                      date, '6.7%' becomes 0.067 re-rendered as '6.70%'. There are
                      ~6,100 such strings, and silent type changes in a live
                      financial sheet are exactly what this pipeline exists to
                      avoid. RAW stores them verbatim.
      USER_ENTERED  — formulas (~6,400) and dates (~115) ONLY, which MUST be
                      parsed: a formula written RAW lands as literal text, and an
                      ISO date string stays text instead of becoming a real date.
    """
    from gspread.utils import ValueInputOption
    from gspread.exceptions import WorksheetNotFound

    n_rows = len(rows)
    n_cols = max((len(r) for r in rows), default=0)
    if n_rows == 0 or n_cols == 0:
        return "skipped (no data in the workbook)"

    try:
        ws = sh.worksheet(title)
        existed = True
    except WorksheetNotFound:
        existed = False
        if dry_run:
            return f"would CREATE ({n_rows}x{n_cols})"
        ws = sh.add_worksheet(title=title, rows=max(n_rows, 100), cols=max(n_cols, 26))

    if dry_run:
        return f"would write {n_rows}x{n_cols}" + ("" if existed else " (new tab)")

    # Grow the tab if the workbook now needs more room. Deliberately never SHRINK:
    # narrowing a tab deletes columns, and with them the formatting and any manual
    # notes to the right of the data.
    if ws.row_count < n_rows or ws.col_count < n_cols:
        ws.resize(rows=max(ws.row_count, n_rows), cols=max(ws.col_count, n_cols))

    # clear() wipes VALUES only — fills, borders, merges and column widths
    # survive, which is the whole point (decision 1 in the docstring).
    ws.clear()

    # Pad ragged rows: the API rejects a jagged 2-D range.
    padded = [list(r) + [""] * (n_cols - len(r)) for r in rows]

    # Blank the parse-me cells in the raw grid; pass 2 fills them in.
    parsed = {rc: v for rc, v in parse_cells.items()
              if rc[0] <= n_rows and rc[1] <= n_cols}
    for (ri, ci) in parsed:
        padded[ri - 1][ci - 1] = ""

    ws.update(padded, "A1", value_input_option=ValueInputOption.raw)

    restored = 0
    if parsed:
        ranges = _parsed_ranges(title, {rc: v for rc, (v, _) in parsed.items()})
        sh.values_batch_update({
            "valueInputOption": "USER_ENTERED",
            "data": [{"range": rng, "values": vals} for rng, vals in ranges],
        })
        # USER_ENTERED wiped the number format on every cell it just wrote.
        want = {rc: nf for rc, (_, f) in parsed.items()
                if (nf := sheets_number_format(f))}
        restored = _restore_formats(sh, ws, want)

    extra = f", {len(parsed)} parsed" if parsed else ""
    extra += f", {restored} formats restored" if restored else ""
    return f"wrote {n_rows}x{n_cols}{extra}" + ("" if existed else " (new tab)")


def append_run_log(sh, line, dry_run=False):
    """One dated line at the BOTTOM of the ClaudeCode tab, as the manual sync did.

    Not `append_row`: with table_range='A1' it treats the first contiguous block
    as "the table" and INSERTS after it, which put five test entries at the top of
    the log and pushed the whole history down. The next free row is computed
    explicitly instead, and written with RAW so a line beginning with a date is
    never parsed into something else.
    """
    if dry_run:
        return
    try:
        ws = sh.worksheet(LOG_TAB)
        used = len(ws.get_all_values())          # rows with any content
        target = used + 2                        # blank spacer row, matching the log's style
        if target > ws.row_count:
            ws.resize(rows=target + 10)
        ws.update([[line]], f"A{target}", value_input_option="RAW")
    except Exception as e:                                   # noqa: BLE001
        # A failed log line must never fail the sync — the data is already written.
        print(f"  (could not append to {LOG_TAB}: {e})")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dry-run", action="store_true",
                    help="report what would be written; touch nothing")
    ap.add_argument("--tabs", default="",
                    help="comma-separated tab names to sync (default: all)")
    ap.add_argument("--workbook", default=None, help="override the source xlsx")
    args = ap.parse_args()

    path = args.workbook or downloads_file("masterWorkbook")
    wanted = {t.strip() for t in args.tabs.split(",") if t.strip()} or None

    print(f"Source : {path}")
    print(f"Sheet  : {SHEET_ID}")
    if args.dry_run:
        print("MODE   : dry run — nothing will be written\n")

    tabs = read_workbook(path, wanted)
    if not tabs:
        raise SystemExit("Nothing to sync (check --tabs against the workbook's tab names).")

    sh, account = open_sheet()
    print(f"Account: {account}")
    print(f"Opened : {sh.title}\n")

    results = []
    for title, rows, parse_cells in tabs:
        status = sync_tab(sh, title, rows, parse_cells, dry_run=args.dry_run)
        print(f"  {title:26s} {status}")
        results.append((title, status))

    skipped = sorted(PROTECTED)
    print(f"\n  {'(protected)':26s} left untouched: {', '.join(skipped)}")

    written = [t for t, s in results if s.startswith("wrote")]
    if written and not args.dry_run:
        stamp = dt.date.today().isoformat()
        append_run_log(sh, (
            f"{stamp} API sync: wrote {len(written)} tab(s) in place via the Sheets API "
            f"({', '.join(written)}) — no tab deletion, no file picker. Values only; "
            f"formatting is whatever the last manual xlsx import established. "
            f"{LOG_TAB} and Stats untouched (Stats holds charts the API cannot rebuild)."
        ))

    print(f"\n{'Would sync' if args.dry_run else 'Synced'} {len(results)} tab(s).")


if __name__ == "__main__":
    main()
