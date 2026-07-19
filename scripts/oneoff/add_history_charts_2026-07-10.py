#!/usr/bin/env python3
"""One-off (2026-07-10): rebuild the History sheet's monthly summary tables and charts.

Fixes a bug from the earlier RELX correction: inserting a row at 56 shifted the
Capital Deployed / Realised P&L data down by one row, but the two chart objects
already embedded in the sheet still referenced their old (now stale) cell ranges.

Also implements the user's request:
  - Realised P&L by Month rebuilt to cover all 10 months (Oct 2025 - Jul 2026),
    matching Capital Deployed's month list, instead of only the 4 months that
    happened to have a closed trade (the other 6 evaluate to 0 via the same
    SUMPRODUCT pattern, not left blank).
  - New Unrealised P&L by Month table: sum of (Value if Sold Today - Cost) for
    still-OPEN positions, grouped by Buy Date's month. Value if Sold Today (col
    K) only calculates in real Google Sheets (GOOGLEFINANCE-backed), same as
    the rest of this sheet, so this is a live formula, not a static snapshot.
  - Charts for both, matching the existing Capital Deployed chart's style
    (value labels shown above each bar) via deepcopy of the existing Realised
    P&L chart object rather than hand-building chart XML from scratch.

Not part of the periodic pipeline — run once, by hand, against the live file.
"""
import openpyxl
from copy import deepcopy
from openpyxl.chart import Reference

MASTER_PATH = r'C:\Users\Paul\Downloads\Stocks_Buy_Strategy.xlsx'
SHEET = 'History'

MONTHS = [
    ('Oct 2025', '2025-10'), ('Nov 2025', '2025-11'), ('Dec 2025', '2025-12'),
    ('Jan 2026', '2026-01'), ('Feb 2026', '2026-02'), ('Mar 2026', '2026-03'),
    ('Apr 2026', '2026-04'), ('May 2026', '2026-05'), ('Jun 2026', '2026-06'),
    ('Jul 2026', '2026-07'),
]
LAST_DATA_ROW = 56  # trade rows are D2:D56 etc, per the RELX fix


def main():
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    ws = wb[SHEET]

    # --- 1. Fix the Capital Deployed chart's stale range (62:71 -> 63:72) ---
    cap_chart = ws._charts[0]
    assert 'Capital Deployed' in str(cap_chart.title)
    s = cap_chart.series[0]
    s.val.numRef.f = 'History!$B$63:$B$72'
    if s.cat.strRef is not None:
        s.cat.strRef.f = 'History!$A$63:$A$72'
    else:
        s.cat.numRef.f = 'History!$A$63:$A$72'

    # --- 2. Rebuild Realised P&L by Month with all 10 months (rows 76-85) ---
    ws.cell(row=74, column=1, value='Realised P&L by Month (Sell Date)')
    ws.cell(row=75, column=1, value='Month')
    ws.cell(row=75, column=2, value='Realised P&L (£)')
    for i, (label, ym) in enumerate(MONTHS):
        r = 76 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=(
            f'=SUMPRODUCT((P2:P{LAST_DATA_ROW}="Closed")*'
            f'(TEXT(E2:E{LAST_DATA_ROW},"YYYY-MM")="{ym}")*M2:M{LAST_DATA_ROW})'
        ))

    # --- 3. New Unrealised P&L by Month table (rows 89-98) ---
    ws.cell(row=87, column=1, value='Unrealised P&L by Month (Buy Date, still-Open positions)')
    ws.cell(row=88, column=1, value='Month')
    ws.cell(row=88, column=2, value='Unrealised P&L (£)')
    for i, (label, ym) in enumerate(MONTHS):
        r = 89 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=(
            f'=SUMPRODUCT((P2:P{LAST_DATA_ROW}="Open")*'
            f'(TEXT(D2:D{LAST_DATA_ROW},"YYYY-MM")="{ym}")*'
            f'IFERROR(K2:K{LAST_DATA_ROW}-J2:J{LAST_DATA_ROW},0))'
        ))

    # --- 4. Fix the Realised P&L chart's stale/sparse range (75:78 -> 76:85) ---
    realised_chart = ws._charts[1]
    assert 'Realised P&L' in str(realised_chart.title)
    s = realised_chart.series[0]
    s.val.numRef.f = 'History!$B$76:$B$85'
    if s.cat.strRef is not None:
        s.cat.strRef.f = 'History!$A$76:$A$85'
    else:
        s.cat.numRef.f = 'History!$A$76:$A$85'

    # --- 5. New Unrealised P&L chart, cloned from the Realised P&L chart so it
    #    matches its style exactly (value labels shown above each bar, same
    #    green/red P&L number format, same fonts) rather than rebuilding from
    #    scratch and risking a style mismatch.
    unrealised_chart = deepcopy(realised_chart)
    unrealised_chart.title.tx.rich.p[0].r[0].t = 'Unrealised P&L by Month'
    s2 = unrealised_chart.series[0]
    s2.val.numRef.f = 'History!$B$89:$B$98'
    if s2.cat.strRef is not None:
        s2.cat.strRef.f = 'History!$A$89:$A$98'
    else:
        s2.cat.numRef.f = 'History!$A$89:$A$98'
    # Re-anchor below the Realised P&L chart (which spans rows 73-89) so the
    # two don't overlap.
    unrealised_chart.anchor._from.row = 90
    unrealised_chart.anchor.to.row = 106
    ws.add_chart(unrealised_chart)

    wb.save(MASTER_PATH)
    print('Saved.')


if __name__ == '__main__':
    main()
