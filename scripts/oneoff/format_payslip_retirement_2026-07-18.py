"""One-off (2026-07-18): tidy the Payslip Summary + Retirement Income Plan tabs.

User requests:
  - Payslip Summary: reduce the font to 9 (was a mix of Arial 10/12 + Cambria)
    and make the text display across all columns (fit widths).
  - Retirement Income Plan: widen column A so all its text shows (long lines like
    "Paul born 29 Mar ..." were clipped at width 22).
  - Arial 9 as the base text size for the Google-Sheet tabs generally.

These two tabs are hand-maintained and PRESERVED across pipeline runs from
spending_summary.xlsx (preserve_manual_sheets), then copied into the master
Stocks_Buy_Strategy.xlsx by integrate_spending_tabs.py. So the source of truth is
spending_summary.xlsx — but we format BOTH files here so the change is immediate in
the master AND persists on the next run (the preserved copy is already formatted).
Backs up each workbook first; preserves bold/italic/colour/fill.
"""
import datetime
import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_FONT = "Arial"
BASE_SIZE = 9
WIDTH_CAP = 60
WIDTH_FLOOR = 8


def fit_widths(ws, cols=None, cap=WIDTH_CAP):
    """Set column widths to fit content (header + non-formula data)."""
    cols = cols or range(1, ws.max_column + 1)
    for c in cols:
        L = get_column_letter(c)
        longest = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v)
            if s.startswith("="):
                continue
            longest = max(longest, *(len(line) for line in s.split("\n")))
        ws.column_dimensions[L].width = min(max(longest + 2, WIDTH_FLOOR), cap)


def set_base_font(ws, size=BASE_SIZE):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            f = cell.font
            cell.font = Font(name=BASE_FONT, size=size, bold=f.bold, italic=f.italic,
                             underline=f.underline, strike=f.strike, color=f.color,
                             vertAlign=f.vertAlign)


def process(path):
    if not os.path.exists(path):
        print("  (skip, not found):", path)
        return
    bak = path + ".bak-payretfmt-" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copyfile(path, bak)
    wb = openpyxl.load_workbook(path)
    touched = []
    if "Payslip Summary" in wb.sheetnames:
        ws = wb["Payslip Summary"]
        set_base_font(ws, BASE_SIZE)   # everything to Arial 9
        fit_widths(ws)                 # display across all columns
        touched.append("Payslip Summary")
    if "Retirement Income Plan" in wb.sheetnames:
        ws = wb["Retirement Income Plan"]
        set_base_font(ws, BASE_SIZE)   # Arial 9 base
        # User asked specifically to widen column 1 to show all its text; also fit
        # the rest so nothing else collides after the font change.
        fit_widths(ws)
        touched.append("Retirement Income Plan")
    if touched:
        wb.save(path)
        print(f"  {os.path.basename(path)}: formatted {touched} (backup {os.path.basename(bak)})")
    else:
        os.remove(bak)
        print(f"  {os.path.basename(path)}: neither tab present, nothing done")


def main():
    home = os.path.expanduser("~")
    for path in (os.path.join(home, "Downloads", "spending_summary.xlsx"),
                 os.path.join(home, "Downloads", "Stocks_Buy_Strategy.xlsx")):
        process(path)


if __name__ == "__main__":
    main()
