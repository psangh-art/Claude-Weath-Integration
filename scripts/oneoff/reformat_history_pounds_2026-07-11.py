#!/usr/bin/env python3
"""One-off reformat of Stocks_Buy_Strategy.xlsx (2026-07-11, user-requested):

1. History sheet: colour-group its header row using the same semantic palette
   Investments uses (identity=navy FF1F3864, per-share prices=amber FFB9770E,
   money/£ and P&L=purple FF6C3483) — everything else about History already
   matched Investments (navy header, FF2E5077 month bands), verified before
   writing this.
2. Whole-pound display for aggregate £ amounts across ALL sheets ("round up the
   values removing numbers after the decimal point" — implemented as number
   FORMAT '#,##0', display-only and reversible, never touching cell values):
   - Investments  D,E,F,H  (Holdings/Gain-Loss/Target/Sell Result £)
   - History      I,J,K,L  (Fees/Cost/Value/Proceeds £) and M (P&L £, keeping
                  its green/red positive-negative colouring)
   - Income Funds I,J      (Monthly/Annual Revenue — B,C,D were already whole £)
   - Stocks of Interest G,H rows 5-38 (below-alert block Holdings/Target £)
                  and M,N rows 41+ (old-content Holdings/Target £)
   Deliberately NOT touched: per-share prices (pence), % columns, and
   £bn columns (Market Cap/Net Debt — rounding billions to whole £bn would
   destroy the value).

Usage: python reformat_history_pounds_2026-07-11.py <in.xlsx> <out.xlsx>
"""
import sys
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

NAVY = 'FF1F3864'
AMBER = 'FFB9770E'
PURPLE = 'FF6C3483'

HISTORY_HEADER_GROUPS = {
    'G': AMBER, 'H': AMBER,                                  # Buy/Sell Price
    'I': PURPLE, 'J': PURPLE, 'K': PURPLE, 'L': PURPLE,      # £ money
    'M': PURPLE, 'N': PURPLE,                                # P&L £ / P&L %
}

WHOLE_POUNDS = '#,##0'
WHOLE_POUNDS_PNL = '[GREEN]#,##0;[RED]\\(#,##0\\);0'


def set_col_format(ws, col_letter, fmt, first_row, last_row):
    changed = 0
    for r in range(first_row, last_row + 1):
        cell = ws[f'{col_letter}{r}']
        if cell.number_format != fmt:
            cell.number_format = fmt
            changed += 1
    return changed


def main():
    src_path, out_path = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(src_path)

    # ── 1. History header colour groups ─────────────────────────────────────
    hist = wb['History']
    for col, rgb in HISTORY_HEADER_GROUPS.items():
        cell = hist[f'{col}1']
        cell.fill = PatternFill(fill_type='solid', fgColor=rgb)
    print(f'History header: recoloured {len(HISTORY_HEADER_GROUPS)} columns into '
          'Investments-style groups')

    # ── 2. Whole-pound formats ───────────────────────────────────────────────
    total = 0
    inv = wb['Investments']
    for col in ('D', 'E', 'F', 'H'):
        total += set_col_format(inv, col, WHOLE_POUNDS, 3, inv.max_row)

    for col in ('I', 'J', 'K', 'L'):
        total += set_col_format(hist, col, WHOLE_POUNDS, 2, hist.max_row)
    total += set_col_format(hist, 'M', WHOLE_POUNDS_PNL, 2, hist.max_row)

    inc = wb['Income Funds']
    for col in ('I', 'J'):
        total += set_col_format(inc, col, '\\£' + WHOLE_POUNDS, 5, inc.max_row)

    soi = wb['Stocks of Interest']
    for col in ('G', 'H'):
        total += set_col_format(soi, col, WHOLE_POUNDS, 5, 38)
    for col in ('M', 'N'):
        total += set_col_format(soi, col, WHOLE_POUNDS, 41, soi.max_row)

    print(f'Whole-pound number format applied to {total} cells')
    wb.save(out_path)
    print(f'Saved -> {out_path}')


if __name__ == '__main__':
    main()
