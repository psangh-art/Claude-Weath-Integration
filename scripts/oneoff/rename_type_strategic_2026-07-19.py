"""One-off (2026-07-19): Investments 'Type' — rename 'Long Term' to 'Strategic'.

User request. Updates the cell values AND any data-validation dropdown that offers
the old wording, so a later edit can't put 'Long Term' back.
"""
import os
import shutil
import datetime
import openpyxl

WB = os.path.join(os.path.expanduser('~'), 'Downloads', 'Stocks_Buy_Strategy.xlsx')
SHEET = 'Investments'
I_TYPE = 39
OLD, NEW = 'long term', 'Strategic'

wb = openpyxl.load_workbook(WB)
ws = wb[SHEET]

changed = []
for row in range(4, ws.max_row + 1):
    c = ws.cell(row=row, column=I_TYPE)
    if isinstance(c.value, str) and c.value.strip().lower() == OLD:
        ticker = ws.cell(row=row, column=4).value
        c.value = NEW
        changed.append((row, ticker))

dvs = []
for dv in ws.data_validations.dataValidation:
    if dv.formula1 and 'long term' in str(dv.formula1).lower():
        before = dv.formula1
        dv.formula1 = str(dv.formula1).replace('Long Term', NEW).replace('long term', NEW)
        dvs.append((before, dv.formula1, str(dv.sqref)))

if changed or dvs:
    backup = WB + '.bak-' + datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    shutil.copyfile(WB, backup)
    wb.save(WB)
    print('backup:', backup)
wb.close()

print(f'{len(changed)} cell(s) renamed to {NEW}:')
for row, ticker in changed:
    print(f'  row {row}  {ticker}')
for before, after, sqref in dvs:
    print(f'  dropdown {sqref}: {before} -> {after}')
if not changed and not dvs:
    print('nothing to change (already renamed)')
