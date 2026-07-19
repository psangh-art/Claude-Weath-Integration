#!/usr/bin/env python3
"""One-off structural migration: insert a 'Marked Up' column at Investments!B.

User request 2026-07-16: after column A ('Chart' = does a chart exist) add a column
confirming whether the user has MARKED UP the chart (drawn channel/trend lines) —
'Yes'/'No', auto-derived from channel_detect. This shifts every Investments data
column one to the right, so ~5,600 formulas in Investments plus the cross-sheet
VLOOKUPs in History and 'Stocks of Interest' that point at Investments must have
their column references bumped. No spreadsheet engine (LibreOffice / Excel COM) is
available on this machine to do that automatically, so the shift is done here with
openpyxl's formula Tokenizer: it touches ONLY real cell references, and only those
that resolve to the Investments sheet — string literals (googlefinance("INDEXFTSE:UKX"),
HYPERLINK urls), 'Base Data'! refs, and each sheet's own refs are left untouched.

Structure handled explicitly (openpyxl moves neither of these on a column insert):
merged title A1:W1 -> A1:X1, conditional-format range, 20 column widths, freeze D3.

The pipeline column constants were bumped to match in the same change
(update_master_sheet.py, refresh_soi_sections.py, build_review_deck.py,
verify_pipeline.py). update_master_sheet.marked_up_flag() is the single source of the
Yes/No rule and maintains the column every run; this script populates it once now.

Usage:  python insert_marked_up_col_2026-07-16.py <workbook.xlsx> [--apply]
  Without --apply it writes to <workbook>.migrated.xlsx so the result can be checked
  before overwriting the real file.
"""
import os
import re
import sys
import json
from copy import copy as _copy

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SCRIPT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')  # scripts/
sys.path.insert(0, SCRIPT_DIR)
from ticker_normalize import normalize
from update_master_sheet import marked_up_flag

INSERT_AT = 2  # new column B
INV = 'Investments'

# ---- formula reference shifter (validated against the real formulas) ----------
_REF = re.compile(r'(\$?)([A-Za-z]{1,3})(\$?)(\d*)')
_PURE = re.compile(r'^\$?[A-Za-z]{1,3}\$?\d*(:\$?[A-Za-z]{1,3}\$?\d*)?$')


def _shift_ref_part(ref):
    def repl(m):
        d1, letters, d2, digits = m.groups()
        idx = column_index_from_string(letters.upper())
        if idx >= INSERT_AT:
            letters = get_column_letter(idx + 1)
        return f'{d1}{letters}{d2}{digits}'
    return _REF.sub(repl, ref)


def _maybe_shift(val, on_inv):
    """Shift a single RANGE operand iff it resolves to the Investments sheet."""
    if '!' in val:
        sheet, ref = val.rsplit('!', 1)
        if sheet.strip().strip("'").strip() != INV:
            return val
        return sheet + '!' + _shift_ref_part(ref)
    # unqualified ref -> the sheet the formula lives on
    if on_inv and _PURE.match(val):
        return _shift_ref_part(val)
    return val


def shift_formula(formula, on_inv):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    out = []
    for t in Tokenizer(formula).items:
        v = t.value
        if t.type == 'OPERAND' and t.subtype == 'RANGE':
            v = _maybe_shift(v, on_inv)
        out.append(v)
    return '=' + ''.join(out)


def _copy_style(dst, src):
    dst.font = _copy(src.font)
    dst.fill = _copy(src.fill)
    dst.border = _copy(src.border)
    dst.alignment = _copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = _copy(src.protection)


def load_patterns():
    """master ticker -> channel_detect record, same mapping the pipeline uses."""
    results = json.load(open(os.path.join(SCRIPT_DIR, 'channel_results_tmp.json'), encoding='utf-8'))
    patterns = {}
    for rec in results:
        t = (rec.get('ticker') or '').strip().upper()
        if not t:
            continue
        patterns[t] = rec
        mt = (normalize(t) or {}).get('master_ticker')
        if mt:
            patterns.setdefault(mt.upper(), rec)
    return patterns


def migrate(path, out_path):
    wb = openpyxl.load_workbook(path, data_only=False)
    inv = wb[INV]
    maxr, maxc = inv.max_row, inv.max_column

    # Capture letter-keyed attributes BEFORE the shift (openpyxl won't move these).
    old_widths = {k: v.width for k, v in inv.column_dimensions.items() if v.width}
    merges = [str(m) for m in inv.merged_cells.ranges]
    cf_ranges = [(str(rng.sqref), list(inv.conditional_formatting[rng]))
                 for rng in inv.conditional_formatting]

    # 1. Shift Investments cells right by one, right-to-left so nothing is clobbered
    #    before it is copied. Formulas are reference-shifted during the copy.
    for m in merges:
        inv.unmerge_cells(m)
    for row in range(1, maxr + 1):
        for col in range(maxc, INSERT_AT - 1, -1):
            s = inv.cell(row=row, column=col)
            d = inv.cell(row=row, column=col + 1)
            v = s.value
            if isinstance(v, str) and v.startswith('='):
                v = shift_formula(v, on_inv=True)
            d.value = v
            _copy_style(d, s)
        # 2. Reset the freshly-opened column B for this row (still holds old B content).
        b = inv.cell(row=row, column=INSERT_AT)
        a = inv.cell(row=row, column=1)
        b.value = None
        _copy_style(b, a)           # inherit the row's border/height styling from col A
        b.fill = PatternFill()      # ...but never the green Chart=Yes fill
        b.font = Font(name='Arial', size=10)
        b.alignment = Alignment(horizontal='center', vertical='center')

    # 3. Header + column width for the new column.
    hdr = inv.cell(row=2, column=INSERT_AT, value='Marked Up')
    _copy_style(hdr, inv.cell(row=2, column=1))   # match the 'Chart' header band
    hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 4. Re-key column widths (>= B shift +1); give B its own width.
    new_widths = {}
    for letter, w in old_widths.items():
        idx = column_index_from_string(letter)
        new_widths[get_column_letter(idx + 1 if idx >= INSERT_AT else idx)] = w
    new_widths['B'] = 9
    for k in list(inv.column_dimensions.keys()):
        del inv.column_dimensions[k]
    for letter, w in new_widths.items():
        inv.column_dimensions[letter].width = w

    # 5. Merged title A1:W1 -> A1:X1 (start col A unchanged, end col W -> X).
    for m in merges:
        a, b = m.split(':')
        def bump(ref):
            mm = re.match(r'([A-Z]+)(\d+)', ref)
            col, rownum = mm.group(1), mm.group(2)
            i = column_index_from_string(col)
            return f'{get_column_letter(i + 1 if i >= INSERT_AT else i)}{rownum}'
        inv.merge_cells(f'{bump(a)}:{bump(b)}')

    # 6. Freeze panes D3 -> E3.
    if inv.freeze_panes:
        mm = re.match(r'([A-Z]+)(\d+)', inv.freeze_panes)
        i = column_index_from_string(mm.group(1))
        inv.freeze_panes = f'{get_column_letter(i + 1 if i >= INSERT_AT else i)}{mm.group(2)}'

    # 7. Conditional formatting: rebuild the list under each range's shifted columns.
    from openpyxl.formatting.formatting import ConditionalFormattingList
    def shift_sqref(sqref):
        return ' '.join(':'.join(_shift_ref_part(p) for p in rng.split(':'))
                        for rng in str(sqref).split())
    new_cf = ConditionalFormattingList()
    for old_sqref, rules in cf_ranges:
        new = shift_sqref(old_sqref)
        for rule in rules:
            new_cf.add(new, rule)
    inv.conditional_formatting = new_cf

    # 8. Fix cross-sheet formulas in OTHER tabs that point at Investments.
    for sh in wb.worksheets:
        if sh.title == INV:
            continue
        for r in sh.iter_rows():
            for c in r:
                if isinstance(c.value, str) and c.value.startswith('=') and INV in c.value:
                    c.value = shift_formula(c.value, on_inv=False)

    # 9. Populate 'Marked Up' from channel_detect (ticker now in col D).
    patterns = load_patterns()
    filled = 0
    for row in range(3, maxr + 1):
        t = inv.cell(row=row, column=4).value          # D = Ticker (post-shift)
        if not isinstance(t, str) or not t.strip():
            continue
        key = t.strip().upper()
        rec = patterns.get(key) or patterns.get((normalize(key) or {}).get('master_ticker', '').upper() if normalize(key) else key)
        inv.cell(row=row, column=INSERT_AT, value=marked_up_flag(rec))
        filled += 1

    wb.save(out_path)
    print(f'migrated -> {out_path}  ({filled} tickers flagged, {maxc}->{maxc+1} cols)')


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    path = args[0] if args else os.path.expanduser('~/Downloads/Stocks_Buy_Strategy.xlsx')
    apply = '--apply' in sys.argv
    out = path if apply else path.replace('.xlsx', '.migrated.xlsx')
    migrate(path, out)


if __name__ == '__main__':
    main()
