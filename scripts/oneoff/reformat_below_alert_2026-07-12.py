#!/usr/bin/env python3
"""One-off (2026-07-12, user-requested): reformat the 'Stocks Trading Below Alert
Low' table at the top of 'Stocks of Interest' so it matches the section tables
lower down the same sheet — Arial, thin borders, navy title band, FF2E5077
subtitle/header bands, a red section band, and pale-red data rows, with the
identity columns reordered to Stock-name-then-Ticker like the tables below.

It reads the 24 rows the last pipeline run wrote (old column order: Ticker,
Share Name, Current Price, Alert Low, Gap %, Alert High, Holdings, Target,
Checked At), coerces any Holdings/Target cells that came through as formulas to
numeric-or-blank, and rewrites the block through the NEW add_below_alert_sheet
.refresh_block() so the live file matches what future runs will produce.

Usage: python reformat_below_alert_2026-07-12.py <workbook.xlsx>
"""
import sys
import shutil

import openpyxl

from add_below_alert_sheet import refresh_block, SHEET_NAME, DATA_START_ROW, MAX_DATA_ROWS

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass


def num_or_none(v):
    return v if isinstance(v, (int, float)) else None


def main():
    path = sys.argv[1]
    backup = path + '.bak-before-below-alert-reformat-2026-07-12'
    shutil.copyfile(path, backup)
    print(f'Backup -> {backup}')

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[SHEET_NAME]

    # Read the existing block (OLD column order) into row dicts.
    rows = []
    for r in range(DATA_START_ROW, DATA_START_ROW + MAX_DATA_ROWS):
        ticker = ws.cell(row=r, column=1).value
        if ticker is None:
            continue
        rows.append({
            'ticker': ticker,
            'share_name': ws.cell(row=r, column=2).value or ticker,
            'price': ws.cell(row=r, column=3).value or 0,
            'alert_low': ws.cell(row=r, column=4).value,
            'gap_pct': ws.cell(row=r, column=5).value or 0,
            'alert_high': num_or_none(ws.cell(row=r, column=6).value),
            'holdings': num_or_none(ws.cell(row=r, column=7).value),
            'target_value': num_or_none(ws.cell(row=r, column=8).value),
            'checked_at': ws.cell(row=r, column=9).value,
        })
    print(f'Read {len(rows)} existing below-alert rows')

    count = refresh_block(ws, rows)
    wb.save(path)
    print(f'Reformatted {count} rows into the styled top block of "{SHEET_NAME}" -> {path}')


if __name__ == '__main__':
    main()
