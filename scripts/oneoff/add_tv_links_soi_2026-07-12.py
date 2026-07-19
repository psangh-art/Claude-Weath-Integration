#!/usr/bin/env python3
"""One-off (2026-07-12, user-requested): add a TradingView layout click-through
to every row in 'Stocks of Interest', matching the Investments sheet's
"TradingView" column (=HYPERLINK("…/chart/<chartId>/","📊 Layout")).

Two tables live on this sheet:
  * the pipeline-rebuilt below-alert block (rows 1-40) — links go in a new
    'TradingView' column J; refresh_block() now reproduces them every run, and
    this script re-applies the block so the live file matches immediately.
  * the hand-maintained section tables (rows 41+) — links go in a new column P;
    nothing regenerates these rows, so this script is their one-off source.

ticker->chartId comes from the layout manifest, matched through
ticker_normalize so master-form tickers (SLVR, BT-A) line up with TradingView's
export names (SILVER, BT.A).

Usage: python add_tv_links_soi_2026-07-12.py <workbook.xlsx> [layout_manifest.json]
"""
import json
import os
import re
import sys
import shutil
from copy import copy

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ticker_normalize import normalize, master_tickers_match
from add_below_alert_sheet import (refresh_block, SHEET_NAME, DATA_START_ROW,
                                    MAX_DATA_ROWS, TV_LAYOUT_URL, TV_LINK_LABEL)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# The section tables already use col P ('Last Updated'); the link goes in the
# next free column, Q. Only rows whose col B is a real ticker get a link —
# this sheet also holds reference tables (FTSE review calendar, dividend-cover
# guide) whose col B is descriptive prose, which must be left untouched.
SECTION_TV_COL = 17          # column Q
_TICKER_RE = re.compile(r'^[A-Z][A-Z0-9]{0,5}([.\-][A-Z])?$')


def looks_like_ticker(v):
    return isinstance(v, str) and v.strip() != 'Ticker' and bool(_TICKER_RE.match(v.strip()))


def build_chart_map(manifest_path):
    """master-ticker (upper, dashes) -> chartId, first chart wins."""
    with open(manifest_path, encoding='utf-8') as f:
        charts = json.load(f)
    m = {}
    for row in charts:
        norm = normalize(row.get('ticker'))
        cid = row.get('chartId')
        if not norm or not norm.get('master_ticker') or not cid:
            continue
        key = norm['master_ticker'].upper().replace('.', '-')
        m.setdefault(key, cid)
    return m


def chart_id_for(chart_map, ticker):
    if not ticker:
        return None
    key = str(ticker).upper().replace('.', '-')
    if key in chart_map:
        return chart_map[key]
    for k, cid in chart_map.items():
        if master_tickers_match(k, ticker):
            return cid
    return None


def link_formula(chart_id):
    return f'=HYPERLINK("{TV_LAYOUT_URL.format(chart_id=chart_id)}","{TV_LINK_LABEL}")'


def apply_below_alert(ws, chart_map):
    """Re-read the reformatted block (A=Stock, B=Ticker, …) and rebuild it through
    refresh_block so column J gets the TradingView links."""
    rows = []
    for r in range(DATA_START_ROW, DATA_START_ROW + MAX_DATA_ROWS):
        ticker = ws.cell(row=r, column=2).value
        if not ticker:
            continue
        rows.append({
            'share_name': ws.cell(row=r, column=1).value or ticker,
            'ticker': ticker,
            'price': ws.cell(row=r, column=3).value or 0,
            'alert_low': ws.cell(row=r, column=4).value,
            'gap_pct': ws.cell(row=r, column=5).value or 0,
            'alert_high': ws.cell(row=r, column=6).value,
            'holdings': ws.cell(row=r, column=7).value,
            'target_value': ws.cell(row=r, column=8).value,
            'checked_at': ws.cell(row=r, column=9).value,
            'chart_id': chart_id_for(chart_map, ticker),
        })
    linked = sum(1 for r in rows if r['chart_id'])
    refresh_block(ws, rows)
    return len(rows), linked


def apply_section_tables(ws, chart_map):
    """Add a 'TradingView' link column (Q) to every hand-maintained stock table
    below row 40. Only real-ticker rows are touched; each link cell mirrors the
    styling of the adjacent 'Last Updated' cell (col P) so it reads as part of the
    table. Reference/note rows (prose in col B) and col P itself are left alone."""
    data_rows = linked = 0
    for r in range(41, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value          # Ticker / 'Ticker' header / prose
        q = ws.cell(row=r, column=SECTION_TV_COL)   # col Q — the new link cell
        src = ws.cell(row=r, column=16)             # col P 'Last Updated' — style source

        if isinstance(b, str) and b.strip() == 'Ticker':
            q.value = 'TradingView'
            q.font = copy(src.font)
            q.fill = copy(src.fill)
            q.border = copy(src.border)
            q.alignment = copy(src.alignment)
        elif looks_like_ticker(b):
            data_rows += 1
            q.fill = copy(src.fill)
            q.border = copy(src.border)
            q.alignment = copy(src.alignment)
            cid = chart_id_for(chart_map, b.strip())
            if cid:
                q.value = link_formula(cid)
                q.font = Font(name='Arial', size=8, color='FF1F3864', underline='single')
                linked += 1
            else:
                q.font = copy(src.font)
    ws.column_dimensions[get_column_letter(SECTION_TV_COL)].width = 12
    return data_rows, linked


def main():
    path = sys.argv[1]
    manifest = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, 'layout_manifest_tmp.json')
    backup = path + '.bak-before-tv-links-2026-07-12'
    shutil.copyfile(path, backup)
    print(f'Backup -> {backup}')

    chart_map = build_chart_map(manifest)
    print(f'Chart map: {len(chart_map)} tickers -> chartId')

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME]

    ba_rows, ba_linked = apply_below_alert(ws, chart_map)
    print(f'Below-alert block: {ba_linked}/{ba_rows} rows linked (col J)')

    sec_rows, sec_linked = apply_section_tables(ws, chart_map)
    print(f'Section tables: {sec_linked}/{sec_rows} data rows linked (col P)')

    wb.save(path)
    print(f'Saved -> {path}')


if __name__ == '__main__':
    main()
