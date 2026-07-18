"""Refresh the master's 'Income Funds' tab Current Value from the latest Fidelity
AccountSummary (user request 2026-07-18: those tables were stale since 3 Jul).

The tab is a PROTECTED, hand-maintained master tab (integrate_spending_tabs never
touches it) and every cell is a literal, so nothing was updating its fund values.
This maps each fund row to its Fidelity Inc holding(s) — a mapping confirmed by the
user (M&G -> the Floating Rate Inc holding; Aegon -> the B-class Inc only) — sums the
current values from AccountSummary, and rewrites Current Value, Difference (Current -
Target), % of Target, and Monthly/Annual Revenue (Current × the row's % Dividend).
The TOTAL (Family) row is re-summed. Every change is logged; a fund with no matching
holding this run is left untouched (never zeroed).

Wired into the pipeline (run each run). Standalone: `python refresh_income_funds.py
[--apply] [workbook.xlsx]` — dry-run by default (prints proposed changes), --apply
writes (backing the workbook up first). Skips cleanly if AccountSummary isn't present.
"""
import csv
import datetime
import io
import os
import shutil
import sys
from collections import defaultdict

import openpyxl

SHEET = "Income Funds"
COL_NAME, COL_TARGET, COL_CUR, COL_DIFF, COL_PCTTGT = 1, 2, 3, 4, 5
COL_PCTDIV, COL_MREV, COL_AREV = 7, 9, 10

# tab-row keyword -> exact AccountSummary Inc holding name(s) to sum (user-confirmed).
FUND_HOLDINGS = {
    "schroder": ["Schroder High Yield Opportunities Fund Z Inc"],
    "m&g":      ["M&G Global Floating Rate High Yield Fund Sterling I-H Inc"],
    "man":      ["Man High Yield Opportunities Fund Prof D Inc"],
    "aegon":    ["Aegon High Yield Bond B Inc"],
}


def account_summary_values(path):
    """holding name -> total current value across accounts (Asset rows)."""
    content = open(path, encoding="utf-8-sig").read()
    lines = content.replace("\r", "").split("\n")
    hi = max(i for i, l in enumerate(lines) if l.startswith("Type,Holdings,Account number"))
    by = defaultdict(float)
    for r in csv.DictReader(io.StringIO("\n".join(lines[hi:]))):
        if r.get("Type", "").strip() != "Asset":
            continue
        h = r.get("Holdings", "").strip().replace('"', '')
        try:
            by[h] += float(r.get("Value (£)", "0").replace(",", ""))
        except (ValueError, TypeError):
            pass
    return by


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def refresh(workbook, account_summary, apply=False):
    by = account_summary_values(account_summary)
    wb = openpyxl.load_workbook(workbook)
    ws = wb[SHEET]

    changes, total_cur = [], 0.0
    total_row = None
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if not isinstance(name, str) or not name.strip():
            continue
        if name.strip().upper().startswith("TOTAL"):
            total_row = r
            break
        key = next((k for k in FUND_HOLDINGS if k in name.lower()), None)
        if not key:
            continue
        new_cur = round(sum(by.get(h, 0.0) for h in FUND_HOLDINGS[key]), 2)
        if new_cur <= 0:
            print(f"  {name[:38]:38} no matching holding this run — left unchanged")
            total_cur += _num(ws.cell(r, COL_CUR).value)
            continue
        old_cur = _num(ws.cell(r, COL_CUR).value)
        target = _num(ws.cell(r, COL_TARGET).value)
        pctdiv = _num(ws.cell(r, COL_PCTDIV).value)
        changes.append((r, name, old_cur, new_cur, target, pctdiv))
        total_cur += new_cur

    for r, name, old_cur, new_cur, target, pctdiv in changes:
        print(f"  {name[:38]:38} {old_cur:>13,.0f} -> {new_cur:>13,.0f}")
        if apply:
            ws.cell(r, COL_CUR).value = new_cur
            ws.cell(r, COL_DIFF).value = round(new_cur - target, 2)
            ws.cell(r, COL_PCTTGT).value = round(new_cur / target, 8) if target else 0
            ws.cell(r, COL_MREV).value = round(new_cur * pctdiv / 12.0, 6)
            ws.cell(r, COL_AREV).value = round(new_cur * pctdiv, 6)

    if apply and total_row is not None:
        tgt_total = _num(ws.cell(total_row, COL_TARGET).value)
        ws.cell(total_row, COL_CUR).value = round(total_cur, 2)
        ws.cell(total_row, COL_DIFF).value = round(total_cur - tgt_total, 2)
        ws.cell(total_row, COL_PCTTGT).value = round(total_cur / tgt_total, 8) if tgt_total else 0
        ws.cell(total_row, COL_MREV).value = round(sum(_num(ws.cell(r, COL_MREV).value)
                                                       for r, *_ in changes), 6)
        ws.cell(total_row, COL_AREV).value = round(sum(_num(ws.cell(r, COL_AREV).value)
                                                       for r, *_ in changes), 6)

    if apply and changes:
        wb.save(workbook)
        print(f"Applied {len(changes)} fund refresh(es); TOTAL current value £{total_cur:,.0f}.")
    elif not apply:
        print(f"(dry run — {len(changes)} fund(s) would change; pass --apply to write)")
    return len(changes)


def main():
    args = [a for a in sys.argv[1:] if a != "--apply"]
    apply = "--apply" in sys.argv
    workbook = args[0] if args else os.path.join(
        os.path.expanduser("~"), "Downloads", "Stocks_Buy_Strategy.xlsx")
    acc = os.path.join(os.path.expanduser("~"), "Downloads", "AccountSummary.csv")
    if not os.path.exists(acc):
        print("AccountSummary.csv not in Downloads — Income Funds refresh skipped.")
        return
    if not os.path.exists(workbook):
        print("Master workbook not found — Income Funds refresh skipped.")
        return
    if apply:
        bak = workbook + ".bak-incfunds-" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copyfile(workbook, bak)
        print("Backed up ->", os.path.basename(bak))
    refresh(workbook, acc, apply=apply)


if __name__ == "__main__":
    main()
