#!/usr/bin/env python3
"""Refresh the 'Stocks Trading Below Alert Low' table at the TOP of the
'Stocks of Interest' sheet in Stocks_Buy_Strategy.xlsx: every ticker whose live
TradingView price (captured during the chart-export step) has dropped below its
Alert Low, sorted worst-gap-first.

History: this table used to live on its own 'Below Alert Low' first-tab sheet
(that was the safe option at the time — openpyxl's insert_rows() had been shown
to corrupt Stocks of Interest's ~121 formulas and 34 merged bands, RELX fix
2026-07-10). On 2026-07-11 the user asked for it to be the top table of Stocks
of Interest instead, so restructure_soi_stats_2026-07-11.py rebuilt that sheet
once with rows 1-40 RESERVED for this table (all pre-existing content, formulas
and merges were shifted to row 41+). This script now rewrites ONLY that reserved
block, never inserting or deleting rows — so the corruption risk that motivated
the separate sheet never applies. Do not write below row 40 here, and do not
add manual content above row 41 in that sheet.

Usage: python add_below_alert_sheet.py <master.xlsx> <below_alert_rows.json>
  rows: [{"ticker","share_name","price","alert_low","alert_high","gap_pct",
          "holdings","target_value","checked_at","chart_id","on_alert","pattern"},
         ...], worst gap first. ("pattern" is the channel_detect label shown in the
  block's Pattern column; absent rows just render blank there.)
"""
import sys
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SHEET_NAME = 'Stocks of Interest'
RESERVED_BLOCK = 40      # rows 1..40 belong to this table; row 41+ is the
                         # pre-existing Stocks of Interest content — never touch it
TITLE_ROW = 1
NOTE_ROW = 2
BAND_ROW = 3             # coloured section band, like the section headers below
HEADER_ROW = 4
DATA_START_ROW = 5
MAX_DATA_ROWS = 34       # rows 5..38; 39-40 stay blank as a separator
LAST_COL = 17            # this table occupies columns A..Q, matching the section
                         # tables below it (2026-07-16 — user asked for column parity
                         # with 'Near Lower Boundary': Pattern, Proximity, Upside %,
                         # P/E, Div Yield, Chart Note, Analyst Rating, Notes, ...)

# Per-row click-through to the stock's TradingView layout, matching the
# Investments sheet's "TradingView" column (=HYPERLINK(chart/<chartId>/)).
from config import CFG as _CFG
from ticker_normalize import normalize
TV_LAYOUT_URL = _CFG['tvLayoutUrlTemplate']
TV_LINK_LABEL = '📊 Layout'

# Palette + typography lifted from the section tables lower down this same sheet
# (rows 41+), so the below-alert table reads as part of the same document rather
# than a foreign block. Verified against the live workbook 2026-07-12.
FONT_NAME = 'Arial'
WHITE = 'FFFFFFFF'
NAVY = 'FF1F3864'          # title band + ticker text (identity colour)
SUBHEAD_FILL = 'FF2E5077'  # subtitle/note band + column-header band
BAND_FILL = 'FFC00000'     # section band — red extends the existing green/amber/
                           # blue priority ladder: "below alert low" is the most
                           # urgent state (change here if a different colour is wanted)
DATA_FILL = 'FFFDF0F0'     # pale-red data tint, matching the per-section pale tints
                           # (green FFF0FFF4 / amber FFFFFDF0 / watchlist FFF7F9FC)
GAP_BAD = 'FFCC0000'       # red-bold gap when a stock is >10% below its alert low

# "On Alert" section (user decision 2026-07-15): price is sitting ON a drawn line
# rather than having fallen through it — the moment the alert exists to catch, so it
# sits ABOVE the below-alert section. Green matches the existing "🟢 AT LOWER
# BOUNDARY — within 5% of alert low (Highest priority)" band in the section tables
# below, which is the same idea; red stays reserved for "already fallen through".
ON_ALERT_BAND_FILL = 'FF1A5733'   # the SAME green as the existing "AT LOWER BOUNDARY" band
                                  # (verified against row 43 of the live sheet) — not the
                                  # FF276221 green, which is a font colour, not a band fill
ON_ALERT_DATA_FILL = 'FFF0FFF4'   # pale-green data tint, as used by the green section below

_THIN = Side(style='thin')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=False)
_TOP_LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
_TOP_RIGHT = Alignment(horizontal='right', vertical='top', wrap_text=False)

# Column schema is IDENTICAL to the section tables below (refresh_soi_sections.py's
# HEADER), so the whole 'Stocks of Interest' sheet reads as one document (user
# request 2026-07-16 — the block used to carry only 10 columns and was "missing
# chart note, div yield etc"). Pattern/Proximity/Upside/P/E/Div Yield are derived
# per run; Chart Note / Analyst Rating / Notes (cols K/L/O) are hand-curated only
# in the section tables and are left BLANK here (user decision 2026-07-16) — the
# block is a full auto-generated list, so inventing per-stock notes for it isn't
# wanted.
HEADER = ['Stock', 'Ticker', 'Pattern', 'Proximity', 'Alert Low', 'Current', 'Alert High',
          'Upside %', 'P/E', 'Div Yield', 'Chart Note', 'Analyst Rating', 'Holdings (£)',
          'Target Value (£)', 'Notes', 'Last Updated', 'TradingView']
# Price columns keep thousands separators; £ holdings/target show whole pounds.
_NUM_FMT = {5: '#,##0.00', 6: '#,##0.00', 7: '#,##0.00', 13: '#,##0', 14: '#,##0'}


def _band(ws, row, text, fill, size, bold, italic=False):
    """Render a full-width (A..J) coloured band on `row`, merged, bordered, with
    `text` in the top-left cell — the same treatment the section headers use."""
    # Unmerge any existing band on this row FIRST (a previous run may have used a
    # different width), so the cells below are writable rather than read-only
    # MergedCells — then re-merge to the current width. Keeps this idempotent.
    for m in list(ws.merged_cells.ranges):
        if m.min_row == row and m.max_row == row:
            ws.unmerge_cells(str(m))
    for c in range(1, LAST_COL + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = None
        cell.fill = PatternFill(fill_type='solid', fgColor=fill)
        cell.border = BORDER
        cell.alignment = _TOP_LEFT
        cell.font = Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=WHITE)
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(f'A{row}:{chr(64 + LAST_COL)}{row}')
    ws.row_dimensions[row].height = 15


def _header_row(ws, row_n):
    """Column header band (FF2E5077, bold white, wrapped) — matches row 44 below."""
    for c, h in enumerate(HEADER, 1):
        cell = ws.cell(row=row_n, column=c, value=h)
        cell.font = Font(name=FONT_NAME, size=8, bold=True, color=WHITE)
        cell.fill = PatternFill(fill_type='solid', fgColor=SUBHEAD_FILL)
        cell.border = BORDER
        cell.alignment = _TOP_LEFT_WRAP
    ws.row_dimensions[row_n].height = 15


def _section(ws, start_row, band_text, band_fill, data_fill, rows):
    """Render one section — coloured band, column headers, data rows — starting at
    start_row. Returns the next free row. Two sections share this block now (On
    Alert above Below Alert Low), so the layout is built sequentially rather than at
    the fixed rows the single-section version used."""
    _band(ws, start_row, band_text, band_fill, size=10, bold=True)
    _header_row(ws, start_row + 1)
    row_n = start_row + 2
    for r in rows:
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=row_n, column=c)
            cell.fill = PatternFill(fill_type='solid', fgColor=data_fill)
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=8)
            cell.alignment = _TOP_LEFT
        ws.row_dimensions[row_n].height = 15

        name_cell = ws.cell(row=row_n, column=1, value=r['share_name'])
        name_cell.font = Font(name=FONT_NAME, size=9, bold=True, color='FF000000')
        ticker_cell = ws.cell(row=row_n, column=2, value=r['ticker'])
        ticker_cell.font = Font(name=FONT_NAME, size=8, bold=True, color=NAVY)

        # C Pattern (channel_detect label, carried on the row); D Proximity is a
        # row-relative formula off Current (F) and Alert Low (E), exactly as the
        # section tables build it. A stock >10% below its alert low still gets the
        # red-bold emphasis the Gap % column used to carry — now on Proximity.
        ws.cell(row=row_n, column=3, value=r.get('pattern') or '')
        prox_cell = ws.cell(row=row_n, column=4,
                            value=f'=IFERROR(TEXT((F{row_n}-E{row_n})/E{row_n},"0.0%")&" above low","")')
        if r['gap_pct'] <= -10:
            prox_cell.font = Font(name=FONT_NAME, size=8, bold=True, color=GAP_BAD)
        ws.cell(row=row_n, column=5, value=r['alert_low'])
        ws.cell(row=row_n, column=6, value=round(r['price'], 2))   # Current — a literal
                                                                   # value (not a VLOOKUP)
                                                                   # so verify_pipeline can
                                                                   # count it as numeric
        ws.cell(row=row_n, column=7, value=r['alert_high'])
        _data_row_tail(ws, row_n, r)
        row_n += 1
    return row_n


def refresh_block(ws, rows):
    # Split into the two sections. "On Alert" (price sitting ON a drawn line) is the
    # live trigger and sorts above the stocks that have already fallen through.
    on_alert_rows = [r for r in rows if r.get('on_alert')]
    below_rows = [r for r in rows if not r.get('on_alert')]

    # Two bands + two header rows cost 4 rows of the reserved block; only the
    # sections actually rendered are charged for.
    overhead = 2 + (2 if on_alert_rows else 0)
    capacity = RESERVED_BLOCK - BAND_ROW + 1 - overhead - 1   # -1 keeps a blank separator
    if len(rows) > capacity:
        print(f'WARNING: {len(rows)} alert rows but only {capacity} fit in the '
              f'reserved block — writing the worst; {len(rows) - capacity} omitted.',
              file=sys.stderr)
        # Never drop an On Alert row to make room for a below-alert one.
        keep = max(0, capacity - len(on_alert_rows))
        below_rows = below_rows[:keep]
        on_alert_rows = on_alert_rows[:capacity]

    # Title / note bands (rows 1-2), mirroring the section-table header stack below
    # (navy title, FF2E5077 italic subtitle, then the coloured section bands).
    _band(ws, TITLE_ROW, 'Stocks On Alert / Trading Below Alert Low', NAVY, size=13, bold=True)
    _band(ws, NOTE_ROW, (
        f'Auto-generated {datetime.now().date().isoformat()} from tradingview_layouts.xlsx live '
        'price capture cross-checked against Alert Low. Rebuilt by the pipeline into rows '
        f'1-{RESERVED_BLOCK} of this sheet each run — do not add manual content above row '
        f'{RESERVED_BLOCK + 1}.'
    ), SUBHEAD_FILL, size=8, bold=False, italic=True)

    # Reset everything below the title/note so a shorter run — or a run with no On
    # Alert section — leaves no styled ghosts or stale band behind. Unmerge first:
    # the section bands are merged A..J, and a previous run's band may sit on any
    # row now being reset (the sections no longer live at fixed rows), whose cells
    # would otherwise be read-only MergedCells.
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= BAND_ROW and m.max_row <= RESERVED_BLOCK:
            ws.unmerge_cells(str(m))
    for r in range(BAND_ROW, RESERVED_BLOCK + 1):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'
        ws.row_dimensions[r].height = None

    cur = BAND_ROW
    if on_alert_rows:
        cur = _section(ws, cur,
                       '🟢  ON ALERT — price has reached a drawn support line  (Highest priority)',
                       ON_ALERT_BAND_FILL, ON_ALERT_DATA_FILL, on_alert_rows)
    cur = _section(ws, cur,
                   '🔴  BELOW ALERT LOW — price has fallen through the alert level  (Action required)',
                   BAND_FILL, DATA_FILL, below_rows)
    return _finish(ws, cur, len(on_alert_rows) + len(below_rows))


def _data_row_tail(ws, row_n, r):
    """The derived + hand-curated + link columns (H..Q), matching the section tables.

    H Upside % and J Div Yield are row-relative formulas; I P/E is a GOOGLEFINANCE
    formula off the stock's Google ticker; K/L/O (Chart Note / Analyst Rating /
    Notes) are hand-curated in the section tables only and stay blank here; M/N are
    the pipeline's Holdings/Target values; P is the rebuild date; Q the TV link."""
    ws.cell(row=row_n, column=8, value=f'=IFERROR((G{row_n}-E{row_n})/E{row_n},"")')

    gt = (normalize(r['ticker']) or {}).get('google_finance_ticker')
    if gt:
        ws.cell(row=row_n, column=9, value=f'=IFERROR(googlefinance("{gt}","pe"),"")')
    ws.cell(row=row_n, column=10,
            value=f"=IFERROR(VLOOKUP(B{row_n},'Base Data'!$A:$Q,9,FALSE()),\"\")")

    # K (11) Chart Note, L (12) Analyst Rating, O (15) Notes — left blank by design.
    # Holdings/Target £ display as whole pounds (user rule, 2026-07-11).
    ws.cell(row=row_n, column=13, value=r['holdings'])
    ws.cell(row=row_n, column=14, value=r['target_value'])
    ws.cell(row=row_n, column=16, value=datetime.now().date())
    ws.cell(row=row_n, column=16).number_format = 'yyyy-mm-dd'

    # Q (17) click-through to this stock's TradingView layout (blank if no chart).
    chart_id = r.get('chart_id')
    tv_cell = ws.cell(row=row_n, column=17)
    if chart_id:
        url = TV_LAYOUT_URL.format(chart_id=chart_id)
        tv_cell.value = f'=HYPERLINK("{url}","{TV_LINK_LABEL}")'
        tv_cell.font = Font(name=FONT_NAME, size=8, color=NAVY, underline='single')
    tv_cell.alignment = _TOP_LEFT

    for c, fmt in _NUM_FMT.items():
        ws.cell(row=row_n, column=c).number_format = fmt


def _finish(ws, first_blank, count):
    """Collapse the unused tail of the reserved block so the section tables below
    sit just under this table instead of after a wide blank gap. The rows stay
    present — the fixed reserved block is what lets the pipeline rewrite rows
    1..RESERVED_BLOCK each run without disturbing the section tables — they're
    only HIDDEN; a run with more alert stocks unhides them as it fills. One blank
    separator row (the first unused row) is kept visible."""
    for r in range(1, RESERVED_BLOCK + 1):
        ws.row_dimensions[r].hidden = r > first_blank
    return count


def main():
    master_path, rows_path = sys.argv[1], sys.argv[2]
    with open(rows_path, 'r', encoding='utf-8') as f:
        rows = json.load(f)

    wb = openpyxl.load_workbook(master_path, data_only=False)
    ws = wb[SHEET_NAME]
    count = refresh_block(ws, rows)
    wb.save(master_path)
    print(f'Wrote {count} below-alert rows into the reserved top block of "{SHEET_NAME}" in {master_path}')


if __name__ == '__main__':
    main()
