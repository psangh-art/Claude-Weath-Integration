#!/usr/bin/env python3
"""Mirror every sheet of spending_summary.xlsx into Stocks_Buy_Strategy.xlsx
as same-named tabs (user request 2026-07-13), so the master workbook — and the
Finance Google Sheet it gets imported into — carries the wealth/spending data
alongside the investment tabs.

Rules:
- spending_summary.xlsx is the source of truth for these tabs: existing
  same-named tabs in the master are replaced in place (same tab position);
  new ones are appended after the master's own tabs.
- Every other master tab (Investments, Stocks of Interest, ...) is never
  touched. A source sheet name that ever collided with one of the master's
  own tabs would clobber it, so the master's core tabs are explicitly
  protected below.
- A missing spending_summary.xlsx is a skip, not a failure — the spending
  build is an optional pipeline stage (see CLAUDE.md).

Usage: python integrate_spending_tabs.py [spending.xlsx] [master.xlsx]
       (defaults: ~/Downloads/spending_summary.xlsx, ~/Downloads/Stocks_Buy_Strategy.xlsx)
"""
import os
import sys

from openpyxl import load_workbook

from xlsx_sheet_copy import replace_sheet

# The master's own tabs — never replaceable by a spending tab, even if a
# future spending_summary.py change emits a colliding sheet name.
PROTECTED_MASTER_TABS = {
    'Investments', 'Income Funds', 'History', 'Stats',
    'Stocks of Interest', 'Base Data',
}


def integrate(spending_path, master_path):
    if not os.path.exists(spending_path):
        print(f"spending_summary.xlsx not found at {spending_path} — skipping "
              f"spending-tab integration (the spending build is optional).")
        return 0
    if not os.path.exists(master_path):
        print(f"Error: master workbook not found at {master_path}", file=sys.stderr)
        return 1

    src_wb = load_workbook(spending_path)
    dst_wb = load_workbook(master_path)

    outcomes = []
    for name in src_wb.sheetnames:
        if name in PROTECTED_MASTER_TABS:
            print(f"  REFUSING to overwrite protected master tab '{name}' — "
                  f"rename the sheet in spending_summary.py instead.", file=sys.stderr)
            continue
        outcomes.append((name, replace_sheet(dst_wb, name, src_wb[name])))

    if not outcomes:
        print("No spending tabs were integrated.")
        return 1

    dst_wb.save(master_path)
    for name, outcome in outcomes:
        print(f"  {outcome}: '{name}'")
    print(f"Integrated {len(outcomes)} spending tab(s) into {master_path}")
    return 0


def main():
    downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
    spending_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(downloads, 'spending_summary.xlsx')
    master_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(downloads, 'Stocks_Buy_Strategy.xlsx')
    sys.exit(integrate(spending_path, master_path))


if __name__ == '__main__':
    main()
