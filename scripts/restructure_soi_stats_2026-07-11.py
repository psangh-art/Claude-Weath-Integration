#!/usr/bin/env python3
"""One-off restructure of Stocks_Buy_Strategy.xlsx (2026-07-11, user-requested):

1. Move the three History-sheet bar charts to a new 'Stats' sheet (charts only —
   their data references still point at History ranges, which stay put).
2. Migrate the 'Below Alert Low' table to the TOP of 'Stocks of Interest'
   (rows 1-40 reserved block, same formatting it had on its own sheet), then
   delete the 'Below Alert Low' sheet.

The Stocks of Interest move deliberately does NOT use ws.insert_rows(): that is
documented (add_below_alert_sheet.py docstring / CLAUDE.md, RELX fix 2026-07-10)
to silently corrupt this sheet's ~121 formula cells and 34 merged section bands.
Instead the whole sheet is rebuilt into a fresh worksheet: below-alert block
written at the top, every old cell copied 40 rows down with styles, merges and
row heights re-anchored, and each formula's unquoted A1 references row-offset by
+40 (all existing formulas are same-row-relative; verified before writing this —
the only cross-sheet refs are column-only ranges like 'Investments'!$C:$I, which
have no row numbers to break). Verified safe additionally because NO other sheet
references 'Stocks of Interest' by name (scanned all formula cells first).

Usage: python restructure_soi_stats_2026-07-11.py <in.xlsx> <out.xlsx>
(run against a copy first; only point at the live file once the copy verifies)
"""
import sys
import re
from copy import copy

import openpyxl

# Rows 1..RESERVED_BLOCK are owned by the below-alert table (title, note, blank,
# header, up to MAX_DATA_ROWS data rows, trailing blank). add_below_alert_sheet.py
# rewrites ONLY this block on each pipeline run.
RESERVED_BLOCK = 40
DATA_START_ROW = 5
MAX_DATA_ROWS = 34  # rows 5..38 inclusive; 39-40 left blank as separator

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass


def offset_formula(formula, offset):
    """Add `offset` to the row number of every unquoted, unqualified A1-style
    reference. Text inside double-quoted string literals is left untouched, as
    are absolute-column-only ranges ($C:$I) and bare numeric arguments."""
    parts = formula.split('"')
    for i in range(0, len(parts), 2):  # even indexes are outside quotes
        parts[i] = re.sub(
            r"(?<![A-Za-z0-9:$!])([A-Z]{1,3})(\d{1,5})",
            lambda m: m.group(1) + str(int(m.group(2)) + offset),
            parts[i],
        )
    return '"'.join(parts)


def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def main():
    src_path, out_path = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(src_path)

    # ── 1. Charts: History → new Stats sheet ─────────────────────────────────
    hist = wb["History"]
    stats = wb.create_sheet("Stats", wb.sheetnames.index("History") + 1)
    moved = 0
    for chart in list(hist._charts):
        stats.add_chart(chart)  # anchor object travels with the chart
        moved += 1
    hist._charts = []
    print(f"Moved {moved} chart(s) from History to Stats")

    # ── 2. Rebuild Stocks of Interest with below-alert block on top ─────────
    soi = wb["Stocks of Interest"]
    bal = wb["Below Alert Low"]
    soi_index = wb.sheetnames.index("Stocks of Interest")

    new = wb.create_sheet("SoI_TMP", soi_index)

    # 2a. Below-alert table into rows 1..RESERVED_BLOCK (straight copy of the
    # old sheet's cells, same formatting).
    bal_rows = min(bal.max_row, RESERVED_BLOCK - 2)
    for r in range(1, bal_rows + 1):
        for c in range(1, bal.max_column + 1):
            copy_cell(bal.cell(row=r, column=c), new.cell(row=r, column=c))
    new.cell(row=2, column=1).value = (
        "Auto-generated from tradingview_layouts.xlsx live price capture cross-checked "
        f"against Alert Low. Rebuilt by the pipeline into rows 1-{RESERVED_BLOCK} of this "
        "sheet each run — do not add manual content above row "
        f"{RESERVED_BLOCK + 1}."
    )

    # 2b. Old Stocks of Interest content shifted down by RESERVED_BLOCK.
    formulas_offset = 0
    for row in soi.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            dst = new.cell(row=cell.row + RESERVED_BLOCK, column=cell.column)
            copy_cell(cell, dst)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                dst.value = offset_formula(cell.value, RESERVED_BLOCK)
                formulas_offset += 1
    print(f"Copied Stocks of Interest content +{RESERVED_BLOCK} rows "
          f"({formulas_offset} formulas row-offset)")

    # merges, column widths, row heights
    for m in soi.merged_cells.ranges:
        new.merge_cells(start_row=m.min_row + RESERVED_BLOCK, end_row=m.max_row + RESERVED_BLOCK,
                        start_column=m.min_col, end_column=m.max_col)
    for col, dim in soi.column_dimensions.items():
        new.column_dimensions[col].width = dim.width
    for r, dim in soi.row_dimensions.items():
        new.row_dimensions[r + RESERVED_BLOCK].height = dim.height
    print(f"Re-anchored {len(soi.merged_cells.ranges)} merged ranges")

    # 2c. Swap sheets: drop originals, rename TMP into place.
    del wb["Stocks of Interest"]
    del wb["Below Alert Low"]
    new.title = "Stocks of Interest"

    wb.save(out_path)
    print(f"Saved -> {out_path}")
    print("Sheets now:", wb.sheetnames)


if __name__ == "__main__":
    main()
