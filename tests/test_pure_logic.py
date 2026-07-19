"""Offline unit tests for the pipeline's pure-logic modules — no TradingView,
no Downloads files, no network. Run with:

    C:\\Users\\Paul\\AppData\\Local\\Python\\bin\\python.exe -m pytest tests/ -q

Covers the four modules whose logic is pure enough to test without the live
environment: ticker_normalize, xlsx_sheet_copy's offset_formula, the below-alert
row builder, and the Fidelity export classifier.
"""
import json
import os
import sys

import pytest

SCRIPTS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'scripts')
sys.path.insert(0, SCRIPTS)

from ticker_normalize import normalize, master_tickers_match  # noqa: E402


# ── ticker_normalize ────────────────────────────────────────────────────────

class TestNormalize:
    def test_commodity_maps_to_master_and_gets_no_formula(self):
        n = normalize('GOLD')
        assert n['kind'] == 'commodity'
        assert n['master_ticker'] == 'GOLD'
        # GOOGLEFINANCE cannot price commodities (verified 2026-07-11)
        assert n['google_finance_ticker'] is None
        assert n['google_finance_formula'] is None

    def test_commodity_tv_name_translates(self):
        assert normalize('SILVER')['master_ticker'] == 'SLVR'
        assert normalize('USOIL')['master_ticker'] == 'OIL'
        assert normalize('NASDAQ')['master_ticker'] == 'NDX'

    def test_brent_maps_to_ukoil_master_row(self):
        # The TV chart is 'Crude Oil Brent Cash' (BRENT); the master row is UKOIL.
        n = normalize('BRENT')
        assert n['kind'] == 'commodity'
        assert n['master_ticker'] == 'UKOIL'
        assert n['google_finance_formula'] is None

    def test_futures_suffix_stripped(self):
        # TradingView continuous futures carry a '1!' suffix (real case: COPPER1!).
        n = normalize('COPPER1!')
        assert n['kind'] == 'commodity'
        assert n['master_ticker'] == 'COPP'

    def test_macro_symbols_never_match_equity_rows(self):
        for t in ('GBPUSD', 'US10Y', 'JP10Y'):
            n = normalize(t)
            assert n['kind'] == 'macro_excluded'
            assert n['master_ticker'] is None

    def test_class_suffix_becomes_dash(self):
        n = normalize('BT.A')
        assert n['kind'] == 'equity'
        assert n['master_ticker'] == 'BT-A'
        assert n['google_finance_ticker'] == 'LON:BT.A'

    def test_trailing_dot_stripped(self):
        n = normalize('SGRO.')
        assert n['master_ticker'] == 'SGRO'
        assert n['google_finance_ticker'] == 'LON:SGRO'

    def test_plain_equity(self):
        n = normalize('ADM')
        assert n['kind'] == 'equity'
        assert n['master_ticker'] == 'ADM'
        assert n['google_finance_formula'] == '=googlefinance("LON:ADM","price")'

    def test_none_and_empty(self):
        assert normalize(None) is None
        assert normalize('') is None


class TestMasterTickersMatch:
    def test_dash_dot_equivalence(self):
        assert master_tickers_match('BT-A', 'BT.A')
        assert master_tickers_match('bt.a', 'BT-A')

    def test_case_and_whitespace(self):
        assert master_tickers_match(' adm ', 'ADM')

    def test_mismatch_and_none(self):
        assert not master_tickers_match('ADM', 'AML')
        assert not master_tickers_match(None, 'ADM')
        assert not master_tickers_match('ADM', '')


# ── offset_formula (row-relative formula re-anchoring) ─────────────────────

from xlsx_sheet_copy import offset_formula  # noqa: E402


class TestOffsetFormula:
    def test_simple_relative_refs_shift(self):
        assert offset_formula('=A1+B2', 40) == '=A41+B42'

    def test_quoted_strings_untouched(self):
        f = '=IFERROR(TEXT((F5-E5)/E5,"0.0%")&" above low","")'
        out = offset_formula(f, 40)
        assert '"0.0%"' in out and '" above low"' in out
        assert 'F45' in out and 'E45' in out

    def test_column_only_absolute_ranges_untouched(self):
        f = "=VLOOKUP(B5,'Investments'!$C:$I,7,FALSE())"
        out = offset_formula(f, 40)
        assert '$C:$I' in out
        assert 'B45' in out

    def test_bare_numbers_untouched(self):
        assert offset_formula('=ROUND(2383*I120/100,2)', 40) == '=ROUND(2383*I160/100,2)'


# ── build_below_alert_rows (gap calc + formula-holdings guard) ──────────────

from openpyxl import Workbook  # noqa: E402
import update_master_sheet as ums  # noqa: E402


def make_master_row(ws, row, ticker, share, alert_low, alert_high, holdings, target):
    ws.cell(row=row, column=ums.COL_TICKER, value=ticker)
    ws.cell(row=row, column=ums.COL_SHARE_NAME, value=share)
    ws.cell(row=row, column=ums.COL_ALERT_LOW, value=alert_low)
    ws.cell(row=row, column=ums.COL_ALERT_HIGH, value=alert_high)
    ws.cell(row=row, column=ums.COL_HOLDINGS, value=holdings)
    ws.cell(row=row, column=ums.COL_TARGET_VALUE, value=target)


class TestBuildBelowAlertRows:
    def _ws(self):
        return Workbook().active

    def test_below_alert_included_with_gap(self):
        ws = self._ws()
        make_master_row(ws, 5, 'WPP', 'WPP plc', 1000.0, 1100.0, 5000.0, 6000.0)
        matches = [{'ticker': 'WPP', 'company': 'WPP plc', 'row': 5,
                    'price': 900.0, 'checked_at': 'T', 'chart_id': 'abc'}]
        rows = ums.build_below_alert_rows(ws, matches)
        assert len(rows) == 1
        r = rows[0]
        assert r['gap_pct'] == pytest.approx(-10.0)
        assert r['holdings'] == 5000.0
        assert r['chart_id'] == 'abc'

    def test_above_alert_excluded(self):
        ws = self._ws()
        make_master_row(ws, 5, 'ADM', 'Admiral', 1000.0, None, None, None)
        matches = [{'ticker': 'ADM', 'company': 'Admiral', 'row': 5,
                    'price': 1500.0, 'checked_at': 'T', 'chart_id': None}]
        assert ums.build_below_alert_rows(ws, matches) == []

    def test_formula_holdings_become_none(self):
        # Holdings/Target read as formula STRINGS with data_only=False must not
        # leak into the below-alert block (real bug, fixed 2026-07-12).
        ws = self._ws()
        make_master_row(ws, 5, 'SGE', 'Sage', 900.0, None,
                        '=ROUND(2383*I120/100,2)', '=IFERROR(D123,"")')
        matches = [{'ticker': 'SGE', 'company': 'Sage', 'row': 5,
                    'price': 800.0, 'checked_at': 'T', 'chart_id': None}]
        rows = ums.build_below_alert_rows(ws, matches)
        assert rows[0]['holdings'] is None
        assert rows[0]['target_value'] is None

    def test_non_numeric_price_skipped(self):
        ws = self._ws()
        make_master_row(ws, 5, 'X', 'X', 100.0, None, None, None)
        matches = [{'ticker': 'X', 'company': 'X', 'row': 5,
                    'price': None, 'checked_at': 'T', 'chart_id': None}]
        assert ums.build_below_alert_rows(ws, matches) == []

    def test_sorted_worst_gap_first(self):
        ws = self._ws()
        make_master_row(ws, 5, 'A', 'A', 100.0, None, None, None)
        make_master_row(ws, 6, 'B', 'B', 100.0, None, None, None)
        matches = [
            {'ticker': 'A', 'company': 'A', 'row': 5, 'price': 95.0, 'checked_at': 'T', 'chart_id': None},
            {'ticker': 'B', 'company': 'B', 'row': 6, 'price': 50.0, 'checked_at': 'T', 'chart_id': None},
        ]
        rows = ums.build_below_alert_rows(ws, matches)
        assert [r['ticker'] for r in rows] == ['B', 'A']


# ── preflight file ageing (optional inputs, 6-week staleness) ───────────────

from datetime import datetime, timedelta  # noqa: E402

import preflight_check as pf  # noqa: E402


class TestFileEntry:
    def test_present_file_uses_mtime(self, tmp_path):
        p = tmp_path / 'activity.csv'
        p.write_text('x', encoding='utf-8')
        e = pf.file_entry('amex', 'Amex', str(p), {})
        assert e['present'] and e['age_days'] == 0 and not e['stale']

    def test_absent_file_falls_back_to_ingestion_state(self):
        as_of = (datetime.now() - timedelta(days=10)).isoformat(timespec='seconds')
        e = pf.file_entry('amex', 'Amex', None, {'amex': {'as_of': as_of}})
        assert not e['present']
        assert e['age_days'] == 10
        assert not e['stale']

    def test_older_than_six_weeks_is_stale(self):
        as_of = (datetime.now() - timedelta(days=pf.STALE_DAYS + 1)).isoformat(timespec='seconds')
        e = pf.file_entry('amex', 'Amex', None, {'amex': {'as_of': as_of}})
        assert e['stale']

    def test_never_seen_is_stale(self):
        e = pf.file_entry('amex', 'Amex', None, {})
        assert e['as_of'] is None and e['age_days'] is None and e['stale']


# ── fidelity_file_classifier ────────────────────────────────────────────────

from fidelity_file_classifier import classify  # noqa: E402


def write_csv(tmp_path, name, text):
    p = tmp_path / name
    p.write_text(text, encoding='utf-8')
    return str(p)


class TestFidelityClassifier:
    def test_historic_export(self, tmp_path):
        p = write_csv(tmp_path, 'TransactionHistory.csv',
                      'Run Date,Action,Symbol,Completion Date\n'
                      '01/07/2026,BUY,ADM,05/07/2026\n'
                      '02/07/2026,SELL,BP,06/07/2026\n')
        assert classify(p) == 'historic'

    def test_no_recognisable_header_is_none(self, tmp_path):
        p = write_csv(tmp_path, 'transactions.csv', 'just,some,noise\n1,2,3\n')
        assert classify(p) is None
